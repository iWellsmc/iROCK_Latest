

from time import strftime, strptime
from webbrowser import get
from django.forms import forms
from django.http.response import JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse_lazy,reverse
from django.forms import formset_factory
from chat.models import Notify_chat
from finance.models import *
from custom_auth.models import *
from invoice.models import Invoice
from django.views.generic import ListView,UpdateView
from InvoiceGuard.models import SignatoriesSettings,SignatoriesUsers
from projects.models import *
from InvoiceGuard.models import *
from projects.forms import *
from django.core.files.storage import FileSystemStorage
from projects.helpers import *
from django.http import HttpResponse
from notifications.signals import notify
from django.forms.models import model_to_dict
from custom_auth.views import markas_read_status
from projects.helpers import create_user_log
import io
from PIL import Image
from .utils.utils import generateImage
import pandas as pd
from django.views.generic import View
import openpyxl
from django.core.serializers import serialize
import xlrd
from .templatetags.custom_tags import get_contract_file
import xlsxwriter
from tablib import Dataset
from cost_code.models import *
from django.core import serializers
from django import forms
from django.db.models import Count
from django.core import mail
import re
from django.utils import timezone
import pytz
from pytz import timezone
import datetime
from datetime import datetime
from datetime import datetime, timedelta
import json
import ast
from io import BytesIO
from reportlab.lib.pagesizes import A4

from easy_pdf.rendering import render_to_pdf_response
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import make_password
from django.db.models import Q
from django.conf import settings
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from ast import literal_eval
from django.contrib import messages
from django import forms
from django.db.models import Count

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from django.core.files.storage import FileSystemStorage

from django.template.loader import render_to_string

from projects.mailer import vendorprimaryapprovermail,vendorsecondaryapprovermail,vendorloginmail,vendorsemail

from django.core.files import File

import os
from django.conf import settings
from custom_auth.forms import *
from wcc.models import *


# Local application/library-specific imports
from custom_auth.helpers import convertdate
from projects.models import VendorContractPriceTable
from invoice.helpers import getinvoiceDate
from invoice.views import notify_invoice_flow
from projects.helpers import getvendorcontracts
from projectflow.models import ProjectFlowlevel







def set_timezone(request):
    if request.method == 'POST':
        request.session['django_timezone'] = request.POST['timezone']
        return redirect('dashboard')
    else:
        return render(request, 'template.html', {'timezones': pytz.common_timezones})


def index(request):
    company=request.company
    country_count=Projects.objects.filter(company=company).values('country').annotate(dcount=Count('country'))
    countries=[]
    for country in country_count:
        country_id=country['country']
        get_countries=Countries.objects.get(id=country_id) 
        countries.append({"country_id":get_countries.id,"country_name":get_countries.nicename})
    get_subtype=ProjectDisciplineSubtype.objects.all()
    get_company=Companies.objects.get(id=request.user.company_id)
    all_projects=Projects.objects.all()
    get_settings = Settings.objects.get_company(request.company).values_list('currency',flat=True).first()
    currency = Basecountries.objects.get_by_id(literal_eval(get_settings))

    datas = [{'id':'1','settings_type':'Default Signatories'},{'id':'2','settings_type':'Project Signatories'}]
    users = User.objects.filter(company=request.company,roles_id__in=[3,2],status=1).exclude(id=request.user.id)

    form=ProjectCreationForm()
    # company_signatory = SignatoriesSettings.objects.get_by_company(request.company,True,'1').count()
    # with_invoice = SignatoriesSettings.objects.get_by_company_type(request.company,True,'1','1')
    # without_invoice = SignatoriesSettings.objects.get_by_company_type(request.company,True,'2','1')
    project_creation_id=''
    data={}
    if request.method=="POST":
        print("form details",request.POST)
        form=ProjectCreationForm(request.POST)
        if form.is_valid():
            projectcreationform=form.save(commit=False)
            projectcreationform.country_id=request.POST['master_country_list']
            projectcreationform.company=request.company
            projectcreationform.projectname_id=request.POST['project_list']
 
            projectcreationform.project_type=request.POST.get('projecttype'," ")
            projectcreationform.entity=request.POST.get('entity'," ")
            # projectcreationform.signatory_type = request.POST.get('sign_settings','1')
            projectcreationform.save()
            project_form_id= projectcreationform
            project_creation_id=projectcreationform

        get_block=request.POST.getlist('block')
        usercreate=request.user.roles_id
        create_user_log(request,project_creation_id,'Project','Create','Project Created',usercreate)
        # print(request.user.id,'hoooo.....')s
        """
        Below code assign users to project & give Project View Rights to users

        """
       
        project_users = request.POST.getlist('project_manager', [])
        if (len(project_users) > 0):
            
            if request.user.roles_id != 3:
                project_users.append(request.user.id)

            [ProjectUser.objects.create(company=request.company,user_id=i,project=project_form_id) for i in project_users]
            
        # new_tr = request.POST.getlist('newtr',False)
        # print(new_tr,'new tr')
        # if new_tr:
        #     for i in range(len(new_tr)):
        #         signatory_create = SignatoriesSettings.objects.create(
        #             currency=None if request.POST.get(f'new_currency{i+1}') == '' else Basecountries.objects.get(id=request.POST.get(f'new_currency{i+1}')),
        #             min_amount=None if request.POST.get(f'new_min_amount{i+1}') == '' else request.POST.get(f'new_min_amount{i+1}').replace(',',''),
        #             max_amount=None if request.POST.get(f'new_max_amount{i+1}') == '' else request.POST.get(f'new_max_amount{i+1}').replace(',',''),
        #             company = request.company,
        #             invoice_type = None if request.POST.get(f'new_invoice_type{i+1}') == '' else request.POST.get(f'new_invoice_type{i+1}'),
        #             signatory_type='2',
        #             project=project_form_id
        #         )
        #         new_user = request.POST.getlist(f'new_newuser{i+1}',False)
        #         if new_user:
        #             for new in new_user:
        #                 SignatoriesUsers.objects.create(signatory=signatory_create,user=User.objects.get_by_id(new.replace(',',''))) if new!='' else False


        UserRights.objects.filter(user_id__in=project_users,module__module_name='Projects').update(view=1)
        for block in get_block:
            print(block)
            create_project_block=ProjectBlock.objects.create(block_id=block,project=project_creation_id)
            get_field=request.POST.getlist('field'+str(block))
            for field in get_field:
                print(field)
                create_project_field=ProjectField.objects.create(block_id=create_project_block.id,field_id=field,project=project_creation_id)
                get_fieldenvironment=request.POST.getlist('fieldenvironment-'+str(field))
                for fieldenvironment in get_fieldenvironment:
                    print(fieldenvironment)
                    create_fieldenvironment=ProjectEnvironment.objects.create(block_id=create_project_block.id,project=project_creation_id,field=create_project_field,field_environment_id=fieldenvironment)
                    get_cluster=request.POST.getlist('cluster-'+str(fieldenvironment))
                    for cluster in get_cluster:
                        print(cluster)
                        create_project_cluster=ProjectCluster.objects.create(field=create_project_field,project=project_creation_id,environment=create_fieldenvironment,clustersubname_id=cluster)
                        development_type=request.POST.getlist('development_type-'+str(fieldenvironment)+'-'+str(cluster)) 
                        for developmenttype in development_type:
                            create_development=ProjectDevelopmentType.objects.create(cluster=create_project_cluster,development_id=developmenttype,project=project_creation_id,environment_id=create_fieldenvironment.id)
                            print("dev type",developmenttype)
                            get_projectdiscipline=request.POST.getlist('project_discipline-'+str(fieldenvironment)+'-'+str(cluster)+'-'+str(developmenttype))
                            for projectdiscipline in get_projectdiscipline:
                                create_discipline=ProjectDevelopmentDiscipline.objects.create(cluster=create_project_cluster,project=project_creation_id,project_discipline=projectdiscipline,development_type=create_development)
                                print("pro dis",projectdiscipline)
                                if projectdiscipline == '3':
                                    get_other_type=request.POST.getlist('otherdisciplinesubtype-'+str(fieldenvironment)+'-'+str(cluster)+'-'+str(developmenttype)+'-'+str(projectdiscipline))
                                    if len(get_other_type) > 0:
                                        for other_type in get_other_type:
                                            create_other_type=ProjectDevelopmentSubType.objects.create(project_discipline=create_discipline,discipline_subtype_id=other_type,project=project_creation_id,development_type=create_development)
                                            print("other_type",other_type)
                                            get_other_activities=request.POST.getlist('otherstypeactivities-'+str(fieldenvironment)+'-'+str(cluster)+'-'+str(developmenttype)+'-'+str(projectdiscipline)+'-'+str(other_type))
                                            for other_activities in get_other_activities:
                                                create_other_activity=DevelopmentSubTypeSub.objects.create(project=project_creation_id,project_discipline=create_discipline,other_type=create_other_type,sub_subtype_sub_id=other_activities)
                                                print("ot act",other_activities)
                                get_disciplinetype=request.POST.getlist('developmentsubtype-'+str(fieldenvironment)+'-'+str(cluster)+'-'+str(developmenttype)+'-'+str(projectdiscipline))
                                for disciplinetype in get_disciplinetype:
                                    create_disciplinetype=ProjectDevelopmentSubType.objects.create(project_discipline=create_discipline,discipline_subtype_id=disciplinetype,project=project_creation_id,development_type=create_development)
                                    print("disciplinetype",disciplinetype)
                                    get_wellsubtype=request.POST.getlist('wellsubtype-'+str(projectdiscipline)+'-'+str(developmenttype)+'-'+str(disciplinetype))
                                    if len(get_wellsubtype) > 0:
                                        wellsubtype=get_wellsubtype[0]
                                        print("brownwellsubtype",wellsubtype)
                                        create_well_discipline_subtype=DevelopmentSubSubType.objects.create(project=project_creation_id,project_discipline=create_discipline,discipline_subtype=create_disciplinetype,disciplinesub_subtype_id=wellsubtype)
                                        get_well_sub_act=request.POST.getlist('wellsubtype_activity-'+str(fieldenvironment)+'-'+str(cluster)+'-'+str(developmenttype)+'-'+str(projectdiscipline)+'-'+str(disciplinetype)+'-'+str(wellsubtype))
                                        for well_sub_act in get_well_sub_act:
                                            create_disciplinesubdiv=DevelopmentSubTypeSub.objects.create(project=project_creation_id,project_discipline=create_discipline,disciplinesub_subtype=create_well_discipline_subtype,sub_subtype_sub_id=well_sub_act)
                                            print("brownget_well_sub_act",well_sub_act)
                                    get_well=request.POST.getlist('wells-'+str(projectdiscipline)+'-'+str(developmenttype)+'-'+str(disciplinetype)) 
                                    if len(get_well) > 0:
                                        for well in get_well:
                                            print("welltype",well)
                                            create_welltype=ProjectWellType.objects.create(project=project_creation_id,welltype_id=well,project_discipline=create_discipline,discipline_type=create_disciplinetype)
                                            get_welltype=request.POST.getlist('welltype'+str(well)+'-'+str(disciplinetype))
                                            for welltype in get_welltype:
                                                create_well=ProjectWellName.objects.create(project=project_creation_id,welltype_id=create_welltype.id,wellname_id=welltype,project_discipline=create_discipline)
                                                print("wellname",welltype)
                                                get_well_activites=request.POST.getlist('typessubtype-'+str(welltype)+'-'+str(disciplinetype))
                                                for wellactivity in get_well_activites:
                                                    print("wellactivity",wellactivity)
                                                    create_wellactivity=DevelopmentSubTypeSub.objects.create(wellname=create_well,sub_subtype_sub_id=wellactivity,project=project_creation_id,project_discipline=create_discipline)
                                    else:
                                        get_discipline_subtype=request.POST.getlist('disciplinesubsubtype-'+str(fieldenvironment)+'-'+str(cluster)+'-'+str(developmenttype)+'-'+str(projectdiscipline)+'-'+str(disciplinetype))
                                        for discipline_subtype in get_discipline_subtype:
                                            create_discipline_subtype=DevelopmentSubSubType.objects.create(project=project_creation_id,project_discipline=create_discipline,discipline_subtype=create_disciplinetype,disciplinesub_subtype_id=discipline_subtype)
                                            print("discipline_subtype",discipline_subtype)
                                            get_discipline_subdiv=request.POST.getlist('typessubtype-'+str(fieldenvironment)+'-'+str(cluster)+'-'+str(developmenttype)+'-'+str(projectdiscipline)+'-'+str(disciplinetype)+'-'+str(discipline_subtype))
                                            for discipline_subdiv in get_discipline_subdiv:
                                                create_disciplinesubdiv=DevelopmentSubTypeSub.objects.create(project=project_creation_id,project_discipline=create_discipline,disciplinesub_subtype=create_discipline_subtype,sub_subtype_sub_id=discipline_subdiv,discipline_type=create_disciplinetype.id)
                                                print("discipline_subdiv",discipline_subdiv)

        return redirect('projects:projectlist')
    else:
        form=ProjectCreationForm()  
    if (request.user.roles_id == 3):
        userrights=UserRights.objects.get(user_id=request.user.id,module_id=16)
        data['rights']=userrights
    data.update({"form":form,"companyname":get_company,'get_subtype':get_subtype,"all_projects":all_projects,"countries":countries,'users':users,'datatypes':datas,'currency':currency,'roleid':request.user.roles_id,'userid':request.user})
    return render(request,"projectcreate.html",data)



def draftprojectcreation(request):
    data={}
    if request.method=="POST":
        print("form details",request.POST)
        form=ProjectCreationForm(request.POST)
        if form.is_valid():
            projectcreationform=form.save(commit=False)
            projectcreationform.country_id=request.POST['master_country_list']
            projectcreationform.company=request.company
            projectcreationform.projectname_id=request.POST['project_list']
            projectcreationform.project_type=request.POST.get('projecttype'," ")
            projectcreationform.entity=request.POST.get('entity'," ")
            projectcreationform.status=0
            projectcreationform.save()
            project_creation_id=projectcreationform
        get_block=request.POST.getlist('block')
        project_users = request.POST.getlist('project_manager', [])
        if (len(project_users)>0):
            project_users.append(request.user.id)
            [ProjectUser.objects.create(company=request.company,user_id=i,project=project_creation_id) for i in project_users]
        for block in get_block:
            print(block)
            create_project_block=ProjectBlock.objects.create(block_id=block,project=project_creation_id)
            get_field=request.POST.getlist('field'+str(block))
            for field in get_field:
                print(field)
                create_project_field=ProjectField.objects.create(block_id=create_project_block.id,field_id=field,project=project_creation_id)
                get_fieldenvironment=request.POST.getlist('fieldenvironment-'+str(field))
                for fieldenvironment in get_fieldenvironment:
                    print(fieldenvironment)
                    create_fieldenvironment=ProjectEnvironment.objects.create(block_id=create_project_block.id,project=project_creation_id,field=create_project_field,field_environment_id=fieldenvironment)
                    get_cluster=request.POST.getlist('cluster-'+str(fieldenvironment))
                    for cluster in get_cluster:
                        print(cluster)
                        create_project_cluster=ProjectCluster.objects.create(field=create_project_field,project=project_creation_id,environment=create_fieldenvironment,clustersubname_id=cluster)
                        development_type=request.POST.getlist('development_type-'+str(fieldenvironment)+'-'+str(cluster)) 
                        for developmenttype in development_type:
                            create_development=ProjectDevelopmentType.objects.create(cluster=create_project_cluster,development_id=developmenttype,project=project_creation_id,environment_id=create_fieldenvironment.id)
                            print("dev type",developmenttype)
                            get_projectdiscipline=request.POST.getlist('project_discipline-'+str(fieldenvironment)+'-'+str(cluster)+'-'+str(developmenttype))
                            for projectdiscipline in get_projectdiscipline:
                                create_discipline=ProjectDevelopmentDiscipline.objects.create(cluster=create_project_cluster,project=project_creation_id,project_discipline=projectdiscipline,development_type=create_development)
                                print("pro dis",projectdiscipline)
                                if projectdiscipline == '3':
                                    get_other_type=request.POST.getlist('otherdisciplinesubtype-'+str(fieldenvironment)+'-'+str(cluster)+'-'+str(developmenttype)+'-'+str(projectdiscipline))
                                    if len(get_other_type) > 0:
                                        for other_type in get_other_type:
                                            create_other_type=ProjectDevelopmentSubType.objects.create(project_discipline=create_discipline,discipline_subtype_id=other_type,project=project_creation_id,development_type=create_development)
                                            print("other_type",other_type)
                                            get_other_activities=request.POST.getlist('otherstypeactivities-'+str(fieldenvironment)+'-'+str(cluster)+'-'+str(developmenttype)+'-'+str(projectdiscipline)+'-'+str(other_type))
                                            for other_activities in get_other_activities:
                                                create_other_activity=DevelopmentSubTypeSub.objects.create(project=project_creation_id,project_discipline=create_discipline,other_type=create_other_type,sub_subtype_sub_id=other_activities)
                                                print("ot act",other_activities)
                                get_disciplinetype=request.POST.getlist('developmentsubtype-'+str(fieldenvironment)+'-'+str(cluster)+'-'+str(developmenttype)+'-'+str(projectdiscipline))
                                for disciplinetype in get_disciplinetype:
                                    create_disciplinetype=ProjectDevelopmentSubType.objects.create(project_discipline=create_discipline,discipline_subtype_id=disciplinetype,project=project_creation_id,development_type=create_development)
                                    print("disciplinetype",disciplinetype)
                                    get_wellsubtype=request.POST.getlist('wellsubtype-'+str(projectdiscipline)+'-'+str(developmenttype)+'-'+str(disciplinetype))
                                    if len(get_wellsubtype) > 0:
                                        wellsubtype=get_wellsubtype[0]
                                        print("brownwellsubtype",wellsubtype)
                                        create_well_discipline_subtype=DevelopmentSubSubType.objects.create(project=project_creation_id,project_discipline=create_discipline,discipline_subtype=create_disciplinetype,disciplinesub_subtype_id=wellsubtype)
                                        get_well_sub_act=request.POST.getlist('wellsubtype_activity-'+str(fieldenvironment)+'-'+str(cluster)+'-'+str(developmenttype)+'-'+str(projectdiscipline)+'-'+str(disciplinetype)+'-'+str(wellsubtype))
                                        for well_sub_act in get_well_sub_act:
                                            create_disciplinesubdiv=DevelopmentSubTypeSub.objects.create(project=project_creation_id,project_discipline=create_discipline,disciplinesub_subtype=create_well_discipline_subtype,sub_subtype_sub_id=well_sub_act)
                                            print("brownget_well_sub_act",well_sub_act)
                                    get_well=request.POST.getlist('wells-'+str(projectdiscipline)+'-'+str(developmenttype)+'-'+str(disciplinetype)) 
                                    if len(get_well) > 0:
                                        for well in get_well:
                                            print("welltype",well)
                                            create_welltype=ProjectWellType.objects.create(project=project_creation_id,welltype_id=well,project_discipline=create_discipline,discipline_type=create_disciplinetype)
                                            get_welltype=request.POST.getlist('welltype'+str(well)+'-'+str(disciplinetype))
                                            for welltype in get_welltype:
                                                create_well=ProjectWellName.objects.create(project=project_creation_id,welltype_id=create_welltype.id,wellname_id=welltype,project_discipline=create_discipline)
                                                print("wellname",welltype)
                                                get_well_activites=request.POST.getlist('typessubtype-'+str(welltype)+'-'+str(disciplinetype))
                                                for wellactivity in get_well_activites:
                                                    print("wellactivity",wellactivity)
                                                    create_wellactivity=DevelopmentSubTypeSub.objects.create(wellname=create_well,sub_subtype_sub_id=wellactivity,project=project_creation_id,project_discipline=create_discipline)
                                    else:
                                        get_discipline_subtype=request.POST.getlist('disciplinesubsubtype-'+str(fieldenvironment)+'-'+str(cluster)+'-'+str(developmenttype)+'-'+str(projectdiscipline)+'-'+str(disciplinetype))
                                        for discipline_subtype in get_discipline_subtype:
                                            create_discipline_subtype=DevelopmentSubSubType.objects.create(project=project_creation_id,project_discipline=create_discipline,discipline_subtype=create_disciplinetype,disciplinesub_subtype_id=discipline_subtype)
                                            print("discipline_subtype",discipline_subtype)
                                            get_discipline_subdiv=request.POST.getlist('typessubtype-'+str(fieldenvironment)+'-'+str(cluster)+'-'+str(developmenttype)+'-'+str(projectdiscipline)+'-'+str(disciplinetype)+'-'+str(discipline_subtype))
                                            for discipline_subdiv in get_discipline_subdiv:
                                                create_disciplinesubdiv=DevelopmentSubTypeSub.objects.create(project=project_creation_id,project_discipline=create_discipline,disciplinesub_subtype=create_discipline_subtype,sub_subtype_sub_id=discipline_subdiv,discipline_type=create_disciplinetype.id)
                                                print("discipline_subdiv",discipline_subdiv)
        data={'data':'success'}
        print(request.POST)
    else:
        data={'data':'error'}
        print(data)
    return JsonResponse(data)




def editprojectcreate(request,pk):
    currentcompany=request.company
    entitytypelist=['Oil_and_Gas_Operators','Oil_and_Gas_Servicing_Companies','Other_Industry']
    country_count=Projects.objects.filter(company=currentcompany).values('country').annotate(dcount=Count('country'))
    countries=[]
    for country in country_count:
        country_id=country['country']
        get_countries=Countries.objects.get(id=country_id) 
        countries.append({"country_id":get_countries.id,"country_name":get_countries.nicename})
    projectdata=Projectcreation.objects.get(id=pk)
    if (projectdata.projectname_id != None):
        projectids=Projectcreation.objects.filter(company=request.company).exclude(projectname_id=projectdata.projectname.id).values_list('projectname',flat=True)
    else:
        projectids=Projectcreation.objects.filter(company=request.company).exclude(projectname__isnull=True).values_list('projectname',flat=True)
    # print('projectids',projectids)
    masterproject=Projects.objects.filter(country_id=projectdata.country_id,company=request.company).exclude(id__in=projectids)
    project_creation=Projectcreation.objects.get(id=pk)
    project_creation=project_creation.signatory_type
    print('project_creation',project_creation)
    # print("masterproject",masterproject)
    blocks=BlockName.objects.filter(project_id=projectdata.projectname_id,status=0)
    projectdiscipline=[{'1':'Green Field Development'},{'2':'Brown Field Development'},{'3':'Others'}]
    data={}
    current_blockids=[]
    if (request.POST):
        print("req",request.POST)
        projectdata=Projectcreation.objects.filter(id=pk).update(country_id=request.POST['master_country_list'],company=request.company,projectname_id=request.POST['project_list'],entity=request.POST.get('entity'," "))
        blocks=request.POST.getlist('block')
        print("req1",request.POST)
        #Block
        for block in blocks:
            getblock=ProjectBlock.objects.filter(block_id=block,project_id=pk).first()
            if(getblock==None):
                create_block=ProjectBlock.objects.create(project_id=pk,block_id=block)
                current_blockids.append(create_block.id)
                print("create new")
            else:
                print("old")
                ProjectBlock.objects.filter(block_id=block,project_id=pk).update(status=1)
                current_blockids.append(getblock.id)
            blockid=ProjectBlock.objects.filter(block_id=block,project_id=pk,status=1).first()
            print(f"blockid {blockid}")
            get_projectfield=request.POST.getlist('field'+str(block))
            current_fieldids=[]
            
            #Field
            for field in get_projectfield:
                getfield=ProjectField.objects.filter(block_id=blockid.id,field_id=field,project_id=pk).first()
                if(getfield==None):
                    create_field=ProjectField.objects.create(project_id=pk,block_id=blockid.id,field_id=field)
                    current_fieldids.append(create_field.id)
                else:
                    ProjectField.objects.filter(block_id=blockid.id,project_id=pk,field_id=field).update(status=1)
                    current_fieldids.append(getfield.id)

                fieldid=ProjectField.objects.filter(block_id=blockid.id,field_id=field,project_id=pk,status=1).first()
                get_fieldenv=request.POST.getlist('fieldenvironment-'+str(field))
                current_environmentids=[]
               
                #Environment
                for environment in get_fieldenv:
                    get_environment=ProjectEnvironment.objects.filter(field_id=fieldid.id,project_id=pk,field_environment_id=environment).first()
                    if(get_environment==None):
                        project_env=ProjectEnvironment.objects.create(block_id=blockid.id,field_id=fieldid.id,project_id=pk,field_environment_id=environment)
                        current_environmentids.append(project_env.id)
                    else:
                        ProjectEnvironment.objects.filter(block_id=blockid.id,field_id=fieldid.id,project_id=pk,field_environment_id=environment).update(status=1)
                        current_environmentids.append(get_environment.id)
                ProjectCluster.objects.filter(field_id=fieldid.id,project_id=pk).exclude(environment_id__in=current_environmentids).update(status=0)
                ProjectEnvironment.objects.filter(project_id=pk,field_id=fieldid.id).exclude(id__in=current_environmentids).update(status=0)
            ProjectField.objects.filter(block_id=blockid.id).exclude(id__in=current_fieldids).update(status=0)
            # print(f"current_fieldids {current_fieldids}")
            ProjectEnvironment.objects.filter(project_id=pk,block_id=blockid.id).exclude(field_id__in=current_fieldids).update(status=0)
        ProjectField.objects.filter(project_id=pk).exclude(block_id__in=current_blockids).update(status=0)
        ProjectBlock.objects.filter(project_id=pk).exclude(id__in=current_blockids).update(status=0)
        data["status"]="success"
        return JsonResponse(data)    
        
    data={'project':projectdata,
         'company':currentcompany,
         'entitytype':entitytypelist,
         'countries':countries,
         'masterprojects':masterproject,
         'blocks':blocks,
         'projectdiscipline':projectdiscipline,'project_creation':project_creation}
    return render(request,"editprojectcreation.html",data)


def editprojectcreatenextstep(request,pk):
    projectdata=Projectcreation.objects.get(id=pk)
    blocks=BlockName.objects.filter(project_id=projectdata.projectname_id,status=0)
    projectblock=ProjectBlock.objects.filter(project_id=projectdata.id,status=1)

    # print(request.POST)
    if (request.POST):
        for blockid in projectblock:
            print("sad",projectblock)
            get_projectfield=ProjectField.objects.filter(block_id=blockid.id,project_id=projectdata.id,status=1)
            for fieldid in get_projectfield:
                # print("c",fieldid)
                get_fieldenv=ProjectEnvironment.objects.filter(field_id=fieldid.id,project_id=projectdata.id,status=1)
                for fieldenv in get_fieldenv:
                    # print(f"fieldenv {fieldenv.field_environment_id}")
                    get_project_cluster=request.POST.getlist('cluster-'+str(fieldenv.field_environment_id))
                    # print("cc",get_project_cluster)
                    currentclusterid=[]
                    #cluster
                    for project_cluster in get_project_cluster:
                        getcluster=ProjectCluster.objects.filter(environment_id=fieldenv.id,clustersubname_id=project_cluster,project_id=projectdata.id).first()
                        if(getcluster==None):
                            create_project_cluster=ProjectCluster.objects.create(field_id=fieldid.id,environment_id=fieldenv.id,clustersubname_id=project_cluster,project_id=projectdata.id,status=1)
                            currentclusterid.append(create_project_cluster.id)
                        else:
                            ProjectCluster.objects.filter(field_id=fieldid.id,environment_id=fieldenv.id,clustersubname_id=project_cluster,project_id=projectdata.id).update(status=1)
                            currentclusterid.append(getcluster.id)
                        clusterid=ProjectCluster.objects.filter(environment_id=fieldenv.id,clustersubname_id=project_cluster,project_id=projectdata.id,status=1).first()     
                        get_projectdevelopment=request.POST.getlist('development_type-'+str(fieldenv.field_environment_id)+'-'+str(project_cluster))
                        currentdevelopmentid=[]
                        # development type
                        for projectdevelopment in get_projectdevelopment:
                            getdevelopment=ProjectDevelopmentType.objects.filter(cluster_id=clusterid.id,development_id=projectdevelopment,project_id=projectdata.id).first()
                            if(getdevelopment==None):
                                create_project_development=ProjectDevelopmentType.objects.create(environment_id=fieldenv.id,cluster_id=clusterid.id,development_id=projectdevelopment,project_id=projectdata.id)
                                currentdevelopmentid.append(create_project_development.id)
                            else:
                                ProjectDevelopmentType.objects.filter(development_id=projectdevelopment,cluster_id=clusterid.id,project_id=projectdata.id).update(status=1)
                                currentdevelopmentid.append(getdevelopment.id)
                            
                            development_id=ProjectDevelopmentType.objects.filter(cluster_id=clusterid.id,development_id=projectdevelopment,project_id=projectdata.id).first()

                            get_projectdiscipline=request.POST.getlist('project_discipline-'+str(fieldenv.field_environment_id)+'-'+str(project_cluster)+'-'+str(projectdevelopment))
                            currentdisciplineids=[]
                            #GFD,BFD,Other discipline Iterations

                            for projectdiscipline in get_projectdiscipline:
                                getdiscipline=ProjectDevelopmentDiscipline.objects.filter(development_type_id=development_id.id,project_id=projectdata.id,project_discipline=projectdiscipline).first()
                                # print(f"getdiscipline {getdiscipline}")
                                if(getdiscipline==None):
                                    create_project_discipline=ProjectDevelopmentDiscipline.objects.create(cluster_id=clusterid.id,development_type_id=development_id.id,project_id=projectdata.id,project_discipline=projectdiscipline)
                                    currentdisciplineids.append(create_project_discipline.id)
                                else:
                                    ProjectDevelopmentDiscipline.objects.filter(development_type_id=development_id.id,project_id=projectdata.id,project_discipline=projectdiscipline).update(status=1)
                                    currentdisciplineids.append(getdiscipline.id)
                                
                                disciplineid=ProjectDevelopmentDiscipline.objects.filter(development_type_id=development_id.id,project_id=projectdata.id,project_discipline=projectdiscipline,status=1).first()

                                if(projectdiscipline=="1"):
                                    # print(f"disciplineid {disciplineid}")
                                    updategreenfield(fieldenv.field_environment_id,project_cluster,projectdevelopment,projectdiscipline,request,disciplineid,projectdata.id,development_id)
                                elif(projectdiscipline=="2"):
                                    updatebrownfield(fieldenv.field_environment_id,project_cluster,projectdevelopment,projectdiscipline,request,disciplineid,projectdata.id,development_id)
                                elif(projectdiscipline=="3"):
                                    updateothers(fieldenv.field_environment_id,project_cluster,projectdevelopment,projectdiscipline,request,disciplineid,projectdata.id,development_id)

                            # print("currentdisciplineids",currentdisciplineids)
                            ProjectDevelopmentDiscipline.objects.filter(development_type_id=development_id.id,project_id=projectdata.id).exclude(id__in=currentdisciplineids).update(status=0)
                            ProjectDevelopmentSubType.objects.filter(development_type_id=development_id.id,project_id=projectdata.id).exclude(project_discipline_id__in=currentdisciplineids).update(status=0)
                        
                        
                        
                        
                        ProjectDevelopmentType.objects.filter(cluster_id=clusterid.id).exclude(id__in=currentdevelopmentid).update(status=0)
                        # print("dis",currentdevelopmentid)
                        ProjectDevelopmentDiscipline.objects.filter(cluster_id=clusterid.id,project_id=projectdata.id).exclude(development_type_id__in=currentdevelopmentid).update(status=0)

                    ProjectCluster.objects.filter(environment_id=fieldenv.id,project_id=projectdata.id).exclude(id__in=currentclusterid).update(status=0)
                    ProjectDevelopmentType.objects.filter(environment_id=fieldenv.id,project_id=projectdata.id).exclude(cluster_id__in=currentclusterid).update(status=0)

        # return redirect('projects:projectlist')       
        if (request.user.roles_id == 3):
            userrights=UserRights.objects.get(user_id=request.user.id,module_id=16)
            if userrights.edit == '1' or userrights.delete == '1':
                return redirect('projects:editprojectuser',pk=pk)
            else:
                return redirect('projects:projectlist')
        else:
            return redirect('projects:editprojectuser',pk=pk)
    data={'blocks':blocks,
          'project':projectdata}
    return render(request,"editprojectcreationstep2.html",data)


def updategreenfield(envid,project_cluster,projectdevelopment,projectdiscipline,request,disciplineid,project_id,development_id):
    currentdisciplintypids=[]
    get_discipline_type=request.POST.getlist('developmentsubtype-'+str(envid)+'-'+str(project_cluster)+'-'+str(projectdevelopment)+'-'+str(projectdiscipline))
    #Surface,drill iterations discipline type etc..
    for disciplinesubype in get_discipline_type:
        getdisciplinesubtype=ProjectDevelopmentSubType.objects.filter(project_discipline_id=disciplineid.id,discipline_subtype_id=disciplinesubype,project_id=project_id).first()
        if(getdisciplinesubtype==None):
            create_disciplinesubtype=ProjectDevelopmentSubType.objects.create(project_discipline_id=disciplineid.id,discipline_subtype_id=disciplinesubype,project_id=project_id,development_type_id=development_id.id)
            currentdisciplintypids.append(create_disciplinesubtype.id)
        else:
            ProjectDevelopmentSubType.objects.filter(project_discipline_id=disciplineid.id,discipline_subtype_id=disciplinesubype,project_id=project_id).update(status=1)
            currentdisciplintypids.append(getdisciplinesubtype.id) 
        discliplinesubtypeid=ProjectDevelopmentSubType.objects.filter(project_discipline_id=disciplineid.id,discipline_subtype_id=disciplinesubype,project_id=project_id,status=1).first()
        # print(f"discliplinesubtypeid {discliplinesubtypeid}")
        if(disciplinesubype=="2"):
            get_disciplinetypewell=request.POST.getlist('wells-'+str(projectdiscipline)+'-'+str(projectdevelopment)+'-'+str(disciplinesubype))
            # print(f"get_disciplinetypewell {get_disciplinetypewell}")
            #iterate well
            currentwellids=[]
            for well in get_disciplinetypewell:
                getwell=ProjectWellType.objects.filter(discipline_type_id=discliplinesubtypeid.id,project_id=project_id,welltype_id=well).first()
                if(getwell == None):
                    createwell=ProjectWellType.objects.create(project_discipline_id=disciplineid.id,discipline_type_id=discliplinesubtypeid.id,project_id=project_id,welltype_id=well)
                    currentwellids.append(createwell.id)
                else:
                    ProjectWellType.objects.filter(discipline_type_id=discliplinesubtypeid.id,project_id=project_id,welltype_id=well).update(status=1) 
                    currentwellids.append(getwell.id)
                wellid=ProjectWellType.objects.filter(discipline_type_id=discliplinesubtypeid.id,project_id=project_id,welltype_id=well,status=1).first()
                get_well_name=request.POST.getlist('welltype'+str(well)+'-'+str(disciplinesubype))
                current_wellnameids=[]
                #well name exp etc..
                for wellname in get_well_name:
                    drillwell_name=ProjectWellName.objects.filter(wellname_id=wellname,project_id=project_id,welltype_id=wellid.id).first()
                    if (drillwell_name==None):
                        create_wellname=ProjectWellName.objects.create(wellname_id=wellname,project_id=project_id,welltype_id=wellid.id)
                        current_wellnameids.append(create_wellname.id)
                    else:
                        ProjectWellName.objects.filter(wellname_id=wellname,project_id=project_id,welltype_id=wellid.id).update(status=1)
                        current_wellnameids.append(drillwell_name.id)
                    get_wellact=request.POST.getlist('typessubtype-'+str(wellname)+'-'+str(disciplinesubype))
                    wellnameid=ProjectWellName.objects.filter(wellname_id=wellname,project_id=project_id,welltype_id=wellid.id,status=1).first()
                    current_wellactids=[]
                    #well act geo etc..
                    for wellact in get_wellact:
                        wellactdata=DevelopmentSubTypeSub.objects.filter(sub_subtype_sub_id=wellact,wellname_id=wellnameid.id,project_id=project_id).first()
                        if (wellactdata==None):
                            create_wellact=DevelopmentSubTypeSub.objects.create(sub_subtype_sub_id=wellact,wellname_id=wellnameid.id,project_id=project_id)
                            current_wellactids.append(create_wellact.id)
                        else:
                            DevelopmentSubTypeSub.objects.filter(sub_subtype_sub_id=wellact,wellname_id=wellnameid.id,project_id=project_id).update(status=1)
                            current_wellactids.append(wellactdata.id)

                    DevelopmentSubTypeSub.objects.filter(wellname_id=wellnameid.id,project_id=project_id).exclude(id__in=current_wellactids).update(status=0)
                
                ProjectWellName.objects.filter(welltype_id=wellid.id,project_id=project_id).exclude(id__in=current_wellnameids).update(status=0)
                
                excluded_wellname=ProjectWellName.objects.filter(welltype_id=wellid.id,project_id=project_id).exclude(id__in=current_wellnameids)
                
                for ex_wellname in excluded_wellname:
                    DevelopmentSubTypeSub.objects.filter(wellname_id=ex_wellname.id,project_id=project_id).update(status=0)

            ProjectWellType.objects.filter(discipline_type_id=discliplinesubtypeid.id,project_id=project_id).exclude(id__in=currentwellids).update(status=0) 

            excluded_well=ProjectWellType.objects.filter(discipline_type_id=discliplinesubtypeid.id,project_id=project_id).exclude(id__in=currentwellids)
            # print(f"excluded_well {excluded_well}")
            for ex_well in excluded_well:
                ProjectWellName.objects.filter(welltype_id=ex_well.id,project_id=project_id).update(status=0)






            
        else:
            get_disciplinesubsubtype=request.POST.getlist('disciplinesubsubtype-'+str(envid)+'-'+str(project_cluster)+'-'+str(projectdevelopment)+'-'+str(projectdiscipline)+'-'+str(disciplinesubype))
            current_dissubsub_ids=[]
            # print(f'get_disciplinesubsubtype {get_disciplinesubsubtype}')
            #Exploration,appraisal discipline activities etc..
            for dissubsubtype in get_disciplinesubsubtype:
                getdisciplinesubsubtype=DevelopmentSubSubType.objects.filter(disciplinesub_subtype_id=dissubsubtype,discipline_subtype_id=discliplinesubtypeid.id,project_id=project_id).first()
                if(getdisciplinesubsubtype==None):
                    create_discipline_subsub_type=DevelopmentSubSubType.objects.create(disciplinesub_subtype_id=dissubsubtype,discipline_subtype_id=discliplinesubtypeid.id,project_id=project_id,project_discipline_id=disciplineid.id)
                    current_dissubsub_ids.append(create_discipline_subsub_type.id)
                else:
                    DevelopmentSubSubType.objects.filter(disciplinesub_subtype_id=dissubsubtype,discipline_subtype_id=discliplinesubtypeid.id,project_id=project_id).update(status=1) 
                    current_dissubsub_ids.append(getdisciplinesubsubtype.id)
                disciplinesubsubtypeid=DevelopmentSubSubType.objects.filter(disciplinesub_subtype_id=dissubsubtype,discipline_subtype_id=discliplinesubtypeid.id,project_id=project_id,status=1).first()
                get_disciplinesubact=request.POST.getlist('typessubtype-'+str(envid)+'-'+str(project_cluster)+'-'+str(projectdevelopment)+'-'+str(projectdiscipline)+'-'+str(disciplinesubype)+'-'+str(dissubsubtype))
                current_dissubsubactid=[]
                #Geo,Asset sub activities etc..
                for disciplinesubact in get_disciplinesubact:
                    get_dissubsubact=DevelopmentSubTypeSub.objects.filter(disciplinesub_subtype_id=disciplinesubsubtypeid.id,sub_subtype_sub_id=disciplinesubact,project_id=project_id).first()
                    if (get_dissubsubact==None):
                        create_dissubsubact=DevelopmentSubTypeSub.objects.create(disciplinesub_subtype_id=disciplinesubsubtypeid.id,sub_subtype_sub_id=disciplinesubact,project_id=project_id,discipline_type=discliplinesubtypeid.id)
                        current_dissubsubactid.append(create_dissubsubact.id)
                        # print(f"disciplinesubact{disciplinesubact}")
                    else:
                        DevelopmentSubTypeSub.objects.filter(disciplinesub_subtype_id=disciplinesubsubtypeid.id,sub_subtype_sub_id=disciplinesubact,project_id=project_id).update(status=1)
                        current_dissubsubactid.append(get_dissubsubact.id)
                DevelopmentSubTypeSub.objects.filter(disciplinesub_subtype_id=disciplinesubsubtypeid.id,project_id=project_id).exclude(id__in=current_dissubsubactid).update(status=0)
            
            DevelopmentSubSubType.objects.filter(discipline_subtype_id=discliplinesubtypeid.id,project_id=project_id).exclude(id__in=current_dissubsub_ids).update(status=0) 
            DevelopmentSubTypeSub.objects.filter(discipline_type=discliplinesubtypeid.id,project_id=project_id).exclude(disciplinesub_subtype_id__in=current_dissubsub_ids).update(status=0)



            


    
    ProjectDevelopmentSubType.objects.filter(project_discipline_id=disciplineid.id,project_id=project_id).exclude(id__in=currentdisciplintypids).update(status=0)

    excludeddevelopment_type=ProjectDevelopmentSubType.objects.filter(project_discipline_id=disciplineid.id,project_id=project_id).exclude(id__in=currentdisciplintypids)
    # print(f"excludeddevelopment_type {excludeddevelopment_type}")
    for developmentsubtype in excludeddevelopment_type:
        # print(f"developmentsubtype.discipline_subtype_id {type(developmentsubtype.discipline_subtype_id)}")
        if(developmentsubtype.discipline_subtype_id==2):
            ProjectWellType.objects.filter(discipline_type_id=developmentsubtype.id,project_id=project_id).update(status=0)
        else:
            DevelopmentSubSubType.objects.filter(discipline_subtype_id=developmentsubtype.id,project_id=project_id).update(status=0)



def updatebrownfield(envid,project_cluster,projectdevelopment,projectdiscipline,request,disciplineid,project_id,development_id):
    currentdisciplintypids=[]
    get_discipline_type=request.POST.getlist('developmentsubtype-'+str(envid)+'-'+str(project_cluster)+'-'+str(projectdevelopment)+'-'+str(projectdiscipline))
    #Surface,drill iterations discipline type etc..
    for disciplinesubype in get_discipline_type:
        # print("disciplinesubype",disciplinesubype)
        getdisciplinesubtype=ProjectDevelopmentSubType.objects.filter(project_discipline_id=disciplineid.id,discipline_subtype_id=disciplinesubype,project_id=project_id).first()
        # print("getdisciplinesubtype",getdisciplinesubtype)
        if(getdisciplinesubtype==None):
            create_disciplinesubtype=ProjectDevelopmentSubType.objects.create(project_discipline_id=disciplineid.id,discipline_subtype_id=disciplinesubype,project_id=project_id,development_type_id=development_id.id)
            currentdisciplintypids.append(create_disciplinesubtype.id)
        else:
            ProjectDevelopmentSubType.objects.filter(project_discipline_id=disciplineid.id,discipline_subtype_id=disciplinesubype,project_id=project_id).update(status=1)
            currentdisciplintypids.append(getdisciplinesubtype.id) 
        discliplinesubtypeid=ProjectDevelopmentSubType.objects.filter(project_discipline_id=disciplineid.id,discipline_subtype_id=disciplinesubype,project_id=project_id,status=1).first()

        if (disciplinesubype=="6"):
            get_browndisciplineact=request.POST.get('wellsubtype-'+str(projectdiscipline)+'-'+str(projectdevelopment)+'-'+str(disciplinesubype))
            # print("dsasad",get_browndisciplineact)
            if (get_browndisciplineact !=None):
                browndissuract=DevelopmentSubSubType.objects.filter(disciplinesub_subtype_id=get_browndisciplineact,discipline_subtype_id=discliplinesubtypeid.id,project_id=project_id).first()
                current_browndissuract=[]
                if (browndissuract==None):
                    create_browndissuract=DevelopmentSubSubType.objects.create(project_discipline_id=disciplineid.id,disciplinesub_subtype_id=get_browndisciplineact,discipline_subtype_id=discliplinesubtypeid.id,project_id=project_id)
                    current_browndissuract.append(create_browndissuract.id)
                else:
                    print("no")
                    DevelopmentSubSubType.objects.filter(project_discipline_id=disciplineid.id,disciplinesub_subtype_id=get_browndisciplineact,discipline_subtype_id=discliplinesubtypeid.id,project_id=project_id).update(status=1)
                    current_browndissuract.append(browndissuract.id)
                print("current_browndissuract",current_browndissuract)
                browndissuractid=DevelopmentSubSubType.objects.filter(disciplinesub_subtype_id=get_browndisciplineact,discipline_subtype_id=discliplinesubtypeid.id,project_id=project_id).first()
                get_brownsuract=request.POST.getlist('wellsubtype_activity-'+str(envid)+'-'+str(project_cluster)+'-'+str(projectdevelopment)+'-'+str(projectdiscipline)+'-'+str(disciplinesubype)+'-'+str(get_browndisciplineact))
                # print("get_brownsuract",get_brownsuract)
                current_dissubsubactid=[]
                for disciplinesubact in get_brownsuract:
                    get_dissubsubact=DevelopmentSubTypeSub.objects.filter(disciplinesub_subtype_id=browndissuractid.id,sub_subtype_sub_id=disciplinesubact,project_id=project_id).first()
                    if (get_dissubsubact==None):
                        create_dissubsubact=DevelopmentSubTypeSub.objects.create(disciplinesub_subtype_id=browndissuractid.id,sub_subtype_sub_id=disciplinesubact,project_id=project_id,discipline_type=discliplinesubtypeid.id)
                        current_dissubsubactid.append(create_dissubsubact.id)
                        # print(f"disciplinesubact{disciplinesubact}")
                    else:
                        DevelopmentSubTypeSub.objects.filter(disciplinesub_subtype_id=browndissuractid.id,sub_subtype_sub_id=disciplinesubact,project_id=project_id).update(status=1)
                        current_dissubsubactid.append(get_dissubsubact.id)
                DevelopmentSubTypeSub.objects.filter(disciplinesub_subtype_id=browndissuractid.id,project_id=project_id).exclude(id__in=current_dissubsubactid).update(status=0)
            else:
                DevelopmentSubSubType.objects.filter(discipline_subtype_id=discliplinesubtypeid.id,project_id=project_id).update(status=0)
                DevelopmentSubTypeSub.objects.filter(discipline_type=discliplinesubtypeid.id,project_id=project_id).update(status=0)

            get_disciplinetypewell=request.POST.getlist('wells-'+str(projectdiscipline)+'-'+str(projectdevelopment)+'-'+str(disciplinesubype))
            # print(f"get_disciplinetypewell {get_disciplinetypewell}")
            #iterate well
            currentwellids=[]
            for well in get_disciplinetypewell:
                getwell=ProjectWellType.objects.filter(discipline_type_id=discliplinesubtypeid.id,project_id=project_id,welltype_id=well).first()
                if(getwell == None):
                    createwell=ProjectWellType.objects.create(project_discipline_id=disciplineid.id,discipline_type_id=discliplinesubtypeid.id,project_id=project_id,welltype_id=well)
                    currentwellids.append(createwell.id)
                else:
                    ProjectWellType.objects.filter(project_discipline_id=disciplineid.id,discipline_type_id=discliplinesubtypeid.id,project_id=project_id,welltype_id=well).update(status=1) 
                    currentwellids.append(getwell.id)
                wellid=ProjectWellType.objects.filter(discipline_type_id=discliplinesubtypeid.id,project_id=project_id,welltype_id=well,status=1).first()
                get_well_name=request.POST.getlist('welltype'+str(well)+'-'+str(disciplinesubype))
                current_wellnameids=[]
                #well name exp etc..
                for wellname in get_well_name:
                    drillwell_name=ProjectWellName.objects.filter(wellname_id=wellname,project_id=project_id,welltype_id=wellid.id).first()
                    if (drillwell_name==None):
                        create_wellname=ProjectWellName.objects.create(wellname_id=wellname,project_id=project_id,welltype_id=wellid.id)
                        current_wellnameids.append(create_wellname.id)
                    else:
                        ProjectWellName.objects.filter(wellname_id=wellname,project_id=project_id,welltype_id=wellid.id).update(status=1)
                        current_wellnameids.append(drillwell_name.id)
                    get_wellact=request.POST.getlist('typessubtype-'+str(wellname)+'-'+str(disciplinesubype))
                    wellnameid=ProjectWellName.objects.filter(wellname_id=wellname,project_id=project_id,welltype_id=wellid.id,status=1).first()
                    current_wellactids=[]
                    #well act geo etc..
                    for wellact in get_wellact:
                        wellactdata=DevelopmentSubTypeSub.objects.filter(sub_subtype_sub_id=wellact,wellname_id=wellnameid.id,project_id=project_id).first()
                        if (wellactdata==None):
                            create_wellact=DevelopmentSubTypeSub.objects.create(sub_subtype_sub_id=wellact,wellname_id=wellnameid.id,project_id=project_id)
                            current_wellactids.append(create_wellact.id)
                        else:
                            DevelopmentSubTypeSub.objects.filter(sub_subtype_sub_id=wellact,wellname_id=wellnameid.id,project_id=project_id).update(status=1)
                            current_wellactids.append(wellactdata.id)

                    DevelopmentSubTypeSub.objects.filter(wellname_id=wellnameid.id,project_id=project_id).exclude(id__in=current_wellactids).update(status=0)
                
                ProjectWellName.objects.filter(welltype_id=wellid.id,project_id=project_id).exclude(id__in=current_wellnameids).update(status=0)
                
                excluded_wellname=ProjectWellName.objects.filter(welltype_id=wellid.id,project_id=project_id).exclude(id__in=current_wellnameids)
                
                for ex_wellname in excluded_wellname:
                    DevelopmentSubTypeSub.objects.filter(wellname_id=ex_wellname.id,project_id=project_id).update(status=0)

            ProjectWellType.objects.filter(discipline_type_id=discliplinesubtypeid.id,project_id=project_id).exclude(id__in=currentwellids).update(status=0) 

            excluded_well=ProjectWellType.objects.filter(discipline_type_id=discliplinesubtypeid.id,project_id=project_id).exclude(id__in=currentwellids)
            # print(f"excluded_well {excluded_well}")
            for ex_well in excluded_well:
                ProjectWellName.objects.filter(welltype_id=ex_well.id,project_id=project_id).update(status=0)

        #brown drill
        elif (disciplinesubype=="7"):
            get_disciplinetypewell=request.POST.getlist('wells-'+str(projectdiscipline)+'-'+str(projectdevelopment)+'-'+str(disciplinesubype))
            # print(f"get_disciplinetypewell {get_disciplinetypewell}")
            #iterate well
            currentwellids=[]
            for well in get_disciplinetypewell:
                getwell=ProjectWellType.objects.filter(discipline_type_id=discliplinesubtypeid.id,project_id=project_id,welltype_id=well).first()
                if(getwell == None):
                    createwell=ProjectWellType.objects.create(project_discipline_id=disciplineid.id,discipline_type_id=discliplinesubtypeid.id,project_id=project_id,welltype_id=well)
                    currentwellids.append(createwell.id)
                else:
                    ProjectWellType.objects.filter(project_discipline_id=disciplineid.id,discipline_type_id=discliplinesubtypeid.id,project_id=project_id,welltype_id=well).update(status=1) 
                    currentwellids.append(getwell.id)
                wellid=ProjectWellType.objects.filter(discipline_type_id=discliplinesubtypeid.id,project_id=project_id,welltype_id=well,status=1).first()
                get_well_name=request.POST.getlist('welltype'+str(well)+'-'+str(disciplinesubype))
                current_wellnameids=[]
                #well name exp etc..
                for wellname in get_well_name:
                    drillwell_name=ProjectWellName.objects.filter(wellname_id=wellname,project_id=project_id,welltype_id=wellid.id).first()
                    if (drillwell_name==None):
                        create_wellname=ProjectWellName.objects.create(project_discipline_id=disciplineid.id,wellname_id=wellname,project_id=project_id,welltype_id=wellid.id)
                        current_wellnameids.append(create_wellname.id)
                    else:
                        ProjectWellName.objects.filter(project_discipline_id=disciplineid.id,wellname_id=wellname,project_id=project_id,welltype_id=wellid.id).update(status=1)
                        current_wellnameids.append(drillwell_name.id)
                    get_wellact=request.POST.getlist('typessubtype-'+str(wellname)+'-'+str(disciplinesubype))
                    wellnameid=ProjectWellName.objects.filter(wellname_id=wellname,project_id=project_id,welltype_id=wellid.id,status=1).first()
                    current_wellactids=[]
                    #well act geo etc..
                    for wellact in get_wellact:
                        wellactdata=DevelopmentSubTypeSub.objects.filter(sub_subtype_sub_id=wellact,wellname_id=wellnameid.id,project_id=project_id).first()
                        if (wellactdata==None):
                            create_wellact=DevelopmentSubTypeSub.objects.create(sub_subtype_sub_id=wellact,wellname_id=wellnameid.id,project_id=project_id)
                            current_wellactids.append(create_wellact.id)
                        else:
                            DevelopmentSubTypeSub.objects.filter(sub_subtype_sub_id=wellact,wellname_id=wellnameid.id,project_id=project_id).update(status=1)
                            current_wellactids.append(wellactdata.id)

                    DevelopmentSubTypeSub.objects.filter(wellname_id=wellnameid.id,project_id=project_id).exclude(id__in=current_wellactids).update(status=0)
                
                ProjectWellName.objects.filter(welltype_id=wellid.id,project_id=project_id).exclude(id__in=current_wellnameids).update(status=0)
                
                excluded_wellname=ProjectWellName.objects.filter(welltype_id=wellid.id,project_id=project_id).exclude(id__in=current_wellnameids)
                
                for ex_wellname in excluded_wellname:
                    DevelopmentSubTypeSub.objects.filter(wellname_id=ex_wellname.id,project_id=project_id).update(status=0)

            ProjectWellType.objects.filter(discipline_type_id=discliplinesubtypeid.id,project_id=project_id).exclude(id__in=currentwellids).update(status=0) 

            excluded_well=ProjectWellType.objects.filter(discipline_type_id=discliplinesubtypeid.id,project_id=project_id).exclude(id__in=currentwellids)
            # print(f"excluded_well {excluded_well}")
            for ex_well in excluded_well:
                ProjectWellName.objects.filter(welltype_id=ex_well.id,project_id=project_id).update(status=0)
        else:
            get_disciplinesubsubtype=request.POST.getlist('disciplinesubsubtype-'+str(envid)+'-'+str(project_cluster)+'-'+str(projectdevelopment)+'-'+str(projectdiscipline)+'-'+str(disciplinesubype))
            current_dissubsub_ids=[]
            # print(f'get_disciplinesubsubtype {get_disciplinesubsubtype}')
            #Exploration,appraisal discipline activities etc..
            for dissubsubtype in get_disciplinesubsubtype:
                getdisciplinesubsubtype=DevelopmentSubSubType.objects.filter(disciplinesub_subtype_id=dissubsubtype,discipline_subtype_id=discliplinesubtypeid.id,project_id=project_id).first()
                if(getdisciplinesubsubtype==None):
                    create_discipline_subsub_type=DevelopmentSubSubType.objects.create(disciplinesub_subtype_id=dissubsubtype,discipline_subtype_id=discliplinesubtypeid.id,project_id=project_id,project_discipline_id=disciplineid.id)
                    current_dissubsub_ids.append(create_discipline_subsub_type.id)
                else:
                    DevelopmentSubSubType.objects.filter(disciplinesub_subtype_id=dissubsubtype,discipline_subtype_id=discliplinesubtypeid.id,project_id=project_id).update(status=1) 
                    current_dissubsub_ids.append(getdisciplinesubsubtype.id)
                disciplinesubsubtypeid=DevelopmentSubSubType.objects.filter(disciplinesub_subtype_id=dissubsubtype,discipline_subtype_id=discliplinesubtypeid.id,project_id=project_id,status=1).first()
                get_disciplinesubact=request.POST.getlist('typessubtype-'+str(envid)+'-'+str(project_cluster)+'-'+str(projectdevelopment)+'-'+str(projectdiscipline)+'-'+str(disciplinesubype)+'-'+str(dissubsubtype))
                current_dissubsubactid=[]
                #Geo,Asset sub activities etc..
                for disciplinesubact in get_disciplinesubact:
                    get_dissubsubact=DevelopmentSubTypeSub.objects.filter(disciplinesub_subtype_id=disciplinesubsubtypeid.id,sub_subtype_sub_id=disciplinesubact,project_id=project_id).first()
                    if (get_dissubsubact==None):
                        create_dissubsubact=DevelopmentSubTypeSub.objects.create(disciplinesub_subtype_id=disciplinesubsubtypeid.id,sub_subtype_sub_id=disciplinesubact,project_id=project_id,discipline_type=discliplinesubtypeid.id)
                        current_dissubsubactid.append(create_dissubsubact.id)
                        # print(f"disciplinesubact{disciplinesubact}")
                    else:
                        DevelopmentSubTypeSub.objects.filter(disciplinesub_subtype_id=disciplinesubsubtypeid.id,sub_subtype_sub_id=disciplinesubact,project_id=project_id).update(status=1)
                        current_dissubsubactid.append(get_dissubsubact.id)
                DevelopmentSubTypeSub.objects.filter(disciplinesub_subtype_id=disciplinesubsubtypeid.id,project_id=project_id).exclude(id__in=current_dissubsubactid).update(status=0)
            
            DevelopmentSubSubType.objects.filter(discipline_subtype_id=discliplinesubtypeid.id,project_id=project_id).exclude(id__in=current_dissubsub_ids).update(status=0) 
            DevelopmentSubTypeSub.objects.filter(discipline_type=discliplinesubtypeid.id,project_id=project_id).exclude(disciplinesub_subtype_id__in=current_dissubsub_ids).update(status=0)

    ProjectDevelopmentSubType.objects.filter(project_discipline_id=disciplineid.id,project_id=project_id).exclude(id__in=currentdisciplintypids).update(status=0)
    excludedbrowndevelopment_type=ProjectDevelopmentSubType.objects.filter(project_discipline_id=disciplineid.id,project_id=project_id).exclude(id__in=currentdisciplintypids)
    # print(f"excludedbrowndevelopment_type {excludedbrowndevelopment_type}")
    for developmentsubtype in excludedbrowndevelopment_type:
        # print(f"developmentsubtype.discipline_subtype_id {type(developmentsubtype.discipline_subtype_id)}")
        if(developmentsubtype.discipline_subtype_id==6):
            ProjectWellType.objects.filter(discipline_type_id=developmentsubtype.id,project_id=project_id).update(status=0)
            DevelopmentSubSubType.objects.filter(discipline_subtype_id=developmentsubtype.id,project_id=project_id).update(status=0)
        else:
            DevelopmentSubSubType.objects.filter(discipline_subtype_id=developmentsubtype.id,project_id=project_id).update(status=0)
            print("discipline_subtype_id")

def updateothers(envid,project_cluster,projectdevelopment,projectdiscipline,request,disciplineid,project_id,development_id):
    get_othertype=request.POST.getlist('otherdisciplinesubtype-'+str(envid)+'-'+str(project_cluster)+'-'+str(projectdevelopment)+'-'+str(projectdiscipline))
    currentotherypids=[]
    #Surface,drill iterations discipline type etc..
    for disciplinesubype in get_othertype:
        getdisciplinesubtype=ProjectDevelopmentSubType.objects.filter(project_discipline_id=disciplineid.id,discipline_subtype_id=disciplinesubype,project_id=project_id).first()
        if(getdisciplinesubtype==None):
            create_disciplinesubtype=ProjectDevelopmentSubType.objects.create(project_discipline_id=disciplineid.id,discipline_subtype_id=disciplinesubype,project_id=project_id,development_type_id=development_id.id)
            currentotherypids.append(create_disciplinesubtype.id)
        else:
            ProjectDevelopmentSubType.objects.filter(project_discipline_id=disciplineid.id,discipline_subtype_id=disciplinesubype,project_id=project_id).update(status=1)
            currentotherypids.append(getdisciplinesubtype.id)
        get_othertype_act=request.POST.getlist('otherstypeactivities-'+str(envid)+'-'+str(project_cluster)+'-'+str(projectdevelopment)+'-'+str(projectdiscipline)+'-'+str(disciplinesubype))
        discliplinesubtypeid=ProjectDevelopmentSubType.objects.filter(project_discipline_id=disciplineid.id,discipline_subtype_id=disciplinesubype,project_id=project_id,status=1).first()
        current_otheractids=[]
        for othertype_act in get_othertype_act:
            get_other_act=DevelopmentSubTypeSub.objects.filter(sub_subtype_sub_id=othertype_act,other_type_id=discliplinesubtypeid.id,project_id=project_id).first()
            if (get_other_act==None):
                create_otheract=DevelopmentSubTypeSub.objects.create(sub_subtype_sub_id=othertype_act,other_type_id=discliplinesubtypeid.id,project_id=project_id,status=1)
                current_otheractids.append(create_otheract.id)
            else:
                DevelopmentSubTypeSub.objects.filter(sub_subtype_sub_id=othertype_act,other_type_id=discliplinesubtypeid.id,project_id=project_id).update(status=1)
                current_otheractids.append(get_other_act.id)
        DevelopmentSubTypeSub.objects.filter(other_type_id=discliplinesubtypeid.id,project_id=project_id).exclude(id__in=current_otheractids).update(status=0)

    ProjectDevelopmentSubType.objects.filter(project_discipline_id=disciplineid.id,project_id=project_id).exclude(id__in=currentotherypids).update(status=0)   
    excluded_other=ProjectDevelopmentSubType.objects.filter(project_discipline_id=disciplineid.id,project_id=project_id).exclude(id__in=currentotherypids)
    for developmentsubtype in excluded_other:
        DevelopmentSubTypeSub.objects.filter(other_type_id=developmentsubtype.id,project_id=project_id).update(status=0)
        
def Checkfieldenvironmentcluster(request):
    get_fieldenvironment_id=request.GET.get('field_environmentid',None)
    print(get_fieldenvironment_id)
    projectid=request.GET.get('projectid')
    if get_fieldenvironment_id != None: 
        clusters=Clusters.objects.filter(field_environment_id=get_fieldenvironment_id,status=1)
        data=[]
        for cluster in clusters:
            get_clustersub=ClusterSubnames.objects.filter(cluster_id=cluster.id,status=1)
            clusters_list=[]
            for cluster_sub in get_clustersub:
                if ProjectCluster.objects.filter(clustersubname_id=cluster_sub.id,status=1,project_id=projectid).exists():
                    clusters_list.append({'id':cluster_sub.id,'cluster_subname':cluster_sub.cluster_subname,'check':True})
                else:
                    clusters_list.append({'id':cluster_sub.id,'cluster_subname':cluster_sub.cluster_subname,'check':False})
            data.append({cluster.cluster_name:clusters_list})
        print("data",data)
        return JsonResponse({'data':data,'environmentid':get_fieldenvironment_id})

def Checkdevelopmenttype(request):
    get_cluster_id=request.GET.get('clustersub_id')
    projectid=request.GET.get('projectid')
    devlist=[]
    development_type=MasterDevelopmentType.objects.filter(clustersubname_id=get_cluster_id,status=1)
    for develop in development_type:
        if  ProjectDevelopmentType.objects.filter(development_id=develop.id,project_id=projectid,status=1).exists():
            devlist.append({'id':develop.id,'development_type':develop.development_type,'check':True})
        else:
            devlist.append({'id':develop.id,'development_type':develop.development_type,'check':False})
        # cluster=ProjectCluster.objects.filter(clustersubname_id=get_cluster_id,project_id=projectid,status=1)
    print("devlist",devlist)
    return JsonResponse({'data':devlist})

def showmasterproject(request):
    request.session['mainmenu'] = 'masterprojects'
    get_country_id=request.GET['country_id']
    # print(get_country_id)
    projectcreation=list(Projectcreation.objects.filter(company=request.company,status=0).values_list('projectname',flat=True))
    print(projectcreation)
    projects=Projects.objects.filter(company_id=request.company,country_id=get_country_id).exclude(id__in=projectcreation).values()

    return JsonResponse({"projects":list(projects)})

def masterdelete(request): 
    project_id=request.GET["id"]
    print("dfssdf",project_id)
    delete_project_master=Projects.objects.filter(id=project_id)
    project_user_log =  Projects.objects.filter(id=project_id).first()
    create_user=project_user_log.name
    delete_project_master.delete()
    usercreate=request.user.roles_id
    create_user_log(request,create_user,'Project Master','Create','Project Master has been Deleted',usercreate)
    return JsonResponse({'data':'delete'})
    
def Getprojectblock(request):
    get_project_id=request.GET['project_id']
    # print(get_country_id)
    blocks=BlockName.objects.filter(project_id=get_project_id,status=0).values()
    # print(projects)
    return JsonResponse({"projects":list(blocks)})

def Getprojectfield(request):
    get_block_id=request.GET['block_id']
    blocks=FieldName.objects.filter(block_id=get_block_id,status=1).values()
    return JsonResponse({"projects":list(blocks)})

def Getfieldenvironment(request):
    field_environment_id=request.GET['field_id']
    field_environment=FieldEnvironment.objects.filter(field_id=field_environment_id,status=1).values()
    print(field_environment)
    return JsonResponse({'data':list(field_environment)})

def getmastercluster(request):
    get_fieldenvironment_id=request.GET.get('field_environmentid',None)
    print(get_fieldenvironment_id)
    if get_fieldenvironment_id != None: 
        clusters=Clusters.objects.filter(field_environment_id=get_fieldenvironment_id,status=1)
    # print(f"clusters {clusters}")
        data=[]
        for cluster in clusters:
    #     print(f"clusterid {cluster.id}")
            get_clustersub=ClusterSubnames.objects.filter(cluster_id=cluster.id,status=1)
            clusters_list=[]
            for cluster_sub in get_clustersub:
                clusters_list.append({'id':cluster_sub.id,'cluster_subname':cluster_sub.cluster_subname})
            data.append({cluster.cluster_name:clusters_list})
        # print("data",data)
        return JsonResponse({'data':data,'environmentid':get_fieldenvironment_id})

def Wells(request):
    cluster_id=request.GET.get('clusetrid',None)
    if cluster_id != None:
        wells=Well.objects.filter(clustersubname_id=cluster_id,status=1).values()
        return JsonResponse({'data':list(wells)})

def Welltypes(request):
    well_id=request.GET.get('well_id',None)
    if well_id != None:
        welltype=WellSub.objects.filter(well_id=well_id,status=1).values()
    return JsonResponse({'data':list(welltype)})


@login_required(login_url='/')
def projectlist(request):
    markas_read_status(request.get_full_path())
    data={}
    request.session['mainmenu'] = 'main_project'
    request.session['submenu'] = 'projects'
    projects = None
    get_countries = []  
    if (request.user.roles_id == 3):
        try:
            userrights = UserRights.objects.get(user_id=request.user.id, module_id=1)
            data['rights']=userrights
            data['project_user']=UserRights.objects.get(user_id=request.user.id,module_id=16)
            assigned_projects = ProjectUser.objects.filter(user=request.user,status=True).values_list('project_id',flat=True)
            get_countries=Projectcreation.objects.filter(company=request.company,id__in=assigned_projects,status=0).values('country').annotate(dcount=Count('country'))
            projects=Projectcreation.objects.filter(company=request.company,id__in=assigned_projects,status=0).order_by('-id')
        except UserRights.DoesNotExist:
            pass

    else:
        get_countries=Projectcreation.objects.filter(company=request.company,status=0).values('country').annotate(dcount=Count('country'))
        projects=Projectcreation.objects.filter(company=request.company,status=0).order_by('-id')
    projects_count = projects.count() if projects is not None else 0
    # print("aaa",len(get_countries))
    countries=[]
    country=''
    project_name=''
    project_creation=Projectcreation.objects.filter(company=request.company,status=0)
    projects_count=project_creation.count()
    project_country=[]
    project_id=''
    
                
    get_settings = Settings.objects.get_company(request.company).values_list('currency',flat=True).first()
    if get_settings!=None:
        currency = Basecountries.objects.get_by_id(literal_eval(get_settings))
        currency_count=currency.count()
    else:
        currency_count=0
        currency=[]
    
    if (len(get_countries) > 0):
        for country in get_countries:
                country_id=country['country']
                # print("dsf",country)
                print(f'country {country_id}')
                try:
                    countr=Countries.objects.get(id=country_id)
                    countries.append({"country_id":countr.id,"country_name":countr.nicename})
                except Countries.DoesNotExist:
                    print()
    # print(countries)
    if request.method=="POST":
        print(request.POST)
        countval=request.POST['country_name']
        if (countval == ' '):
            country=''
        else:
            country=int(countval)
        project_name=request.POST['project_name']
        projects=Projectcreation.objects.filter(company=request.company,status=0)
        print("s")

        if(country):
            project_country=Projectcreation.objects.filter(country_id=country,company=request.company,status=0)
            projects=projects.filter(country_id=country)
        else:
            projects=projects

        if(project_name!=' '):
            project_id=int(project_name)
            projects=projects.filter(id=project_name)
        else:
            project_id=project_name
            

    
            
    data.update({'projects':projects,
    'projects_count':projects_count,
    'get_countries':countries,
    'post_country':country,
    'post_project':project_id,
    'project_country':project_country,
    'currency_count':currency_count
    # 'projectcreation_status':projectcreation_status,
    })
    # print(data)
    return render(request,"projectlist.html",data)
    

def changeprojectcreation_status(request,pk):

    currentproject=Projectcreation.objects.get(id=pk)
    if (currentproject.projectcreation_status == 0):
        Projectcreation.objects.filter(id=pk).update(projectcreation_status=1)
    else:
        Projectcreation.objects.filter(id=pk).update(projectcreation_status=0)


    return redirect('projects:projectlist') 
    





def getprojectcreation(request):
    try:
        get_country_id=request.GET.get('country_id')
        
        # projects=Projectcreation.objects.filter(country_id=int(get_country_id),company=request.company).values()
        projectdata=Projectcreation.objects.filter(country_id=int(get_country_id),company=request.company)
        print(f'projectdata {projectdata}')
        projectlist=[]
        for project in projectdata:
            projectname=Projects.objects.get(id=project.projectname_id)
            projectlist.append({'id':project.id,'name':projectname.name})
        # print(projectlist)
        try:
            total_project=len(Projectcreation.objects.filter(country_id=int(get_country_id),company=request.company))
            active_project=len(Projectcreation.objects.filter(country_id=int(get_country_id),company=request.company,status=0))
            Inactive_project=len(Projectcreation.objects.filter(country_id=int(get_country_id),company=request.company,status=1))

            project_count = {"total_project":total_project, 'active_project':active_project, 'inactive_project': Inactive_project }
        except:
            project_count = {"total_project": 0, 'active_project': 0, 'inactive_project': 0 }
        
        
        return JsonResponse({'data':projectlist , 'project_count':project_count})
    except:
        projectlist=[]
        project_count = {"total_project": 0, 'active_project': 0, 'inactive_project': 0 }
        
        return JsonResponse({'data':projectlist , 'project_count':project_count})
    
def getprojectvendorls(request):
    
    get_project_id=request.GET.get('project_id')
    
    total_contracts = ContractMaster.objects.filter(projects_id=get_project_id).count()
    active_contracts = ContractMaster.objects.filter(projects_id=get_project_id,status=1).count()
    Inactive_contracts = ContractMaster.objects.filter(projects_id=get_project_id,status=0).count()
    
    contract_count = {"total_contracts":total_contracts, 'active_contracts':active_contracts, 'Inactive_contracts': Inactive_contracts }

    
    
    
    
    return JsonResponse({'contract_count':contract_count })

        

def Projectcreationfield(request):
    projectid=request.GET.get('projectid')
    blockid=request.GET.get('blockid')
    if (blockid == "Not Applicable" or blockid == ""):
        fieldlist=[]
        fielddata=ProjectField.objects.filter(project_id=projectid,status=1)
        for field in fielddata:
            fieldlist.append({'id':field.id,'field':field.field.field_name})
        fieldid=ProjectField.objects.filter(status=1).values_list('id',flat=True)
        getallfieldenvironment=ProjectEnvironment.objects.filter(project_id=projectid,field__id__in=fieldid,status=1).values_list('id',flat=True)
        getallcluster=ProjectCluster.objects.filter(project_id=projectid,environment__id__in=getallfieldenvironment,status=1).values_list('id',flat=True)
        getalldevelopmentype=ProjectDevelopmentType.objects.filter(project_id=projectid,cluster__id__in=getallcluster,status=1).values_list('id',flat=True)
        getalldiscipline=ProjectDevelopmentDiscipline.objects.filter(project_id=projectid,development_type__id__in=getalldevelopmentype,status=1).values_list('id',flat=True)
        getalldisciplinetype=ProjectDevelopmentSubType.objects.filter(project_id=projectid,project_discipline__id__in=getalldiscipline,status=1).values_list('id',flat=True)
        getallwelltype=ProjectWellType.objects.filter(project_id=projectid,discipline_type__id__in=getalldisciplinetype,status=1).values_list('id',flat=True)
        getallwellname=ProjectWellName.objects.filter(project_id=projectid,welltype__id__in=getallwelltype,status=1)
        welllist=[]
        for well in getallwellname:
            welllist.append({'id':well.id,'wellname':well.wellname.well_subname})
        # print('welllist',welllist)
        return JsonResponse({'field':fieldlist,'data':welllist})
    else:
        fields=ProjectField.objects.filter(project_id=projectid,block_id=blockid,status=1)
        print('aq',fields)
        data={}
        fieldlist=[]
        for field in fields:
            projectfield=FieldName.objects.get(id=field.field_id,status=1)
            data={'id':field.id,'field':projectfield.field_name}
            fieldlist.append(data)
        print('fields',fieldlist)
        return JsonResponse({'field':fieldlist})

def Projectcreationfieldenv(request):
    fieldid=request.GET.get('fieldid')
    print("fieldid",fieldid) 
    environment=ProjectEnvironment.objects.filter(field_id=fieldid,status=1)
    data={}
    envlist=[]
    for env in environment:
        field_environment=FieldEnvironment.objects.get(id=env.field_environment_id)
        project_env=field_environment.project_environment
        project_env_subtype=field_environment.project_environment_subtype
        data={'id':env.id,'environment':project_env+'-'+project_env_subtype}
        envlist.append(data)
    return JsonResponse({'data':envlist})

def Projectcreationcluster(request):
    fieldenvid=request.GET.get('fieldenvid')
    clusters=ProjectCluster.objects.filter(environment_id=fieldenvid,status=1)
    data={}
    clusterlist=[]
    for cluster in clusters:
        master_cluster=ClusterSubnames.objects.get(id=cluster.clustersubname_id)
        data={'id':cluster.id,'cluster':master_cluster.cluster_subname}
        clusterlist.append(data)
    return JsonResponse({'data':clusterlist})

def Projectdevelopment(request):
    clusterid=request.GET.get('clusterid')
    development=ProjectDevelopmentType.objects.filter(cluster_id=clusterid,status=1)
    data={}
    develop_list=[]
    for develop in development:
        develop_type=MasterDevelopmentType.objects.get(id=develop.development_id)
        replace_dev=(develop_type.development_type).replace("_"," ")
        data={'id':develop.id,'development':replace_dev}
        develop_list.append(data)
    print(develop_list)
    return JsonResponse({'data':develop_list})

def Projectdiscipline(request):
    devid=request.GET.get('devid')
    projectdiscipline=ProjectDevelopmentDiscipline.objects.filter(development_type_id=devid,status=1)
    print("pd",projectdiscipline)
    data={}
    disciplinelist=[]
    for disciplines in projectdiscipline:
        discipline=ProjectDiscipline.objects.get(id=disciplines.project_discipline)
        data={'id':disciplines.id,'discipline':discipline.name}
        disciplinelist.append(data)
    return JsonResponse ({'data':disciplinelist})


def Disciplinetype(request):
    disciplineid=request.GET.get('disciplineid')
    disciplinetypes=ProjectDevelopmentSubType.objects.filter(project_discipline_id=disciplineid,status=1)
    data={}
    disciplinetypelist=[]
    for disciplinetype in disciplinetypes:
        discipline_type=DisciplineSubtype.objects.get(id=disciplinetype.discipline_subtype_id)
        data={'id':disciplinetype.id,'disciplinetype':discipline_type.discipline_subtype}
        disciplinetypelist.append(data)
    return JsonResponse ({'data':disciplinetypelist})

def devdisciplinesubtype(request):
    disciplinetypeid=request.GET.get('typeid')
    # print("sub id",disciplinetypeid)
    welltypes=ProjectWellType.objects.filter(discipline_type_id=disciplinetypeid,status=1)
    data={}
    datalist=[]
    datas={}
    if len(welltypes) > 0:
        checkwelltype=ProjectWellType.objects.filter(discipline_type_id=disciplinetypeid,status=1).first()
        if (checkwelltype != None):
            project_discipline_id=checkwelltype.project_discipline.project_discipline
            discipline=ProjectDiscipline.objects.get(id=project_discipline_id)
            discipline_name=discipline.name
            # print("name",discipline_name)
            if discipline_name =="Green Field Development":
                for welltype in welltypes:
                    well=Well.objects.get(id=welltype.welltype_id)
                    # print(well)
                    data={'id':welltype.id,'welltype':well.well_name}
                    datalist.append(data)
                datas={'well':datalist,"project_discipline":discipline_name}
            else:
                brownwelllist=[]
                browndatas={}
                subtypes=DevelopmentSubSubType.objects.filter(discipline_subtype_id=disciplinetypeid,status=1)
                for subtype in subtypes:
                    sub_types=DisciplineSubSubTypes.objects.get(id=subtype.disciplinesub_subtype_id)
                    data={'id':subtype.id,'subtype':sub_types.sub_subtype_name}
                    datalist.append(data)
                    # print("datalist",datalist)
                datas={'data':datalist,"project_discipline":discipline_name}
                if len(welltypes) > 0:
                    for welltype in welltypes:
                        well=Well.objects.get(id=welltype.welltype_id)
                        browndatas={'id':welltype.id,'welltype':well.well_name}
                        brownwelllist.append(browndatas)
                        print("well have")

                    datas.update({'brownwell':brownwelllist})    
                    print("brownwelllist",brownwelllist)
                    print("all",datas)
                else:
                    print("no wells")
        else:
            print("no data")
            # for welltype in welltypes:
            #     well=Well.objects.get(id=welltype.welltype_id)
            #     print(well)
            #     data={'id':welltype.id,'welltype':well.well_name}
            #     brownwelllist.append(data)
            # datas={'':,'well':brownwelllist,"project_discipline":discipline_name}
        
    else:
        get_disciplinetype=ProjectDevelopmentSubType.objects.filter(id=disciplinetypeid,status=1).first()
        print("sads",get_disciplinetype)
        subtypes=DevelopmentSubSubType.objects.filter(discipline_subtype_id=disciplinetypeid,status=1)
        print("subtypes",subtypes)
        for subtype in subtypes:
            sub_types=DisciplineSubSubTypes.objects.get(id=subtype.disciplinesub_subtype_id)
            data={'id':subtype.id,'subtype':sub_types.sub_subtype_name}
            datalist.append(data)
        datas={'data':datalist,'disciplinetype':get_disciplinetype.discipline_subtype_id}
    return JsonResponse(datas)


def Projectwellname(request):
    welltypeid=request.GET.get('welltypeid')
    well_names=ProjectWellName.objects.filter(welltype_id=welltypeid,status=1)
    data={}
    wellnamelist=[]
    for well_name in well_names:
        well_sub=WellSub.objects.get(id=well_name.wellname_id)
        data={'id':well_name.id,'wellname':well_sub.well_subname}
        wellnamelist.append(data)
    return JsonResponse({'data':wellnamelist})


def Wellnameactivities(request):
    wellnameid=request.GET.get('wellnameid')
    well_activities=DevelopmentSubTypeSub.objects.filter(wellname_id=wellnameid,status=1)
    data={}
    wellactlist=[]
    for wellactivities in well_activities:
        subdiv=Subtypes.objects.get(id=wellactivities.sub_subtype_sub_id)
        data={'id':wellactivities.id,'wellactivities':subdiv.name}
        wellactlist.append(data)
    return JsonResponse({'data':wellactlist})

def subtypeactivities(request):
    discipline_subtypeid=request.GET.get('discipline_subtypeid')
    disname=request.GET.get('disname',None)
    data={}
    subdivlist=[]
    print("disname",disname)
    if disname == "Others":
        disciplinesubdiv=DevelopmentSubTypeSub.objects.filter(other_type_id=discipline_subtypeid,status=1)
        for discipline_subdiv in disciplinesubdiv:
            subdiv=Subtypes.objects.get(id=discipline_subdiv.sub_subtype_sub_id)
            data={'id':discipline_subdiv.id,'activities':subdiv.name}
            subdivlist.append(data)
    else:
        disciplinesubdiv=DevelopmentSubTypeSub.objects.filter(disciplinesub_subtype_id=discipline_subtypeid,status=1)
        for discipline_subdiv in disciplinesubdiv:
            subdiv=Subtypes.objects.get(id=discipline_subdiv.sub_subtype_sub_id)
            data={'id':discipline_subdiv.id,'activities':subdiv.name}
            subdivlist.append(data)
    return JsonResponse({'data':subdivlist})

def view(request,pk):
    request.session['mainmenu'] = 'projects'
    project=Projects.objects.get(pk=pk)
    projects_subtype=DisciplineSubtypelist.objects.filter(project=pk)
    return render(request, "view.html", {'project':project,"projects_subtype":projects_subtype})

def edit(request,pk):
    request.session['mainmenu'] = 'projects'
    get_company=Companies.objects.get(id=request.user.company_id)
    project=Projects.objects.get(pk=pk)
    country=Countries.objects.all()
    project_discipline=ProjectDiscipline.objects.all()
    get_checkbox=DisciplineSubtypelist.objects.filter(project=project.id)
    if(request.POST):
        form=ProjectForm(request.POST,instance=project)
        print(request.POST)
        if form.is_valid():
            forms=form.save(commit=False)
            forms.company=get_company
            get_environment=forms.project_environment.split("-")
            project_environment=get_environment[0]
            subtype_environment=get_environment[1]
            forms.project_environment=project_environment
            forms.project_environment_subtype=subtype_environment
            forms.status=0
            forms.save()
            DisciplineSubtypelist.objects.filter(project=project).delete()
            get_discipline_subtype=request.POST.getlist('project_dicipline_subtype')
            for get_subtype in get_discipline_subtype:
                discipline_subtype_form=DisciplineSubtypelist.objects.create(project=project,discipline_subtype_id=get_subtype,status=0)
            return redirect('projects:projectlist')  
        else:
            form=ProjectForm()
    form=ProjectForm()
    return render(request, "edit.html", {'project':project,"countries":country,"project_discipline":project_discipline,"get_checkbox":get_checkbox})

def delete(request):
    request.session['mainmenu'] = 'projects'
    get_delete_id=request.GET.get("id")
    print("print",get_delete_id)
    Projectcreation.objects.filter(id=get_delete_id).update(status=1)
    return JsonResponse({'delete':'success'})


def lock_project(request,pk):
    projects_status=Projects.objects.get(pk=pk)
    if projects_status.status == 0:
        projects_status=Projects.objects.filter(pk=pk).update(status=1)
    else:
        projects_status=Projects.objects.filter(pk=pk).update(status=0)
    print(projects_status)
    return redirect("projects:projectlist")

def check_email(request):
    get_email=request.GET.get('id')
    print(get_email)
    # email_id=User.objects.filter(email=get_email)
    # print(email_id)
    return JsonResponse({'success':"success"})

def sample(request):
    return render(request,"sample.html") 

def discipline_sub_subtype(request):
    get_subtype_id=request.GET.get("id")
    get_sub_name=ProjectDisciplineSubtype.objects.get(id=get_subtype_id)
    sub_type_data=DisciplineSubSubtype.objects.filter(discipline_subtype=get_subtype_id).values()
    return JsonResponse({'data':list(sub_type_data),'name':get_sub_name.name})

#main
def getdisciplinesub_subtypes(request):
    ids=request.GET.get("id")
    get_sub_subtypes=DisciplineSubSubTypes.objects.filter(discipline_subtype_id=ids).values()
    # print("get_subtypes",get_subtypes)
    return JsonResponse({"data":list(get_sub_subtypes)})

#main
def masterfieldcluster(request):
    clusters=[]
    cluster_field=[]
    get_fields=request.POST.getlist('field[]',None)
    # print(request.POST)
    if len(get_fields) != 0:
        for field in get_fields:
            get_clusters=Clusters.objects.get(field_id=field,status=1)
            clustersub=[]
            get_clustersub=ClusterSubnames.objects.filter(cluster=get_clusters,status=1)
            for cluster_sub in get_clustersub:
                clustersub.append({'id':cluster_sub.id,'clustersubname':cluster_sub.cluster_subname})
            clusters.append({get_clusters.cluster_name:clustersub})
        print(clusters)
    return JsonResponse({'data':clusters})

#main
def Wellslist(request):
    wellsname=request.POST.getlist('wells[]',None)
    print(request.POST)
    return JsonResponse({'data':'success'})

def masterclusterwell(request):
    get_clustersub_id=request.GET.get('clustersub_id',None)
    get_well_name=request.GET.get('wellname',None)
    print(get_well_name)
    get_all_subwell=''
    data={}
    if Well.objects.filter(clustersubname_id=get_clustersub_id,well_name=get_well_name,status=1).exists():
        get_wells=Well.objects.get(clustersubname_id=get_clustersub_id,well_name=get_well_name,status=1)
        get_all_subwell=WellSub.objects.filter(well=get_wells,status=1).values()
        data={"data":list(get_all_subwell)}
        print("yes")

    else:
        data={'no_wells':'no wells in this cluster'}
    return JsonResponse(data)


#main
def get_types_subtype(request):
    ids=request.GET.get("id")
    # get_discipline_subtype_name=DisciplineSubtype.objects.get(id=ids)
    get_subtypes=Subtypes.objects.filter(discipline_subtype_id=ids).values()
    return JsonResponse({"data":list(get_subtypes)})

def ajax_posting(request):
    if request.is_ajax():
        well_name = request.GET.get('well', None) 
        print(well_name)
        # create_well=Well.objects.create(well_name=well_name) 
        return JsonResponse({"data":"success"})

## master project ##
def MasterProject(request):
    
    if request.method=="POST":
        project_name=request.POST['name']
        block_name=request.POST['block_name']
        field_name=request.POST['field_name']
        create_project=Projects.objects.create(name=project_name,block_name=block_name,field_name=field_name)
        get_last_project=Projects.objects.last()
        print(get_last_project)
        get_wells=request.POST.getlist("well_name")
        print(get_wells)
        for wells in get_wells:
            create_wells=Well.objects.create(well_name=wells,project=get_last_project) 
        return redirect("projects:list_master")

    else:
        form=ProjectForm()
    return render(request,"masterproject.html",{"form":form})



## project details ##

def getprojectdetails(request):
    get_id=request.GET.get("project_id")
    projects_details=Projects.objects.filter(id=get_id).values()
    print(projects_details)
    # get_wells=Well.objects.filter(project_id=get_id).values()
    # print(get_wells)
    return JsonResponse({"data":list(projects_details)})
#main
def getprojectdiscipline(request):
    get_id=request.GET.get("id")
    get_project_discipline=DisciplineSubtype.objects.getdiscipline_bydevelopment(get_id)

    return JsonResponse({"data":list(get_project_discipline)})


def getprojectdiscipline_and_costcode(request):
    discipline_id=request.GET.get("discipline_id")  
    development_id=request.GET.get("development_id")
    get_project_discipline=DisciplineSubtype.objects.getdiscipline_bydevelopment(discipline_id)
    level1=CostCodeMaster.objects.getfirst_level(request.company,1)
    check_costcode=CostCodeMain.objects.getcostcode_bydevelopment_discipline(discipline_id,development_id,request.company)
    check_costcode_main=CostCodeMain.objects.getallcostcode(request.company).count()
    getlevel2=CostCodeMaster.objects.filter_by_level_id(request.company,2)

    if(check_costcode_main==0):
        cost_code=level1.start_with
    else:
        if(check_costcode):
            cost_code=check_costcode.level1_cost_code
        else:
            costcode_main=CostCodeMain.objects.getallcostcode(request.company).order_by('-id').first()
            cost_code_add=int(costcode_main.level1_cost_code)+int(level1.sequence_gap)
            cost_code=''
            cost_code_len=len(str(cost_code_add))
            if(int(cost_code_len)<=int(level1.no_digits)):
                remaining_length=int(level1.no_digits)-int(cost_code_len)
                for k in range(remaining_length):
                    cost_code +='0'
                cost_code +=str(cost_code_add)
            else:
                cost_code = cost_code_add
        print('cost_code',cost_code)
    return JsonResponse({"data":list(get_project_discipline),"cost_code":cost_code,'start_with':getlevel2.start_with,'sequence':getlevel2.sequence_gap})



def viewmaster(request,pk):
    get_project=Projects.objects.get(pk=pk)
    # project_well=Well.objects.filter(project=pk)
    return render(request,"masterview.html",{"get_project":get_project,"project_well":"project_well"})

def createmaster(request):
    # request.session['mainmenu'] = 'masterprojects'
    form=ProjectForm(request.POST)

    return render(request,"createmasterproject.html",{"form":form})

#Ajax for master data post
def postmasterproject(request):
    data={}
    if (request.POST):
        get_form_data=request.POST
        get_country=get_form_data['country']
        get_project_name=get_form_data['name']
        print(get_form_data)
        data={}
        if Projects.objects.filter(name__iexact=get_project_name).exists():
            data={'is_taken':'project name already exists'}   
        else:
            print("yes")
            company=User.objects.get(id=request.user.id)
            current_company=company.company.id
            create_project=Projects.objects.create(name=get_project_name,country_id=get_country,project_status=0,company_id=current_company)
            get_created_project=Projects.objects.filter(name=get_project_name).values()
            data={'project':list(get_created_project)}
            print(data)
        # print(get_project_name)
    return JsonResponse(data)
#main
def masterprojectcreate(request):
    project_name=request.POST.get('name')
    country_id=request.POST.get('country')
    result={}
    if Projects.objects.filter(name__exact=project_name,company=request.company).exists():
        return JsonResponse({"duplicate":"duplicate"})
    else:
        try:
            project=Projects.objects.create(name=project_name,country_id=country_id,project_status=1,company=request.company)
            result["status"]="success"
            print('if')
            result["project_id"]=project.pk
        except:
            result["status"]="failure"
            print('else')
        return JsonResponse(result)

def createmasterblock(request,project_id):
    project_details=Projects.objects.filter(id=project_id).first()
    if(request.POST):
        result={}
        block_name=request.POST.getlist('block_name')
        for block in block_name:
            print(block)
            create_block=BlockName.objects.create(block_name=block,project_id=project_id)
        result["status"]="success"
        return JsonResponse(result)
        

    return render(request, "createmasterblock.html", {"project_details":project_details})


def createmasterfield(request,project_id):
    project_details=Projects.objects.filter(id=project_id).first()
    project_block=BlockName.objects.filter(project_id=project_id)
    if(request.POST):
        print(request.POST)
        result={}
        blockname=request.POST.getlist('block_name')
        for block in blockname:
            fieldname=request.POST.getlist('field_name_block'+str(block))
            # print(f"fieldname {fieldname}")
            for index,field in enumerate(fieldname):
                if(field):
                    create_field=FieldName.objects.create(field_name=field,block_id=block,project_id=project_id)
                    project_environment=request.POST.getlist('project_environment-'+str(block)+'-'+str(index))
                    for environment in project_environment:
                        if environment !='Others':
                            split_environment=environment.split("-")
                            environment_type=split_environment[0]
                            environment_subtype=split_environment[1]
                            create_field_environment=FieldEnvironment.objects.create(field=create_field,project_id=project_id,project_environment=environment_type,project_environment_subtype=environment_subtype,status=1)
                        else:
                            field_environment=request.POST.getlist('environment-'+str(block)+'-'+str(environment)+'-'+str(index))
                            for other_environment in field_environment:
                                print("environment oth",other_environment)
                                create_field_environment=FieldEnvironment.objects.create(field=create_field,project_id=project_id,project_environment=environment,project_environment_subtype=other_environment,status=1)
        result["status"]="success"
        return JsonResponse(result)
    data={
        "project_details":project_details,
        'project_block':project_block

    }
    return render(request, "createmasterfield.html", data)


def createmastercluster(request,project_id):
    project_details=Projects.objects.filter(id=project_id).first()
    project_block=BlockName.objects.filter(project_id=project_id)
    clustertype=dict()
    clustertype["Well Pad"]="wellpad"
    clustertype["Well Platform"]="wellplatform"
    clustertype["Subsea Clusters"]="subseaclusters"
    if(request.POST):
        result={}
        environmentlist=request.POST.getlist('environment_name')
        for environment in environmentlist:
            cluster_type=request.POST.getlist('cluster_type'+str(environment))
            for cluster in cluster_type:
                create_cluster=Clusters.objects.create(cluster_name=cluster,field_environment_id=environment,project_id=project_id)
                cluster_given_name=clustertype[cluster]
                cluster_name=request.POST.getlist('clustername_'+str(cluster_given_name)+str(environment))
                if(len(cluster_name)>0):
                    for clustername in cluster_name:
                        create_clustersub=ClusterSubnames.objects.create(cluster_subname=clustername,cluster_id=create_cluster.id,project_id=project_id)
        result["status"]="success"
        return JsonResponse(result)

    data={
        "project_details":project_details,
        'project_block':project_block

    }
    return render(request, "createmastercluster.html", data)

def createmasterwell(request,project_id): 
    project_details=Projects.objects.filter(id=project_id).first()

    project_block=BlockName.objects.filter(project_id=project_id)
    well_type=dict()
    well_type["Exploration Wells"]="exploration"
    well_type["Appraisal Wells"]="appraisal"
    well_type["Development Wells"]="development"
    well_type["Pilot Holes"]="pilot"
    well_type["Workovers/Well Intervention"]="workovers"
    well_type["Infill Wells"]="infill"
    if(request.POST):
        result={}
        print("req",request.POST)
        cluster_name=request.POST.getlist('cluster_name')
        for cluster in cluster_name:
            print(cluster)
            development_type=request.POST.getlist('development-'+str(cluster))
            for development in development_type:
                print(development)
                create_development_type=MasterDevelopmentType.objects.create(project_id=project_id,clustersubname_id=cluster,development_type=development)
                welltype=request.POST.getlist('well_type'+str(cluster)+'-'+str(development))
                for well in welltype:
                    print(well)  
                    create_well=Well.objects.create(well_name=well,project_id=project_id,development_type_id=create_development_type.id)
                    well_given_name=well_type[well]
                    well_name=request.POST.getlist('wellname_'+str(well_given_name)+str(cluster)+'-'+str(development))
                    for name in well_name:
                        print(name)
                        create_subwell=WellSub.objects.create(well_subname=name,project_id=project_id,well_id=create_well.id)
                        print(create_subwell,create_subwell.id,'gggg...........')
                        usercreate=request.user.roles_id
                        create_user=project_details.name
                        create_user_log(request,create_user,'Project Master','Create','Project Master Created',usercreate)
        result["status"]="success"
        return JsonResponse(result)

    data={
        "project_details":project_details,
        'project_block':project_block

    }
    return render(request, "createmasterwell.html", data)
    

# master project list#
def test(request):
    company=User.objects.get(id=request.user.id)
    current_company=company.company.id
    country_count=Projects.objects.values('country').annotate(dcount=Count('country'))
    countries=[]
    for country in country_count:
        country_id=country['country']
        get_countries=Countries.objects.get(id=country_id) 
        countries.append({"country_id":get_countries.id,"country_name":get_countries.name})
    return render(request, "masterlist.html", {"countries":countries})



def blockcreate(request,pk):
    get_project=Projects.objects.get(id=pk)
    # print(get_project)
    return render(request,"createblock.html",{"get_project":get_project})

def Createblock(request):
    print(request.POST)
    get_project_id=request.POST['project_id']
    get_block_name=request.POST.getlist('block_name')
    for block in get_block_name:
        create_block=BlockName.objects.create(block_name=block,project_id=int(get_project_id))

    return JsonResponse({"data":"block created"})

def fieldcreate(request,pk):
    # print(pk)
    get_project=Projects.objects.get(id=pk)
    get_block=BlockName.objects.filter(project_id=pk)
    return render(request,"createfield.html",{'project':get_project,'blocks':get_block})

def Createfield(request):
    print(request.POST)
    get_project_id=request.POST['project_id']
    get_checked_block=request.POST.getlist('block_name_check')
    for checked_block in get_checked_block:
        get_block=BlockName.objects.get(block_name=checked_block,project_id=int(get_project_id))
        update_block=BlockName.objects.filter(block_name=checked_block,project_id=int(get_project_id)).update(status=1)
        if get_block:
            get_field=request.POST.getlist('field_name'+str(get_block.id))
            for field in get_field:
                create_field=FieldName.objects.create(field_name=field,block_id=get_block.id,project_id=int(get_project_id),status=0)
    return JsonResponse({'data':'field create'})

def clustercreate(request,pk):
    get_project=Projects.objects.get(id=pk)
    get_block=BlockName.objects.filter(project_id=pk)
    get_field=FieldName.objects.filter(project_id=pk)
    return render(request,"createcluster.html",{"project":get_project,"blocks":get_block,"fields":get_field})

def Createcluster(request):
    print(request.POST)
    get_project_id=request.POST['project_id']
    get_checked_field=request.POST.getlist('field_checkbox')
    for checked_field in get_checked_field:
        get_field_instance=FieldName.objects.get(field_name=checked_field,project=int(get_project_id))
        update_field=FieldName.objects.filter(field_name=checked_field,project=int(get_project_id)).update(status=1)
        get_clusters=request.POST.getlist('field_subtype'+str(get_field_instance.id))
        for clusters in get_clusters:
            # print("cluster",clusters)
            create_cluster=Clusters.objects.create(cluster_name=clusters,field_id=get_field_instance.id,project_id=int(get_project_id),status=0)
            get_cluster_sub=request.POST.getlist('fieldsub_name-'+str(create_cluster.cluster_name)+str(get_field_instance.id))
            for clustersub in get_cluster_sub:
                create_clustersub=ClusterSubnames.objects.create(cluster_subname=clustersub,cluster_id=create_cluster.id,project_id=int(get_project_id),status=0)
            # print("subname",get_cluster_sub)
    return JsonResponse({'data':'cluster create'})

def wellcreate(request,pk):
    get_project=Projects.objects.get(id=pk)
    get_block=BlockName.objects.filter(project_id=pk)
    get_field=FieldName.objects.filter(project_id=pk)
    get_clustersub=ClusterSubnames.objects.filter(project_id=pk)
    return render(request,"wellcreate.html",{"project":get_project,"blocks":get_block,"fields":get_field,"clustersub":get_clustersub})
    
def Createwell(request):
    get_project_id=request.POST['project_id']
    get_checked_cluster=request.POST.getlist('sub_checkbox')
    print(request.POST)
    for checked_cluster in get_checked_cluster:
        get_clustersub=ClusterSubnames.objects.get(cluster_subname=checked_cluster,project_id=int(get_project_id))
        # print(get_clustersub)
        get_wells=request.POST.getlist('wells'+str(get_clustersub.cluster_subname).replace(" ","_"))
        for wells in get_wells:
            # print(wells)
            create_well=Well.objects.create(well_name=wells,project_id=int(get_project_id),clustersubname_id=get_clustersub.id)
            get_sub_wells=request.POST.getlist('wellsub_name-'+str(get_clustersub.cluster_subname).replace(" ","_")+'-'+wells)
            for subwell in get_sub_wells:
                # print(subwell)
                create_subwell=WellSub.objects.create(well_subname=subwell,project_id=int(get_project_id),well_id=create_well.id)
    return JsonResponse({"data":"well created"})

def editmasterproject(request,pk):
    get_project=Projects.objects.get(id=pk)
    country=Countries.objects.all()
    if(request.POST):
        result={}
        print(request.POST)
        project=Projects.objects.filter(id=pk).update(name=request.POST.get('name'),country_id=request.POST.get('country'))
        result["status"]="success"
        return JsonResponse(result)

    return render(request,"editmasterproject.html",{"project":get_project,"countries":country})

def Editmasterproject(request):
    print(request.POST)
    projectid=request.POST['project_id']
    get_country=request.POST['country']
    get_project_name=request.POST['name']
    convert_projectid=int(projectid)
    project=Projects.objects.filter(id=convert_projectid).update(name=get_project_name,country_id=get_country)
    return JsonResponse({'data':convert_projectid})

def editmasterblock(request,project_id):
    project_details=Projects.objects.filter(id=project_id).first()
    project_block=BlockName.objects.filter(project_id=project_id,status=0)
    if(request.POST):
        result={}
        print(request.POST)
        block_id=request.POST.getlist('blockid')
        block_name=request.POST.getlist('blockname')
        i=0
        current_blockids=[]
        while i<len(block_id):
            if(block_id[i]):
                 block=BlockName.objects.filter(id=block_id[i]).update(block_name=block_name[i])
                 current_blockids.append(block_id[i])
            else:
                create_block=BlockName.objects.create(block_name=block_name[i],project_id=project_id)
                current_blockids.append(create_block.id)
            i +=1
        # print("current_blockids",current_blockids)
        BlockName.objects.filter(project_id=project_id).exclude(id__in=current_blockids).update(status=1)
        result["status"]="success"
        usercreate=request.user.roles_id
        create_user=project_details.name
        create_user_log(request,create_user,'Project Master','Create','Project Master Block has been Edited',usercreate)
        return JsonResponse(result)

    data={
        'project_details':project_details,
        'project_block':project_block
    }
    return render(request,"editmasterblock.html",data)

def checkblockexists(request):
    get_block_id=request.GET.get("block_id", None)
    print('get_block_id',get_block_id)
    data={}
    if get_block_id !='':
        blocks=FieldName.objects.filter(block_id=get_block_id,status=1)
        print('blocks',len(blocks))
        if len(blocks)>0:
            data={'data':'exists'}
        else:
            data={'data':'no_fields'}
    else:
        data={'data':'no_block'}
    return JsonResponse(data)
    

def checkfieldedexists(request):
    get_field_id=request.GET.get("field_id")
    data={}
    if get_field_id !='':
        fields=FieldEnvironment.objects.filter(field_id=get_field_id,status=1)

        print('fields',fields)
        if len(fields)>0:
            data={'data':'exists'}
        else:
            data={'data':'no_fieldenvironment'}
    else:
        data={'data':'no_field'}
    return JsonResponse(data)


def checkfieldexists(request):
    get_fieldenv_id=request.GET.get("field_env_id")
    data={}
    if get_fieldenv_id != 'None':
        fields=Clusters.objects.filter(field_environment_id=get_fieldenv_id)
        if len(fields)>0:
            data={'data':'exists'}
        else:
            data={'data':'no_clusters'}
    else:
        data={'data':'no_field_env'}
    return JsonResponse(data)

def checkfieldenvexists(request):
    get_fieldenv_id=request.GET.get("other_field_env_id",None)
    print("sdf",get_fieldenv_id)
    data={}
    if  get_fieldenv_id != None:
        if Clusters.objects.filter(field_environment_id=get_fieldenv_id,status=1).exists():
            data={'data':'exists'}
        else:
            data={'data':'no_clusters'}
    return JsonResponse(data)

def checkfenvexist(request):
    get_field_envid=request.GET.get('field_env_id',None)
    print(type(get_field_envid))
    get_field_id=request.GET.get('fieldid',None)
    data={}
    if (get_field_envid != 'None' and get_field_envid != None):
        if Clusters.objects.filter(field_environment_id=get_field_envid,status=1).exists():
            data={'data':'exists'}
        else:
            data={'data':'no_clusters'}
    else:
        data={'data':'no_data'}
    print(data)
    return JsonResponse(data)

def checkclusterexists(request):
    get_cluster_id=request.GET.get("cluster_id", None)
    print('get_cluster_id',get_cluster_id)
    data={}
    if get_cluster_id !='':
        clusters=MasterDevelopmentType.objects.filter(clustersubname__field_environment=get_cluster_id)
        print('clusters',len(clusters))
        if len(clusters)>0:
            data={'data':'exists'}
        else:
            data={'data':'no_well'}
    else:           
        data={'data':'no_cluster'}
    return JsonResponse(data)


def welldupcheck(request):
    well_id=request.GET.get("well_id", None)
    projectid=request.GET.get("projectid", None)
    wellcheckid=request.GET.get("wellcheckid", None)
    print("sadass",len(wellcheckid))
    
    data={}
    if (wellcheckid != ''):
        print("whid",wellcheckid)
        data={'data':'status change'}
    else:
        print("whid",wellcheckid)
        if ProjectWellName.objects.filter(wellname_id=well_id,project_id=projectid,status=1).exists():
            data={'data':'exists','status':False}
        else:
            data={'data':'no_data','status':True}
    # else:
    #     ProjectWellName.objects.filter(wellname_id=well_id,project_id=projectid).update(status=0)
    return JsonResponse(data)


def checkwellexists(request):
    get_well_id=request.GET.get("well_id", None)
    data={}
    well=WellSub.objects.filter(well_id=get_well_id)
    if len(well)>0:
        data={'data':'exists'}
    else:
        data={'data':'no_wellsub'}
    return JsonResponse(data)



def editmasterfield(request,project_id):
    project_details=Projects.objects.filter(id=project_id).first()
    project_block=BlockName.objects.filter(project_id=project_id,status=0)
    if(request.POST):
        result={}
        blockname=request.POST.getlist('block_name')
        for block in blockname:
            fieldname=request.POST.getlist('fieldname'+str(block))
            fieldid=request.POST.getlist('fieldid'+str(block))

            i=0
            currentfieldids=[]
            while i<len(fieldid):
                if(fieldid[i]):
                    FieldName.objects.filter(id=fieldid[i]).update(field_name=fieldname[i])
                    currentfieldids.append(fieldid[i])
                else:
                    create_field=FieldName.objects.create(field_name=fieldname[i],block_id=block,project_id=project_id)
                    # currentfieldids.append(create_field.id)
                    currentfieldids.append(create_field.id)
                i +=1
            # print("currentfieldids",currentfieldids)

            FieldName.objects.filter(block_id=block,project_id=project_id).exclude(id__in=currentfieldids).update(status=0)
            for i,j in enumerate(currentfieldids):
                field_environment=request.POST.getlist('project_environment-'+str(block)+'-'+str(i))
                currentenvironmentids=[]
                field_environmentsubtypes = []    
                for environment in field_environment:
                    if "-" in environment:
                        spiltenvironment=environment.split("-")
                        field_environment=spiltenvironment[0]
                        field_environmentsubtype=spiltenvironment[1]
                        environment_field=FieldEnvironment.objects.filter(field_id=j,project_environment=field_environment,project_environment_subtype=field_environmentsubtype)
                        if environment_field.exists():
                            currentenvironmentids.append(environment_field[0].id)
                            # print("env exists")
                        else:
                            create_field_environment=FieldEnvironment.objects.create(field_id=j,project_environment=field_environment,project_environment_subtype=field_environmentsubtype,status=1,project_id=project_id)
                            currentenvironmentids.append(create_field_environment.id)
                    else:
                        get_others_environment=request.POST.getlist('environment-'+str(block)+'-'+str(environment)+'-'+str(i))
                        environment_id=request.POST.getlist('environmentid-'+str(block)+'-'+str(environment)+'-'+str(i))
                        print(f'get_others_environment {get_others_environment} environment_id {environment_id}')
                        i=0
                        while i<len(get_others_environment):
                            try:
                                if (environment_id[i] !=' '):
                                    update_fieldenvironment=FieldEnvironment.objects.filter(id=environment_id[i]).update(project_environment_subtype=get_others_environment[i])
                                    currentenvironmentids.append(environment_id[i])
                                else:
                                    create_environment=FieldEnvironment.objects.create(field_id=j,project_environment=environment,project_environment_subtype=get_others_environment[i],project_id=project_id,status=1)
                                    currentenvironmentids.append(create_environment.id)
                            except Exception as e:  
                                print(e)                       
                                create_environment=FieldEnvironment.objects.create(field_id=j,project_environment=environment,project_environment_subtype=get_others_environment[i],project_id=project_id,status=1)
                                currentenvironmentids.append(create_environment.id)
                            i +=1
                FieldEnvironment.objects.filter(field_id=j).exclude(id__in=currentenvironmentids).update(status=0)
        FieldName.objects.filter(project_id=project_id).exclude(block_id__in=blockname).update(status=0)
        result["status"]="success"
        usercreate=request.user.roles_id
        create_user=project_details.name
        create_user_log(request,create_user,'Project Master','Edit','Project Master Field has been Edited',usercreate)
        return JsonResponse(result)
    data={
        "project_details":project_details,
        'project_block':project_block

    }
    print("data",data)
    return render(request, "editmasterfield.html", data)

def editmastercluster(request,project_id):
    project_details=Projects.objects.filter(id=project_id).first()
    project_block=BlockName.objects.filter(project_id=project_id,status=0)
    clustertype=dict()
    clustertype["Well Pad"]="wellpad"
    clustertype["Well Platform"]="wellplatform"
    clustertype["Subsea Clusters"]="subseaclusters"
    if(request.POST):
        result={}
        # print(request.POST)
        environmentlist=request.POST.getlist('environment_name')
        # print(environmentlist,'envlist')
        currentfieldids=[]
        for environment in environmentlist:
            cluster_type=request.POST.getlist('cluster_type'+str(environment))
            print('cluster_type',cluster_type)
            # print(cluster_type,'cluster_type')
            currentfieldids.append(environment)
            # print(f"environment{environment}")
            current_cluster_ids=[]
            for cluster in cluster_type:
                checkcluster=Clusters.objects.filter(field_environment_id=environment,cluster_name=cluster,status=1).first()
                print(f"checkcluster {checkcluster}")
                if(checkcluster != None):
                    cluster_id=checkcluster.id
                    current_cluster_ids.append(cluster_id)
                else:
                   create_cluster=Clusters.objects.create(cluster_name=cluster,field_environment_id=environment,project_id=project_id,status=1) 
                   cluster_id=create_cluster.id
                   current_cluster_ids.append(cluster_id)
                cluster_given_name=clustertype[cluster]
                cluster_name=request.POST.getlist('clustername_'+str(cluster_given_name)+str(environment))
                cluster_subid=request.POST.getlist('clusterid_'+str(cluster_given_name)+str(environment))
                i=0
                print('list',current_cluster_ids)
                currentclustersubids=[]
                while i<len(cluster_subid):
                    if(cluster_subid[i]):
                        ClusterSubnames.objects.filter(id=cluster_subid[i]).update(cluster_subname=cluster_name[i],status=1)
                        currentclustersubids.append(cluster_subid[i])
                      
                    else:
                        create_clustersub=ClusterSubnames.objects.create(cluster_subname=cluster_name[i],cluster_id=cluster_id,project_id=project_id)
                        currentclustersubids.append(create_clustersub.id)
                    i +=1
                print('li2',)
                # print('clustersubname',ClusterSubnames.objects.filter(cluster_id=cluster_id))
                ClusterSubnames.objects.filter(cluster_id=cluster_id).exclude(id__in=currentclustersubids).update(status=0)
                
            # print('')   
            Clusters.objects.filter(field_environment_id=environment).exclude(id__in=current_cluster_ids).update(status=0)
           

        Clusters.objects.filter(project_id=project_id).exclude(field_environment_id__in=currentfieldids).update(status=0)          
        result["status"]="success"
        print('project details',project_details)
        print('project block',project_block)
        usercreate=request.user.roles_id
        create_user=project_details.name
        create_user_log(request,create_user,'Project Master','Edit','Project Master Cluster has been Edited',usercreate)
        # return JsonResponse(result)

    data={
        "project_details":project_details,
        'project_block':project_block

    }
    return render(request, "editmastercluster.html", data)

def editmasterwell(request,project_id):
    project_details=Projects.objects.filter(id=project_id).first()
    project_block=BlockName.objects.filter(project_id=project_id,status=0)
    development=['Oil_Development','Gas_Development','Unconventional_Oil','Unconventional_Gas']
    welltypes=['Exploration Wells','Appraisal Wells','Development Wells','Pilot Holes','Workovers/Well Intervention','Infill Wells']
    well_type=dict()
    well_type["Exploration Wells"]="exploration"
    well_type["Appraisal Wells"]="appraisal"
    well_type["Development Wells"]="development"
    well_type["Pilot Holes"]="pilot"
    well_type["Workovers/Well Intervention"]="workovers"
    well_type["Infill Wells"]="infill"
    if(request.POST):
        result={}
        print(request.POST)
        cluster_name=request.POST.getlist('cluster_name')
        # currentclustersubids=[]
        for cluster in cluster_name:
            developmentid=request.POST.getlist('developmenthdn-'+str(cluster))
            developmenttype=request.POST.getlist('development-'+str(cluster))
            i=0
            development_list=[]
            while i< len(developmenttype):
                if MasterDevelopmentType.objects.filter(development_type=developmenttype[i],clustersubname_id=cluster,status=1).exists():
                    update_developmenttype=MasterDevelopmentType.objects.filter(development_type=developmenttype[i],clustersubname_id=cluster,status=1).first()
                    development_list.append(update_developmenttype.id)
                else:
                    create_developmenttype=MasterDevelopmentType.objects.create(development_type=developmenttype[i],clustersubname_id=cluster,status=1,project_id=project_id)
                    development_list.append(create_developmenttype.id)
                i+=1
            
            for development in developmenttype:
                develop=MasterDevelopmentType.objects.filter(development_type=development,clustersubname_id=cluster,status=1).first()
                welltype=request.POST.getlist('well_type'+str(cluster)+'-'+str(development))
                currentwellids=[]
                for well in welltype:
                    checkwell=Well.objects.filter(development_type_id=develop.id,well_name=well).first()
                    
                    if(checkwell != None):
                        print("checkwell",checkwell,"devid",develop.id)
                        well_id=checkwell.id
                        currentwellids.append(well_id)
                    else:
                        print("create new well",well,"devid",develop.id)
                        create_well=Well.objects.create(well_name=well,project_id=project_id,development_type_id=develop.id,status=1)
                        well_id=create_well.id
                        currentwellids.append(well_id)
                        print('well_id',well_id)
                
                    well_given_name=well_type[well]
                    well_name=request.POST.getlist('wellname_'+str(well_given_name)+str(cluster)+'-'+str(development))
                    well_subid=request.POST.getlist('wellid_'+str(well_given_name)+str(cluster)+'-'+str(development))
                    print("wellname",well_name)
                    print("wellid",well_subid)
                    j=0
                    currentwellsubids=[]
                    while j<len(well_subid):
                        if(well_subid[j]):
                            WellSub.objects.filter(id=well_subid[j]).update(well_subname=well_name[j])
                            currentwellsubids.append(well_subid[j])
                        else:
                            print('new wellsub',well_name[j],'well_id',well_id)
                            create_subwell=WellSub.objects.create(well_subname=well_name[j],project_id=project_id,well_id=well_id)
                            currentwellsubids.append(create_subwell.id)
                        j +=1
                    # print("wellsub",currentwellsubids)
                    WellSub.objects.filter(well_id=well_id).exclude(id__in=currentwellsubids).update(status=0)
                print("well",currentwellids)
                Well.objects.filter(development_type_id=develop.id).exclude(id__in=currentwellids).update(status=0)
            # print(development_list,cluster)
            MasterDevelopmentType.objects.filter(clustersubname_id=cluster).exclude(id__in=development_list).update(status=0)
        result["status"]="success"
        usercreate=request.user.roles_id
        create_user=project_details.name
        create_user_log(request,create_user,'Project Master','Edit','Project Master Wells has been Edited',usercreate)
        return JsonResponse(result)
    
    data={
        "project_details":project_details,
        'project_block':project_block,
        'development':development,
        'welltypes':welltypes

    }
    
    return render(request, "editmasterwell.html", data)

#create back button#
def projectmaster(request,project_id):
    get_project=Projects.objects.get(id=project_id)
    country=Countries.objects.all()
    if(request.POST):
        result={}
        print(request.POST)
        project=Projects.objects.filter(id=project_id).update(name=request.POST.get('name'),country_id=request.POST.get('country'))
        result["status"]="success"
        return JsonResponse(result)

    return render(request,"backmasterproject.html",{'project':get_project,'countries':country})


def backmasterblock(request,project_id):
    project_details=Projects.objects.filter(id=project_id).first()
    project_block=BlockName.objects.filter(project_id=project_id,status=0)
    if(request.POST):
        result={}
        print(request.POST)
        block_id=request.POST.getlist('blockid')
        block_name=request.POST.getlist('blockname')
        i=0
        current_blockids=[]
        while i<len(block_id):
            if(block_id[i]):
                 block=BlockName.objects.filter(id=block_id[i]).update(block_name=block_name[i])
                 current_blockids.append(block_id[i])
            else:
                create_block=BlockName.objects.create(block_name=block_name[i],project_id=project_id)
                current_blockids.append(create_block.id)
            i +=1
        BlockName.objects.filter(project_id=project_id).exclude(id__in=current_blockids).update(status=1)
        result["status"]="success"
        return JsonResponse(result)

    data={
        'project_details':project_details,
        'project_block':project_block
    }
    return render(request,"backmasterblock.html",data)

def backmasterfield(request,project_id):    
    project_details=Projects.objects.filter(id=project_id).first()
    project_block=BlockName.objects.filter(project_id=project_id,status=0)
    if(request.POST):
        result={}
        print(request.POST)
        blockname=request.POST.getlist('block_name')
        for block in blockname:
            fieldname=request.POST.getlist('fieldname'+str(block))
            fieldid=request.POST.getlist('fieldid'+str(block))

            i=0
            currentfieldids=[]
            while i<len(fieldid):
                if(fieldid[i]):
                    FieldName.objects.filter(id=fieldid[i]).update(field_name=fieldname[i])
                    currentfieldids.append(fieldid[i])
                else:
                    create_field=FieldName.objects.create(field_name=fieldname[i],block_id=block,project_id=project_id)
                    # currentfieldids.append(create_field.id)
                    currentfieldids.append(create_field.id)
                i +=1
            print("currentfieldids",currentfieldids)

            FieldName.objects.filter(block_id=block,project_id=project_id).exclude(id__in=currentfieldids).update(status=0)
            for i,j in enumerate(currentfieldids):
                field_environment=request.POST.getlist('project_environment-'+str(block)+'-'+str(i))
                currentenvironmentids=[]    
                for environment in field_environment:
                    if "-" in environment:
                        spiltenvironment=environment.split("-")
                        field_environment=spiltenvironment[0]
                        field_environmentsubtype=spiltenvironment[1]
                        environment_field=FieldEnvironment.objects.filter(field_id=j,project_environment=field_environment,project_environment_subtype=field_environmentsubtype)
                        if environment_field.exists():
                            currentenvironmentids.append(environment_field[0].id)
                            print("env exists")
                        else:
                            create_field_environment=FieldEnvironment.objects.create(field_id=j,project_environment=field_environment,project_environment_subtype=field_environmentsubtype,status=1,project_id=project_id)
                            currentenvironmentids.append(create_field_environment.id)
                    else:
                        get_others_environment=request.POST.getlist('environment-'+str(block)+'-'+str(environment)+'-'+str(i))
                        environment_id=request.POST.getlist('environmentid-'+str(block)+'-'+str(environment)+'-'+str(i))
                        print("id",environment_id)
                        print("data",get_others_environment)
                        i=0
                        while i<len(environment_id):
                            if (environment_id[i] !=' '):
                                update_fieldenvironment=FieldEnvironment.objects.filter(id=environment_id[i]).update(project_environment_subtype=get_others_environment[i])
                                currentenvironmentids.append(environment_id[i])
                            else:
                                create_environment=FieldEnvironment.objects.create(field_id=j,project_environment=environment,project_environment_subtype=get_others_environment[i],project_id=project_id,status=1)
                                currentenvironmentids.append(create_environment.id)
                            i +=1
                FieldEnvironment.objects.filter(field_id=j).exclude(id__in=currentenvironmentids).update(status=0)
        FieldName.objects.filter(project_id=project_id).exclude(block_id__in=blockname).update(status=0)
        result["status"]="success"
        return JsonResponse(result)
    data={
        "project_details":project_details,
        'project_block':project_block

    }
    return render(request, "backmasterfield.html", data)
    

def backmastercluster(request,project_id):
    project_details=Projects.objects.filter(id=project_id).first()
    project_block=BlockName.objects.filter(project_id=project_id,status=0)
    clustertype=dict()
    clustertype["Well Pad"]="wellpad"
    clustertype["Well Platform"]="wellplatform"
    clustertype["Subsea Clusters"]="subseaclusters"
    if(request.POST):
        result={}
        print(request.POST)
        environmentlist=request.POST.getlist('environment_name')
        currentfieldids=[]
        for environment in environmentlist:
            cluster_type=request.POST.getlist('cluster_type'+str(environment))
            currentfieldids.append(environment)
            current_cluster_ids=[]
            for cluster in cluster_type:
                checkcluster=Clusters.objects.filter(field_environment_id=environment,cluster_name=cluster).first()
                print(f"checkcluster {checkcluster}")
                if(checkcluster != None):
                    cluster_id=checkcluster.id
                    current_cluster_ids.append(cluster_id)
                else:
                   create_cluster=Clusters.objects.create(cluster_name=cluster,field_environment_id=environment,project_id=project_id,status=1) 
                   cluster_id=create_cluster.id
                   current_cluster_ids.append(cluster_id)
                cluster_given_name=clustertype[cluster]
                cluster_name=request.POST.getlist('clustername_'+str(cluster_given_name)+str(environment))
                cluster_subid=request.POST.getlist('clusterid_'+str(cluster_given_name)+str(environment))
                i=0
                currentclustersubids=[]
                while i<len(cluster_subid):
                    if(cluster_subid[i]):
                        ClusterSubnames.objects.filter(id=cluster_subid[i]).update(cluster_subname=cluster_name[i],status=1)
                        currentclustersubids.append(cluster_subid[i])
                    else:
                        create_clustersub=ClusterSubnames.objects.create(cluster_subname=cluster_name[i],cluster_id=cluster_id,project_id=project_id)
                        currentclustersubids.append(create_clustersub.id)
                    i +=1
                ClusterSubnames.objects.filter(cluster_id=cluster_id).exclude(id__in=currentclustersubids).update(status=0)
            Clusters.objects.filter(field_environment_id=environment).exclude(id__in=current_cluster_ids).update(status=0)

        Clusters.objects.filter(project_id=project_id).exclude(field_environment_id__in=currentfieldids).update(status=0)
          
        result["status"]="success"
        return JsonResponse(result)

    data={
        "project_details":project_details,
        'project_block':project_block

    }
    return render(request, "backmastercluster.html", data)

def editblock(request,pk):
    get_project=Projects.objects.get(id=pk)
    get_block=BlockName.objects.filter(project_id=pk)
    block_count=get_block.count()
    return render(request,"editblock.html",{"project":get_project,"blocks":get_block,"block_count":block_count})

def Editblock(request):
    print(request.POST)
    projectid=request.POST['project_id']
    get_block_id=request.POST.getlist('block_id')
    get_block_name=request.POST.getlist('block_name')
    for ind,block_id in enumerate(get_block_id):
        if block_id != ' ':
            # print(block_id)
            block=BlockName.objects.filter(id=int(block_id))

            for index,block_name in enumerate(get_block_name):
                if ind==index:
                    block.update(block_name=block_name)
                else:
                    print()       
        else:
            create_block=BlockName.objects.create(block_name=block_name,project_id=int(projectid),status=0)
    return JsonResponse({"data":"block edited"})

def editfield(request,pk):
    get_project=Projects.objects.get(id=pk)
    get_block=BlockName.objects.filter(project_id=pk)
    get_field=FieldName.objects.filter(project_id=pk)
    return render(request,"editfield.html",{"project":get_project,"blocks":get_block,"fields":get_field})

def Editfield(request):
    print(request.POST)
    projectid=request.POST['project_id']
    get_block_id=request.POST.getlist('block_hdid')
    get_field_id=request.POST.getlist('field_hdid')
    for get_block in get_block_id:
        block=BlockName.objects.get(id=int(get_block),project_id=int(projectid))
        if block.status == 1:
            get_field=request.POST.getlist('field_hd'+str(block.id))
            get_fieldname=request.POST.getlist('field_name'+str(block.id))
            i=0
            allfieldid=[]
            while i<len(get_field):
                if(get_field[i]):
                    FieldName.objects.filter(id=get_field[i]).update(field_name=get_fieldname[i])
                    allfieldid.append(get_field[i])

                else:
                    fielddata=FieldName.objects.create(field_name=get_fieldname[i],project_id=projectid,block_id=block.id,status=0)
                    allfieldid.append(fielddata.id)
                i +=1
            print(f"allfieldid {allfieldid}")
            
        else:
            get_field=request.POST.getlist('field_name'+str(block.id))
            if len(get_field)!=0:
                for fields in get_field:   
                    print("0",fields)
    return JsonResponse({'data':'field created'})

def editcluster(request,pk):
    get_project=Projects.objects.get(id=pk)
    get_block=BlockName.objects.filter(project_id=pk)
    get_field=FieldName.objects.filter(project_id=pk)
    # get_clusters=
    return render(request,"editcluster.html",{"project":get_project,"blocks":get_block,"fields":get_field})

def Editcluster(request):
    print(request.POST)
    get_checked_field=request.POST.getlist('field_checkbox')
    for checked_field in get_checked_field:
        get_field=FieldName.objects.get(field_name=checked_field)
        get_cluster=request.POST.getlist('clustersubtype_'+str(get_field.id))
        print(get_cluster)

    return JsonResponse({'data':'cluster edited'})



@login_required(login_url='/')
def listmaster(request):
    request.session['mainmenu'] = 'masters'
    request.session['submenu'] = 'master_projects'
    company=User.objects.get(id=request.user.id)
    current_company=company.company.id
    data={}
    if (request.user.roles_id == 3):
        userrights=UserRights.objects.get(user_id=request.user.id,module_id=3)
        # print("wew",userrights)
        data['rights']=userrights
    get_countries=Projects.objects.filter(company=request.company).values('country').annotate(dcount=Count('country'))
    master_count=Projects.objects.filter(company=request.company).count()
    master=Projects.objects.filter(company=request.company)
    # paginator = Paginator(master, 1)
    # # users = paginator.page(1)
    # page = request.GET.get('page', 1)
    # try:
    #     users = paginator.page(page)
    # except PageNotAnInteger:
    #     users = paginator.page(1)
    # except EmptyPage:
    #     users = paginator.page(paginator.num_pages)
    countries=[]
    country=''
    project_name=''
    project_country=[]
    project_id=''
    for country in get_countries:
            country_id=country['country']
            get_countries=Countries.objects.get(id=country_id) 
            countries.append({"country_id":get_countries.id,"country_name":get_countries.nicename})
    if request.method=="POST":
        countval=request.POST['country_name']
        if (countval == ' '):
            country=''
        else:
            country=int(countval)
        project_name=request.POST['project_name']
        projects=Projects.objects.filter(company=request.company)

        if(country):
            project_country=Projects.objects.filter(country_id=country,company=request.company)
            projects=projects.filter(country_id=country)
        else:
            projects=projects
        
        if(project_name!=' '):
            project_id=int(project_name)
            projects=projects.filter(id=project_name)
            # print("dssa",projects)
        else:
            project_id=project_name
            
    else:
        projects=Projects.objects.filter(company=request.company).order_by('-id')
    
    
    data.update({
        "listmaster":projects,
        "get_countries":countries,
        'post_country':country,
        'post_project':project_id,
        'project_country':project_country,
        'master_count':master_count,
        # 'users':users
    })
    return render(request,"listmasterproject.html",data)

def editproject():
    print("vh")

def getproject(request):
    get_country_id=request.GET.get('country_id')
    projects=Projects.objects.filter(country_id=int(get_country_id),company=request.company).values()
    return JsonResponse({'data':list(projects)})


def getblocks(request):
    get_project_id=request.GET.get('project_id')
    blocks=BlockName.objects.filter(project_id=int(get_project_id)).values()
    return JsonResponse({'data':list(blocks)})
    
    
def getfield(request):
    get_block_id=request.GET.get('block_id')
    fields=FieldName.objects.filter(block_id=int(get_block_id),status=1).values()
    print('dsff')
    return JsonResponse({'data':list(fields)})
    
def getcluster(request):
    get_field_id=request.GET.get('field_id')
    field_env=FieldEnvironment.objects.filter(field_id=get_field_id,status=1).values()
    # clusters=Clusters.objects.filter(field_id=get_field_id,status=1)
    # print(f"clusters {clusters}")
    data=[]
    # for cluster in clusters:
    #     print(f"clusterid {cluster.id}")
    #     get_clustersub=ClusterSubnames.objects.filter(cluster_id=cluster.id)
    #     clusters_list=[]
    #     for cluster_sub in get_clustersub:
    #       clusters_list.append({'id':cluster_sub.id,'cluster_subname':cluster_sub.cluster_subname})
    #     data.append({
    #         cluster.cluster_name:clusters_list
    #     })
    environment=list(field_env)

    return JsonResponse({'environment':environment})


def fieldenvironment(request):
    environment_id=request.GET.get('environment_id')
    clusters=Clusters.objects.filter(field_environment_id=environment_id,status=1)
    data=[]
    for cluster in clusters:
        get_clustersub=ClusterSubnames.objects.filter(cluster_id=cluster.id,status=1)
        clusters_list=[]
        for cluster_sub in get_clustersub:
          clusters_list.append({'id':cluster_sub.id,'cluster_subname':cluster_sub.cluster_subname})
        data.append({
            cluster.cluster_name:clusters_list
        })
    return JsonResponse({'data':data})

def developmenttype(request):
    get_cluster_id=request.GET.get('clustersub_id')
    development_type=MasterDevelopmentType.objects.filter(clustersubname_id=get_cluster_id,status=1).values()
    print(development_type)
    return JsonResponse({'data':list(development_type)})


#green field drill wells
def getwell(request):
    get_development_type_id=request.GET.get('development_id')
    # wells=Well.objects.filter(development_type_id=get_development_type_id,status=1).values()
    wells=Well.objects.filter(development_type_id=get_development_type_id,status=1)
    wellsub=wells.filter(well_name__iexact="Pilot Holes").values() | wells.filter(well_name__iexact="Development Wells").values() | wells.filter(well_name__iexact="Appraisal Wells").values() | wells.filter(well_name__iexact="Exploration Wells").values()
    # print("wells",wellsub)
    return JsonResponse({'data':list(wellsub)})

def getmasterwell(request):
    get_development_type_id=request.GET.get('development_id')
    wells=Well.objects.filter(development_type_id=get_development_type_id,status=1).values()
    # wells=Well.objects.filter(development_type_id=get_development_type_id,status=1)
    # wellsub=wells.filter(well_name__iexact="Pilot Holes").values() | wells.filter(well_name__iexact="Development Wells").values() | wells.filter(well_name__iexact="Appraisal Wells").values() | wells.filter(well_name__iexact="Exploration Wells").values()
    # print("wells",wellsub)
    return JsonResponse({'data':list(wells)})


#brown subsurface
def Getwellbrownfield(request):
    get_development_type_id=request.GET.get('development_id')
    wells=Well.objects.filter(development_type_id=get_development_type_id,status=1)
    data={}
    welltypelist=[]
    val={}
    for well in wells:
        if well.well_name == "Infill Wells":
            data={'id':well.id,'well_name':well.well_name}
            welltypelist.append(data)
    val={'data':welltypelist}
    return JsonResponse(val)

#brown Drilling

def Getwellbrownfielddrill(request):
    get_development_type_id=request.GET.get('development_id')
    wells=Well.objects.filter(development_type_id=get_development_type_id,status=1)
    data={}
    welltypelist=[]
    val={}
    for well in wells:
        print(well)
        if well.well_name == "Pilot Holes" or well.well_name == 'Development Wells' or well.well_name == 'Workovers/Well Intervention':
            data={'id':well.id,'well_name':well.well_name}
            welltypelist.append(data)
    val={'data':welltypelist}
    print("welldata",val)
    return JsonResponse(val)

def getwellsubname(request):
    get_well_id=request.GET.get('well_id')
    wellsublist=WellSub.objects.filter(well_id=get_well_id,status=1).values()
    return JsonResponse({'data':list(wellsublist)})



def invitevendorlist(request):
    request.session['mainmenu'] = 'vendors'
    request.session['submenu']='invitevendors'
    invitevendors=VendorInvitationModel.objects.filter(company=request.company).order_by('-id')
    # print(invitevendors.values())
    invitevendorcount=invitevendors.count()
    data={
        "invitevendor_count":invitevendorcount,
        "invitevendors":invitevendors
    }
    return render(request,'vendorinvitelist.html',data)

def vendorinvite(request):
    request.session['mainmenu'] = 'vendors'
    request.session['submenu']='invitevendors'
    companyid=request.company.id
    print(companyid)
    # print(request.company.image.url)
    # a=request.build_absolute_uri('/')
    get_ip=request.get_raw_uri()
    replace_url=get_ip.replace('/projects/vendorinvite','')
    # print(replace_url)
    datetimedata=Settings.objects.filter(company_id=companyid).first()
    # timedata=datetimedata.timezone

    # now_asia = pytz.timezone(timedata)
    # utc = pytz.utc
    # local_datetime = now_asia.localize(datetime.now())
    # timedata=datetimedata.timezone
    tzname = request.session.get('companytimezone')
    local_datetime=datetime.now(pytz.timezone(tzname))
    print('lt',local_datetime)

    convert_string=datetime.strftime(local_datetime, "%d %b %Y  %H:%M:%S")
    date_object=datetime.strptime(convert_string,"%d %b %Y  %H:%M:%S")
    print(convert_string)
    print(date_object)
    datetimestamp=datetime.timestamp(date_object)
    print(datetimestamp)
    # dt_object = datetime.fromtimestamp(datetimestamp)
    # print(dt_object)
    # print(datetime.now())

    # print(local_datetime)
    if request.method=="POST":
        # print(request.POST)
        title=request.POST['Tile']
        companyname=request.POST['companyname']
        email=request.POST['emailaddress']
        cxname=request.POST['personname']
        companyid=request.POST['companyid']

        create_vendor_invite=VendorInvitationModel.objects.create(user_id=request.user.id,title=title,contactpersonname=cxname,company_id=request.user.company.id,vendorcompany=companyname,contact_person_email=email) 
        VendorInvitationModel.objects.filter(id=create_vendor_invite.id).update(created_at=datetimestamp)
        # local link

        # url=replace_url+'/projects/'+str(request.user.id)+'/'+str(create_vendor_invite.id)+'/outervendorbasicinfo'

        #(live link )
        # url='https://irockinfo.mo.vc/projects/'+str(request.user.id)+'/'+str(create_vendor_invite.id)+'/outervendorbasicinfo'
        current_url = f'{request.scheme}://{request.get_host()}'
        url = f'{current_url}/projects/{request.user.id}/{create_vendor_invite.id}/outervendorbasicinfo'
        # print("saddsd",url)
        imageurl=''
        if Companies.objects.filter(pk=companyid).filter(Q(image__isnull=True) | Q(image='')):
            imageurl=''
        else:
            imageurl=request.company.image.url
        f_name=request.company.first_name
        l_name=request.company.last_name
        fullname=f_name+' '+l_name
        invitetime=datetimedata.vendor_invite_time
        inviteunit=datetimedata.vendor_invite_unit
        subject = 'iROCK Vendor Registration'
        if (request.user.roles_id == 2):
            context={'invitetime':invitetime,'inviteunit':inviteunit,'title':title,'url':url,'name': companyname,'companyname':request.company.company_name,'contact':cxname,'role':request.user.designation_role,'landline_countrycode':request.company.phone_countrycode,'lanline':request.company.phonenumber,'mobile_countrycode':request.company.mobile_countrycode,'mobile':request.company.mobile,'website':request.company.website,'address':request.company.address,'image':imageurl,'email':request.company.email,'fullname':fullname,'areacode':request.company.areacode,'current_url':current_url}
        else:
            user=User.objects.get(id=request.user.id)
            firstname=user.name
            lastname=user.lastname
            username=firstname+''+lastname
            context={'invitetime':invitetime,'inviteunit':inviteunit,'title':title,'url':url,'name': companyname,'companyname':request.company.company_name,'contact':cxname,'role':request.user.designation_role,'landline_countrycode':user.phone_countrycode,'lanline':user.phone,'mobile_countrycode':user.mobile_countrycode,'mobile':user.mobile,'website':request.company.website,'address':request.company.address,'image':imageurl,'email':user.email,'fullname':username,'areacode':user.areacode,'current_url':current_url}

        html_message = render_to_string('invitemail.html', context)
        plain_message = strip_tags(html_message)
        from_email = settings.EMAIL_HOST_USER

        to = email

        msg=mail.send_mail(subject,plain_message, from_email, [to], html_message=html_message)
        if msg:
            messages.success(request, 'Mail has been sent to '+str(cxname))
        else:
            messages.success(request, 'please check enterd details!!!')
        return redirect('projects:invitevendorlist')

        

    data={'companyid':companyid}
    return render(request ,'invitevendor.html',data)





def Getstate(request):
    countryid=request.GET.get('countryid')
    data=States.objects.filter(country_id=countryid).values()
    print(data,'getvalues')
    return JsonResponse({'data':list(data)})

# def vendorbasicinfo(request):
    
#     return render(request, 'vendorbasicinfo.html', {'countries':country})

#outer vendor creation form 1

def outervendorbasicinfo(request,userid,invitevendorid):
    user=User.objects.get(id=userid)
    datetimedata=Settings.objects.filter(company_id=user.company.id).first()
    invitetime=VendorInvitationModel.objects.get(id=invitevendorid)
    # variable =timezone.localtime(timezone.now()
    #)
    if (invitetime.invite_link_status == 1):
        if (datetimedata.vendor_invite_unit == 'Hours'):
            invitetimedata=invitetime.created_at

            dt_object = datetime.fromtimestamp(float(invitetimedata))
            # print("a",dt_object)
            # a=dt_object.replace(second=0, microsecond=0)
            # print(datetime.now())
            # local_datetime=pytz.timezone(tzname)
            # print(local_datetime)
            # localized_timestamp = local_datetime.localize(a)
            # print(localized_timestamp)
            # new_timezone_timestamp = localized_timestamp.astimezone(local_datetime)
            # print(new_timezone_timestamp)

            timedata=datetimedata.vendor_invite_time
            #change hours dynamically
            hours_added = dt_object + timedelta(hours=int(timedata))
            # print('c_d',dt_object)
            # print("add_d",hours_added)
            # invite_dateadd_time=hours_added.time().replace(second=0, microsecond=0)
            # print(invite_dateadd_time)

            tzname = request.session.get('companytimezone')
            local_datetime=datetime.now(pytz.timezone(tzname))
            c=local_datetime.replace(second=0, microsecond=0,tzinfo=None)
            # print('c',c)
            # current_datetime=local_datetime.time().replace(second=0, microsecond=0)
            # print(current_datetime)

            invitedatetimestamp=datetime.timestamp(hours_added)
            print(invitedatetimestamp)

            currentdatetimestamp=datetime.timestamp(c)
            print(currentdatetimestamp)

            if (currentdatetimestamp <= invitedatetimestamp):
                print("hours link working")
                inviteid=VendorInvitationModel.objects.get(id=invitevendorid,company_id=user.company.id)
            # print(inviteid.vendorcompany)
                countries=Basecountries.objects.all()
                countrylist=[]
                for countri in countries:
                    if States.objects.filter(country_id=countri.id).exists():
                        countrylist.append({'id':countri.id,'name':countri.name})
                maincountry=Countries.objects.all()
                return render(request, 'outervendorbasicinfo.html', {'countries':countrylist,'maincountry':maincountry,'invite_vid':inviteid})

            else:
                # invitetime.invite_link_status=0
                # invitetime.save()
                invitetime=VendorInvitationModel.objects.filter(id=invitevendorid).update(invite_link_status=0)
                print("link expire")
                return redirect('projects:expirepage')
    

        else:
            print("days")
            invitetimedata=invitetime.created_at
            timedata=datetimedata.vendor_invite_time
            dt_object = datetime.fromtimestamp(float(invitetimedata))
            convert_strdatetime=dt_object.strftime('%Y-%m-%d %H:%M:%S')
            # print(convert_strdatetime)
            a=datetime.strptime(convert_strdatetime,'%Y-%m-%d %H:%M:%S')
            future_date = a + \
                            timedelta(days = int(timedata))
            # print(future_date)
            tzname = request.session.get('companytimezone')
            local_datetime=datetime.now(pytz.timezone(tzname))
            current_datetime=local_datetime.replace(microsecond=0,tzinfo=None)
            # print(current_datetime)
            # print(future_date)
            # custome_date_check=datetime(2022, 2, 24, 15,38,10)
            # print(custome_date_check)
            if (current_datetime < future_date):
                print("days link working")

                inviteid=VendorInvitationModel.objects.get(id=invitevendorid,company_id=user.company.id)
                # print(inviteid.vendorcompany)
                countries=Basecountries.objects.all()
                countrylist=[]
                for countri in countries:
                    if States.objects.filter(country_id=countri.id).exists():
                        countrylist.append({'id':countri.id,'name':countri.name})
                maincountry=Countries.objects.all()
                return render(request, 'outervendorbasicinfo.html', {'countries':countrylist,'maincountry':maincountry,'invite_vid':inviteid})
            else:
                # invitetime.invite_link_status=0
                # invitetime.save()
                invitetime=VendorInvitationModel.objects.filter(id=invitevendorid).update(invite_link_status=0)
                print("days link expire")
                return redirect('projects:expirepage')
    else:
        print("days link expire")
        return redirect('projects:expirepage')
        

def expirepage(request):
    return render(request,'expirepage.html')



def createvendor(request):
    request.session['mainmenu'] = 'vendors'
    request.session['submenu']='createvendors'
    companyid=request.company.id
    userid=request.user.id
    print(userid)
    countrylist=[]
    if (request.POST):
        vendor_name=request.POST['vendorname']
        country_of_operation=request.POST['vendor_country_list']
        country_of_incorporation=request.POST['vendor_country_list_incorporation']
        # print()
        state=request.POST['vendor_state_list']
        Address=request.POST['address']
        website=request.POST['website']
        Registernumber=request.POST['Registernumber']
        invite_vid=request.POST.get('invitevid','')
        # email=request.POST['v_email']
        datetimedata=Settings.objects.filter(company_id=companyid).first()
        timedata=datetimedata.timezone

        # now_asia = pytz.timezone(timedata)
        # utc = pytz.utc
        # local_datetime = now_asia.localize(datetime.now())
        tzname = request.session.get('companytimezone')
        local_datetime=datetime.now(pytz.timezone(tzname))
        convert_string=datetime.strftime(local_datetime, "%d %b %Y  %H:%M:%S")
        date_object=datetime.strptime(convert_string,"%d %b %Y  %H:%M:%S")
        print(convert_string)
        print(date_object)
        datetimestamp=datetime.timestamp(date_object)
        createbasic_vendor=VendorRegistraion.objects.create(invite_vendor_id=invite_vid,vendor_name=vendor_name,country_of_operation_id=country_of_operation,country_of_incorporation_id=country_of_incorporation,state_id=state,Address=Address,website=website,RegisterNumber=Registernumber,company_id=companyid,user_id=userid)
        VendorRegistraion.objects.filter(id=createbasic_vendor.id).update(created_at=datetimestamp)
        taxname=request.POST.getlist('taxname')
        taxnumber=request.POST.getlist('taxnumber')
        try:
            for tax,num in zip(taxname,taxnumber):
                create_vendor_tax=VendorTaxDetails.objects.create(company_id=companyid,vendor=createbasic_vendor,Tax_Name=tax,Tax_Number=num)
        except:
            pass
        data={'vendorid':createbasic_vendor.id}
        return JsonResponse(data)

    country=Basecountries.objects.all()
    maincountry=Countries.objects.all()
    for contri in country:
        if States.objects.filter(country_id=contri.id).exists():
            countrylist.append({'id':contri.id,'country_name':contri.name})
    # print(countrylist)
    context={'countries':countrylist,
            'maincountry':maincountry}
    return render(request,'vendorbasicinfo.html',context) 

        


def createvendoroutside(request):
    if (request.POST):
        print(request.POST)
        vendor_name=request.POST['vendorname']
        country_of_operation=request.POST['vendor_country_list']
        country_of_incorporation=request.POST['vendor_country_list_incorporation']
        companyid=request.POST['companyid']
        userid=request.POST['userid']
        state=request.POST['vendor_state_list']
        Address=request.POST['address']
        website=request.POST['website']
        Registernumber=request.POST['Registernumber']
        invite_vid=request.POST.get('invitevid','')
        # email=request.POST['v_email']
        # tzname = request.session.get('companytimezone')
        company=Settings.objects.filter(company_id=companyid).first()
        if (company != None):
            tzname=company.timezone
            local_datetime=datetime.now(pytz.timezone(tzname))
            convert_string=datetime.strftime(local_datetime, "%d %b %Y  %H:%M:%S")
            date_object=datetime.strptime(convert_string,"%d %b %Y  %H:%M:%S")
            print(convert_string)
            print(date_object)
            datetimestamp=datetime.timestamp(date_object)
        else:
            datetimestamp=''
        createbasic_vendor=VendorRegistraion.objects.create(user_id=userid,invite_vendor_id=invite_vid,vendor_name=vendor_name,country_of_operation_id=country_of_operation,country_of_incorporation_id=country_of_incorporation,state_id=state,Address=Address,website=website,RegisterNumber=Registernumber,company_id=companyid,created_at=datetimestamp)
        taxname=request.POST.getlist('taxname')
        taxnumber=request.POST.getlist('taxnumber')
        try:
            for tax,num in zip(taxname,taxnumber):
                create_vendor_tax=VendorTaxDetails.objects.create(company_id=companyid,vendor=createbasic_vendor,Tax_Name=tax,Tax_Number=num)
        except:
            pass
        data={'vendorid':createbasic_vendor.id}
        return JsonResponse(data)

def vendorregistration(request,pk):
    datetimedata=Settings.objects.filter(company_id=1).first()
    timedata=datetimedata.timezone
    request.session['mainmenu'] = 'vendors'
    request.session['submenu']='createvendors'
    phonecodelist=[]
    phonecode = Basecountries.objects.all().values('iso3','id','phonecode').order_by('phonecode').distinct()
    for codes in phonecode:
        id=codes.get('id')
        num=codes.get('phonecode')
        iso=codes.get('iso3')
        x = re.findall("\+",num)
        if x:
            phonecodelist.append({'id':id,'phonecode':str(num),'iso':iso})
        else:
            phonecodelist.append({'id':id,'phonecode':'+'+str(num),'iso':iso})
    newlist=sorted(phonecodelist, key = lambda i:  (i['phonecode']))
    # print (newlist)
    vendorid=VendorRegistraion.objects.get(id=pk)
    return render(request ,'vendorRegistration.html',{'vendorid':vendorid ,'phone':newlist})

def Outervendorregistration(request,userid,pk):
    user=User.objects.get(id=userid)
    phonecodelist=[]
    phonecode = Basecountries.objects.all().values('iso3','id','phonecode').order_by('phonecode').distinct()
    for codes in phonecode:
        id=codes.get('id')
        num=codes.get('phonecode')
        iso=codes.get('iso3')
        x = re.findall("\+",num)
        if x:
            phonecodelist.append({'id':id,'phonecode':str(num),'iso':iso})
        else:
            phonecodelist.append({'id':id,'phonecode':'+'+str(num),'iso':iso})
    newlist=sorted(phonecodelist, key = lambda i:  (i['phonecode']))
    # print (newlist)
    vendorid=VendorRegistraion.objects.get(id=pk)
    return render(request ,'outervendorRegistration.html',{'vendorid':vendorid ,'phone':newlist,'companyid':user.company.id})

def checkuseremailexists(request):
    # data={}
    emailid=request.GET.get('email')
    print(emailid)
    if User.objects.filter(email__iexact=emailid,status=1).exists():
        data={'data':'exists'}
    else:
        data={'data':'success'}
    return JsonResponse(data)


def createvendornext(request):

    if (request.POST):
        vendorid=request.POST['vendorid']
        contact_primary_title=request.POST['Tile']
        contact_primary_first_name=request.POST['firstname']
        contact_primary_middle_name=request.POST['middlename']
        contact_primary_last_name=request.POST['lastname']
        
        contact_secondary_title=request.POST['Title']
        contact_secondary_first_name=request.POST['alternatefirstname']
        contact_secondary_middle_name=request.POST['alternatemiddlename']
        contact_secondary_last_name=request.POST['alternatelastname']

        secondary_designation=request.POST['alternateDesignation']
        primary_designation=request.POST['Designation']
        contact_primary_phone_code=request.POST['phonecountrycode']
        contact_primary_area_code=request.POST['phoneareacode']
        contact_primary_phone=request.POST['phone']
        contact_primary_mobile_code=request.POST['mobilecoutrycode']
        contact_primary_mobile=request.POST['mobile']
        contact_primary_phone_number_extenstion=request.POST['phonenumber_extenstion']


        contact_secondary_phone_code=request.POST['phonecountryalternatecode']
        contact_secondary_phone=request.POST['alternatephone']
        contact_secondary_mobile_code=request.POST['alternate_country_code']
        contact_secondary_mobile=request.POST['alternatemobile']
        contact_secondary_area_code=request.POST['alternatephoneareacode']
        contact_secondary_phone_number_extenstion=request.POST['alternatephonenumber_extenstion']

        official_primary_email=request.POST['Email']
        official_secondary_email=request.POST['alternateEmail']

        create_registrtion_vendor=VendorRegistraion.objects.filter(id=int(vendorid)).update(contact_primary_first_name=contact_primary_first_name,contact_primary_middle_name=contact_primary_middle_name,contact_primary_last_name=contact_primary_last_name,contact_secondary_first_name=contact_secondary_first_name,contact_secondary_middle_name= contact_secondary_middle_name,contact_secondary_last_name=contact_secondary_last_name,secondary_designation=secondary_designation,primary_designation=primary_designation,contact_primary_phone_code=contact_primary_phone_code,contact_primary_phone=contact_primary_phone,contact_primary_mobile_code=contact_primary_mobile_code,contact_primary_mobile=contact_primary_mobile, contact_secondary_phone_code=contact_secondary_phone_code,contact_secondary_phone=contact_secondary_phone,contact_secondary_mobile_code=contact_secondary_mobile_code,contact_secondary_mobile=contact_secondary_mobile,official_primary_email=official_primary_email,official_secondary_email=official_secondary_email,contact_primary_area_code=contact_primary_area_code,contact_secondary_area_code=contact_secondary_area_code,contact_primary_title=contact_primary_title,contact_secondary_title=contact_secondary_title,contact_primary_phone_number_extenstion=contact_primary_phone_number_extenstion,contact_secondary_phone_number_extenstion=contact_secondary_phone_number_extenstion)
        vendorid=VendorRegistraion.objects.get(id=vendorid)
        data={'vendorid':vendorid.id}
        return JsonResponse(data)



def vendorservice(request,pk):
    date_format={}
    if Settings.objects.filter(company_id=request.company).exists():
        date_format=Settings.objects.filter(company_id=request.company).first()
    else:
        date_format
    companyrole=request.user.designation_role
    vendorid=VendorRegistraion.objects.get(id=pk)
    # print()

    request.session['mainmenu'] = 'vendors'
    request.session['submenu']='createvendors'
    companyid=request.company.id

    services=request.POST.getlist('types_service')
    if (request.POST):

        for types in services:
            get_contract=request.POST.getlist(str(types)+'Contract')
            get_Scope=request.POST.getlist(str(types)+'Abridged_Scope')
            get_service_contract=request.POST.getlist(str(types)+'contractnumber')
            get_date=request.POST.getlist(str(types)+'Contract_Executed_Date')
            # print("date",get_date)
            # print(date_format)
            i=0
            current_contract_id=[]
            all_dateformat = {'dd-M-yy':"%d-%b-%Y",'dd-mm-yy':"%d-%m-%Y",'dd/mm/yy':"%d/%m/%Y",'mm-dd-yy':'%m-%d-%Y','mm/dd/yy':'%m/%d/%Y','yy-mm-dd':'%Y-%m-%d','yy/mm/dd':'%Y/%m/%d'}
            while i<len(get_contract):
                if (date_format.dateformat == 'dd-M-yy'):
                    convert_date=datetime.strptime(get_date[i],"%d-%b-%Y").date()
                    # print('1',get_date[i])
                elif (date_format.dateformat == 'dd-mm-yy'):
                    convert_date=datetime.strptime(get_date[i],"%d-%m-%Y").date()
                    # print('2',get_date[i])
                elif (date_format.dateformat == 'dd/mm/yy'):
                    convert_date=datetime.strptime(get_date[i],"%d/%m/%Y").date()
                    # print('3',get_date[i])
                elif (date_format.dateformat == 'mm-dd-yy'):
                    convert_date=datetime.strptime(get_date[i],"%m-%d-%Y").date()
                    # print('4',get_date[i])
                elif (date_format.dateformat == 'mm/dd/yy'):
                    convert_date=datetime.strptime(get_date[i],"%m/%d/%Y").date()
                    # print('5',get_date[i])
                elif (date_format.dateformat == 'yy-mm-dd'):
                    convert_date=datetime.strptime(get_date[i],"%Y-%m-%d").date()
                    # print('6',get_date[i])
                elif (date_format.dateformat == 'yy/mm/dd'):
                    convert_date=datetime.strptime(get_date[i],"%Y/%m/%d").date()
                    # print('7',get_date[i])
                else:
                    convert_date=datetime.strptime(get_date[i],"%d-%b-%Y").date()
                    # print('8',get_date[i])
                # print(get_contract[i],get_Scope[i],)
                create_service=Vendorcompanyserviceinfo.objects.create(type_services=types,nameofcontract=get_contract[i],abrigdge_scope_service=get_Scope[i],contract_reference_number=get_service_contract[i],contract_date=convert_date,vendor_id=vendorid.id,company_id=companyid)
                current_contract_id.append(create_service.id)
                i+=1
            vendor_service=Vendorcompanyserviceinfo.objects.filter(vendor_id=vendorid.id,type_services=types)
            print("c_id",vendor_service)
            for i,j in enumerate(vendor_service):
                get_contractfile=request.FILES.getlist(types+'file'+str(i))
                # print(j)
                for contractfile in get_contractfile:
                    create_contract_file=VendorFileUpload.objects.create(contract_id=j.id,vendor_id=vendorid.id,files=contractfile)
                    print(j,'aaa',contractfile)
        VendorRegistraion.objects.filter(id=pk).update(status=3)
        if User.objects.filter(company_id=request.company,is_primary=1,primary_active_status=1).exists():
            approver=User.objects.filter(company_id=request.company,is_primary=1,primary_active_status=1).first()
            print('p',approver)
            VendorRegistraion.objects.filter(id=pk).update(approver_status=1)
        elif User.objects.filter(company_id=request.company,is_secondary=1).exists():
            approver=User.objects.filter(company_id=request.company,is_secondary=1).first()
            print('s',approver)
            VendorRegistraion.objects.filter(id=pk).update(approver_status=2)
        # vendorregistartionapprovermail(vendorid.company.id,approver.id,companyrole)
        return redirect('projects:vendorview')

    return render(request ,'vendorservice.html',{'vendorid':vendorid,'date_format':date_format})


def Outervendorservice(request,userid,pk):
    user=User.objects.get(id=userid)
    companyid=user.company.id
    date_format={}

    if Settings.objects.filter(company_id=companyid).exists():
        date_format=Settings.objects.filter(company_id=companyid).first()
    else:
        date_format
    vendorid=VendorRegistraion.objects.get(id=pk)
    # print(vendorid.invite_vendor.title)

    services=request.POST.getlist('types_service')
    if (request.POST):

        for types in services:
            get_contract=request.POST.getlist(str(types)+'Contract')
            get_Scope=request.POST.getlist(str(types)+'Abridged_Scope')
            get_service_contract=request.POST.getlist(str(types)+'contractnumber')
            get_date=request.POST.getlist(str(types)+'Contract_Executed_Date')
            # print("date",get_date)
            # print(date_format)
            i=0
            current_contract_id=[]
            all_dateformat = {'dd-M-yy':"%d-%b-%Y",'dd-mm-yy':"%d-%m-%Y",'dd/mm/yy':"%d/%m/%Y",'mm-dd-yy':'%m-%d-%Y','mm/dd/yy':'%m/%d/%Y','yy-mm-dd':'%Y-%m-%d','yy/mm/dd':'%Y/%m/%d'}
            while i<len(get_contract):
                if (date_format.dateformat == 'dd-M-yy'):
                    convert_date=datetime.strptime(get_date[i],"%d-%b-%Y").date()
                    # print('1',get_date[i])
                elif (date_format.dateformat == 'dd-mm-yy'):
                    convert_date=datetime.strptime(get_date[i],"%d-%m-%Y").date()
                    # print('2',get_date[i])
                elif (date_format.dateformat == 'dd/mm/yy'):
                    convert_date=datetime.strptime(get_date[i],"%d/%m/%Y").date()
                    # print('3',get_date[i])
                elif (date_format.dateformat == 'mm-dd-yy'):
                    convert_date=datetime.strptime(get_date[i],"%m-%d-%Y").date()
                    # print('4',get_date[i])
                elif (date_format.dateformat == 'mm/dd/yy'):
                    convert_date=datetime.strptime(get_date[i],"%m/%d/%Y").date()
                    # print('5',get_date[i])
                elif (date_format.dateformat == 'yy-mm-dd'):
                    convert_date=datetime.strptime(get_date[i],"%Y-%m-%d").date()
                    # print('6',get_date[i])
                elif (date_format.dateformat == 'yy/mm/dd'):
                    convert_date=datetime.strptime(get_date[i],"%Y/%m/%d").date()
                    # print('7',get_date[i])
                else:
                    convert_date=datetime.strptime(get_date[i],"%d-%b-%Y").date()
                    # print('8',get_date[i])
                # print(get_contract[i],get_Scope[i],)
                create_service=Vendorcompanyserviceinfo.objects.create(type_services=types,nameofcontract=get_contract[i],abrigdge_scope_service=get_Scope[i],contract_reference_number=get_service_contract[i],contract_date=convert_date,vendor_id=vendorid.id,company_id=companyid)
                current_contract_id.append(create_service.id)
                i+=1
            vendor_service=Vendorcompanyserviceinfo.objects.filter(vendor_id=vendorid.id,type_services=types)
            print("c_id",vendor_service)
            for i,j in enumerate(vendor_service):
                get_contractfile=request.FILES.getlist(types+'file'+str(i))
                # print(j)
                for contractfile in get_contractfile:
                    create_contract_file=VendorFileUpload.objects.create(contract_id=j.id,vendor_id=vendorid.id,files=contractfile)
                    print(j,'aaa',contractfile)
        VendorRegistraion.objects.filter(id=pk).update(status=3)
        company=Companies.objects.get(id=companyid)
        if Companies.objects.filter(id=companyid).filter(Q(image__isnull=True) | Q(image='')):
            print("no images")
            imageurl=''
        else:
            imageurl=company.image.url
            print(imageurl)
        # company=Companies.objects.get(id=companyid)
        f_name=company.first_name
        l_name=company.last_name
        fullname=f_name+' '+l_name
        messages.success(request, "Vendor Details Registered Successfully.Please Wait for Admin Approval")
        subject = 'Acknowledgement of Registration'
        url= f'{request.scheme}://{request.get_host()}'
        if (user.roles_id == 2):
            context={'areacode':company.areacode,'title':vendorid.invite_vendor.title,
                    'vendorname':vendorid.invite_vendor.contactpersonname,
                    'companyname':vendorid.company.company_name,'role':user.designation_role,
                    'landline_countrycode':company.phone_countrycode,
                    'lanline':company.phonenumber,'mobile_countrycode':company.mobile_countrycode,
                    'mobile':company.mobile,'website':company.website,
                    'address':company.address,'image':imageurl,
                    'email':company.email,'fullname':fullname,'vendor_processingdays':date_format.vendor_registartion_time,'url':url}
        else:
            firstname=user.name
            lastname=user.lastname
            username=firstname+' '+lastname
            context={'areacode':company.areacode,'title':vendorid.invite_vendor.title,
                    'vendorname':vendorid.invite_vendor.contactpersonname,
                    'companyname':vendorid.company.company_name,'role':user.designation_role,
                    'landline_countrycode':user.phone_countrycode,
                    'lanline':user.phone,'mobile_countrycode':user.mobile_countrycode,
                    'mobile':user.mobile,'website':company.website,
                    'address':company.address,'image':imageurl,
                    'email':user.email,'fullname':username,'vendor_processingdays':date_format.vendor_registartion_time,'url':url}


        html_message = render_to_string('acknowledgement.html', context)
        plain_message = strip_tags(html_message)
        from_email = settings.EMAIL_HOST_USER
        to = vendorid.invite_vendor.contact_person_email

        msg=mail.send_mail(subject,plain_message, from_email, [to], html_message=html_message)      
        if (user.roles_id == 2):
            print('company')
            companyrole=user.designation_role
            if User.objects.filter(company_id=user.company.id,is_primary=1,primary_active_status=1).exists():
                approver=User.objects.filter(company_id=user.company.id,is_primary=1,primary_active_status=1).first()
                print('p',approver)
                VendorRegistraion.objects.filter(id=pk).update(approver_status=1)
            elif User.objects.filter(company_id=user.company.id,is_secondary=1).exists():
                approver=User.objects.filter(company_id=user.company.id,is_secondary=1).first()
                print('s',approver)
                VendorRegistraion.objects.filter(id=pk).update(approver_status=2)
            # vendorregistartionapprovermail(user.company.id,approver.id,companyrole)
        else:
            print('user')
            # if User.objects.filter(company_id=request.company,is_primary=1,primary_active_status=1).exists():
            #     approver=User.objects.filter(company_id=request.company,is_primary=1,primary_active_status=1).first()
            #     print('p',approver)
            # elif User.objects.filter(company_id=request.company,is_secondary=1).exists():
            #     approver=User.objects.filter(company_id=request.company,is_secondary=1).first()
            #     print('s',approver)
            # vendorregistartionapprovermail(vendorid.company.id,approver.id,companyrole)
        return redirect('custom_auth:vendorlogin')

    return render(request ,'outervendorservice.html',{'vendorid':vendorid,'companyid':companyid,'dateformat':date_format})



@login_required(login_url='/')
def vendorview(request):
    request.session['mainmenu'] = 'vendors'
    request.session['submenu']='createvendors'
    data={}
    if (request.user.roles_id == 3):
        userrights=UserRights.objects.get(user_id=request.user.id,module_id=2)
        print("ssg",userrights)
        data['rights']=userrights
    vendordetails=VendorRegistraion.objects.filter(company=request.company,status=3).order_by('-id')
    vendor_count=vendordetails.count()

    data.update({
        'vendors':vendordetails,
        'vendor_count':vendor_count,
        })

    return render(request, "vendorview.html",data)


def ReSendEmail(request,pk):
    vendor=VendorRegistraion.objects.get(id=pk)
    data={'vendor':vendor}
    return render(request,"resendemail.html",data)

def Usercreate(request):
    companyid=request.company.id
    general_setting=Settings.objects.filter(company_id=request.company.id)
    request.session['mainmenu']='users'
    request.session['submenu']='userscreate' 
    phonecodelist=[]
    phonecode = Basecountries.objects.all().values('iso3','id','phonecode').order_by('phonecode').distinct()
    for codes in phonecode:
        id=codes.get('id')
        num=codes.get('phonecode')
        iso=codes.get('iso3')
        x = re.findall("\+",num)
        if x:
            phonecodelist.append({'id':id,'phonecode':str(num),'iso':iso})
        else:
            phonecodelist.append({'id':id,'phonecode':'+'+str(num),'iso':iso})
    newlist=sorted(phonecodelist, key = lambda i:  (i['phonecode']))
    # user_department = UserDepartment.objects.filter(company_id=request.company.id)
    # get userdepartment by company and null value company
    user_department = UserDepartment.objects.filter(Q(company_id=request.company.id) | Q(company_id=None),status=1)
    user_group= UserGroup.objects.filter(Q(company_id=request.company.id) | Q(company_id=None),status=1)
    form = Usercreation()
    if request.method == 'POST':  
        print(request.POST)
        firstname=request.POST['name']
        lastname=request.POST['lastname']
        email=request.POST['email']
        designation=request.POST['designation_role']
        phone=request.POST['phone']
        mobile=request.POST['mobile']         
        get_password=request.POST['password']
        phone_countrycode=request.POST['phone_countrycode']  
        mobile_countrycode=request.POST['mobile_countrycode']
        title=request.POST['title']  
        areacode= request.POST['areacode']
        phone_number_extenstion=request.POST.get('phone_number_extenstion')
        userDepartment=request.POST.get('user_department')
        userGroup=request.POST.get('user_group')
        print(f'user_department {user_department} user_group {user_group}')
        password = make_password(get_password)
        user_department = UserDepartment.objects.get(id=userDepartment) 
        user_group = UserGroup.objects.get(id=userGroup)     
        create_user=User.objects.create(phone_number_extenstion=phone_number_extenstion,Title=title,areacode=areacode,mobile_countrycode=mobile_countrycode,phone_countrycode=phone_countrycode,mobile=mobile,phone=phone,designation_role=designation,name=firstname,lastname=lastname,email=email,password=password,company_id=companyid,roles_id=3,cin_number=request.company.cin_number,user_department=user_department,user_group=user_group)
        usercreate = request.user.roles_id
        username = create_user.name +" " + create_user.lastname
        create_user_log(request,username,'User','Create','User Created',usercreate)


        # usercreate = request.user.roles_id  
        # create_user=create_user.id

        
        # (request,create_user,'User','Create','User Created',usercreate)
         

        if Companies.objects.filter(id=companyid).filter(Q(image__isnull=True) | Q(image='')):
            print("no images")
            imageurl=''
        else:
            imageurl=request.company.image.url
            print(imageurl)
        company=Companies.objects.get(id=companyid)
        fullname=firstname+' '+lastname
        companyusername=company.first_name+' '+company.last_name
        print(companyusername)
        # url='https://irockinfo.mo.vc/'
        url=f'{request.scheme}://{request.get_host()}'
        #messages.success(request, "Vendor Details Registered Successfully. Please Wait for Admin Approval")
        subject = 'iROCK Login Credentials'
   
         
        usercreate = request.user.roles_id
        print('usercreate',usercreate)

       
        context={'areacode':company.areacode,'userrole':designation,'role':request.user.designation_role,
        'landline_countrycode':company.phone_countrycode,
        'lanline':company.phonenumber,'mobile_countrycode':company.mobile_countrycode,
        'mobile':company.mobile,'website':company.website,
        'address':company.address,'image':imageurl,
        'companyname':company.company_name,
        'cin':company.cin_number,
        'email':company.email,'fullname':fullname,'password':get_password,'url':url,'title':title,'companyusername':companyusername,'useremail':email}

        



    
        html_message = render_to_string('usercreatemail.html', context)
        plain_message = strip_tags(html_message)
        from_email = settings.EMAIL_HOST_USER
        to = email

        msg=mail.send_mail(subject,plain_message, from_email, [to], html_message=html_message)      
        
        return redirect('projects:userlist')
    else:  
        form = Usercreation()   
    data={'form':form,
          'phone':newlist,'general_setting':general_setting.count(),'company_id':companyid,'user_department':user_department,'user_group':user_group}
    return render(request,"usercreate.html",data)


def checkcompanyuseremailexists(request):
    get_email=request.GET.get('email')
    if (User.objects.filter(email__exact=get_email,status=1,company=request.company).filter(Q(roles_id=2) | Q(roles_id=3) | Q(roles_id=4)).exists()):
        data={'data':'exists'}
    else:
        data={'data':'success'}
    return JsonResponse(data)



def userview(request,pk):
    markas_read_status(request.get_full_path())
    user=User.objects.get(id=pk)
    data={'user':user}
    return render(request,'userview.html',data)

def useredit(request,pk):
    phonecodelist=[]
    phonecode = Basecountries.objects.all().values('iso3','id','phonecode').order_by('phonecode').distinct()
    for codes in phonecode:
        id=codes.get('id')
        num=codes.get('phonecode')
        iso=codes.get('iso3')
        x = re.findall("\+",num)
        if x:
            phonecodelist.append({'id':str(id),'phonecode':str(num),'iso':iso})
        else:
            phonecodelist.append({'id':str(id),'phonecode':'+'+str(num),'iso':iso})
    newlist=sorted(phonecodelist, key = lambda i:  (i['phonecode']))
    datas=User.objects.get(id=pk)
    old_email=datas.email
    form=Usercreation(instance=datas)
    user_department = UserDepartment.objects.filter(Q(company_id=request.company.id) | Q(company_id=None),status=1)
    user_group= UserGroup.objects.filter(Q(company_id=request.company.id) | Q(company_id=None),status=1)
    titles=['Mr','Mrs','Ms']
    if request.method == "POST":
        print(request.POST)
        companyid=request.company.id
        firstname=request.POST.get('name')
        lastname=request.POST.get('lastname')
        email=request.POST.get('email')
        title=request.POST.get('title')
        phone_number_extenstion=request.POST.get('phone_number_extenstion')
        new_email=request.POST.get('email')
        designation=request.POST.get('designation_role')
        phone=request.POST.get('phone')
        mobile=request.POST.get('mobile')         
        get_password='Hello@123'
        phone_countrycode=request.POST.get('phone_countrycode')  
        mobile_countrycode=request.POST.get('mobile_countrycode')
        title=request.POST.get('title')  
        areacode= request.POST.get('areacode')
        userdepartment = request.POST.get('userdepartment')
        usergroup = request.POST.get('usergroup')
        print(f'userdepartment {userdepartment} usergroup {usergroup}')
        # password = make_password('Hello@123') 
        form=Usercreation(request.POST,instance=datas)
        sender = User.objects.get(id=request.user.id)
        recipient = User.objects.filter(id=pk)
        scheme=request.scheme
        gethost=request.get_host()
        url=f"{scheme}://{gethost}/userdetails/{pk}"
        notify.send(sender, recipient=recipient,data=url, verb='Your Details Have Been Updated', description=f'Your details have been updated by {sender.name} {sender.lastname if sender.lastname != None else ""}')
        usercreate = request.user.roles_id
        fullname=firstname+' '+lastname
        create_user_log(request,fullname,'User','Edit','User Edited',usercreate)
        if form.is_valid():
            # print("yes")
            # print(form)
            forms=form.save(commit=False)
            forms.Title=title
            forms.phone_number_extenstion=phone_number_extenstion
            # forms.user_department = userdepartment
            # forms.user_group = usergroup
            forms.save()
            if(old_email != new_email):
                if Companies.objects.filter(id=companyid).filter(Q(image__isnull=True) | Q(image='')):
                    print("no images")
                    imageurl=''
                else:
                    imageurl=request.company.image.url
                    print(imageurl)
                company=Companies.objects.get(id=companyid)
                fullname=firstname+' '+lastname
                companyusername=company.first_name+' '+company.last_name
                print(companyusername)
                # url='https://irockinfo.mo.vc/'
                current_url = f"{scheme}://{gethost}"
                #messages.success(request, "Vendor Details Registered Successfully. Please Wait for Admin Approval")
                subject = 'iROCK Login Credentials'
                context={'areacode':company.areacode,'userrole':designation,'role':request.user.designation_role,
                        'landline_countrycode':company.phone_countrycode,
                        'lanline':company.phonenumber,'mobile_countrycode':company.mobile_countrycode,
                        'mobile':company.mobile,'website':company.website,
                        'address':company.address,'image':imageurl,
                        'companyname':company.company_name,
                        'cin':company.cin_number,
                        'email':company.email,'fullname':fullname,'password':get_password,'url':current_url,'title':title,'companyusername':companyusername,'useremail':email}


            
                html_message = render_to_string('usercreatemail.html', context)
                plain_message = strip_tags(html_message)
                from_email = settings.EMAIL_HOST_USER
                to = email

                msg=mail.send_mail(subject,plain_message, from_email, [to], html_message=html_message) 
            return redirect('projects:userlist')
        else:
            print(form.errors)
    data={'form':form,
          'phone':newlist,
          'datas':datas,
          'titles':titles,
          'user_department':user_department,
          'user_group':user_group
          }
    # print("data",data,'form',form)
    return render(request,'useredit.html',data)


def userdelete(request):
    userid=request.GET.get('id',None)
    if (userid != None):
        user = User.objects.filter(id=userid).update(status=0)
        usercreate = request.user.roles_id
        the_user = User.objects.get(id=userid)
       
        create_user_log(request,the_user,'User','Delete','User Deleted',usercreate)
    return JsonResponse({'data':'success'})


@login_required(login_url='/')
def Userlist(request):
    if (request.POST):
        print('asd',request.POST)
        context = {}
        search_query = request.POST.get('q',False)
        if search_query =='':
            process_list = User.objects.filter(company_id=request.company,roles_id=3,status=1).order_by('-id')
            context['query'] = search_query
        else:
            process_list = User.objects.filter((Q(name__icontains=search_query) | Q(lastname__icontains=search_query)),company=request.company,roles_id=3,status=1)
        context['query'] = search_query
        page = request.POST.get('page', 1)
        pageper_data = request.POST.get('pageperdata',10) 
        paginator = Paginator(process_list, pageper_data)
        context['process_list'] = paginator.page(page)  
        context['pageper_data'] = pageper_data
        context['scheme']=request.scheme
        context['gethost']=request.get_host()
        # render to template string
        html = render_to_string('search_user.html',context,request)
        return JsonResponse({'status':True,'html':html})
    else:
        request.session['mainmenu']='users'
        request.session['submenu']='userscreate' 
        process_list=User.objects.filter(company_id=request.company,roles_id=3,status=1).order_by('-id')
        usercount=process_list.count()
        department_count = UserDepartment.objects.filter(Q(company_id=request.company.id) | Q(company_id=None),status=1).count()
        group_count = UserGroup.objects.filter(Q(company_id=request.company.id) | Q(company_id=None),status=1).count()
        page = request.GET.get('page', 1)
        pageper_data = request.GET.get('pageperdata',10)
        paginator = Paginator(process_list, 10)
        data={"process_list":process_list,"usercount":usercount,'department_count':department_count,'group_count':group_count,'pageper_data':10,'process_list':paginator.page(page),'scheme':request.scheme,'gethost':request.get_host()}
        return render(request,"userslist.html",data)

  

def Userrights(request,pk):
    status = []
    request.session['mainmenu']='users'
    request.session['submenu']='userscreate' 
    users=User.objects.get(id=pk,roles_id=3)
    modules=Modules.objects.all()
    moduleslist=Modules.objects.all().values_list('module_name',flat=True)
    # print('moduleslist',moduleslist)
    if (request.POST):
        # print("ads------",request.POST)
        remove_comma=request.POST.get('userid')
        userid=remove_comma.replace(',','')
        for module in moduleslist:
            # print(module)
            moduleid=Modules.objects.filter(module_name=module).first()
            if module == 'Vendors':
                user=UserRights.objects.filter(module_id=moduleid.id,user_id=userid).first()
                if user == None:
                    # print("create v")
                    create_access=request.POST.get(module+'create') 
                    view_access=request.POST.get(module+'view')
                    edit_access=request.POST.get(module+'edit')
                    delete_access=request.POST.get(module+'delete')
                    # invite_access=request.POST.get(module+'vendor_invite')
                    # approve_mail_access=request.POST.get(module+'vendor_approvemail')
                    create_userrights=UserRights.objects.create(user_id=userid,module_id=moduleid.id,create=create_access,view=view_access,delete=delete_access,edit=edit_access)
                    # status.append(module)
                else:
                    # print("exists")
                    create_access=request.POST.get(module+'create') 
                    view_access=request.POST.get(module+'view')
                    edit_access=request.POST.get(module+'edit')
                    delete_access=request.POST.get(module+'delete')
                    # invite_access=request.POST.get(module+'vendor_invite')
                    # approve_mail_access=request.POST.get(module+'vendor_approvemail')
                    UserRights.objects.filter(user_id=userid,module_id=moduleid.id).update(create=create_access,view=view_access,delete=delete_access,edit=edit_access)

            else:
                user=UserRights.objects.filter(module_id=moduleid.id,user_id=userid).first()
                if user == None:
                    print('modulename1',module)
                    create_access=request.POST.get(module+'create')
                    view_access=request.POST.get(module+'view')
                    edit_access=request.POST.get(module+'edit')
                    delete_access=request.POST.get(module+'delete')
                    lock_access=request.POST.get(module+'lock')
                    unlock_access=request.POST.get(module+'unlock') 
                    create_userrights=UserRights.objects.create(unlock=unlock_access,user_id=userid,module_id=moduleid.id,create=create_access,view=view_access,edit=edit_access,delete=delete_access,lock=lock_access)
                else:
                    # print("exists")
                    print('modulename2',module)
                    create_access=request.POST.get(module+'create')
                    view_access=request.POST.get(module+'view')
                    edit_access=request.POST.get(module+'edit')
                    delete_access=request.POST.get(module+'delete')
                    lock_access=request.POST.get(module+'lock')
                    unlock_access=request.POST.get(module+'unlock')
                    UserRights.objects.filter(user_id=userid,module_id=moduleid.id).update(unlock=unlock_access,create=create_access,view=view_access,edit=edit_access,delete=delete_access,lock=lock_access)
        status = list(UserRights.objects.filter(user_id=userid).values_list('module_id__module_name',flat=True))
        sender = User.objects.get(id=request.user.id)
        # recipient = User.objects.filter(company=request.company.id,roles_id=2).first()
        scheme=request.scheme
        gethost=request.get_host()
        url=f"{scheme}://{gethost}/dashboard/dashboard"
        # General Rights Assigned To You f'General rights assigned the Modules Are {str(status)[1:-1]} ' 
        # user_name=users.name+''+users.lastname
        notify.send(sender,recipient=users,data=url, verb='General Rights Updated', description=f'{sender.name} {sender.lastname if sender.lastname != None else ""} has updated your general rights')
        return redirect('projects:userlist')
    data={'users':users,
          'modules':modules}
    return render(request,"userrights.html",data)

def userprojectrights(request,pk):
    user=User.objects.get(id=pk,roles_id=3)
    projects=Projectcreation.objects.filter(company=request.company)
    noprojects=ProjectUser.objects.filter(company=request.company,user=pk)
    projectlist=[]
   
    for i in noprojects:
        projectdetails=Projectcreation.objects.filter(company=request.company,id=i.project_id)
        projectlist.append(projectdetails)
        print('projectdetails',projectdetails)
    print('projectlist',projectlist)
    if (request.POST):
        print(request.POST)
        projectids=request.POST.getlist('projects')
        list_projects = [ProjectRights(project_id=project,user_id=pk) for project in projectids]
        if ProjectRights.objects.filter(user_id=pk).exists():
            ProjectRights.objects.filter(user_id=pk).delete()
        # print(f"project name: {[i.project.projectname.name for i in list_projects]}")
        create_projectrights=ProjectRights.objects.bulk_create(list_projects)
        if projectids:
            assigned_projects=', '.join(str(e) for e in [i.project.projectname.name for i in list_projects])
            print(f'Assigned Project {[i.project.projectname.name for i in list_projects]}')
            sender = User.objects.filter(id=request.user.id,roles_id=2).first()
            recipient = User.objects.get(id=pk)
            scheme=request.scheme
            gethost=request.get_host()
            url=f"{scheme}://{gethost}/projects/projectlist"
            notify.send(sender, recipient=recipient,data=url, verb='Project Assigned', description=f'{assigned_projects} Project Assigned by {sender.name} {sender.lastname if sender.lastname != None else ""}')
        return redirect('projects:specificprojectrights',pk=pk)
    data={'users':user,
          'projects':projects,
          'projectlist':projectlist}
    return render(request,'userprojectrights.html',data)

def Specificprojectrights(request,pk):
    projectrights=ProjectRights.objects.filter(user_id=pk)
    if (request.POST):
        return redirect('projects:userlist')
    data={'projectrights':projectrights,'id':pk}
    return render(request,'specificprojects.html',data)
 
def editoutervendorbasicinfo(request,userid,pk):
    user=User.objects.get(id=userid)
    companyid=user.company.id
    # print(companyid)
    countries=Basecountries.objects.all()
    countrylist=[]
    for countri in countries:
        if States.objects.filter(country_id=countri.id).exists():
            countrylist.append({'id':countri.id,'name':countri.name})
    maincountry=Countries.objects.all()
    vendorform=VendorRegistraion.objects.get(company_id=companyid,id=pk)
    states=States.objects.filter(country_id=vendorform.country_of_operation_id)
    vendortaxdetails=VendorTaxDetails.objects.filter(company_id=companyid,vendor_id=pk,status=1)
    if (request.POST):
        print(request.POST)
        vendor_name=request.POST['vendorname']
        country_of_operation=request.POST['vendor_country_list']
        country_of_incorporation=request.POST['vendor_country_list_incorporation']
        state=request.POST['vendor_state_list']
        Address=request.POST['address']
        website=request.POST['website']
        Registernumber=request.POST['Registernumber']
        VendorRegistraion.objects.filter(company_id=companyid,id=pk).update(vendor_name=vendor_name,country_of_operation_id=country_of_operation,country_of_incorporation_id=country_of_incorporation,state_id=state,Address=Address,website=website,RegisterNumber=Registernumber,company_id=companyid)
        get_taxids=request.POST.getlist('taxnamehdid')
        get_taxname=request.POST.getlist('taxname')
        get_taxnum=request.POST.getlist('taxnumber')
        i=0
        currenttaxids=[]
        while i < len(get_taxids):
            if (get_taxids[i]):
                VendorTaxDetails.objects.filter(id=get_taxids[i]).update(company_id=companyid,vendor=pk,Tax_Name=get_taxname[i],Tax_Number=get_taxnum[i])
                currenttaxids.append(get_taxids[i])
                print("old",get_taxids[i],get_taxname[i],get_taxnum[i])
            else:
                create_vendortax=VendorTaxDetails.objects.create(company_id=companyid,vendor_id=pk,Tax_Name=get_taxname[i],Tax_Number=get_taxnum[i])
                currenttaxids.append(create_vendortax.id)
                print("no val",get_taxids[i],get_taxname[i],get_taxnum[i])
            i+=1
        VendorTaxDetails.objects.filter(company_id=companyid,vendor_id=pk).exclude(id__in=currenttaxids).update(status=0)
        return JsonResponse({"data":"success"})
    data={"vendorform":vendorform,
         "countries":countrylist,
         "maincountry":maincountry,
         "states":states,
         "userid":user.id,
         "vendortaxdetails":vendortaxdetails,
         }
    return render(request,"editvendorform1.html",data)




def editoutervendorregistration(request,userid,pk):
    user=User.objects.get(id=userid)
    companyid=user.company.id
    titles=['Mr','Mrs','Ms']
    phonecodelist=[]
    phonecode = Basecountries.objects.all().values('iso3','id','phonecode').order_by('phonecode').distinct()
    for codes in phonecode:
        id=codes.get('id')
        num=codes.get('phonecode')
        iso=codes.get('iso3')
        x = re.findall("\+",num)
        if x:
            phonecodelist.append({'id':id,'phonecode':str(num),'iso':iso})
        else:
            phonecodelist.append({'id':str(id),'phonecode':'+'+str(num),'iso':iso})
    newlist=sorted(phonecodelist, key = lambda i:  (i['phonecode']))
    vendor=VendorRegistraion.objects.get(company_id=companyid,id=pk)
    if ():
        contact_primary_title=request.POST['Tile']
        contact_primary_first_name=request.POST['firstname']
        contact_primary_middle_name=request.POST['middlename']
        contact_primary_last_name=request.POST['lastname']
        
        contact_secondary_title=request.POST['Title']
        contact_secondary_first_name=request.POST['alternatefirstname']
        contact_secondary_middle_name=request.POST['alternatemiddlename']
        contact_secondary_last_name=request.POST['alternatelastname']

        secondary_designation=request.POST['alternateDesignation']
        primary_designation=request.POST['Designation']
        contact_primary_phone_code=request.POST['phonecountrycode']
        contact_primary_area_code=request.POST['phoneareacode']
        contact_primary_phone=request.POST['phone']
        contact_primary_mobile_code=request.POST['mobilecoutrycode']
        contact_primary_mobile=request.POST['mobile']
        contact_primary_phone_number_extenstion=request.POST['phonenumber_extenstion']

        contact_secondary_phone_code=request.POST['phonecountryalternatecode']
        contact_secondary_phone=request.POST['alternatephone']
        contact_secondary_mobile_code=request.POST['alternate_country_code']
        contact_secondary_mobile=request.POST['alternatemobile']
        contact_secondary_area_code=request.POST['alternatephoneareacode']
        contact_secondary_phone_number_extenstion=request.POST['alternatephonenumber_extenstion']


        official_primary_email=request.POST['Email']
        official_secondary_email=request.POST['alternateEmail']

        create_registrtion_vendor=VendorRegistraion.objects.filter(id=pk).update(contact_primary_first_name=contact_primary_first_name,contact_primary_middle_name=contact_primary_middle_name,contact_primary_last_name=contact_primary_last_name,contact_secondary_first_name=contact_secondary_first_name,contact_secondary_middle_name= contact_secondary_middle_name,contact_secondary_last_name=contact_secondary_last_name,secondary_designation=secondary_designation,primary_designation=primary_designation,contact_primary_phone_code=contact_primary_phone_code,contact_primary_phone=contact_primary_phone,contact_primary_mobile_code=contact_primary_mobile_code,contact_primary_mobile=contact_primary_mobile, contact_secondary_phone_code=contact_secondary_phone_code,contact_secondary_phone=contact_secondary_phone,contact_secondary_mobile_code=contact_secondary_mobile_code,contact_secondary_mobile=contact_secondary_mobile,official_primary_email=official_primary_email,official_secondary_email=official_secondary_email,contact_primary_area_code=contact_primary_area_code,contact_secondary_area_code=contact_secondary_area_code,contact_primary_title=contact_primary_title,contact_secondary_title=contact_secondary_title,contact_primary_phone_number_extenstion=contact_primary_phone_number_extenstion,contact_secondary_phone_number_extenstion=contact_secondary_phone_number_extenstion)
        data={'vendorid':"success"}
        return JsonResponse(data)

    data={"vendor":vendor,
          "titles":titles,
          "phone":newlist,
          'userid':user.id,
          }
    return render(request,"editvendorform2.html",data)



def Vendorpanel(request,companyid):
    request.session['mainmenu']='panel'
    request.session['submenu']='vendorpanel'
    userlistid=User.objects.filter(company_id=companyid,status=1).values_list('id',flat=True)
    # print(userlistid)
    if User.objects.filter(company_id=companyid,status=1,id__in=userlistid,is_primary=1):
        # print('if')
        return redirect('projects:vendorappprovepanellist',companyid=companyid)
    else:
        # print('else')
        users=User.objects.filter(company_id=companyid,status=1).values('id','name','lastname')
        userlist=[]
        companyrole=request.user.designation_role
        # print(companyrole)
        for user in users:
            userid=user['id']
            firstname=user['name']
            lastname=user['lastname']
            fullname=firstname+' '+lastname
            userdata=UserRights.objects.filter(user_id=userid,vendor_approve=1).first()
            if (userdata != None):
                # print(userid)          
                userlist.append({'id':userid,'username':fullname}) 
        if (request.POST):
            print(request.POST)
            postvalues=request.POST
            primary=postvalues.get('primarymember')
            secondary=postvalues.get('secondarymember')
            current_url = f'{request.scheme}://{request.get_host()}'
            User.objects.filter(id=primary,company_id=companyid).update(is_primary=1,primary_active_status=1)
            User.objects.filter(id=secondary,company_id=companyid).update(is_secondary=1)
            User.objects.filter(company_id=companyid).exclude(id=primary).update(is_primary=0,primary_active_status=0)
            User.objects.filter(company_id=companyid).exclude(id=secondary).update(is_secondary=0)
            vendorprimaryapprovermail(companyid,primary,companyrole,current_url)
            vendorsecondaryapprovermail(companyid,secondary,companyrole,current_url)
            return redirect('projects:vendorappprovepanellist',companyid=companyid)
        data={'users':userlist}
        return render(request,'vendorapprovepanel.html',data)

def Vendorappprovepanellist(request,companyid):
    primary=User.objects.get(company_id=companyid,is_primary=1)
    secondary=User.objects.get(company_id=companyid,is_secondary=1)
    # print(secondary)
    data={'primary':primary,
          'secondary':secondary}
    return render(request,'vendorapprovepanelist.html',data)

def EditVendorapprovepanel(request,companyid):
    companyrole=request.user.designation_role
    users=User.objects.filter(company_id=companyid,status=1).values('id','name','lastname')
    userlist=[]
    for user in users:
        userid=user['id']
        firstname=user['name']
        lastname=user['lastname']
        fullname=firstname+' '+lastname
        userdata=UserRights.objects.filter(user_id=userid,vendor_approve=1).first()
        if (userdata != None):         
            userlist.append({'id':userid,'username':fullname}) 
    primary=User.objects.get(company_id=companyid,is_primary=1)
    # print(primary)
    secondary=User.objects.get(company_id=companyid,is_secondary=1)
    current_url = f'{request.scheme}://{request.get_host()}'
    if (request.POST):
        # print(request.POST)
        postvalues=request.POST
        primary=postvalues.get('primarymember')
        secondary=postvalues.get('secondarymember')
        if (User.objects.filter(company_id=companyid,id=primary,is_primary=1).exists()):
            print("yes")
        else:
            User.objects.filter(company_id=companyid,id=primary).update(is_primary=1,primary_active_status=1)
            vendorprimaryapprovermail(companyid,primary,companyrole,current_url)
        User.objects.filter(company_id=companyid).exclude(id=primary).update(is_primary=0,primary_active_status=0)
        if (User.objects.filter(company_id=companyid,id=secondary,is_secondary=1).exists()):
            print("yes")
        else:
            User.objects.filter(company_id=companyid,id=secondary).update(is_secondary=1)
            vendorsecondaryapprovermail(companyid,secondary,companyrole,current_url)
        User.objects.filter(company_id=companyid).exclude(id=secondary).update(is_secondary=0)
        return redirect('projects:vendorappprovepanellist',companyid=companyid)
    data={'users':userlist,
          'primary':primary,
          'secondary':secondary}
    return render(request,'editvendorapprovepanel.html',data)

def Getapproveusers(request):
    id=request.GET.get('id',None)
    users=User.objects.filter(company_id=request.company.id,status=1).values('id','name','lastname')
    userlist=[]
    for user in users:
        userid=user['id']
        firstname=user['name']
        lastname=user['lastname']
        fullname=firstname+' '+lastname
        userdata=UserRights.objects.filter(user_id=userid,vendor_approve=1).first()
        if (userdata != None):
            if (int(id) != userid):
                userlist.append({'id':userid,'username':fullname})
    print(userlist)
    data={'data':userlist}
    return JsonResponse(data)

def Getactivestatus(request):
    status=request.GET.get('activestatus',None)
    print(status)
    # print(currentuser)
    if (status == '1'):
        User.objects.filter(id=request.user.id).update(primary_active_status=1)
        data={'data':'active'}
    else:
        User.objects.filter(id=request.user.id).update(primary_active_status=0)
        data={'data':'inactive'}
    return JsonResponse(data)



def vendormasterlist(request,patt=None):

    if request.method == 'POST':
        if (request.FILES):
            print('if')
            vendorfile=request.FILES.get('vendorfile',None)
            if (vendorfile):
                print("vendorfile",vendorfile)
                dataset= Dataset()
                imported_data = dataset.load(vendorfile.read(),format='xlsx')
                i=0
                all_data=[]
                while i<len(imported_data):
                    print('len',len(imported_data))
                    print('i',i)
                    print(imported_data[i])
                    if (imported_data[i][1] == None and imported_data[i][2] == None and imported_data[i][3] == None and imported_data[i][4] == None and imported_data[i][5] == None):
                        print('empty row')
                    else:
                        email_text=imported_data[i][5]
                        all_data.append({'vendor_name':imported_data[i][1],'title':imported_data[i][2],'primary_name':imported_data[i][3],'primary_last_name':imported_data[i][4],'email':email_text.strip()})
                    i+=1
                request.session['allimportvendor']=all_data
            return JsonResponse({'data':'success'})
        # call get context data
        context = {}
        try:
            userrights=UserRights.objects.get(user_id=request.user.id,module_id=2)
        except UserRights.DoesNotExist:
            userrights=''
        # if (self.request.user.roles_id == 3):
        #     userrights=UserRights.objects.get_by_module(self.request.user.id,6)
        #     context['rights']=userrights
        search_query = request.POST.get('q',False)
        if search_query != '':
            vendors = ContractMasterVendor.objects.filter(company=request.user.company,status=True,vendor_name__icontains=search_query).order_by('-id')
            context['query'] = search_query
        else:
            vendors = ContractMasterVendor.objects.filter(company=request.user.company,status=True).order_by('-id')
        context['request'] = request  
        context['query'] = search_query
        page = request.POST.get('page', 1)
        pageper_data = request.POST.get('pageperdata',10)
        paginator = Paginator(vendors, pageper_data)
        context['vendor_master_count'] =  vendors.count()
        context['rights']=userrights
        context['vendors'] = paginator.page(page)
        context['pageper_data'] = pageper_data
        context['scheme']=request.scheme
        context['gethost']=request.get_host()

        # render to template string
        html = render_to_string('search_vendormaster.html',context,request)
        return JsonResponse({'status':True,'html':html})

    else:
        markas_read_status(request.get_full_path())
        request.session['submenu']='vendor_masters'
        list_vendor_master=ContractMasterVendor.objects.filter(company_id=request.company.id,status=1).order_by('-id')
        vendor_master_count=ContractMasterVendor.objects.filter(company_id=request.company.id,status=1).count()
        
        page = request.GET.get('page', 1)

        paginator = Paginator(list_vendor_master, 10)
        vendor_master_list = paginator.page(page)
        try:
            userrights=UserRights.objects.get(user_id=request.user.id,module_id=2)
        except UserRights.DoesNotExist:
            userrights=''
    
    
    
    get_settings = Settings.objects.get_company(request.company).values_list('currency',flat=True).first()
    if get_settings!=None:
        currency = Basecountries.objects.get_by_id(literal_eval(get_settings))
        currency_count=currency.count()
    else:
        currency_count=0
        currency=[]
    if (request.GET):
        getvendorname=request.GET.get('vendor','')
        value=request.GET.get('sortname','')   
        sortname=request.GET.get('sortvendorname','')   
        print('getvendorname',getvendorname)
        print('value',value)
        print('sn',sortname)
        
        if (getvendorname !='' and value == '' and sortname == ''):
            vendor_master_list=ContractMasterVendor.objects.filter(vendor_name__icontains=getvendorname,company=request.company,status=1)
            vendor_master_count=vendor_master_list.count()
            
            page = request.GET.get('page', 1)

            paginator = Paginator(vendor_master_list, 10)
            paginator_vendor_master_list = paginator.page(page)


            data={'vendors':paginator_vendor_master_list,'vendor_master_count':vendor_master_count,'se_vendor':getvendorname,'currency_count':currency_count}
            if (request.user.roles_id == 3):
                data['rights']=userrights
            return render(request,'vendormasterlist.html',data)

        elif (value != '' or sortname != '' and getvendorname != ''):
            print('both')
            print(len(value),len(sortname),len(getvendorname))
            print({'value':value,'sortname':sortname,'getvendorname':getvendorname})
            if (value != ''):
                print('vin')
                if (value == 'descending'):
                
                    vendor_master_list=ContractMasterVendor.objects.filter(vendor_name__icontains=getvendorname,company_id=request.company.id,status=1).order_by('-id')
                    
                                
                    page = request.GET.get('page', 1)

                    paginator = Paginator(vendor_master_list, 10)
                    paginator_vendor_master_list = paginator.page(page)

                    data={'vendors':paginator_vendor_master_list,'vendor_master_count':vendor_master_count,'se_vendor':getvendorname,'currency_count':currency_count}
                    if (request.user.roles_id == 3):
                        data['rights']=userrights
                    return render(request,'vendormasterlist.html',data) 
                else:
                    vendor_master_list=ContractMasterVendor.objects.filter(vendor_name__icontains=getvendorname,company_id=request.company.id,status=1).order_by('id')
                    vendor_master_count=vendor_master_list.count()
                    print(vendor_master_list)
                    #code for pagenation
                    page = request.GET.get('page', 1)
                    paginator = Paginator(vendor_master_list, 10)
                    paginator_vendor_master_list = paginator.page(page)
                    data={'vendors':paginator_vendor_master_list,'vendor_master_count':vendor_master_count,'se_vendor':getvendorname,'currency_count':currency_count}
                    if (request.user.roles_id == 3):
                        data['rights']=userrights
                    return render(request,'vendormasterlist.html',data)  
            elif (sortname):
                print('vname')
                if (sortname == 'descending'):
                    
                    vendor_master_list=ContractMasterVendor.objects.filter(vendor_name__icontains=getvendorname,company_id=request.company.id,status=1).order_by('-vendor_name')
                    #code for pagenation
                    page = request.GET.get('page', 1)
                    paginator = Paginator(vendor_master_list, 10)
                    paginator_vendor_master_list = paginator.page(page)
                    print('pagination working desc') 
                    data={'vendors':paginator_vendor_master_list,'vendor_master_count':vendor_master_count,'se_vendor':getvendorname,'currency_count':currency_count}
                    if (request.user.roles_id == 3):
                        data['rights']=userrights
                    return render(request,'vendormasterlist.html',data) 
                else:
                    vendor_master_list=ContractMasterVendor.objects.filter(vendor_name__icontains=getvendorname,company_id=request.company.id,status=1).order_by('vendor_name')
                    vendor_master_count=vendor_master_list.count()
                    print(vendor_master_list)
                    #code for pagenation
                    page = request.GET.get('page', 1)
                    paginator = Paginator(vendor_master_list, 10)
                    paginator_vendor_master_list = paginator.page(page)
                    data={'vendors':paginator_vendor_master_list,'vendor_master_count':vendor_master_count,'se_vendor':getvendorname,'currency_count':currency_count}
                    if (request.user.roles_id == 3):
                        data['rights']=userrights
                    return render(request,'vendormasterlist.html',data)  
        
        elif (getvendorname == '' and value != '' and sortname == ''):
            print('only sort',value)
            if (value == 'descending'):
                vendor_master_list=ContractMasterVendor.objects.filter(company_id=request.company.id,status=1).order_by('-id')
                page = request.GET.get('page', 1)
                paginator = Paginator(vendor_master_list, 10)
                paginator_vendor_master_list = paginator.page(page)
                data={'vendors':paginator_vendor_master_list,'vendor_master_count':vendor_master_count,'currency_count':currency_count}
                if (request.user.roles_id == 3):
                    data['rights']=userrights
                return render(request,'vendormasterlist.html',data) 
            else:
                vendor_master_list=ContractMasterVendor.objects.filter(company_id=request.company.id,status=1).order_by('id')
                vendor_master_count=vendor_master_list.count()
                print(vendor_master_list)
                page = request.GET.get('page', 1)
                paginator = Paginator(vendor_master_list, 10)
                paginator_vendor_master_list = paginator.page(page)
                data={'vendors':paginator_vendor_master_list,'vendor_master_count':vendor_master_count,'currency_count':currency_count}
                if (request.user.roles_id == 3):
                    data['rights']=userrights
                return render(request,'vendormasterlist.html',data)  

#sortvendorname
        elif (getvendorname == '' and value == '' and sortname != ''):
            print('vendor name sort')
            if (sortname == 'descending'):
                print('yes')
                vendor_master_list=ContractMasterVendor.objects.filter(vendor_name__icontains=getvendorname,company_id=request.company.id,status=1).order_by('-vendor_name')
                vendor_master_count=vendor_master_list.count()
                print('vn',vendor_master_list)
                page = request.GET.get('page', 1)
                paginator = Paginator(vendor_master_list, 10)
                paginator_vendor_master_list = paginator.page(page)
                data={'vendors':paginator_vendor_master_list,'vendor_master_count':vendor_master_count,'currency_count':currency_count}
                if (request.user.roles_id == 3):
                    data['rights']=userrights
                return render(request,'vendormasterlist.html',data) 
            else:
                vendor_master_list=ContractMasterVendor.objects.filter(vendor_name__icontains=getvendorname,company_id=request.company.id,status=1).order_by('vendor_name')
                vendor_master_count=vendor_master_list.count()
                page = request.GET.get('page', 1)
                paginator = Paginator(vendor_master_list, 10)
                paginator_vendor_master_list = paginator.page(page)
                data={'vendors':paginator_vendor_master_list,'vendor_master_count':vendor_master_count,'currency_count':currency_count}
                if (request.user.roles_id == 3):
                    data['rights']=userrights
                return render(request,'vendormasterlist.html',data)

    data={}
    if (request.user.roles_id == 3):
        data['rights']=userrights
        
                  
    data.update({'vendors':vendor_master_list,'vendor_master_count':vendor_master_count,'currency_count':currency_count})
    return render(request,'vendormasterlist.html',data)





#vendor master
def createvendorexcelsheet(request):

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()

    txt = 'Select a value from a drop down list'
    locked   = workbook.add_format({'locked': True})
    unlocked = workbook.add_format({'locked': False})
    worksheet.protect()
    worksheet.write(0,0,'S/N',locked )
    worksheet.write('B1','Vendor Name',locked)
    worksheet.write('C1','Title',locked)
    worksheet.write('D1','First Name',locked)
    worksheet.write('E1','Last Name',locked)
    worksheet.write('F1','Email Address',locked)

    worksheet.set_column('B:B', 40)
    worksheet.set_column('D:D', 40)
    worksheet.set_column('E:E', 40)
    worksheet.set_column('F:F', 40)

    format1 = workbook.add_format({'bg_color': '#FFC7CE',
                               'font_color': '#9C0006'})
    worksheet.conditional_format('B2:B1000', {'type':   'duplicate',
                                              'format': format1})
    # worksheet.conditional_format('E2:E1000', {'type':   'duplicate',
                                            #   'format': format1})
    for i in range(2,1001):
        worksheet.write('A'+str(i),'', unlocked)
        worksheet.write('B'+str(i),'', unlocked)
        worksheet.write('C'+str(i),'', unlocked)
        worksheet.write('D'+str(i),'', unlocked)
        worksheet.write('E'+str(i),'', unlocked)
        worksheet.write('F'+str(i),'', unlocked)
        worksheet.data_validation('C'+str(i)+'', {'validate': 'list',
                                  'source': ['Mr', 'Mrs', 'Ms']})
    workbook.close()
    output.seek(0)
    filename = 'vendormaster.xlsx'
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=%s' % filename

    return response


def checkvendorimport(request):
    all_data=request.session['allimportvendor']
    # print(all_data)
    error_list=[]
    countlist=0
    for index,vendor in enumerate(all_data): 
        if(ContractMasterVendor.objects.filter(vendor_name__exact=vendor['vendor_name'],company=request.company,status=1).exists() or vendor['vendor_name'] == None):
            countlist+=1
            # print(vendor['vendor_name'],vendor['email'])
            error_list.append(index+1)
        elif (vendor['title'] ==  None):
            countlist+=1
            error_list.append(index+1)
        elif (vendor['primary_name'] ==  None):
            countlist+=1
            error_list.append(index+1)
            # print("in",index)
        # elif (User.objects.filter(email__exact=vendor['email'],company=request.company,status=1).exists() or vendor['email'] == None):
            # countlist+=1
            # error_list.append(index+1)
            # print(vendor['vendor_name'],vendor['email'])

    convert_str = str(error_list)[1:-1]
    # print(convert_str)

    if(request.POST):
        # print(request.POST)
        vendorname=request.POST.getlist('vendorname')
        title=request.POST.getlist('title')
        first_name=request.POST.getlist('first_name')
        last_name=request.POST.getlist('last_name')
        email=request.POST.getlist('email')
        user =request.user
        userroles = request.user.roles_id
        i=0
        while i<len(vendorname):
            # print('c',i,vendorname[i],title[i],contactperson[i],email[i])
            vin_number=generatecin(request.company.id)
            vin=vin_number[0]
            setvin=vin_number[1]
            create_contract_vendor=ContractMasterVendor.objects.create(set_vin_id=setvin,company_id=request.company.id,contactpersontitle=title[i],vendor_name=vendorname[i],vin=vin,contactpersonname=first_name[i],contactpersonlastname=last_name[i],contactpersonemail=email[i])
            password = make_password('Hello@123')
            create_user=User.objects.create(email=email[i],name=first_name[i],lastname=last_name[i],company_id=request.company.id,roles_id=4,Title=title[i],password=password,designation_role='Vendor',cin_number=vin,contactpersonstatus=1)
            i+=1
            vendor_id=create_contract_vendor.id
            company=request.company
            companyrole=request.user.designation_role
            subject="iROCK Vendor Registration"
            current_url = f'{request.scheme}://{request.get_host()}'
            vendorloginmail(subject,company,companyrole,vendor_id,user,userroles ,current_url)
        return redirect('projects:vendormasterlist')
                
    titles=['Mr','Mrs','Ms']        
    data={'all_data':all_data,
          'all_data_count':len(all_data),
          'titles':titles,
          'companyid':request.company.id,
          'countlist':countlist,
          'error_list':convert_str
          }
    # print(len(same_vendor))
    # print(emptydata_vendor)
    # print(create_datas)
    return render(request,'checkvendorimport.html',data)


def sendmailforvendor(request):
    vendor_id=request.GET.get('data')
    company=request.company
    companyrole=request.user.designation_role
    subject="iROCK Vendor Registration"
    user =request.user
    userroles = request.user.roles_id
    current_url = f'{request.scheme}://{request.get_host()}'

    vendorloginmail(subject,company,companyrole,vendor_id,user,userroles,current_url)
    return JsonResponse({'data':'success'})

def downloadimportvendorexcel(request):
    alldata_vendor=request.session['allimportvendor']
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()



    worksheet.write('A1','S/N')
    worksheet.write('B1','Vendor Name')
    worksheet.write('C1','Title')
    worksheet.write('D1','First Name')
    worksheet.write('E1','Last Name')
    worksheet.write('F1','Email Address')
    # print(alldata_vendor)
    i=0
    s_num=1
    count=2
    while i<len(alldata_vendor):
        worksheet.write('A'+str(count)+'',s_num)
        if (alldata_vendor[i]['vendor_name'] == None):
            worksheet.write_comment('B'+str(count)+'', 'This Field is Required')
        elif (ContractMasterVendor.objects.filter(vendor_name__exact=alldata_vendor[i]['vendor_name']).exists()):
            worksheet.write('B'+str(count)+'',alldata_vendor[i]['vendor_name'])
            worksheet.write_comment('B'+str(count)+'', 'Vendor name already exists')
        else:
            worksheet.write('B'+str(count)+'',alldata_vendor[i]['vendor_name'])

        if (alldata_vendor[i]['title'] == None):
            worksheet.write_comment('C'+str(count)+'', 'This Field is Required')
        else:
            worksheet.write('C'+str(count)+'',alldata_vendor[i]['title'])
        worksheet.data_validation('C'+str(count)+'', {'validate': 'list',
                              'source': ['Mr', 'Mrs', 'Ms']})

        if (alldata_vendor[i]['primary_name'] == None):
            worksheet.write_comment('D'+str(count)+'', 'This Field is Required')
        else:
            worksheet.write('D'+str(count)+'',alldata_vendor[i]['primary_name'])

        if (alldata_vendor[i]['primary_name'] == None):
            worksheet.write_comment('E'+str(count)+'', 'This Field is Required')
        else:
            worksheet.write('E'+str(count)+'',alldata_vendor[i]['primary_last_name'])

        if (alldata_vendor[i]['email'] == None):
            worksheet.write_comment('F'+str(count)+'', 'This Field is Required')
        elif(User.objects.filter(email__exact=alldata_vendor[i]['email']).exists()):
            worksheet.write('F'+str(count)+'',alldata_vendor[i]['email'])
            worksheet.write_comment('F'+str(count)+'', 'Email already exists')
        else:
            worksheet.write('F'+str(count)+'',alldata_vendor[i]['email'])

        count +=1
        s_num +=1
        i+=1

    workbook.close()
    output.seek(0)
    filename = 'Re-vendormaster.xlsx'
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=%s' % filename

    return response


# def createimportvendordetails(request):

#     if(request.POST):
#         print(request.POST)
#         vendorname=request.POST.getlist('vendorname')
#         title=request.POST.getlist('title')
#         contactperson=request.POST.getlist('contactperson')
#         email=request.POST.getlist('email')

#         i=0
#         while i<len(vendorname):
#             print(i,vendorname[i],title[i],contactperson[i],email[i])
#             # vin_number=generatecin(request.company.id)
#             # vin=vin_number[0]
#             # setvin=vin_number[1]
#             # create_contract_vendor=ContractMasterVendor.objects.create(set_vin_id=setvin,company_id=request.company.id,contact_primary_title=title[i],vendor_name=vendorname[i],vin=vin,contact_primary_first_name=contactperson[i],official_primary_email=email[i])
#             # password = make_password('hello@123')
#             # create_user=User.objects.create(email=email[i],name=contactperson[i],company_id=request.company.id,roles_id=4,Title=title[i],password=password,designation_role='Vendor')
#             i+=1

#         return JsonResponse({'data':'success'})


def createvendormaster(request): 
    request.session['submenu']='vendor_masters'
    general_setting=Settings.objects.filter(company_id=request.company.id)
    company_id=request.company.id
    if (request.POST):
        print(request.POST)
        title=request.POST.get('title')
        vendorname=request.POST.get('vendor_name')
        primary_first_name=request.POST.get('primary_first_name')
        # primary_middle_name=request.POST.get('primary_middle_name')   
        primary_last_name=request.POST.get('primary_last_name')
        primary_email=request.POST.get('email')
        vin_number=generatecin(request.company.id)
        vin=vin_number[0]
        setvin=vin_number[1]
        # print(vin,setvin)
        create_contract_vendor=ContractMasterVendor.objects.create(set_vin_id=setvin,company_id=request.company.id,contactpersontitle=title,vendor_name=vendorname,vin=vin,contactpersonname=primary_first_name,contactpersonlastname=primary_last_name,contactpersonemail=primary_email,created_by=request.user)
        password = make_password('Hello@123')
        create_user=User.objects.create(email=primary_email,name=primary_first_name,lastname=primary_last_name,company_id=request.company.id,roles_id=4,Title=title,password=password,designation_role='Vendor',cin_number=vin,contactpersonstatus=1)
        messages.success(request, 'Vendor Added Successfully') 
        vendor_id=create_contract_vendor.id
        company=request.company
        companyrole=request.user.designation_role
        subject="iROCK Vendor Registration"
        user =request.user
        userroles = request.user.roles_id
        current_url = f'{request.scheme}://{request.get_host()}'
        
        vendorloginmail(subject,company,companyrole,vendor_id,user,userroles,current_url)
        usercreate=request.user.roles_id
        full_name = primary_first_name + " " + primary_last_name
        create_user_log(request,full_name,'Vendor Master','Create',"Vendor has been Created",usercreate)
        return redirect('projects:vendormasterlist')
    titles=['Mr','Mrs','Ms']
    data={'titles':titles,'general_setting':general_setting.count(),'company_id':company_id} 
    return render(request,'createvendormaster.html',data)

def editvendormaster(request,pk):
    vendor=ContractMasterVendor.objects.get(id=pk)
    if (request.POST):
        title=request.POST.get('title')
        vendorname=request.POST.get('vendor_name')
        primary_first_name=request.POST.get('primary_first_name')
        # primary_middle_name=request.POST.get('primary_middle_name')
        primary_last_name=request.POST.get('primary_last_name')
        primary_email=request.POST.get('email')
        pervious_email = ContractMasterVendor.objects.filter(id=pk).first()
        ContractMasterVendor.objects.filter(id=pk).update(contactpersontitle=title,vendor_name=vendorname,contactpersonname=primary_first_name,contactpersonlastname=primary_last_name,contactpersonemail=primary_email)
        get_vendordetails = User.objects.filter(company=request.company,cin_number=vendor.vin,contactpersonstatus=1).update(Title=title,name=primary_first_name,lastname=primary_last_name,email=primary_email)
        vendor_id = User.objects.filter(company=request.company,cin_number=vendor.vin,contactpersonstatus=1).values_list('id',flat=True).first()
        print(f"vendor_id {vendor_id}")
        vendor_id=pk
        company=request.company
        companyrole=request.user.designation_role
        # if Client User update Vendor name or Contact person or Title, notification will send to the vendor
        subject="iROCK Vendor Registration"
        current_email = ContractMasterVendor.objects.filter(id=pk).first()
        user =request.user
        userroles = request.user.roles_id
        current_url = f'{request.scheme}://{request.get_host()}'

        if pervious_email != None:
            if pervious_email.contactpersonemail != current_email.contactpersonemail:
                vendorloginmail(subject,company,companyrole,vendor_id,user,userroles,current_url)
        # if Client User update Vendor name or Contact person or Title, notification will send to the vendor
        sender = User.objects.get(id=request.user.id)
        recipient = User.objects.filter(company=request.company,cin_number=vendor.vin,contactpersonstatus=1)
        scheme=request.scheme
        gethost=request.get_host()
        vendor_user=User.objects.filter(company=request.company,cin_number=vendor.vin,contactpersonstatus=1).first()
        url=f"{scheme}://{gethost}/userdetails/{vendor_user.id}"
        notify.send(sender, recipient=recipient,data=url, verb='Your Details Have Been Updated', description=f'Your details have been updated by {sender.name} {sender.lastname if sender.lastname != None else ""}')
        usercreate=request.user.roles_id
        full_name = primary_first_name + " " +primary_last_name
        create_user_log(request,full_name,'Vendor Master','Edit','Vendor Edited',usercreate)
        return redirect('projects:vendormasterlist')
    titles=['Mr','Mrs','Ms']
    data={'vendor':vendor,'titles':titles}
    return render(request,'editvendormaster.html',data)

    

def Vendordelete(request):
    vendor_id=request.GET["id"]
    # print("vendor_id",vendor_id)
    delete_vendor_master=ContractMasterVendor.objects.filter(id=vendor_id)
    delete_vendor_master.update(status=0)
    vendoruser=ContractMasterVendor.objects.get(id=vendor_id)
    user_id = User.objects.filter(cin_number=vendoruser.vin,company=request.company)
    user_id.update(status=0)
    deleted_user =  User.objects.filter(cin_number=vendoruser.vin,company=request.company).first()
    print(deleted_user,'deleted_user''')
    full_name = deleted_user.name + " " + deleted_user.lastname
    usercreate=request.user.roles_id
    create_user_log(request,full_name,'Vendor Master','Deleted','Vendor Deleted',usercreate)
    Notify_chat.objects.filter(sender_id__in=[i.id for i in user_id]).delete()
    return JsonResponse({'data':'success'})


def checkcontractduplicate(request):
    vendorname=request.GET.get('name',None)
    vendor_id = request.GET.get('id',None)
    data = {'data':'exists'} if ContractMasterVendor.objects.filter(company_id=request.company.id,vendor_name__exact=vendorname,status=1).exclude(id=vendor_id).exists() else {'data':'not exist'}       
    return JsonResponse(data)

class contract_detailed_view(View):
    def get(self,request,contract_id):
        
        Contract=ContractMaster.objects.filter(id=contract_id).first()
        data={'contract':Contract}
        if (request.user.roles_id == 3):
            userrights=UserRights.objects.get(user_id=request.user.id,module_id=4)
            data['rights']=userrights

        
        
        
        
        return render(request,'contract_detailed_view.html',data)
    
 
    
def contractmasterlist(request,patt=None):
    request.session['submenu']='contract_projects'
    # scheme=request.scheme
    # gethost=request.get_host
    # print("currentdomain",currentdomain)
    values=ContractMaster.objects.filter(status=1).values_list('contractvendor_id', flat=True).distinct()
    company=Settings.objects.filter(company_id=request.company.id).first()
    projects=Projectcreation.objects.filter(company=request.user.company,status=0)    
    contractlist=ContractMasterVendor.objects.filter(company_id=request.user.company_id,status=1).order_by('-id').values_list('id',flat=True)
    contractmastercount= ContractMaster.objects.filter(company_id=request.user.company_id,status=1).count()
    vendormasterlist=[]
    for vendorid in contractlist:
        if (ContractMaster.objects.filter(contractvendor_id=vendorid,status=1).exists()):
            getvendormasterdetails=ContractMasterVendor.objects.filter(company_id=request.user.company_id,status=1,id=vendorid).first()
            vendormasterlist.append({'id':getvendormasterdetails.id,'vin':getvendormasterdetails.vin,'vendor_name':getvendormasterdetails.vendor_name,'active_status':getvendormasterdetails.active_status})
    page = request.GET.get('page', 1)
    paginator = Paginator(vendormasterlist, 10)
    vendor_master_list = paginator.page(page)
    if(request.POST):
        context = {}
        if (request.user.roles_id == 3):
            userrights=UserRights.objects.get_by_module(request.user.id,6)
            context['rights']=userrights
        search_query = request.POST.get('q',False)
        if search_query:
            contractlist = ContractMasterVendor.objects.filter(id__in=values,company=request.user.company,status=True,vendor_name__icontains=search_query).order_by('-id')
            context['query'] = search_query
        else:
            contractlist = ContractMasterVendor.objects.filter(id__in=values,company=request.user.company,status=True).order_by('-id')
        context['query'] = search_query
        page = request.POST.get('page', 1)
        pageper_data = request.POST.get('pageperdata',10)
        paginator = Paginator(contractlist, pageper_data)
        context['contractmastercount'] =  contractlist.count()
        context['contractlist'] = paginator.page(page)
        context['pageper_data'] = pageper_data
        context['scheme']=request.scheme
        context['gethost']=request.get_host()

        if (request.FILES):
            project=request.POST.get('projects')
            project_discipline=request.POST.get('project_discipline')
            disciplinetype=request.POST.get('disciplinetype')
            contractfile=request.FILES.get('contractfile',None)
            if (contractfile != None):
                dataset= Dataset()
                imported_data = dataset.load(contractfile.read(),format='xlsx')
                emptyuploadcontractlist=[]
                existscontracts=[]
                original_contract=[]
                amendment_list=[]
                new_vendor=[]
                all_contracts=[]
                i=0
                while i<len(imported_data):
                    if (imported_data[i][1] == None and imported_data[i][2] == None and imported_data[i][3] == None and imported_data[i][4] == None and imported_data[i][5] == None and imported_data[i][6] == None and imported_data[i][7] == None and imported_data[i][8] == None and imported_data[i][9] == None):
                        pass
                    elif (i != 0):
                        date_time=imported_data[i][7]
                        # print(date_time)
                        if (date_time):
                            # date_time_obj = datetime.strptime(str(date_time), '%Y-%m-%d')
                            # date_string = date_time_obj.strftime('%Y-%m-%d')
                            convert_date=json.dumps(str(date_time))
                        else:
                            convert_date=None
                        all_contracts.append({'vendor_name': imported_data[i][1],'type_contract':imported_data[i][2],'type_service':imported_data[i][3],'name_contract':imported_data[i][4],'con_num':imported_data[i][5],'amendment_refnum':imported_data[i][6],'executed_date':convert_date,'max_value_currency':imported_data[i][8],'max_value_amount':imported_data[i][9]})
                        print(f'all_contracts {all_contracts}')
                    i+=1
                date_file={'project':project,'project_discipline':project_discipline,'disciplinetype':disciplinetype,
                'all_contracts':all_contracts}
                request.session['allcontratdata']=date_file
                # request.session['allcontracts']=all_contracts
                return redirect('projects:checkvendorcontractimport')
   
      
        html = render_to_string('search_contractmaster.html',context,request)
        return JsonResponse({'status':True,'html':html})
    
    if (request.GET):
        print('3-----')
        print('Get method ---')
        getvendorname=request.GET.get('vendor','')
        value=request.GET.get('sortname','')   
        sortname=request.GET.get('sortvendorname','')   
        print('getvendorname',getvendorname)
        print('value',value)
        print('sn',sortname)
        if (getvendorname !='' and value == '' and sortname == ''):
            print('asd')
            contractlist=ContractMasterVendor.objects.filter(vendor_name__icontains=getvendorname,company_id=request.user.company_id,status=1).order_by('-id').values_list('id',flat=True)

            vendormasterlist=[]
            for vendorid in contractlist:
                if (ContractMaster.objects.filter(contractvendor_id=vendorid,status=1).exists()):
                    getvendormasterdetails=ContractMasterVendor.objects.filter(company_id=request.user.company_id,status=1,id=vendorid).first()
                    vendormasterlist.append({'id':getvendormasterdetails.id,'vin':getvendormasterdetails.vin,'vendor_name':getvendormasterdetails.vendor_name,'active_status':getvendormasterdetails.active_status})
            print('1',vendormasterlist)
            data={}
            if (request.user.roles_id == 3):
                userrights=UserRights.objects.get(user_id=request.user.id,module_id=4)
                data['rights']=userrights
            page = request.GET.get('page', 1)
            paginator = Paginator(vendormasterlist,10)
            paginator_contractvendor_master_list = paginator.page(page)
        
            # print(paginator)
            
            data.update({'contractlist':paginator_contractvendor_master_list,'contractmastercount':len(paginator_contractvendor_master_list),'se_vendor':getvendorname,'projects':projects,'company':company})
            return render(request,'contractmasterlist.html',data)

        elif (value != '' or sortname != '' and getvendorname != ''):
            print('both')
            print(len(value),len(sortname),len(getvendorname))
            if (value != ''):
                
                if (value == 'descending'):
                    #checking id by descending
                    contractlist=ContractMasterVendor.objects.filter(vendor_name__icontains=getvendorname,company_id=request.user.company_id,status=1).order_by('-id').values_list('id',flat=True)
                    vendormasterlist=[]
                    for vendorid in contractlist:
                        if (ContractMaster.objects.filter(contractvendor_id=vendorid,status=1).exists()):
                            getvendormasterdetails=ContractMasterVendor.objects.filter(company_id=request.user.company_id,status=1,id=vendorid).first()
                            vendormasterlist.append({'id':getvendormasterdetails.id,'vin':getvendormasterdetails.vin,'vendor_name':getvendormasterdetails.vendor_name,'active_status':getvendormasterdetails.active_status})
                    data={}
                    if (request.user.roles_id == 3):
                        userrights=UserRights.objects.get(user_id=request.user.id,module_id=4)
                        data['rights']=userrights
                    page = request.GET.get('page', 1)
                    paginator = Paginator(vendormasterlist,5)
                    paginator_contractvendor_master_list = paginator.page(page)
                
        
                    data.update({'contractlist':paginator_contractvendor_master_list,'contractmastercount':len(vendormasterlist),'se_vendor':getvendorname,'projects':projects,'company':company})
                    return render(request,'contractmasterlist.html',data)

                else:
                    #checking id by asc
                    contractlist=ContractMasterVendor.objects.filter(vendor_name__icontains=getvendorname,company_id=request.user.company_id,status=1).order_by('id').values_list('id',flat=True)
                    vendormasterlist=[]
                    for vendorid in contractlist:
                        if (ContractMaster.objects.filter(contractvendor_id=vendorid,status=1).exists()):
                            getvendormasterdetails=ContractMasterVendor.objects.filter(company_id=request.user.company_id,status=1,id=vendorid).first()
                            vendormasterlist.append({'id':getvendormasterdetails.id,'vin':getvendormasterdetails.vin,'vendor_name':getvendormasterdetails.vendor_name,'active_status':getvendormasterdetails.active_status})
                    data={}
                    if (request.user.roles_id == 3):
                        userrights=UserRights.objects.get(user_id=request.user.id,module_id=4)
                        data['rights']=userrights
                    page = request.GET.get('page', 1)
                    paginator = Paginator(vendormasterlist,5)
                    paginator_contractvendor_master_list = paginator.page(page)
                
                  
                    data.update({'contractlist':paginator_contractvendor_master_list,'contractmastercount':len(vendormasterlist),'se_vendor':getvendorname,'projects':projects,'company':company})
                    return render(request,'contractmasterlist.html',data)
                            
            elif (sortname):
             #sort by vendor name 
                if (sortname == 'descending'):
                    contractlist=ContractMasterVendor.objects.filter(vendor_name__icontains=getvendorname,company_id=request.user.company_id,status=1).order_by('-vendor_name').values_list('id',flat=True)
                    vendormasterlist=[]
                    for vendorid in contractlist:
                        if (ContractMaster.objects.filter(contractvendor_id=vendorid,status=1).exists()):
                            getvendormasterdetails=ContractMasterVendor.objects.filter(company_id=request.user.company_id,status=1,id=vendorid).first()
                            vendormasterlist.append({'id':getvendormasterdetails.id,'vin':getvendormasterdetails.vin,'vendor_name':getvendormasterdetails.vendor_name,'active_status':getvendormasterdetails.active_status})
                    data={}
                    if (request.user.roles_id == 3):
                        userrights=UserRights.objects.get(user_id=request.user.id,module_id=4)
                        data['rights']=userrights
                    page = request.GET.get('page', 1)
                    paginator = Paginator(vendormasterlist,5)
                    paginator_contractvendor_master_list = paginator.page(page)
                


                    
                    data.update({'contractlist':paginator_contractvendor_master_list,'contractmastercount':len(vendormasterlist),'se_vendor':getvendorname,'projects':projects,'company':company})
                    return render(request,'contractmasterlist.html',data)
    
                else:
                    print('vname')
                    contractlist=ContractMasterVendor.objects.filter(vendor_name__icontains=getvendorname,company_id=request.user.company_id,status=1).order_by('vendor_name').values_list('id',flat=True)
                    vendormasterlist=[]
                    for vendorid in contractlist:
                        if (ContractMaster.objects.filter(contractvendor_id=vendorid,status=1).exists()):
                            getvendormasterdetails=ContractMasterVendor.objects.filter(company_id=request.user.company_id,status=1,id=vendorid).first()
                            vendormasterlist.append({'id':getvendormasterdetails.id,'vin':getvendormasterdetails.vin,'vendor_name':getvendormasterdetails.vendor_name,'active_status':getvendormasterdetails.active_status})
                    data={}
                    if (request.user.roles_id == 3):
                        userrights=UserRights.objects.get(user_id=request.user.id,module_id=4)
                        data['rights']=userrights
                    page = request.GET.get('page', 1)
                    paginator = Paginator(vendormasterlist,5)
                    paginator_contractvendor_master_list = paginator.page(page)
                      
                    data.update({'contractlist':paginator_contractvendor_master_list,'contractmastercount':len(vendormasterlist),'se_vendor':getvendorname,'projects':projects,'company':company})
                    return render(request,'contractmasterlist.html',data)
                
        elif (getvendorname == '' and value != '' and sortname == ''):
            print('only sort',value)
            if (value == 'descending'):
                contractlist=ContractMasterVendor.objects.filter(vendor_name__icontains=getvendorname,company_id=request.user.company_id,status=1).order_by('-id').values_list('id',flat=True)
                vendormasterlist=[]
                for vendorid in contractlist:
                    if (ContractMaster.objects.filter(contractvendor_id=vendorid,status=1).exists()):
                        getvendormasterdetails=ContractMasterVendor.objects.filter(company_id=request.user.company_id,status=1,id=vendorid).first()
                        vendormasterlist.append({'id':getvendormasterdetails.id,'vin':getvendormasterdetails.vin,'vendor_name':getvendormasterdetails.vendor_name,'active_status':getvendormasterdetails.active_status})
                data={}
                if (request.user.roles_id == 3):
                    userrights=UserRights.objects.get(user_id=request.user.id,module_id=4)
                    data['rights']=userrights
                page = request.GET.get('page', 1)
                paginator = Paginator(vendormasterlist,5)
                paginator_contractvendor_master_list = paginator.page(page)
                  
               
                data.update({'contractlist':paginator_contractvendor_master_list,'contractmastercount':len(vendormasterlist),'se_vendor':getvendorname,'projects':projects,'company':company})
                return render(request,'contractmasterlist.html',data)
    
            else:
                contractlist=ContractMasterVendor.objects.filter(vendor_name__icontains=getvendorname,company_id=request.user.company_id,status=1).order_by('id').values_list('id',flat=True)
                vendormasterlist=[]
                for vendorid in contractlist:
                    if (ContractMaster.objects.filter(contractvendor_id=vendorid,status=1).exists()):
                        getvendormasterdetails=ContractMasterVendor.objects.filter(company_id=request.user.company_id,status=1,id=vendorid).first()
                        vendormasterlist.append({'id':getvendormasterdetails.id,'vin':getvendormasterdetails.vin,'vendor_name':getvendormasterdetails.vendor_name,'active_status':getvendormasterdetails.active_status})
                data={}
                if (request.user.roles_id == 3):
                    userrights=UserRights.objects.get(user_id=request.user.id,module_id=4)
                    data['rights']=userrights


                page = request.GET.get('page', 1)
                paginator = Paginator(vendormasterlist,5)
                paginator_contractvendor_master_list = paginator.page(page)
                data.update({'contractlist':paginator_contractvendor_master_list,'contractmastercount':len(vendormasterlist),'se_vendor':getvendorname,'projects':projects,'company':company})
                return render(request,'contractmasterlist.html',data)

      # #sortvendorname
        elif (getvendorname == '' and value == '' and sortname != ''):
            print('vendor name sort')
            if (sortname == 'descending'):
                contractlist=ContractMasterVendor.objects.filter(vendor_name__icontains=getvendorname,company_id=request.user.company_id,status=1).order_by('-vendor_name').values_list('id',flat=True)
                vendormasterlist=[]
                for vendorid in contractlist:
                    if (ContractMaster.objects.filter(contractvendor_id=vendorid,status=1).exists()):
                        getvendormasterdetails=ContractMasterVendor.objects.filter(company_id=request.user.company_id,status=1,id=vendorid).first()
                        vendormasterlist.append({'id':getvendormasterdetails.id,'vin':getvendormasterdetails.vin,'vendor_name':getvendormasterdetails.vendor_name,'active_status':getvendormasterdetails.active_status})
                data={}
                if (request.user.roles_id == 3):
                    userrights=UserRights.objects.get(user_id=request.user.id,module_id=4)
                    data['rights']=userrights
                page = request.GET.get('page', 1)

                paginator = Paginator(vendormasterlist, 5)
                paginator_vendor_master_list = paginator.page(page)
                data.update({'contractlist':paginator_vendor_master_list,'contractmastercount':len(vendormasterlist),'se_vendor':getvendorname,'projects':projects,'company':company})
                return render(request,'contractmasterlist.html',data)
        
            else:
                contractlist=ContractMasterVendor.objects.filter(vendor_name__icontains=getvendorname,company_id=request.user.company_id,status=1).order_by('vendor_name').values_list('id',flat=True)
                vendormasterlist=[]
                for vendorid in contractlist:
                    if (ContractMaster.objects.filter(contractvendor_id=vendorid,status=1).exists()):
                        getvendormasterdetails=ContractMasterVendor.objects.filter(company_id=request.user.company_id,status=1,id=vendorid).first()
                        vendormasterlist.append({'id':getvendormasterdetails.id,'vin':getvendormasterdetails.vin,'vendor_name':getvendormasterdetails.vendor_name,'active_status':getvendormasterdetails.active_status})
                data={}
                if (request.user.roles_id == 3):
                    userrights=UserRights.objects.get(user_id=request.user.id,module_id=4)
                    data['rights']=userrights
                page = request.GET.get('page', 1)

                paginator = Paginator(vendormasterlist, 5)
                paginator_vendor_master_list = paginator.page(page)
                data.update({'contractlist':paginator_vendor_master_list,'contractmastercount':len(vendormasterlist),'se_vendor':getvendorname,'projects':projects,'company':company})
                return render(request,'contractmasterlist.html',data)
 
    data={}
    if (request.user.roles_id == 3):
        userrights=UserRights.objects.get(user_id=request.user.id,module_id=4)
        data['rights']=userrights
    values=ContractMaster.objects.filter(status=1).values_list('contractvendor_id', flat=True).distinct()
    contractlist = ContractMasterVendor.objects.filter(id__in=values,company=request.user.company,status=True).order_by('-id')
    page = request.GET.get('page', 1)
    paginator = Paginator(contractlist, 10)
    paginator_vendor_master_list = paginator.page(page)
    data.update({'contractlist':paginator_vendor_master_list,'contractmastercount':contractmastercount,'projects':projects,'company':company})
    return render(request,'contractmasterlist.html',data)

#vendor tax#
def createvendortax(request,contracttype,id):
    contractmaster=None
    amendmentmaster=None
    if (contracttype == "original"):
        contractmaster=ContractMaster.objects.get(id=id)
        contractId = contractmaster.id
        conid=contractmaster.id
        draft_count=VendorCompanyTaxDetails.objects.filter(contract_id=contractmaster.id,status=1,draft_status=1).count() + VendorInvoiceSplit.objects.filter(contract_id=contractmaster.id,status=1,draft_status=1).count() + VendorPaymentTerms.objects.filter(contract_id=contractmaster.id,status=1,draft_status=1).count()
        vendorid=contractmaster.contractvendor_id
        companytax=[{'id':'not_applicable','Tax_Name':'Not Applicable'}]
        companyid=contractmaster.company.id
        company=Settings.objects.filter(company_id=companyid).first()
        getcurrencylist=ast.literal_eval(company.currency)
        companycurrency=Basecountries.objects.filter(id__in=getcurrencylist)
        allcompanytax=CompanyTaxDetails.objects.filter(company_id=companyid,status=1)
        check_count=CompanyTaxDetails.objects.filter(company_id=companyid,status=1).count()
        addcompanytax=list(CompanyTaxDetails.objects.filter(company_id=companyid,status=1).values('id','Tax_Name'))
        companytax.extend(addcompanytax)
        vendorcompanytaxdetails=VendorCompanyTaxDetails.objects.filter(company_id=companyid,vendor_id=contractmaster.contractvendor.id,contract_id=contractmaster.id,status=1)#add contractid
        getinclusivetax=vendorcompanytaxdetails.filter(Tax_Type="Inclusive")
        getexclusivetax=vendorcompanytaxdetails.filter(Tax_Type="Exclusive")
        companytaxlist=companytax
        if (vendorcompanytaxdetails.count() > 0):
            vtax=list(vendorcompanytaxdetails.values_list('tax',flat=True))
            filtered_list = list(filter(None, vtax))
            allcompanytax=allcompanytax.exclude(id__in=filtered_list)
            companytaxlist=[{'id':'not_applicable','Tax_Name':'Not Applicable'}]
            companytaxlist.extend(list(allcompanytax.values('id','Tax_Name')))
        else:
            companytaxlist=companytaxlist
        # print("companytaxlist",companytaxlist)
        invoicesplitlist=VendorInvoiceSplit.objects.filter(contract_id=id,status=1)
        count_invoice=invoicesplitlist.count()
        # print("count_invoice",count_invoice)
        if  count_invoice > 0:
            get_no_invoice=invoicesplitlist.values('no_invoice','exchange_rate').distinct()
            invoicenumber=int(get_no_invoice[0]['no_invoice'])
            exchangerate=int(get_no_invoice[0]['exchange_rate'])
        else:
            invoicenumber=None
            exchangerate=None
        vendorpaymentterms=VendorPaymentTerms.objects.filter(company_id=companyid,vendor_id=contractmaster.contractvendor.id,contract_id=id,status=1)
        print('vendorpaymentterms',vendorpaymentterms)
        if (request.POST):
            vendor_name = request.POST.get('vendor_name_user_log')
            print(vendor_name,"VENDOR NAME")
            reference_name = request.POST.get('reference_name_user_log')
            print(reference_name,"REFERNECE NAME")
            print("a",request.POST)
            form_submit=request.POST.get('submit_type')
            inclusivetax=request.POST.getlist('inclusivecompanytax')
            updatelist=[]
            for  inclusive in inclusivetax:
                if (inclusive == "not_applicable"):
                    inclusivetaxdata=VendorCompanyTaxDetails.objects.filter(Tax_Type='Inclusive',tax_id=None,company_id=companyid,contract_id=id,vendor_id=contractmaster.contractvendor.id).first()
                    print('a',inclusivetaxdata)
                    if (inclusivetaxdata != None):
                        VendorCompanyTaxDetails.objects.filter(id=inclusivetaxdata.id).update(status=1,draft_status=int(form_submit))
                        #Tax_Type='Inclusive',tax_id="",company_id=companyid,contract_id=id,vendor_id=contractmaster.contractvendor.id)
                        updatelist.append(inclusivetaxdata.id)
                    else:
                        create_vendor_tax=VendorCompanyTaxDetails.objects.create(Tax_Type='Inclusive',tax_id="",company_id=companyid,contract_id=id,vendor_id=contractmaster.contractvendor.id,draft_status=int(form_submit))
                        updatelist.append(create_vendor_tax.id)
                else:
                    inclusivetaxdata=VendorCompanyTaxDetails.objects.filter(Tax_Type='Inclusive',tax_id=inclusive,company_id=companyid,contract_id=id,vendor_id=contractmaster.contractvendor.id).first()
                    if (inclusivetaxdata != None):
                        VendorCompanyTaxDetails.objects.filter(Tax_Type='Inclusive',tax_id=inclusive,company_id=companyid,contract_id=id,vendor_id=contractmaster.contractvendor.id).update(status=1,draft_status=int(form_submit))
                        updatelist.append(inclusivetaxdata.id)
                    else:
                        inclusivetaxdata=VendorCompanyTaxDetails.objects.create(Tax_Type='Inclusive',tax_id=inclusive,company_id=companyid,contract_id=id,vendor_id=contractmaster.contractvendor.id,draft_status=int(form_submit))
                        updatelist.append(inclusivetaxdata.id)
                    getinclusivetaxpercentage=request.POST.getlist('inclusivepercentage'+str(inclusive))
                    hdnvalues=request.POST.getlist('hdninclusivepercentage'+str(inclusive))
                    taxupdatelist=[]
                    for hdnid,taxpercentage in zip(hdnvalues,getinclusivetaxpercentage):
                        if (hdnid):
                            VendorCompanyTaxPercentage.objects.filter(id=hdnid).update(taxpercentage=taxpercentage)
                            taxupdatelist.append(hdnid)
                            # print('exists',taxpercentage)
                        else:
                            # print('new',taxpercentage)
                            create_tax_percentage=VendorCompanyTaxPercentage.objects.create(taxpercentage=taxpercentage,vendortax_id=inclusivetaxdata.id)
                            taxupdatelist.append(create_tax_percentage.id)
                    VendorCompanyTaxPercentage.objects.filter(vendortax_id=inclusivetaxdata.id).exclude(id__in=taxupdatelist).update(status=0)
            VendorCompanyTaxDetails.objects.filter(Tax_Type='Inclusive',contract_id=id,vendor_id=contractmaster.contractvendor.id).exclude(id__in=updatelist).update(status=0)

            exclusivetax=request.POST.getlist('exclusivecompanytax')
            exclusiveupdatelist=[]
            for  exclusive in exclusivetax:
                if (exclusive == "not_applicable"):
                    exclusivetaxdata=VendorCompanyTaxDetails.objects.filter(Tax_Type='Exclusive',tax_id=None,company_id=companyid,contract_id=id,vendor_id=contractmaster.contractvendor.id).first()
                    if (exclusivetaxdata != None):
                        VendorCompanyTaxDetails.objects.filter(id=exclusivetaxdata.id).update(status=1,draft_status=int(form_submit))
                        #Tax_Type='Exclusive',tax_id="",company_id=companyid,contract_id=id,vendor_id=contractmaster.contractvendor.id
                        exclusiveupdatelist.append(exclusivetaxdata.id)
                    else:
                        create_vendor_tax=VendorCompanyTaxDetails.objects.create(Tax_Type='Exclusive',tax_id="",company_id=companyid,contract_id=id,vendor_id=contractmaster.contractvendor.id,draft_status=int(form_submit))
                        exclusiveupdatelist.append(create_vendor_tax.id)
                else:
                    exclusivetaxdata=VendorCompanyTaxDetails.objects.filter(Tax_Type='Exclusive',tax_id=exclusive,company_id=companyid,contract_id=id,vendor_id=contractmaster.contractvendor.id).first()
                    print("b",exclusivetaxdata)
                    if (exclusivetaxdata != None):
                        VendorCompanyTaxDetails.objects.filter(Tax_Type='Exclusive',tax_id=exclusive,company_id=companyid,contract_id=id,vendor_id=contractmaster.contractvendor.id).update(status=1,draft_status=int(form_submit))
                        exclusiveupdatelist.append(exclusivetaxdata.id)
                    else:
                        exclusivetaxdata=VendorCompanyTaxDetails.objects.create(Tax_Type='Exclusive',tax_id=exclusive,company_id=companyid,contract_id=id,vendor_id=contractmaster.contractvendor.id,draft_status=int(form_submit))
                        exclusiveupdatelist.append(exclusivetaxdata.id)
                    getexclusivetaxpercentage=request.POST.getlist('exclusivepercentage'+str(exclusive))
                    hdnvalues=request.POST.getlist('hdnexclusivepercentage'+str(exclusive))
                    print("hdnvalues",hdnvalues)
                    exclusivetaxupdatelist=[]
                    for hdnid,taxpercentage in zip(hdnvalues,getexclusivetaxpercentage):
                        if (hdnid):
                            VendorCompanyTaxPercentage.objects.filter(id=hdnid).update(taxpercentage=taxpercentage)
                            exclusivetaxupdatelist.append(hdnid)
                            # print('exists',taxpercentage)
                        else:
                            # print('new',taxpercentage)
                            create_tax_percentage=VendorCompanyTaxPercentage.objects.create(taxpercentage=taxpercentage,vendortax_id=exclusivetaxdata.id)
                            exclusivetaxupdatelist.append(create_tax_percentage.id)
                    print("exclusivetaxupdatelist",exclusivetaxupdatelist)
                    VendorCompanyTaxPercentage.objects.filter(vendortax_id=exclusivetaxdata.id).exclude(id__in=exclusivetaxupdatelist).update(status=0)
            VendorCompanyTaxDetails.objects.filter(Tax_Type='Exclusive',contract_id=id,vendor_id=contractmaster.contractvendor.id).exclude(id__in=exclusiveupdatelist).update(status=0)

            no_invoice=request.POST.get('totalinvoice','')
            invoicecurrency=request.POST.getlist('invoicecurrency')
            invoicepercentage=request.POST.getlist('invoicepercentage')
            invoiceexchangerate=request.POST.get('exchangerate')
            hdninvoicesplit=request.POST.getlist('hdninvoicesplit')
            print("hdninvoicesplit",hdninvoicesplit)
            newlist=[]
            if (no_invoice):
                if (int(no_invoice) == invoicenumber):
                    print('a')
                    for currency,percentage,hdid in zip(invoicecurrency,invoicepercentage,hdninvoicesplit):
                        print('asd',hdid)
                        VendorInvoiceSplit.objects.filter(id=hdid).update(currency_id=currency,percentage=percentage,exchange_rate=invoiceexchangerate,draft_status=int(form_submit))
                        newlist.append(hdid)
                    VendorInvoiceSplit.objects.filter(vendor_id=contractmaster.contractvendor.id,contract_id=id).exclude(id__in=newlist).update(status=0)
                else:
                    print('b')
                    for currency,percentage in zip(invoicecurrency,invoicepercentage):
                        createinvoicesplit=VendorInvoiceSplit.objects.create(no_invoice=no_invoice,currency_id=currency,percentage=percentage,exchange_rate=invoiceexchangerate,contract_id=id,company_id=companyid,vendor_id=contractmaster.contractvendor.id,draft_status=int(form_submit))
                        newlist.append(createinvoicesplit.id)
                    VendorInvoiceSplit.objects.filter(vendor_id=contractmaster.contractvendor.id,contract_id=id).exclude(id__in=newlist).update(status=0)
            
            getpaymentype=request.POST.getlist('paymentype')
            getpaymentday=request.POST.getlist('paymentday')
            getpaymentperentage=request.POST.getlist('paymentpercentage')
            hdnvalueslist=request.POST.getlist('hdnpaymentvalues')
            hdnselected=request.POST.getlist('hdnselected')
            print('hdnselected',hdnselected)
            pamentupdatelist=[]
            for hdnvalues,paymenttype,day,percentage,selectedname in zip(hdnvalueslist,getpaymentype,getpaymentday,getpaymentperentage,hdnselected):
                if hdnvalues:
                    VendorPaymentTerms.objects.filter(id=hdnvalues).update(payment_type=paymenttype,payment_day=day,payment_percentage=percentage,payment_name=selectedname,draft_status=int(form_submit))
                    pamentupdatelist.append(hdnvalues)
                else:
                    create_payment=VendorPaymentTerms.objects.create(payment_type=paymenttype,payment_day=day,payment_percentage=percentage,company_id=companyid,vendor_id=contractmaster.contractvendor.id,contract_id=id,payment_name=selectedname,draft_status=int(form_submit))
                    pamentupdatelist.append(create_payment.id)
            VendorPaymentTerms.objects.filter(vendor_id=contractmaster.contractvendor.id,contract_id=id).exclude(id__in=pamentupdatelist).update(status=0)
            if reference_name is not None and vendor_name is not None:
                usercreate = request.user.roles_id
                create_user_log(
                    request,
                    reference_name,  # <-- Corrected variable name
                    'Vendor Tax',
                    'Create',
                    f'Vendor Tax has been Created for {vendor_name}',
                    usercreate
                )

            return redirect('projects:contractmasterlist')
        # print('check_count',check_count)
        original_contract=ContractMaster.objects.filter(contractvendor_id=vendorid,status=1).exclude(Q(upload_contract='') | (Q(upload_pricetable='')))
        # print('asd',original_contract)
        original_contract_id=id
        remove_current_contract=original_contract.exclude(id=original_contract_id)
        original_contracts=remove_current_contract.exclude(Q(name_service__exact = '') | Q(name_service__isnull=True) | Q(reference_number__exact='') | Q(reference_number__isnull=True) | Q(executed_date__isnull=True) | Q(amount__exact='') | Q(amount__isnull=True) | Q(upload_contract__exact='') | Q(upload_contract__isnull=True) | Q(upload_pricetable__exact='') | Q(upload_pricetable__isnull=True) | Q(projects_id__isnull=True) | Q(projectdiscipline_id__isnull=True) | Q(projectdisciplinetype_id__isnull=True))
        print('12q',original_contracts)
        data={'contract':contractmaster,'companytaxlist':companytaxlist,'companycurrency':companycurrency,'getinclusivetax':getinclusivetax,'getexclusivetax':getexclusivetax,'invoicesplitlist':invoicesplitlist,'invoicesplitlistcount':invoicesplitlist.count(),'invoicenumber':invoicenumber,'exchangerate':exchangerate,'vendorpaymentterms':vendorpaymentterms,'vendorpaymenttermscount':vendorpaymentterms.count(),'range':range(1, 61),'check_count':check_count,'vendorcompanytaxdetailscount':vendorcompanytaxdetails.count(),'original_contracts':original_contracts,'draft_count':draft_count,'contracttype':contracttype,'contractId':contractId , "id":id}
    else:
        print("else",contracttype)
        form_submit=request.POST.get('submit_type')
        amendmentmaster=Amendment.objects.get(id=id)
        contractId = amendmentmaster.id
        conid=amendmentmaster.service.id
        draft_count=VendorCompanyTaxDetails.objects.filter(amendment_id=amendmentmaster.id,status=1,draft_status=1).count() + VendorInvoiceSplit.objects.filter(amendment_id=amendmentmaster.id,status=1,draft_status=1).count() + VendorPaymentTerms.objects.filter(amendment_id=amendmentmaster.id,status=1,draft_status=1).count()
        vendorid=amendmentmaster.contractvendor_id
        companytax=[{'id':'not_applicable','Tax_Name':'Not Applicable'}]
        companyid=amendmentmaster.company.id
        company=Settings.objects.filter(company_id=companyid).first()
        getcurrencylist=ast.literal_eval(company.currency)
        companycurrency=Basecountries.objects.filter(id__in=getcurrencylist)
        allcompanytax=CompanyTaxDetails.objects.filter(company_id=companyid,status=1)
        check_count=CompanyTaxDetails.objects.filter(company_id=companyid,status=1).count()
        addcompanytax=list(CompanyTaxDetails.objects.filter(company_id=companyid,status=1).values('id','Tax_Name'))
        companytax.extend(addcompanytax)
        vendorcompanytaxdetails=VendorCompanyTaxDetails.objects.filter(company_id=companyid,amendment_id=amendmentmaster.id,status=1)#add contractid
        getinclusivetax=vendorcompanytaxdetails.filter(Tax_Type="Inclusive")
        getexclusivetax=vendorcompanytaxdetails.filter(Tax_Type="Exclusive")
        companytaxlist=companytax
        if (vendorcompanytaxdetails.count() > 0):
            vtax=list(vendorcompanytaxdetails.values_list('tax',flat=True))
            filtered_list = list(filter(None, vtax))
            allcompanytax=allcompanytax.exclude(id__in=filtered_list)
            # print(filtered_list)
            companytaxlist=[{'id':'not_applicable','Tax_Name':'Not Applicable'}]
            companytaxlist.extend(list(allcompanytax.values('id','Tax_Name')))
        else:
            companytaxlist=companytaxlist
        invoicesplitlist=VendorInvoiceSplit.objects.filter(amendment_id=id,status=1)
        count_invoice=invoicesplitlist.count()
        print("count_invoice",count_invoice)
        if  count_invoice > 0:
            get_no_invoice=invoicesplitlist.values('no_invoice','exchange_rate').distinct()
            print("a",get_no_invoice)
            invoicenumber=int(get_no_invoice[0]['no_invoice'])
            exchangerate=int(get_no_invoice[0]['exchange_rate'])
        else:
            invoicenumber=None
            exchangerate=None

        vendorpaymentterms=VendorPaymentTerms.objects.filter(company_id=companyid,vendor_id=amendmentmaster.contractvendor.id,amendment_id=id,status=1)

        if (request.POST):
            print("a",request.POST)
            inclusivetax=request.POST.getlist('inclusivecompanytax')
            updatelist=[]
            for  inclusive in inclusivetax:
                if (inclusive == "not_applicable"):
                    inclusivetaxdata=VendorCompanyTaxDetails.objects.filter(Tax_Type='Inclusive',tax_id=None,company_id=companyid,amendment_id=id,vendor_id=amendmentmaster.contractvendor.id).first()
                    if (inclusivetaxdata != None):
                        VendorCompanyTaxDetails.objects.filter(id=inclusivetaxdata.id).update(status=1,draft_status=int(form_submit))
                        #Tax_Type='Inclusive',tax_id="",company_id=companyid,amendment_id=id,vendor_id=amendmentmaster.contractvendor.id
                        updatelist.append(inclusivetaxdata.id)
                    else:
                        create_vendor_tax=VendorCompanyTaxDetails.objects.create(Tax_Type='Inclusive',tax_id="",company_id=companyid,amendment_id=id,vendor_id=amendmentmaster.contractvendor.id,draft_status=int(form_submit))
                        updatelist.append(create_vendor_tax.id)
                else:
                    inclusivetaxdata=VendorCompanyTaxDetails.objects.filter(Tax_Type='Inclusive',tax_id=inclusive,company_id=companyid,amendment_id=id,vendor_id=amendmentmaster.contractvendor.id).first()
                    if (inclusivetaxdata != None):
                        VendorCompanyTaxDetails.objects.filter(Tax_Type='Inclusive',tax_id=inclusive,company_id=companyid,amendment_id=id,vendor_id=amendmentmaster.contractvendor.id).update(status=1,draft_status=int(form_submit))
                        updatelist.append(inclusivetaxdata.id)
                    else:
                        inclusivetaxdata=VendorCompanyTaxDetails.objects.create(Tax_Type='Inclusive',tax_id=inclusive,company_id=companyid,amendment_id=id,vendor_id=amendmentmaster.contractvendor.id,draft_status=int(form_submit))
                        updatelist.append(inclusivetaxdata.id)
                    getinclusivetaxpercentage=request.POST.getlist('inclusivepercentage'+str(inclusive))
                    hdnvalues=request.POST.getlist('hdninclusivepercentage'+str(inclusive))
                    taxupdatelist=[]
                    for hdnid,taxpercentage in zip(hdnvalues,getinclusivetaxpercentage):
                        if (hdnid):
                            VendorCompanyTaxPercentage.objects.filter(id=hdnid).update(taxpercentage=taxpercentage)
                            taxupdatelist.append(hdnid)
                            # print('exists',taxpercentage)
                        else:
                            # print('new',taxpercentage)
                            create_tax_percentage=VendorCompanyTaxPercentage.objects.create(taxpercentage=taxpercentage,vendortax_id=inclusivetaxdata.id)
                            taxupdatelist.append(create_tax_percentage.id)
                    VendorCompanyTaxPercentage.objects.filter(vendortax_id=inclusivetaxdata.id).exclude(id__in=taxupdatelist).update(status=0)
            VendorCompanyTaxDetails.objects.filter(Tax_Type='Inclusive',amendment_id=id,vendor_id=amendmentmaster.contractvendor.id).exclude(id__in=updatelist).update(status=0)

            exclusivetax=request.POST.getlist('exclusivecompanytax')
            exclusiveupdatelist=[]
            for  exclusive in exclusivetax:
                if (exclusive == "not_applicable"):
                    exclusivetaxdata=VendorCompanyTaxDetails.objects.filter(Tax_Type='Exclusive',tax_id=None,company_id=companyid,amendment_id=id,vendor_id=amendmentmaster.contractvendor.id).first()
                    if (exclusivetaxdata != None):
                        VendorCompanyTaxDetails.objects.filter(id=exclusivetaxdata.id).update(status=1,draft_status=int(form_submit))
                        #,Tax_Type='Exclusive',tax_id="",company_id=companyid,amendment_id=id,vendor_id=amendmentmaster.contractvendor.id
                        exclusiveupdatelist.append(exclusivetaxdata.id)
                    else:
                        create_vendor_tax=VendorCompanyTaxDetails.objects.create(Tax_Type='Exclusive',tax_id="",company_id=companyid,amendment_id=id,vendor_id=amendmentmaster.contractvendor.id,draft_status=int(form_submit))
                        exclusiveupdatelist.append(create_vendor_tax.id)
                else:
                    exclusivetaxdata=VendorCompanyTaxDetails.objects.filter(Tax_Type='Exclusive',tax_id=exclusive,company_id=companyid,amendment_id=id,vendor_id=amendmentmaster.contractvendor.id).first()
                    if (exclusivetaxdata != None):
                        VendorCompanyTaxDetails.objects.filter(Tax_Type='Exclusive',tax_id=exclusive,company_id=companyid,amendment_id=id,vendor_id=amendmentmaster.contractvendor.id).update(status=1,draft_status=int(form_submit))
                        exclusiveupdatelist.append(exclusivetaxdata.id)
                    else:
                        exclusivetaxdata=VendorCompanyTaxDetails.objects.create(Tax_Type='Exclusive',tax_id=exclusive,company_id=companyid,amendment_id=id,vendor_id=amendmentmaster.contractvendor.id,draft_status=int(form_submit))
                        exclusiveupdatelist.append(exclusivetaxdata.id)
                    getexclusivetaxpercentage=request.POST.getlist('exclusivepercentage'+str(exclusive))
                    hdnvalues=request.POST.getlist('hdnexclusivepercentage'+str(exclusive))
                    print("hdnvalues",hdnvalues)
                    exclusivetaxupdatelist=[]
                    for hdnid,taxpercentage in zip(hdnvalues,getexclusivetaxpercentage):
                        if (hdnid):
                            VendorCompanyTaxPercentage.objects.filter(id=hdnid).update(taxpercentage=taxpercentage)
                            exclusivetaxupdatelist.append(hdnid)
                            # print('exists',taxpercentage)
                        else:
                            # print('new',taxpercentage)
                            create_tax_percentage=VendorCompanyTaxPercentage.objects.create(taxpercentage=taxpercentage,vendortax_id=exclusivetaxdata.id)
                            exclusivetaxupdatelist.append(create_tax_percentage.id)
                    print("exclusivetaxupdatelist",exclusivetaxupdatelist)
                    VendorCompanyTaxPercentage.objects.filter(vendortax_id=exclusivetaxdata.id).exclude(id__in=exclusivetaxupdatelist).update(status=0)
            VendorCompanyTaxDetails.objects.filter(Tax_Type='Exclusive',amendment_id=id,vendor_id=amendmentmaster.contractvendor.id).exclude(id__in=exclusiveupdatelist).update(status=0)

            no_invoice=request.POST.get('totalinvoice','')
            invoicecurrency=request.POST.getlist('invoicecurrency')
            invoicepercentage=request.POST.getlist('invoicepercentage')
            invoiceexchangerate=request.POST.get('exchangerate')
            hdninvoicesplit=request.POST.getlist('hdninvoicesplit')
            newlist=[]
            if (no_invoice):
                if (int(no_invoice) == invoicenumber):
                    for currency,percentage,hdid in zip(invoicecurrency,invoicepercentage,hdninvoicesplit):
                        VendorInvoiceSplit.objects.filter(id=hdid).update(currency_id=currency,percentage=percentage,exchange_rate=invoiceexchangerate,draft_status=int(form_submit))
                        newlist.append(hdid)
                    VendorInvoiceSplit.objects.filter(vendor_id=amendmentmaster.contractvendor.id,amendment_id=id).exclude(id__in=newlist).update(status=0)
                else:
                    for currency,percentage in zip(invoicecurrency,invoicepercentage):
                        createinvoicesplit=VendorInvoiceSplit.objects.create(no_invoice=no_invoice,currency_id=currency,percentage=percentage,exchange_rate=invoiceexchangerate,amendment_id=id,company_id=companyid,vendor_id=amendmentmaster.contractvendor.id,draft_status=int(form_submit))
                        newlist.append(createinvoicesplit.id)
                    VendorInvoiceSplit.objects.filter(vendor_id=amendmentmaster.contractvendor.id,amendment_id=id).exclude(id__in=newlist).update(status=0)
            
            getpaymentype=request.POST.getlist('paymentype')
            getpaymentday=request.POST.getlist('paymentday')
            getpaymentperentage=request.POST.getlist('paymentpercentage')
            # hdnvalues=request.POST.getlist('hdnpayment')
            hdnselected=request.POST.getlist('hdnselected')
            hdnvalueslist=request.POST.getlist('hdnpaymentvalues')
            pamentupdatelist=[]

            for hdnvalues,paymenttype,day,percentage,selectedname in zip(hdnvalueslist,getpaymentype,getpaymentday,getpaymentperentage,hdnselected):
                if hdnvalues:
                    VendorPaymentTerms.objects.filter(id=hdnvalues).update(payment_type=paymenttype,payment_day=day,payment_percentage=percentage,payment_name=selectedname,draft_status=int(form_submit))
                    pamentupdatelist.append(hdnvalues)
                else:
                    create_payment=VendorPaymentTerms.objects.create(payment_type=paymenttype,payment_day=day,payment_percentage=percentage,company_id=companyid,vendor_id=amendmentmaster.contractvendor.id,amendment_id=id,payment_name=selectedname,draft_status=int(form_submit))
                    pamentupdatelist.append(create_payment.id)
            VendorPaymentTerms.objects.filter(vendor_id=amendmentmaster.contractvendor.id,amendment_id=id).exclude(id__in=pamentupdatelist).update(status=0)

            return redirect('projects:contract_detailed_view',conid)

        original_contract=ContractMaster.objects.filter(contractvendor_id=vendorid,status=1).exclude(Q(upload_contract='') | (Q(upload_pricetable='')))
        print('original_contract',original_contract)
        data={'amendment':amendmentmaster,'companytaxlist':companytax,'companycurrency':companycurrency,'getinclusivetax':getinclusivetax,'getexclusivetax':getexclusivetax,'invoicesplitlist':invoicesplitlist,'invoicesplitlistcount':invoicesplitlist.count(),'invoicenumber':invoicenumber,'exchangerate':exchangerate,'vendorpaymentterms':vendorpaymentterms,'vendorpaymenttermscount':vendorpaymentterms.count(),'range':range(1, 61),'check_count':check_count,'contracttype':contracttype,'vendorcompanytaxdetailscount':vendorcompanytaxdetails.count(),'original_contracts':original_contract,'draft_count':draft_count,'contractId':contractId, "id":id}
    return render(request,'createvendortax.html',data)

def viewvendortax(request,contracttype,id):
    if (contracttype == "original"):
        contractmaster=ContractMaster.objects.get(id=id)
        vendor_name=contractmaster.contractvendor.vendor_name
        reference_number=contractmaster.reference_number
        vendorid=contractmaster.contractvendor_id
        vendorcompanytaxdetails=VendorCompanyTaxDetails.objects.filter(company_id=request.company.id,vendor_id=contractmaster.contractvendor.id,contract_id=contractmaster.id,status=1)#add contractid
        getinclusivetax=vendorcompanytaxdetails.filter(Tax_Type="Inclusive")
        getexclusivetax=vendorcompanytaxdetails.filter(Tax_Type="Exclusive")
        invoicesplitlist=VendorInvoiceSplit.objects.filter(company_id=request.company.id,vendor_id=contractmaster.contractvendor.id,contract_id=id,status=1)
        vendorpaymentterms=VendorPaymentTerms.objects.filter(company_id=request.company.id,vendor_id=contractmaster.contractvendor.id,contract_id=id,status=1)
        # print('a',getinclusivetax)
    else:
        amendmentmaster=Amendment.objects.get(id=id)
        vendor_name=amendmentmaster.contractvendor.vendor_name
        reference_number=amendmentmaster.amendment_reference_number
        vendorid=amendmentmaster.contractvendor_id
        vendorcompanytaxdetails=VendorCompanyTaxDetails.objects.filter(company_id=request.company.id,amendment_id=amendmentmaster.id,status=1)#add contractid
        getinclusivetax=vendorcompanytaxdetails.filter(Tax_Type="Inclusive")
        getexclusivetax=vendorcompanytaxdetails.filter(Tax_Type="Exclusive")
        invoicesplitlist=VendorInvoiceSplit.objects.filter(company_id=request.company.id,vendor_id=amendmentmaster.contractvendor.id,amendment_id=id,status=1)
        vendorpaymentterms=VendorPaymentTerms.objects.filter(company_id=request.company.id,vendor_id=amendmentmaster.contractvendor.id,amendment_id=id,status=1)
        # print('b',getinclusivetax)

    data={'inclusive_tax':getinclusivetax,'exclusive_tax':getexclusivetax,'invoice_splits':invoicesplitlist,'vendor_payment_terms':vendorpaymentterms,'vendor_name':vendor_name,'reference_number':reference_number}
    return render(request,'viewvendortax.html',data)
# Cretae a xlsx file
def createexcelsheet(request):
    currencies=Settings.objects.filter(company=request.company).first()
    if (currencies != None):
        currency_convert=ast.literal_eval(currencies.currency)
    #print(currency_convert)
    get_currencies=list(Basecountries.objects.filter(id__in=currency_convert).values_list('currency',flat=True))
    vendors=list(ContractMasterVendor.objects.filter(company=request.company,status=1,active_status=1).values_list('vendor_name',flat=True))
    # print(vendors)
    vendor_len=len(vendors)
    if vendor_len <= 999:
        set_zeros=str(vendor_len).zfill(3)
        set_cell_num='5'+str(set_zeros)
        print(set_cell_num)
    elif vendor_len >= 1000:
        num_add=5000+vendor_len
        set_cell_num=num_add
        print(set_cell_num)
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()
    locked   = workbook.add_format({'locked': True})
    unlocked = workbook.add_format({'locked': False})
    worksheet.protect()
    worksheet.write(0,0,'S/N',locked)
    worksheet.write(0,1,'Vendor Name',locked)
    worksheet.write(0,2,'Type of Contract',locked)
    worksheet.write(0,3,'Type of Service',locked)
    worksheet.write(0,4,'Name of Contract',locked)
    worksheet.write(0,5,'Contract Reference Number',locked)
    worksheet.write(0,6,'For Amendment/Addendum',locked)
    worksheet.write(0,7,'Executed Date',locked)
    
    merge_format = workbook.add_format({
    'align': 'center',
    'valign': 'vcenter'})
    format1 = workbook.add_format({'bg_color': '#FFC7CE',
                               'font_color': '#9C0006'})
    worksheet.merge_range('I1:J1', 'Maximum Value of Contract', merge_format)
    worksheet.conditional_format('F2:F1000', {'type':   'duplicate',
                                              'format': format1})
    # worksheet.conditional_format('I2:I1000', {'type':   'duplicate',
    #                                           'format': format1})
    worksheet.set_column('B:B', 75)
    worksheet.set_column('C:C', 20)
    worksheet.set_column('D:D', 20)    
    worksheet.set_column('E:E', 50)
    worksheet.set_column('F:F', 40)
    worksheet.set_column('G:G', 40)
    worksheet.set_column('H:H', 14)
    worksheet.set_column('I:I', 30)
    worksheet.set_column('J2:J2', 60)   
    worksheet.write(1,0,'',locked)
    worksheet.write(1,1,'(Select from the vendor list)',locked)
    worksheet.write(1,2,'(Select from Options)',locked)
    worksheet.write(1,3,'(Select from Options)',locked)
    worksheet.write(1,4,'',locked)
    worksheet.write(1,5,'',locked)
    worksheet.write(1,6,'(Type Original Contract Ref. No)',locked)
    worksheet.write(1,7,'(dd-mm-yyyy)',locked)
    worksheet.write(1,8,'Currency',locked)
    worksheet.write(1,9,'Value (If there is no maximum value of contract give No Max Limit)',locked)
    worksheet.write_column('AZ5001', vendors)

    # worksheet.write(1,5,'(Select from Options)',locked)
    # worksheet.write(1,6,'(Select from Options)',locked)

    # worksheet.write(1,9,'(dd/mm/yyyy)',locked)
    for i in range(3,1001):
        worksheet.write('A'+str(i),'', unlocked)
        worksheet.write('B'+str(i),'', unlocked)
        worksheet.write('C'+str(i),'', unlocked)
        worksheet.write('D'+str(i),'', unlocked)
        worksheet.write('E'+str(i),'', unlocked)
        worksheet.write('F'+str(i),'', unlocked)
        worksheet.write('G'+str(i),'', unlocked)
        worksheet.write('H'+str(i),'', unlocked)
        worksheet.write('I'+str(i),'', unlocked)
        worksheet.write('J'+str(i),'', unlocked)
        # worksheet.write('K'+str(i),'', unlocked)
        # worksheet.write('L'+str(i),'', unlocked)
        worksheet.data_validation('B'+str(i)+'',
                            {'validate': 'list',
                            'source': '$AZ5001:$AZ'+str(set_cell_num)})

        worksheet.data_validation('C'+str(i)+'', {'validate': 'list',
                                  'source': ['Original Contract', 'Amendment', 'Addendum']})
        worksheet.data_validation('D'+str(i)+'', {'validate': 'list',
                                  'source': ['Service', 'Supply', 'Service and Supply']})
        # worksheet.data_validation('F'+str(i)+'', {'validate': 'list',
        #                           'source': ['Green Field Development', 'Brown Field Development', 'Others']})
        # worksheet.data_validation('G'+str(i)+'', {'validate': 'list',
        #                           'source': ['Subsurface and Reservoir Engineering', 'Drilling and Completions', 'Facilities and Projects','Production and Operations','HR and Administration','Legal','ICT','Finance','QHSE & Community Development','Business Development & Sales and Marketing','Insurance','Miscellaneous']})

        worksheet.data_validation('I'+str(i)+'', {'validate': 'list',
                                  'source': get_currencies})

    workbook.close()
    output.seek(0)
    filename = 'Contractmaster.xlsx'
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=%s' % filename

    return response

def checkvendorcontractimport(request):
    try:
        data_file=request.session['allcontratdata']
    except KeyError:
        return HttpResponse('<h2 style="text-align:center">Session Expired</h2>')
    contracts=data_file.get('all_contracts')
    project=Projectcreation.objects.filter(id=data_file.get('project')).first()
    project_discipline=ProjectDevelopmentDiscipline.objects.filter(id=data_file.get('project_discipline')).first()
    disciplinetype=ProjectDevelopmentSubType.objects.filter(id=data_file.get('disciplinetype')).first()
    currencies=Settings.objects.filter(company=request.company).first()
    companydateformat=''
    if (currencies.dateformat != None):
        companydateformat=currencies.dateformat
    else:
        companydateformat=''

    if (currencies != None):
        currency_convert=ast.literal_eval(currencies.currency)
    #print(currency_convert)
    get_currencies=list(Basecountries.objects.filter(id__in=currency_convert).values('id','currency'))
    currenices=[{'id':cur['id'],'currency':cur['currency']} for cur in get_currencies]
    if (request.POST):
        print(request.POST)
        vendorname=request.POST.getlist('vendorname')
        type_contract=request.POST.getlist('type_contract')
        type_service=request.POST.getlist('type_service')
        name_contract=request.POST.getlist('name_contract')
        # project_discipline=request.POST.getlist('project_discipline')
        # disciplinetype=request.POST.getlist('disciplinetype')
        ref_num=request.POST.getlist('ref_num')
        amendment=request.POST.getlist('amendment')
        date=request.POST.getlist('date')
        currency=request.POST.getlist('currency')
        amount=request.POST.getlist('amount')
        i=0
        amendmentlist=[]
        contractdata = []
        amendmentdata = []
        while i < len(vendorname):
            if (type_contract[i] == 'Original Contract'):
                if (companydateformat == 'dd-M-yy'):
                    convert_date=datetime.strptime(date[i],"%d-%b-%Y").date()
                    # print('1',date[i])
                elif (companydateformat == 'dd-mm-yy'):
                    convert_date=datetime.strptime(date[i],"%d-%m-%Y").date()
                    # print('2',date[i])
                elif (companydateformat == 'dd/mm/yy'):
                    convert_date=datetime.strptime(date[i],"%d/%m/%Y").date()
                    # print('3',date[i])
                elif (companydateformat == 'mm-dd-yy'):
                    convert_date=datetime.strptime(date[i],"%m-%d-%Y").date()
                    # print('4',date[i])
                elif (companydateformat == 'mm/dd/yy'):
                    convert_date=datetime.strptime(date[i],"%m/%d/%Y").date()
                    # print('5',date[i])
                elif (companydateformat == 'yy-mm-dd'):
                    convert_date=datetime.strptime(date[i],"%Y-%m-%d").date()
                    # print('6',date[i])
                elif (companydateformat == 'yy/mm/dd'):
                    convert_date=datetime.strptime(date[i],"%Y/%m/%d").date()
                    # print('7',date[i])
                else:
                    convert_date=datetime.strptime(date[i],"%d-%b-%Y").date()
                vendor=ContractMasterVendor.objects.filter(vendor_name=vendorname[i],status=1,company=request.company).first()
                savecontractmaster=ContractMaster(projects_id=project.id,projectdiscipline_id =project_discipline.id,projectdisciplinetype_id=disciplinetype.id,contractvendor_id=vendor.id,company_id=request.company.id,types_service=type_service[i],name_service=name_contract[i],reference_number=ref_num[i],executed_date=convert_date,currency_id=currency[i],amount=amount[i])
                savecontractmaster.save()
                contractdata.append((vendor.id,ref_num[i]))
            else:
                amendmentlist.append({'vendor':vendorname[i],'type':type_contract[i],'service':type_service[i],'contractname':name_contract[i],'projectdiscipline':project_discipline.id,'disciplinetype':disciplinetype.id,'refnum':ref_num[i],'amendment':amendment[i],'date':date[i],'currency':currency[i],'amount':amount[i]})
            i+=1
        contractdict = {}
        contractdict = {vendor: [contract_number for v, contract_number in contractdata if v == vendor] for vendor in set([v for v, _ in contractdata])}
        url = f'/projects/contractlist'            
        sender = User.objects.get(id=request.user.id)
        for i,j in contractdict.items():
            vendorcontract=ContractMasterVendor.objects.get(id=i)
            get_userby_vin = User.objects.filter(cin_number=vendorcontract.vin)
            notify.send(sender, recipient=get_userby_vin,data=url, verb='New Contract Imported', description=f'New Contract {",".join(map(str,j))} has been imported by {sender.name} {sender.lastname if sender.lastname != None else ""}')            
        # print('a',amendmentlist)
        for amendments in amendmentlist:
            vendor=ContractMasterVendor.objects.filter(vendor_name=amendments['vendor'],status=1,company=request.company).first()
            original_contract_num=amendments['amendment']
            contractid=ContractMaster.objects.filter(contractvendor=vendor,reference_number=original_contract_num,status=1).first()
            if (companydateformat == 'dd-M-yy'):
                amendmnet_convert_date=datetime.strptime(amendments['date'],"%d-%b-%Y").date()
                # print('1',date)
            elif (companydateformat == 'dd-mm-yy'):
                amendmnet_convert_date=datetime.strptime(amendments['date'],"%d-%m-%Y").date()
                # print('2',date)
            elif (companydateformat == 'dd/mm/yy'):
                amendmnet_convert_date=datetime.strptime(amendments['date'],"%d/%m/%Y").date()
                # print('3',date)
            elif (companydateformat == 'mm-dd-yy'):
                amendmnet_convert_date=datetime.strptime(amendments['date'],"%m-%d-%Y").date()
                # print('4',date)
            elif (companydateformat == 'mm/dd/yy'):
                amendmnet_convert_date=datetime.strptime(amendments['date'],"%m/%d/%Y").date()
                # print('5',date)
            elif (companydateformat == 'yy-mm-dd'):
                amendmnet_convert_date=datetime.strptime(amendments['date'],"%Y-%m-%d").date()
                # print('6',date)
            elif (companydateformat == 'yy/mm/dd'):
                amendmnet_convert_date=datetime.strptime(amendments['date'],"%Y/%m/%d").date()
                # print('7',date)
            else:
                amendmnet_convert_date=datetime.strptime(amendments['date'],"%d-%b-%Y").date()
            save_amendment=Amendment(contractvendor=vendor,company=request.company,service_id =contractid.id,amendment_type=amendments['type'],amendment_reference_number=amendments['refnum'],amendment_executed_date=amendmnet_convert_date,amendment_currency_id =amendments['currency'],amendment_amount=amendments['amount'])
            save_amendment.save()
            amendmentdata.append((vendor.id,amendments['refnum']))
        amendmentdict = {}
        amendmentdict = {vendor: [contract_number for v, contract_number in amendmentdata if v == vendor] for vendor in set([v for v, _ in amendmentdata])}
        url = f'/projects/contractlist'            
        sender = User.objects.get(id=request.user.id)
        for i,j in amendmentdict.items():
            vendorcontract=ContractMasterVendor.objects.get(id=i)
            get_userby_vin = User.objects.filter(cin_number=vendorcontract.vin)
            notify.send(sender, recipient=get_userby_vin,data=url, verb='New Amendment/Addendum Imported', description=f'New Amendment/Addendum {",".join(map(str,j))} has been imported by {sender.name} {sender.lastname if sender.lastname != None else ""}')  
        return redirect('projects:contractmasterlist')
    contract_snum=[]
    for key,contract in enumerate(contracts):
        if (contract['vendor_name'] == None):
            contract_snum.append(key+1)
            print('a')
        elif (contract['type_contract'] == None):
            contract_snum.append(key+1)
            print('b')

        elif (contract['type_service'] == None):
            contract_snum.append(key+1)
            print('c')

        elif (contract['name_contract'] == None):
            contract_snum.append(key+1)
            print('d')

        # elif (contract['project_discipline'] == None):
        #     contract_snum.append(key+1)

        # elif (contract['discipline_type'] == None):
        #     contract_snum.append(key+1)
    
        elif (contract['con_num'] == None or ContractMaster.objects.filter(company=request.company,reference_number__exact=contract['con_num'],status=1).exists()):
            contract_snum.append(key+1)
            print('e')
        elif (contract['executed_date'] == None):
            contract_snum.append(key+1)
            print('g')

        elif (contract['max_value_currency'] == None):
            contract_snum.append(key+1)
            print('h')

        elif (contract['max_value_amount'] == None):
            contract_snum.append(key+1)
            print('j')
        if (contract['type_contract'] ==  None or contract['type_contract'] != 'Original Contract'):
            if (contract['amendment_refnum'] == None):
                contract_snum.append(key+1)
                print('dsf',contract_snum)


    converttuple=tuple(contract_snum)
    sortlist=list(converttuple)
    print(sortlist)
    sortlist.sort()
    if (len(sortlist) > 0):
        convert_str = str(sortlist)[1]
    else:
        convert_str=''
    print('convert_str',convert_str)

    all_vendors=list(ContractMasterVendor.objects.filter(company=request.company,status=1).values_list('vendor_name',flat=True))
    contract_type=['Original Contract','Amendment','Addendum']
    service_type={'Service':'service','Supply':'supply','Service and Supply':'service_supply'}
    # projectdiscpline={'Green Field Development':'green_field_development','Brown Field Development':'brown_field_development','Others':'others'}
    # dis_type={'Subsurface and Reservoir Engineering':'subsurface_and_reservoir_engineering', 'Drilling and Completions':'drilling_and_completions','Facilities and Projects':'facilities_and_projects','Production and Operations':'production_and_operations','HR and Administration':'human_resources_and_administration','Legal':'legal','ICT':'ICT','Finance':'finance','QHSE & Community Development':'QHSE_&_community_development','Business Development & Sales and Marketing':'business_development_&_sales_and_marketing','Insurance':'insurance','Miscellaneous':'miscellaneous'}
    data={'contracts':contracts,
        'all_vendors':all_vendors,
        'contract_types':contract_type,
        'service_type':service_type,
        # 'projectdiscpline':projectdiscpline,
        # 'discplinetypes':dis_type,
        'currencies':currenices,'companyid':request.company.id,
        'companydateformat':companydateformat,
        'convert_str':convert_str,
        'project':project,'project_discipline':project_discipline,
        'disciplinetype':disciplinetype}
    return render(request,'checkvendorcontractimport.html',data)


def checkcontractrefduplicate(request):
    refnum=request.GET.get('refnum')
    if ContractMaster.objects.filter(company=request.company,reference_number__exact=refnum,status=1).exists() or Amendment.objects.filter(company=request.company,amendment_reference_number__exact=refnum,status=1).exists():
        data={'data':'exists'}
    else:
        data={'data':'success'}
    return JsonResponse(data)

def checkcontractamendment(request):
    vname=request.GET.get('v_name')
    refnum=request.GET.get('refnum')
    amendment=request.GET.get('amendment')
    print(refnum,refnum)
    print('amendment',amendment)
    vendorname=ContractMasterVendor.objects.filter(company=request.company,vendor_name=vname,status=1).first()
    if (ContractMaster.objects.filter(contractvendor=vendorname,reference_number=refnum,status=1).exists()):
        contractid=ContractMaster.objects.filter(contractvendor=vendorname,reference_number=refnum,status=1).first()
        if (Amendment.objects.filter(amendment_reference_number=amendment,contractvendor=vendorname,service_id=contractid,status=1).exists()):
            data={'data':'exists'}
        else:
            data={'data':'new'}
    else:
        data={'data':'no_data'}
    return JsonResponse(data)

def checkcontractamendmentref(request):
    vname=request.GET.get('v_name')
    refnum=request.GET.get('refnum')
    vendorname=ContractMasterVendor.objects.filter(company=request.company,vendor_name=vname,status=1).first()
    if (ContractMaster.objects.filter(contractvendor=vendorname,reference_number=refnum,status=1).exists()):
        contractid=ContractMaster.objects.filter(contractvendor=vendorname,reference_number=refnum,status=1).first()
        data={'data':'success'}
    else:
        data={'data':'no_data'}
    return JsonResponse(data)

def addfilescontract(request):
    without_file_contract=ContractMaster.objects.filter(company=request.company,status=1)
    contract_empty_files=without_file_contract.filter(Q(upload_contract='') | Q(upload_pricetable='')).count()
    get_contract_ids=without_file_contract.values_list('id',flat=True)
    amendments_empty_files=Amendment.objects.filter(Q(service_id__in=get_contract_ids),Q(status=1),Q(amendment_upload_contract='') | Q(amendment_upload_pricetable='')).count()
    print('ewfewf',contract_empty_files,amendments_empty_files)
    if (contract_empty_files  == 0  and  amendments_empty_files == 0):
        contract_count = 0 
    else:
        contract_count = 1
    print(contract_count)
    # without_file_contract=files.filter(Q(upload_contract='') | Q(upload_pricetable=''))[:15]
    # without_file_contract=ContractMaster.objects.filter(Q(company=request.company) | Q(upload_contract='') | Q (upload_pricetable='') | Q(status=1)).exclude(Q(name_service__exact = '') | Q(name_service__isnull=True) | Q(reference_number__exact='') | Q(reference_number__isnull=True) | Q(executed_date__isnull=True) | Q(amount__exact='') | Q(amount__isnull=True) | Q(projects_id__isnull=True) | Q(projectdiscipline_id__isnull=True) | Q(projectdisciplinetype_id__isnull=True))[:15]
    if (request.POST):
        print('request',request.POST)
        contractfileids=request.POST.getlist('contractfileids')
        for contract in contractfileids:
            contractfile=request.FILES.get('contractfile'+str(contract))
            pricetablefile=request.FILES.get('pricetable'+str(contract))
            fs = FileSystemStorage()
            if (contractfile != None):
                ContractMaster.objects.filter(id=contract).update(upload_contract=contractfile)
                file1 = fs.save(contractfile.name, contractfile)
            if (pricetablefile != None):
                ContractMaster.objects.filter(id=contract).update(upload_pricetable=pricetablefile)
                file2 = fs.save(pricetablefile.name, pricetablefile)
            # print(contract,contractfile,pricetablefile)
            getamendmentids=request.POST.getlist('amendmentfileids'+str(contract))
            if (len(getamendmentids) > 0):
                for amendment in getamendmentids:
                    amendmentcontractfile=request.FILES.get('amendmentcontract'+str(amendment))
                    amendmentpricefile=request.FILES.get('amendmentprice'+str(amendment))
                    fs = FileSystemStorage()
                    print('amendmentcontractfile',amendmentcontractfile)
                    print('amendmentpricefile',amendmentpricefile)
                if (amendmentcontractfile != None):
                    Amendment.objects.filter(id=amendment).update(amendment_upload_contract=amendmentcontractfile)
                    file3 = fs.save(amendmentcontractfile.name, amendmentcontractfile)
                if (amendmentpricefile != None):
                    Amendment.objects.filter(id=amendment).update(amendment_upload_pricetable=amendmentpricefile)
                    file4 = fs.save(amendmentpricefile.name, amendmentpricefile)
        return redirect('projects:contractmasterlist')
                    # print('getamendmentids',getamendmentids,amendmentcontractfile,amendmentpricefile)
    data={'contract_file':without_file_contract,'contract_count':contract_count}
    # print(data)
    return render(request,'addcontractfiles.html',data)

def downloaduploadcontractdata(request):
    missed_contracts=request.GET.get('missed')
    exists_contract=request.GET.get('exists_contract')
    # convert_misseddatas=ast.literal_eval(missed_contracts)
    convert_missedcontracts=json.loads(missed_contracts)
    convert_existscontract=json.loads(exists_contract)
    # print('EX_CON',convert_existscontract)
    datas={'missdata':convert_missedcontracts,
            'ex_data':convert_existscontract}
    return JsonResponse({'data':'success'})


def createcontractmaster(request):
    company=Settings.objects.filter(company_id=request.company.id).first()
    if (company != None):
        getcurrencylist=ast.literal_eval(company.currency)
        companycurrency=Basecountries.objects.filter(id__in=getcurrencylist)
    else:
        companycurrency=None
    vendormasterlist=ContractMasterVendor.objects.filter(company_id=request.company.id,status=1)
    request.session['submenu']='contract_projects'
    reference_number = []
    if (request.POST):
        print(request.POST)
        vendorname=request.POST.get('vendorname')
        submit_type=request.POST.get('submit_type')
        save_type=1
        if submit_type == '1':
            save_type=2
        types_service=request.POST.getlist('types_service')
        print('types_service',types_service)
        for service in types_service:
            nameservice=request.POST.getlist(service+'n_service')
            projectid=request.POST.getlist(service+'project')
            discipline=request.POST.getlist(service+'project_discipline')
            disciplinetype=request.POST.getlist(service+'disciplinetype')
            referencenumber=request.POST.getlist(service+'ref_number')
            wcclist=request.POST.getlist(service+'wcc')
            execute_date=request.POST.getlist(service+'execute_date')
            currency=request.POST.getlist(service+'currency')
            amount=request.POST.getlist(service+'amount')
            contractfile=request.FILES.getlist(service+'contract_file',None)
            price_table=request.FILES.getlist(service+'contract_price_file',None)
            contract_file_name=request.FILES.getlist(service+'contract_file',None)
            price_table_file_name=request.FILES.getlist(service+'contract_price_file',None)
            new_amendments=[]
            i=0
            while i<len(nameservice):
                if (projectid[i] != '' or discipline[i] != '' or disciplinetype[i] != '' or  nameservice[i] != '' or referencenumber[i] != '' or execute_date[i] != '' or currency[i] != '' or amount[i] != ''):
                    contractfile=None if(len(contractfile) == 0)  else contractfile[i]
                    convert_date=convertdate(company.dateformat,execute_date[i])
                    try:
                        contract_file_name=contract_file_name[i].name
                    except:
                        contract_file_name=contract_file_name[i]
                    create_contracts=ContractMaster.objects.create(projects_id=projectid[i],projectdiscipline_id =discipline[i],projectdisciplinetype_id =disciplinetype[i],contractvendor_id=vendorname,company_id=request.company.id,types_service=service,name_service=nameservice[i],reference_number=referencenumber[i],executed_date=convert_date,currency_id=currency[i],amount=amount[i],upload_contract=contractfile,contract_file_name=contract_file_name,upload_pricetable=price_table[i],price_table_file_name=price_table_file_name[i],created_by=request.user,wcc=wcclist[i],save_type=save_type)

                    if(len(price_table) > 0):   
                        for pricetable in price_table:
                            VendorContractPriceTable.objects.uploadpricetablefile(pricetable,request.company.id,vendorname,create_contracts.id)
                    usercreate = request.user.roles_id
                    create_user_log(request,create_contracts.reference_number,'Contract Master','Create','Contract Master Created',usercreate)

                    reference_number.append(referencenumber[i])
                    a_wcc=request.POST.getlist(service+'wcc'+str(i))
                    a_type=request.POST.getlist(service+'extrafile'+str(i))
                    ref_num=request.POST.getlist(service+'ref_num'+str(i))
                    a_date=request.POST.getlist(service+'dateformat'+str(i))
                    a_currency=request.POST.getlist(service+'currency'+str(i))
                    a_amount=request.POST.getlist(service+'amount'+str(i))
                    a_con_file=request.FILES.getlist(service+'changed_con_file'+str(i))
                    a_price_table=request.FILES.getlist(service+'changed_price_file'+str(i))
                    j=0
                    if (len(a_type) > 0):
                        while j<len(a_type):
                            if (a_type[j] != '' or ref_num[j] != '' or a_date[j] != '' or a_currency[j] != '' or a_amount[j] != '' or a_con_file[j] != None or a_price_table[j] != None):
                                amendmnet_convert_date=convertdate(company.dateformat,a_date[j])
                                amendment_contract_file=None if len(a_con_file) == 0 else a_con_file[j]
                                create_amendment=Amendment.objects.create(contractvendor_id=vendorname,company_id=request.company.id,service_id =create_contracts.id,amendment_type=a_type[j],amendment_reference_number=ref_num[j],amendment_executed_date=amendmnet_convert_date,amendment_currency_id =a_currency[j],amendment_amount=a_amount[j],wcc=a_wcc[j],amendment_upload_contract=amendment_contract_file,save_type=save_type)

                                if(len(a_price_table) > 0):
                                    for pricetable in a_price_table:
                                        VendorContractPriceTable.objects.uploadpricetablefile(pricetable,request.company.id,vendorname,create_contracts.id,create_amendment.id)
                                new_amendments.append(ref_num[j])

                            j += 1
                i+=1
        sender = User.objects.get(id=request.user.id)
        for i in vendormasterlist:
            if i.id == int(vendorname):
                get_vinnumber = list(ContractMasterVendor.objects.filter(id=vendorname,company=request.company.id).values_list('vin',flat=True))
                get_userby_vin = User.objects.filter(cin_number=get_vinnumber[0])
        scheme=request.scheme
        gethost=request.get_host()  
        url=f"{scheme}://{gethost}/projects/contractlist"
        # submit_value=request.POST.get('submit_type')
        # if (submit_value == '1'):
        contract_ref=','.join(map(str,reference_number))  
        notify.send(sender, recipient=get_userby_vin,data=url, verb='New Contract Added', description=f'New Contract {contract_ref} has been added by {sender.name} {sender.lastname if sender.lastname != None else ""}')    
        # create_user_log(create_contracts.id,"Contract",request.company.id,"Create","Contract Created")

        if (len(new_amendments) > 0):
            amendment_ref=','.join(map(str,new_amendments))
            notify.send(sender, recipient=get_userby_vin,data=url, verb='New Amendment/Addendum Created', description=f'New Amendment/Addendum {amendment_ref} has been created by {sender.name} {sender.lastname if sender.lastname != None else ""}') 
        return redirect('projects:contractmasterlist')
    projects=Projectcreation.objects.filter(company=request.company,status=0)
    
    data={'company':company,'companycurrency':companycurrency,'vendormasterlist':vendormasterlist,'projects':projects}
    return render(request,'createcontractmaster.html',data)

def getprojectdevelopment(request):
    projectid=request.GET.get('projectid')
    projectflow_level=ProjectFlowlevel.objects.getprojectflowlevel_by_project_id(projectid)
    if projectflow_level.count() == 0:
        return JsonResponse({'data': 'invoiceflow'})
    # wccflow_level=WccFlow.objects.filter_by_project(project_id=projectid).count()
    # if wccflow_level == 0:
    #     return JsonResponse({'data': 'wccflow'})
    projectdisciplineid=request.GET.get('projectdisciplineid')
    if(projectdisciplineid != None):
        projectdisciplinesubtype=ProjectDevelopmentSubType.objects.filter(project_discipline_id=projectdisciplineid,status=1)
        disciplinetypelist=[]
        for disciplinetype in projectdisciplinesubtype:
            # disctype=disciplinetype.discipline_subtype.discipline_subtype
            # replace_word=disctype.replace(' ','_')
            disciplinetypelist.append({'value':disciplinetype.id,'text':disciplinetype.discipline_subtype.discipline_subtype})
        return JsonResponse({'data':disciplinetypelist})
    elif (projectid != None):
        projectdevelopment=ProjectDevelopmentDiscipline.objects.filter(project_id=projectid,status=1)
        developmentlist=[]
        for development in projectdevelopment:
            convert_title=development.development_type.cluster.environment.field.field.field_name.title()
            convert_cluster_title=development.development_type.cluster.clustersubname.cluster_subname.title()
            # print('q',convert_cluster_title)
            if (development.project_discipline == '1'):
                developmentlist.append({'id':development.id,'fieldname':'Green Field Development-'+convert_title+'-'+convert_cluster_title})
            elif (development.project_discipline == '2'):
                developmentlist.append({'id':development.id,'fieldname':'Brown Field Development-'+convert_title+'-'+convert_cluster_title})
            elif (development.project_discipline == '3'):
                developmentlist.append({'id':development.id,'fieldname':'Others-'+convert_title+'-'+convert_cluster_title})
        check_project_wcc_flow=WccFlow.objects.filter_by_project(project_id=projectid).count()
        return JsonResponse({'data':developmentlist,'project_wcc_flow':check_project_wcc_flow})
    
    else:
        return JsonResponse({'data':'none'})


    


def editcontractmaster(request,pk):
    company=Settings.objects.filter(company_id=request.company.id).first()
    getcurrencylist=ast.literal_eval(company.currency)
    companycurrency=Basecountries.objects.filter(id__in=getcurrencylist)
    vendorcontract=ContractMasterVendor.objects.get(id=pk)
    discipline=['green_field_development','brown_field_development','others']
    disciplinetype=['subsurface_and_reservoir_engineering','drilling_and_completions','facilities_and_projects','production_and_operations','human_resources_and_administration','legal','ICT','finance','QHSE_&_community_development','business_development_&_sales_and_marketing','insurance','miscellaneous']
    types=['Amendment','Addendum']
    orginal_amendment = list(Amendment.objects.filter(contractvendor=pk,status=1).values_list('amendment_reference_number',flat=True))
    original_contract = list(ContractMaster.objects.filter(contractvendor=pk,status=1).values_list('reference_number',flat=True))
    if (request.POST):
        editdetails = []
        amendment_edit=[]
        # print(request.POST)
        # print(request.FILES)
        types_service=request.POST.getlist('types_service')

        currentcontractids=[]
        new_contracts=[]
        new_amendments=[]
        for service in types_service:
            conttracthdnid=request.POST.getlist(service+'contracthdnid')
            nameservice=request.POST.getlist(service+'n_service')
            projects=request.POST.getlist(service+'project')
            discipline=request.POST.getlist(service+'project_discipline')
            disciplinetype=request.POST.getlist(service+'disciplinetype')
            referencenumber=request.POST.getlist(service+'ref_number')
            execute_date=request.POST.getlist(service+'execute_date')
            currency=request.POST.getlist(service+'currency')
            amount=request.POST.getlist(service+'amount')
            wcc=request.POST.getlist(service+'wcc')
            # contractfile=request.FILES.get(service+'contract_file')
            checkcontractfile=request.POST.getlist(service+'contract_file')
            price_table=request.FILES.getlist(service+'contract_price_file')
            i=0

            while i<len(conttracthdnid):

                if (conttracthdnid[i]):
                    # print('conttracthdnid',conttracthdnid[i])
                    if (execute_date[i] == ''):
                        convert_date= None
                    elif (company.dateformat == 'dd-M-yy'):
                        convert_date=datetime.strptime(execute_date[i],"%d-%b-%Y").date()
                        # print('1',execute_date[i])
                    elif (company.dateformat == 'dd-mm-yy'):
                        convert_date=datetime.strptime(execute_date[i],"%d-%m-%Y").date()
                        # print('2',execute_date[i])
                    elif (company.dateformat == 'dd/mm/yy'):
                        convert_date=datetime.strptime(execute_date[i],"%d/%m/%Y").date()
                        # print('3',execute_date[i])
                    elif (company.dateformat == 'mm-dd-yy'):
                        convert_date=datetime.strptime(execute_date[i],"%m-%d-%Y").date()
                        # print('4',execute_date[i])
                    elif (company.dateformat == 'mm/dd/yy'):
                        convert_date=datetime.strptime(execute_date[i],"%m/%d/%Y").date()
                        # print('5',execute_date[i])
                    elif (company.dateformat == 'yy-mm-dd'):
                        convert_date=datetime.strptime(execute_date[i],"%Y-%m-%d").date()
                        # print('6',execute_date[i])
                    elif (company.dateformat == 'yy/mm/dd'):
                        convert_date=datetime.strptime(execute_date[i],"%Y/%m/%d").date()
                        # print('7',execute_date[i])
                    else:
                        convert_date= None
                        convert_date=datetime.strptime(execute_date[i],"%d-%b-%Y").date()
                    # print('old',conttracthdnid[i])
                    contractfile=request.FILES.get(service+'contract_file'+str(i))
                    pricefile=request.FILES.get(service+'contract_price_file'+str(i))
                    contract_file_name = request.FILES.get(service+'contract_file'+str(i))
                    price_table_file_name = request.FILES.get(service+'contract_price_file'+str(i))
                    # print(contractfile,pricefile)
                    previous_data = ContractMaster.objects.get(id=conttracthdnid[i])
                    # for old_data in previous_data:
                        # print('old_data',i.reference_number)
                    if (previous_data.name_service != nameservice[i]):
                        # print(f'name of service edited {previous_data.name_service} to {nameservice[i]}')
                        editdetails.append(referencenumber[i])
                        # editdetails.append(f'{previous_data.reference_number} name of service edited {previous_data.name_service} to {nameservice[i]}')
                    # elif previous_data.projects_id != None:
                    #     if previous_data.projects_id != int(projects[i]):
                    #         editdetails.append(referencenumber[i])
                            # print(f'project edited',type(projects[i]),type(previous_data.projects_id))
                            # editdetails.append(f'{previous_data.reference_number} project edited')
                    # elif (previous_data.projectdiscipline_id != None):
                    #     if previous_data.projectdiscipline_id != int(discipline[i]):
                    #         editdetails.append(referencenumber[i])
                            # print(f'project discipline edited',type(discipline[i]),type(previous_data.projectdiscipline_id))
                            # editdetails.append(f'{previous_data.reference_number} project discipline edited')
                    # elif (previous_data.projectdisciplinetype_id != None):
                    #     if previous_data.projectdisciplinetype_id != int(disciplinetype[i]):
                    #         editdetails.append(referencenumber[i])
                            # print(f'discipline type edited',type(discipline[i]),type(previous_data.projectdisciplinetype_id))
                            # editdetails.append(f'{previous_data.reference_number} discipline type edited')
                    elif (previous_data.reference_number != referencenumber[i]):
                        editdetails.append(referencenumber[i])
                        # print(f'reference number edited')
                        # editdetails.append(f'{previous_data.reference_number} reference number edited {previous_data.reference_number} to {referencenumber[i]}')
                    elif (previous_data.executed_date != convert_date):
                        editdetails.append(referencenumber[i])
                        # print(f'execute date edited')
                        # print(previous_data.executed_date,convert_date)
                    #     editdetails.append(f'{previous_data.reference_number} execute date edited {previous_data.executed_date} to {convert_date}')
                    elif (previous_data.currency_id != None):
                        if previous_data.currency_id != int(currency[i]):
                            editdetails.append(referencenumber[i])
                            # print(f'currency edited')
                            # editdetails.append(f'{previous_data.reference_number} currency edited')
                    elif (previous_data.amount != amount[i]):
                        editdetails.append(referencenumber[i])
                        # print(f'amount edited')
                        # editdetails.append(f'{previous_data.reference_number} amount edited {previous_data.amount} to {amount[i]}')
                    # elif (previous_data.upload_contract != contractfile):
                    #     print(f'contract file edited')
                    #     editdetails.append(f'{previous_data.reference_number} contract file edited {previous_data.upload_contract} to {contractfile}')
                    # elif (previous_data.upload_pricetable != pricefile):
                    #     print(f'price table edited')
                    #     editdetails.append(f'{previous_data.reference_number} price table edited {previous_data.upload_pricetable} to {pricefile}')
                    else:
                        print()
                        # print(f'no changes')
                    if (contractfile == None and pricefile == None):
                        # print("no change in old data")
                        bothnone = ContractMaster.objects.get(id=conttracthdnid[i])
                        bothnone.projects_id=projects[i]
                        bothnone.projectdiscipline_id=discipline[i]
                        bothnone.projectdisciplinetype_id=disciplinetype[i]
                        bothnone.name_service=nameservice[i]
                        bothnone.reference_number=referencenumber[i]
                        bothnone.executed_date=convert_date
                        bothnone.currency_id=currency[i]
                        bothnone.amount=amount[i]
                        bothnone.wcc=wcc[i]
                        bothnone.updated_by=request.user
                        bothnone.save()
                        # ContractMaster.objects.filter(id=conttracthdnid[i]).update(projects_id=projects[i],projectdiscipline_id=discipline[i],projectdisciplinetype_id=disciplinetype[i],name_service=nameservice[i],reference_number=referencenumber[i],executed_date=convert_date,currency_id=currency[i],amount=amount[i])
                        currentcontractids.append(conttracthdnid[i])

                    elif (contractfile == None):
                        # print('price table change')
                        # print(pricefile)
                        contractfilenone = ContractMaster.objects.get(id=conttracthdnid[i])
                        contractfilenone.upload_pricetable=pricefile
                        contractfilenone.price_table_file_name=price_table_file_name
                        contractfilenone.projects_id=projects[i]
                        contractfilenone.projectdiscipline_id=discipline[i]
                        contractfilenone.projectdisciplinetype_id=disciplinetype[i]
                        contractfilenone.name_service=nameservice[i]
                        contractfilenone.reference_number=referencenumber[i]
                        contractfilenone.executed_date=convert_date
                        contractfilenone.currency_id=currency[i]
                        contractfilenone.wcc=wcc[i]
                        contractfilenone.amount=amount[i]
                        contractfilenone.updated_by=request.user
                        contractfilenone.save()
                        # ContractMaster.objects.filter(id=conttracthdnid[i]).update(upload_pricetable=pricefile,projects_id=projects[i],projectdiscipline_id=discipline[i],projectdisciplinetype_id=disciplinetype[i],name_service=nameservice[i],reference_number=referencenumber[i],executed_date=convert_date,currency_id=currency[i],amount=amount[i])
                        # currentcontractids.append(conttracthdnid[i])
                        fs = FileSystemStorage()
                        file = fs.save(pricefile.name, pricefile)
                        editdetails.append(referencenumber[i])
                        # editdetails.append(f'{previous_data.reference_number} price table edited {previous_data.upload_pricetable} to {pricefile}')
                        # print("file",file)
                    elif (pricefile == None):
                        # print('contract file change')
                        pricefilenone = ContractMaster.objects.get(id=conttracthdnid[i])
                        pricefilenone.upload_contract=contractfile
                        pricefilenone.contract_file_name=contract_file_name.name
                        pricefilenone.projects_id=projects[i]
                        pricefilenone.projectdiscipline_id=discipline[i]
                        pricefilenone.projectdisciplinetype_id=disciplinetype[i]
                        pricefilenone.name_service=nameservice[i]
                        pricefilenone.reference_number=referencenumber[i]
                        pricefilenone.executed_date=convert_date
                        pricefilenone.currency_id=currency[i]
                        pricefilenone.wcc=wcc[i]
                        pricefilenone.amount=amount[i]
                        pricefilenone.updated_by=request.user
                        pricefilenone.save()
                        # ContractMaster.objects.filter(id=conttracthdnid[i]).update(upload_contract=contractfile,projects_id=projects[i],projectdiscipline_id=discipline[i],projectdisciplinetype_id=disciplinetype[i],name_service=nameservice[i],reference_number=referencenumber[i],executed_date=convert_date,currency_id=currency[i],amount=amount[i])
                        currentcontractids.append(conttracthdnid[i])
                        fs = FileSystemStorage()
                        file = fs.save(contractfile.name, contractfile)
                        editdetails.append(referencenumber[i])
                        # editdetails.append(f'{previous_data.reference_number} contract file edited {previous_data.upload_contract} to {contractfile}')
                    else:
                        # print("file change in old data",contractfile,pricefile)
                        contractnone=ContractMaster.objects.get(id=conttracthdnid[i])
                        contractnone.upload_contract=contractfile
                        contractnone.contract_file_name=contract_file_name.name
                        contractnone.upload_pricetable=pricefile
                        contractnone.price_table_file_name=price_table_file_name
                        contractnone.projects_id=projects[i]
                        contractnone.projectdiscipline_id=discipline[i]
                        contractnone.projectdisciplinetype_id=disciplinetype[i]
                        contractnone.name_service=nameservice[i]
                        contractnone.reference_number=referencenumber[i]
                        contractnone.wcc=wcc[i]
                        contractnone.executed_date=convert_date
                        contractnone.currency_id=currency[i]
                        contractnone.amount=amount[i]
                        contractnone.updated_by=request.user
                        contractnone.save()
                        # ContractMaster.objects.filter(id=conttracthdnid[i]).update(upload_contract=contractfile,upload_pricetable=pricefile,projects_id=projects[i],projectdiscipline_id=discipline[i],projectdisciplinetype_id=disciplinetype[i],name_service=nameservice[i],reference_number=referencenumber[i],executed_date=convert_date,currency_id=currency[i],amount=amount[i])
                        fs = FileSystemStorage()
                        file = fs.save(contractfile.name, contractfile)
                        pfs = FileSystemStorage()
                        pricefile = pfs.save(pricefile.name, pricefile)
                        currentcontractids.append(conttracthdnid[i])
                        editdetails.append(referencenumber[i])
                        # editdetails.append(f'{previous_data.reference_number} contract file edited {previous_data.upload_contract} to {contractfile} and {previous_data.reference_number} price table edited {previous_data.upload_pricetable} to {pricefile}')

                    get_amendment_ids=request.POST.getlist(str(service)+str(i))
                    # print("get_amendment_ids",get_amendment_ids)

                    current_amendemnt_ids=[]
                    for index,values in enumerate(get_amendment_ids):
                        # print('index',index)
                        if (values):
                            old_amendment_data=Amendment.objects.get(id=values)
                            amendment_type=request.POST.get(service+'extrafile'+str(i)+str(index))
                            amendment_ref_num=request.POST.get(service+'ref_num'+str(i)+str(index))
                            amedment_date=request.POST.get(service+'dateformat'+str(i)+str(index))
                            amedment_wcc=request.POST.get(service+'wcc'+str(i)+str(index))
                            amendment_currency=request.POST.get(service+'currency'+str(i)+str(index))
                            amendment_amount=request.POST.get(service+'amount'+str(i)+str(index))
                            amendment_con_file=request.FILES.get(service+'con_file'+str(i)+str(index))
                            amendment_price_file=request.FILES.get(service+'price_file'+str(i)+str(index))
                            if (amedment_date == ''):
                                amendmnet_convert_date = None
                            elif (company.dateformat == 'dd-M-yy'):
                                amendmnet_convert_date=datetime.strptime(amedment_date,"%d-%b-%Y").date()
                                # print('1',amedment_date)
                            elif (company.dateformat == 'dd-mm-yy'):
                                amendmnet_convert_date=datetime.strptime(amedment_date,"%d-%m-%Y").date()
                                # print('2',amedment_date)
                            elif (company.dateformat == 'dd/mm/yy'):
                                amendmnet_convert_date=datetime.strptime(amedment_date,"%d/%m/%Y").date()
                                # print('3',amedment_date)
                            elif (company.dateformat == 'mm-dd-yy'):
                                amendmnet_convert_date=datetime.strptime(amedment_date,"%m-%d-%Y").date()
                                # print('4',amedment_date)
                            elif (company.dateformat == 'mm/dd/yy'):
                                amendmnet_convert_date=datetime.strptime(amedment_date,"%m/%d/%Y").date()
                                # print('5',amedment_date)
                            elif (company.dateformat == 'yy-mm-dd'):
                                amendmnet_convert_date=datetime.strptime(amedment_date,"%Y-%m-%d").date()
                                # print('6',amedment_date)
                            elif (company.dateformat == 'yy/mm/dd'):
                                amendmnet_convert_date=datetime.strptime(amedment_date,"%Y/%m/%d").date()
                                # print('7',amedment_date)
                            else:
                                amendmnet_convert_date = None
                                # amendmnet_convert_date=datetime.strptime(amedment_date,"%d-%b-%Y").date()
                            if (old_amendment_data.amendment_type != amendment_type):
                                amendment_edit.append(amendment_ref_num)
                                print('a1')
                                # amendment_edit.append(f'{old_amendment_data.amendment_ref_num} amendment/addendum edited {old_amendment_data.amendment_type} to {amendment_type}')
                            elif (old_amendment_data.amendment_reference_number != amendment_ref_num):
                                print('a2')
                                amendment_edit.append(amendment_ref_num)
                            # elif (old_amendment_data.amendment_executed_date == None  and amendmnet_convert_date == None):
                            #     print('a3',old_amendment_data.amendment_executed_date,'a3-a',amendmnet_convert_date)
                            #     amendment_edit.append(amendment_ref_num)
                            elif (old_amendment_data.amendment_executed_date != None and old_amendment_data.amendment_executed_date.date() != amendmnet_convert_date):
                                print('e-old')
                                amendment_edit.append(amendment_ref_num)
                            elif (old_amendment_data.amendment_executed_date == None and  amendmnet_convert_date != None):
                                print('e-new')  
                                amendment_edit.append(amendment_ref_num)
                            # convert_amendment_currency=amendment_currency
                            elif (old_amendment_data.amendment_currency_id != None and old_amendment_data.amendment_currency_id != int(amendment_currency)):
                                print('a4',type(old_amendment_data.amendment_currency_id),'a4a',type(amendment_currency))
                                amendment_edit.append(amendment_ref_num)
                            elif (old_amendment_data.amendment_currency_id == None and amendment_currency != ''):
                                amendment_edit.append(amendment_ref_num)
                            elif (old_amendment_data.amendment_amount != amendment_amount):
                                print('a5')
                                amendment_edit.append(amendment_ref_num)
                            if (amendment_con_file == None and amendment_price_file == None):
                                # print('no change')
                                amendmentbothnone=Amendment.objects.get(id=values)
                                amendmentbothnone.amendment_type=amendment_type
                                amendmentbothnone.amendment_reference_number=amendment_ref_num
                                amendmentbothnone.amendment_executed_date=amendmnet_convert_date
                                amendmentbothnone.amendment_currency_id=amendment_currency
                                amendmentbothnone.wcc=amedment_wcc
                                amendmentbothnone.amendment_amount=amendment_amount
                                amendmentbothnone.updated_by=request.user
                                amendmentbothnone.save()
                                # Amendment.objects.filter(id=values).update(amendment_type=amendment_type,amendment_reference_number=amendment_ref_num,amendment_executed_date=amendmnet_convert_date,amendment_currency_id=amendment_currency,amendment_amount=amendment_amount)
                                current_amendemnt_ids.append(values)
                            elif (amendment_con_file == None):
                                # print('price table change')
                                print('a6')
                                amendment_con_filenone=Amendment.objects.get(id=values)
                                amendment_con_filenone.amendment_upload_pricetable=amendment_price_file
                                amendment_con_filenone.amendment_type=amendment_type
                                amendment_con_filenone.wcc=amedment_wcc
                                amendment_con_filenone.amendment_reference_number=amendment_ref_num
                                amendment_con_filenone.amendment_executed_date=amendmnet_convert_date
                                amendment_con_filenone.amendment_currency_id=amendment_currency
                                amendment_con_filenone.amendment_amount=amendment_amount
                                amendment_con_filenone.updated_by=request.user
                                amendment_con_filenone.save()
                                # Amendment.objects.filter(id=values).update(amendment_upload_pricetable=amendment_price_file,amendment_type=amendment_type,amendment_reference_number=amendment_ref_num,amendment_executed_date=amendmnet_convert_date,amendment_currency_id=amendment_currency,amendment_amount=amendment_amount)
                                current_amendemnt_ids.append(values)
                                fs = FileSystemStorage()
                                file = fs.save(amendment_price_file.name, amendment_price_file)
                                amendment_edit.append(amendment_ref_num)
                            elif (amendment_price_file == None):
                                print('a7')
                                # print('price table change')
                                amendment_price_filenone=Amendment.objects.get(id=values)
                                amendment_price_filenone.amendment_upload_contract=amendment_con_file
                                amendment_price_filenone.amendment_type=amendment_type
                                amendment_price_filenone.wcc=amedment_wcc
                                amendment_price_filenone.amendment_reference_number=amendment_ref_num
                                amendment_price_filenone.amendment_executed_date=amendmnet_convert_date
                                amendment_price_filenone.amendment_currency_id=amendment_currency
                                amendment_price_filenone.amendment_amount=amendment_amount
                                amendment_price_filenone.updated_by=request.user
                                amendment_price_filenone.save()
                                # Amendment.objects.filter(id=values).update(amendment_upload_contract=amendment_con_file,amendment_type=amendment_type,amendment_reference_number=amendment_ref_num,amendment_executed_date=amendmnet_convert_date,amendment_currency_id=amendment_currency,amendment_amount=amendment_amount)
                                current_amendemnt_ids.append(values)
                                fs = FileSystemStorage()
                                file = fs.save(amendment_con_file.name, amendment_con_file)
                                amendment_edit.append(amendment_ref_num)
                            else:
                                # print('both change')
                                print('a8')
                                amendmentelsenone=Amendment.objects.get(id=values)
                                amendmentelsenone.amendment_upload_contract=amendment_con_file
                                amendmentelsenone.amendment_upload_pricetable=amendment_price_file
                                amendmentelsenone.amendment_type=amendment_type
                                amendmentelsenone.wcc=amedment_wcc
                                amendmentelsenone.amendment_reference_number=amendment_ref_num
                                amendmentelsenone.amendment_executed_date=amendmnet_convert_date
                                amendmentelsenone.amendment_currency_id=amendment_currency
                                amendmentelsenone.amendment_amount=amendment_amount
                                amendmentelsenone.updated_by=request.user
                                amendmentelsenone.save()
                                # Amendment.objects.filter(id=values).update(amendment_upload_contract=amendment_con_file,amendment_upload_pricetable=amendment_price_file,amendment_type=amendment_type,amendment_reference_number=amendment_ref_num,amendment_executed_date=amendmnet_convert_date,amendment_currency_id=amendment_currency,amendment_amount=amendment_amount)
                                current_amendemnt_ids.append(values)
                                fs = FileSystemStorage()
                                file = fs.save(amendment_con_file.name, amendment_con_file)
                                price_fs = FileSystemStorage()
                                file = price_fs.save(amendment_price_file.name, amendment_price_file)
                                amendment_edit.append(amendment_ref_num)
                            print('old',service+'con_file'+str(i)+str(index),amendment_type,amendment_con_file,amendment_price_file)
                        else:
                            amendment_type=request.POST.get(service+'extrafile'+str(i)+str(index))
                            amendment_ref_num=request.POST.get(service+'ref_num'+str(i)+str(index))
                            amedment_date=request.POST.get(service+'dateformat'+str(i)+str(index))
                            amendment_currency=request.POST.get(service+'currency'+str(i)+str(index))
                            amendment_amount=request.POST.get(service+'amount'+str(i)+str(index))
                            amendment_con_file=request.FILES.get(service+'con_file'+str(i)+str(index))
                            amendment_price_file=request.FILES.get(service+'price_file'+str(i)+str(index))
                            amendment_wcc=request.FILES.get(service+'wcc'+str(i)+str(index))
                            if (amedment_date == ''):
                                amendmnet_convert_date = None
                            elif (company.dateformat == 'dd-M-yy'):
                                amendmnet_convert_date=datetime.strptime(amedment_date,"%d-%b-%Y").date()
                                # print('1',amedment_date)
                            elif (company.dateformat == 'dd-mm-yy'):
                                amendmnet_convert_date=datetime.strptime(amedment_date,"%d-%m-%Y").date()
                                # print('2',amedment_date)
                            elif (company.dateformat == 'dd/mm/yy'):
                                amendmnet_convert_date=datetime.strptime(amedment_date,"%d/%m/%Y").date()
                                # print('3',amedment_date)
                            elif (company.dateformat == 'mm-dd-yy'):
                                amendmnet_convert_date=datetime.strptime(amedment_date,"%m-%d-%Y").date()
                                # print('4',amedment_date)
                            elif (company.dateformat == 'mm/dd/yy'):
                                amendmnet_convert_date=datetime.strptime(amedment_date,"%m/%d/%Y").date()
                                # print('5',amedment_date)
                            elif (company.dateformat == 'yy-mm-dd'):
                                amendmnet_convert_date=datetime.strptime(amedment_date,"%Y-%m-%d").date()
                                # print('6',amedment_date)
                            elif (company.dateformat == 'yy/mm/dd'):
                                amendmnet_convert_date=datetime.strptime(amedment_date,"%Y/%m/%d").date()
                                # print('7',amedment_date)
                            else:
                                amendmnet_convert_date=None
                                amendmnet_convert_date=datetime.strptime(amedment_date,"%d-%b-%Y").date()
                            print(f'pk {pk}, conttracthdnid[i] {conttracthdnid[i]}, amendment_con_file {amendment_con_file},amendment_price_file {amendment_price_file}, amendment_type {amendment_type}, amendment_ref_num {amendment_ref_num}, amendmnet_convert_date {amendmnet_convert_date}, amendment_currency {amendment_currency}')
                            create_amendment=Amendment.objects.create(company=request.company,contractvendor_id=pk,service_id=conttracthdnid[i],amendment_upload_contract=amendment_con_file,amendment_upload_pricetable=amendment_price_file,amendment_type=amendment_type,amendment_reference_number=amendment_ref_num,amendment_executed_date=amendmnet_convert_date,amendment_currency_id=amendment_currency,amendment_amount=amendment_amount,wcc=amendment_wcc)
                            current_amendemnt_ids.append(create_amendment.id)
                            # print('new',service+'con_file'+str(i)+str(index),amendment_type,amendment_con_file,amendment_price_file)
                            new_amendments.append(amendment_ref_num)
                    deleteAmendment = Amendment.objects.filter(company=request.company,contractvendor_id=pk,service_id=conttracthdnid[i]).exclude(id__in=current_amendemnt_ids).update(status=0)
                else:
                    print('New',conttracthdnid[i],i)
                    contractfile=request.FILES.get(service+'contract_file'+str(i))
                    pricefile=request.FILES.get(service+'contract_price_file'+str(i))
                    contract_file_name = request.FILES.get(service+'contract_file'+str(i))
                    price_table_file_name = request.FILES.get(service+'contract_price_file'+str(i))
                    if (projects[i] != '' or discipline[i] != '' or disciplinetype[i] != '' or  nameservice[i] != '' or referencenumber[i] != '' or execute_date[i] != '' or currency[i] != '' or amount[i] != '' or contractfile != None or pricefile != None):
                        if (execute_date[i] == ''):
                            convert_date = None
                        elif (company.dateformat == 'dd-M-yy'):
                            convert_date=datetime.strptime(execute_date[i],"%d-%b-%Y").date()
                            # print('1',execute_date[i])
                        elif (company.dateformat == 'dd-mm-yy'):
                            convert_date=datetime.strptime(execute_date[i],"%d-%m-%Y").date()
                            # print('2',execute_date[i])
                        elif (company.dateformat == 'dd/mm/yy'):
                            convert_date=datetime.strptime(execute_date[i],"%d/%m/%Y").date()
                            # print('3',execute_date[i])
                        elif (company.dateformat == 'mm-dd-yy'):
                            convert_date=datetime.strptime(execute_date[i],"%m-%d-%Y").date()
                            # print('4',execute_date[i])
                        elif (company.dateformat == 'mm/dd/yy'):
                            convert_date=datetime.strptime(execute_date[i],"%m/%d/%Y").date()
                            # print('5',execute_date[i])
                        elif (company.dateformat == 'yy-mm-dd'):
                            convert_date=datetime.strptime(execute_date[i],"%Y-%m-%d").date()
                            # print('6',execute_date[i])
                        elif (company.dateformat == 'yy/mm/dd'):
                            convert_date=datetime.strptime(execute_date[i],"%Y/%m/%d").date()
                            # print('7',execute_date[i])
                        else:
                            convert_date= None
                            # convert_date=datetime.strptime(execute_date[i],"%d-%b-%Y").date()
                        create_newcontract=ContractMaster.objects.create(types_service=service,company_id=request.company.id,contractvendor_id=pk,upload_contract=contractfile,contract_file_name=contract_file_name.name,upload_pricetable=pricefile,price_table_file_name=price_table_file_name,projects_id=projects[i],projectdiscipline_id=discipline[i],projectdisciplinetype_id=disciplinetype[i],name_service=nameservice[i],reference_number=referencenumber[i],executed_date=convert_date,currency_id=currency[i],amount=amount[i],created_by=request.user,wcc=wcc[i])
                        currentcontractids.append(create_newcontract.id)
                        new_contracts.append(referencenumber[i])
                        # print(i,service,contractfile,pricefile)
                        get_amendment_ids=request.POST.getlist(str(service)+str(i))
                        print('uncheck new values',get_amendment_ids)
                        for index,values in enumerate(get_amendment_ids):
                            print('index',index)
                            amendment_type=request.POST.get(service+'extrafile'+str(i)+str(index))
                            amendment_ref_num=request.POST.get(service+'ref_num'+str(i)+str(index))
                            amedment_date=request.POST.get(service+'dateformat'+str(i)+str(index))
                            amendment_currency=request.POST.get(service+'currency'+str(i)+str(index))
                            amendment_wcc=request.POST.get(service+'wcc'+str(i)+str(index))
                            amendment_amount=request.POST.get(service+'amount'+str(i)+str(index))
                            amendment_con_file=request.FILES.get(service+'con_file'+str(i)+str(index))
                            amendment_price_file=request.FILES.get(service+'price_file'+str(i)+str(index))
                            if (amendment_type != '' or amendment_ref_num != '' or amedment_date != '' or amendment_currency != '' or amendment_amount != '' or amendment_con_file != None or amendment_price_file != None):
                                if (amedment_date == ''):
                                    amendmnet_convert_date=None
                                elif (company.dateformat == 'dd-M-yy'):
                                    amendmnet_convert_date=datetime.strptime(amedment_date,"%d-%b-%Y").date()
                                    # print('1',amedment_date)
                                elif (company.dateformat == 'dd-mm-yy'):
                                    amendmnet_convert_date=datetime.strptime(amedment_date,"%d-%m-%Y").date()
                                    # print('2',amedment_date)
                                elif (company.dateformat == 'dd/mm/yy'):
                                    amendmnet_convert_date=datetime.strptime(amedment_date,"%d/%m/%Y").date()
                                    # print('3',amedment_date)
                                elif (company.dateformat == 'mm-dd-yy'):
                                    amendmnet_convert_date=datetime.strptime(amedment_date,"%m-%d-%Y").date()
                                    # print('4',amedment_date)
                                elif (company.dateformat == 'mm/dd/yy'):
                                    amendmnet_convert_date=datetime.strptime(amedment_date,"%m/%d/%Y").date()
                                    # print('5',amedment_date)
                                elif (company.dateformat == 'yy-mm-dd'):
                                    amendmnet_convert_date=datetime.strptime(amedment_date,"%Y-%m-%d").date()
                                    # print('6',amedment_date)
                                elif (company.dateformat == 'yy/mm/dd'):
                                    amendmnet_convert_date=datetime.strptime(amedment_date,"%Y/%m/%d").date()
                                    # print('7',amedment_date)
                                else:
                                    amendmnet_convert_date=None
                                    # amendmnet_convert_date=datetime.strptime(amedment_date,"%d-%b-%Y").date()
                                create_amendment=Amendment.objects.create(contractvendor_id=vendorcontract.id,company_id=request.company.id,service_id =create_newcontract.id,amendment_type=amendment_type,amendment_reference_number=amendment_ref_num,amendment_executed_date=amendmnet_convert_date,amendment_currency_id =amendment_currency,amendment_amount=amendment_amount,amendment_upload_contract=amendment_con_file,amendment_upload_pricetable=amendment_price_file,wcc=amendment_wcc)
                                new_amendments.append(amendment_ref_num)
                i+=1
            # # print(currentcontractids)
        ContractMaster.objects.filter(company_id=request.company.id,contractvendor_id=pk).exclude(id__in=currentcontractids).update(status=0)
        #Send notification when data changed
        sender = User.objects.get(id=request.user.id)
        scheme=request.scheme 
        gethost=request.get_host()
        url=f"{scheme}://{gethost}/projects/contractlist"
        get_userby_vin = User.objects.filter(cin_number=vendorcontract.vin)
        # if editdetails not empty send message
        # submit_value=request.POST.get('submit_type')
        # if (submit_value == '1'): 
        # if len(editdetails) > 0:
        #     contract_ref=','.join(map(str,editdetails))
        #     notify.send(sender, recipient=get_userby_vin,data=url, verb='Contract Details Updated', description=f'Contract {contract_ref} has been updated by {sender.name} {sender.lastname if sender.lastname != None else ""}')

        # if len(amendment_edit) > 0:
        #     print('chdcking',amendment_edit)
        #     amendment_ref=','.join(map(str,amendment_edit))
        #     notify.send(sender, recipient=get_userby_vin,data=url, verb='Amendment/Addendum Details Updated', description=f'Amendment/Addendum {amendment_ref} has been updated by {sender.name} {sender.lastname if sender.lastname != None else ""}')

        if len(new_contracts) > 0:
            new_contract_ref=','.join(map(str,new_contracts))
            notify.send(sender, recipient=get_userby_vin,data=url, verb='New Contract Added', description=f'New Contract {new_contract_ref} has been added by {sender.name} {sender.lastname if sender.lastname != None else ""}')

        if len(new_amendments) > 0:
            new_amendment_ref=','.join(map(str,new_amendments))
            notify.send(sender, recipient=get_userby_vin,data=url, verb='New Amendment/Addendum Created', description=f'New Amendment/Addendum {new_amendment_ref} has been created by {sender.name} {sender.lastname if sender.lastname != None else ""}')
        # amendment/addendum deleted send notification to vendor
        post_amendment=request.POST.getlist('serviceamendment')+request.POST.getlist('supplyamendment')+request.POST.getlist('serviceandsupplyamendment')
        amendment_delete = []
        [amendment_delete.append(i) for i in orginal_amendment if i not in post_amendment]
        if amendment_delete:
            new_amendment_reference=', '.join(str(e) for e in amendment_delete)
            notify.send(sender, recipient=get_userby_vin,data=url, verb='Amendment/Addendum Deleted', description=f'Amendment/Addendum {new_amendment_reference} has been deleted by {sender.name} {sender.lastname if sender.lastname != None else ""}')   
        # contract is deleted send notification to vendor
        post_contract=request.POST.getlist('servicerefnumber')+request.POST.getlist('supplyrefnumber')+request.POST.getlist('serviceandsupplyrefnumber')
        print({
            'servicerefnumber':request.POST.getlist('servicerefnumber'),
            'supplyrefnumber':request.POST.getlist('supplyrefnumber'),
            'serviceandsupplyrefnumber':request.POST.getlist('serviceandsupplyrefnumber')
        })
        contract_delete = []
        print(f'original contract {original_contract}')
        [contract_delete.append(i) for i in original_contract if i not in post_contract]
        if contract_delete:
            new_contract_reference=', '.join(str(e) for e in contract_delete)
            notify.send(sender, recipient=get_userby_vin,data=url, verb='Contract Deleted', description=f'Contract {new_contract_reference} has been deleted by {sender.name} {sender.lastname if sender.lastname != None else ""}')    
        # reference_number_changed.connect(reference_number_changed, sender=ContractMaster, weak=False, dispatch_uid='reference_number_changed', request=request)       
        return redirect('projects:contractmasterlist')
    projects=Projectcreation.objects.filter(company=request.company,status=0)
    data={'company':company,
          'companycurrency':companycurrency,
          'vendorcontract':vendorcontract,
          'projectdiscipline':discipline,
          'disciplinetype':disciplinetype,
          'types':types,
          'projects':projects}
    return render(request,'editcontractmaster.html',data)

def contractmasterdelete(request):
    contract_master_id=request.GET.get('id',None)
    if (contract_master_id != None):
        print('contract_master_id',contract_master_id)
        data={'data':'success'}
        # get_vinnumber = list(ContractMasterVendor.objects.filter(id=contract_master_id,company=request.company.id).values_list('vin',flat=True))
        # print('get_vinnumber',get_vinnumber)
        # get_userby_vin = User.objects.filter(cin_number=get_vinnumber[0])
        # print('get_userby_vin',get_userby_vin)
        # scheme=request.scheme
        # gethost=request.get_host()  
        # url=f"{scheme}://{gethost}/projects/contractlist"
        # sender = User.objects.get(id=request.user.id)
        # name_service = list(ContractMaster.objects.filter(contractvendor_id=contract_master_id,company=request.company.id).values_list('name_service',flat=True))
        # notify.send(sender, recipient=get_userby_vin,data=url, verb='Contract Deleted', description=f'{name_service[0]} Contract Deleted')
        print('contract_master_id',contract_master_id)
        delete_contract = ContractMaster.objects.filter(contractvendor_id=contract_master_id,company=request.company.id).update(status=0)
        print(f'delete_contract {delete_contract}')
        return JsonResponse(data)
    #     ContractMaster.objects.filter(contractvendor_id=contract_master_id).update(status=0)
    # return JsonResponse(data)


def vendorapproverlist(request):
    if (request.user.is_primary == 1):
        print('pri')
        vendordetails=VendorRegistraion.objects.filter(company=request.company,status=3,approver_status=1).order_by('-id')
        vendor_count=vendordetails.count()


    elif (request.user.is_secondary == 1):
        print('sec')
        vendordetails=VendorRegistraion.objects.filter(company=request.company,status=3,approver_status=2).order_by('-id')
        vendor_count=vendordetails.count()
    data={
        'vendors':vendordetails,
        'vendor_count':vendor_count,
        }
    return render(request,'vendorapproverlist.html',data)


def vendordetailpage(request):
    # if (request.user.roles_id == 4 and request.user.userfirsttimelogin == 1):
    #         vendorid=ContractMasterVendor.objects.filter(vin=request.user.cin_number).first()
    #         if(request.user.is_primary == 0 and request.user.is_secondary == 0) :
    #             return redirect('projects:editinsidevendorbasicinfo',pk=vendorid.id)
    request.session['mainmenu']='vendor'
    request.session['submenu']='vendordetailpage'
    get_vendor=User.objects.getUser(request.user.id)       
    vendor=ContractMasterVendor.objects.getvendor_byvin(get_vendor.cin_number,request.company)
    data={'vendor':vendor}
    return render(request,'vendordetailpage.html',data)

def getCompanyUserlist(request,sender,url,verb,description):
    clinet_admin_id=User.objects.filter(company_id=request.company.id).first()
    company_users=User.objects.filter(company_id=request.company.id,roles_id=3,status=1).values_list('id',flat=True)
    check_rights_list=list(UserRights.objects.filter(user_id__in=company_users,module_id=4).filter((Q(create=1) & Q(view=1) & Q(edit=1) & Q(delete=1)) | ( Q(view=1) & Q(edit=1) & Q(delete=1)) | (Q(create=1) & Q(view=1)) |(Q(view=1) & Q(edit=1) ) |(Q(create=1) & Q(view=1) & Q(edit=1) |(Q(create=1) & Q(view=1)  & Q(delete=1)))).values_list('user_id',flat=True))
    # check_rights_list.append(clinet_admin_id.id)
    for user_id in check_rights_list:
        recipient=User.objects.get(id=user_id) 
        send_url=url
        notify.send(sender, recipient=recipient,data=send_url, verb=verb, description=description)
    return True
    
def contractlist(request):
    markas_read_status(request.get_full_path())
    request.session['mainmenu']='vendor'
    request.session['submenu']='contracts'
    get_vin=request.user.cin_number
    getvendordetails=ContractMasterVendor.objects.getvendor_byvin(get_vin,request.company)
    contractlist=ContractMasterVendor.objects.filter(company_id=request.company.id,vin=get_vin,status=1).order_by('-id').values_list('id',flat=True)
    vendormasterlist=[]
    for vendorid in contractlist:
        if (ContractMaster.objects.filter(contractvendor_id=vendorid,status=1).exists()):
            getvendormasterdetails=ContractMasterVendor.objects.getvendor_byid(vendorid,request.company.id)
            vendormasterlist.append({'id':getvendormasterdetails.id,'vin':getvendormasterdetails.vin,'vendor_name':getvendormasterdetails.vendor_name})
    if (request.POST or request.FILES):
        query=request.POST.getlist('ref_num')
        newfilerep=request.POST.getlist('repfiles')
        print('newfilerep',newfilerep)
        # Bharath
        
        querychecklist=request.POST.getlist('querycheck')
        utc_timezone = pytz.utc
        utc_time = datetime.now(utc_timezone)
        time_stamp = utc_time.strftime('%Y-%m-%d %H:%M:%S')
        for con_id,val in zip(querychecklist,query):
            get_type=request.POST.get('contract_type'+str(val))
            reason=request.POST.get('reason'+str(val))
            cfile=request.FILES.get('cfile'+str(val))
            rep_type=request.POST.get('type'+str(val))
            newfilerep=request.POST.get('repfiles'+str(val))
            if (get_type == 'Original'):
                create_checkedvendor=CheckVendorContract.objects.create(refnum=val,reason=reason,c_file=cfile,contract_type=get_type,company_id=request.company.id,user_id=request.user.id,vendor_id=getvendordetails.id,contract_id=con_id,contractfile_id=newfilerep)
                # ContractMaster.objects.filter(id=con_id).update(save_type=1)
            else:
                create_checkedvendor=CheckVendorContract.objects.create(refnum=val,reason=reason,c_file=cfile,contract_type=get_type,company_id=request.company.id,user_id=request.user.id,vendor_id=getvendordetails.id,amendment_id=con_id,contractfile_id=newfilerep)
                # Amendment.objects.filter(id=con_id).update(save_type=1)
            CheckVendorContractHistory.objects.create(query_id=create_checkedvendor.id,deniedreason=reason,date=time_stamp,user=request.user,file=create_checkedvendor.c_file,vendor_id=getvendordetails.id,query_status=int(rep_type))
            sender = User.objects.getUser(request.user.id)
            scheme=request.scheme
            gethost=request.get_host()  
            url=f"{scheme}://{gethost}/projects/queryhistory/{create_checkedvendor.id}"
            description=f'Query received from {getvendordetails.vendor_name}'
            create_edit_delete_notfy(request,sender,url,'Query Received',description)
        return redirect('projects:vendorcontracchecklist')
    data={'contractlist':vendormasterlist,'count':len(vendormasterlist),'getvendordetails':getvendordetails}
    return render(request,'contractlist.html',data)

def Editinsidevendorbasicinfo(request,pk):
    markas_read_status(request.get_full_path())
    companyid=request.company.id
    countries=Basecountries.objects.all()
    countrylist=[]
    company=Settings.objects.getcompany_settings(request.user.company_id)
    getcurrencylist=ast.literal_eval(company.currency)
    companycurrency=Basecountries.objects.filter(id__in=getcurrencylist)
    for countri in countries:
        if States.objects.filter(country_id=countri.id).exists():
            countrylist.append({'id':countri.id,'name':countri.name})
    maincountry=Countries.objects.all()
    vendorform=ContractMasterVendor.objects.getvendor_byid(pk,companyid)
    states=States.objects.filter(country_id=vendorform.country_of_operation_id)
    companytax=[{'id':'Not Applicable','Tax_Name':'Not Applicable'}]
    allcompanytax=CompanyTaxDetails.objects.filter(company_id=companyid,status=1)
    addcompanytax=list(CompanyTaxDetails.objects.filter(company_id=companyid,status=1).values('id','Tax_Name'))
    companytax.extend(addcompanytax)
    vendortaxdetails=VendorTaxDetails.objects.filter(company_id=companyid,vendor_id=vendorform.id,status=1)
    vendorcompanytaxdetails=VendorCompanyTaxDetails.objects.getvendortaxdetails(companyid,vendorform.id)
    getinclusivetax=vendorcompanytaxdetails.filter(Tax_Type="Inclusive")
    getexclusivetax=vendorcompanytaxdetails.filter(Tax_Type="Exclusive")
    companytaxlist=companytax
    if (vendorcompanytaxdetails.count() > 0):
        vtax=list(vendorcompanytaxdetails.values_list('tax',flat=True))
        filtered_list = list(filter(None, vtax))
        allcompanytax=allcompanytax.exclude(id__in=filtered_list)
        companytaxlist=[{'id':'Not Applicable','Tax_Name':'Not Applicable'}]
        companytaxlist.extend(list(allcompanytax.values('id','Tax_Name')))
    else:
        companytaxlist=companytaxlist

    taxcount=vendortaxdetails.count()
    vendorbankdetails=BankDetails.objects.getbankdetails_byvendor_company(companyid,vendorform.id)
    bankcount=vendorbankdetails.count()
    taxtypelist=['Inclusive','Exclusive']
    if (request.POST):
        print(f'{request.POST}')
        vendor_name=request.POST['vendorname']
        country_of_operation=request.POST['vendor_country_list']
        country_of_incorporation=request.POST['vendor_country_list_incorporation']
        remove_comma=request.POST['vendor_state_list']
        state=remove_comma.replace(",","")
        Address=request.POST['address']
        website=request.POST['website']
        Registernumber=request.POST['Registernumber']
        ContractMasterVendor.objects.update_contractmastervendor(request.company.id,pk,vendor_name,country_of_operation,country_of_incorporation,state,Address,website,Registernumber)
        get_taxids=request.POST.getlist('taxnamehdid')
        get_taxname=request.POST.getlist('taxname')
        get_taxnum=request.POST.getlist('taxnumber')
        bankdetailids=request.POST.getlist('bankhdid')
        currencybank=request.POST.getlist('currency_bank')
        bankname=request.POST.getlist('bankname')
        accno=request.POST.getlist('accno')
        bankdetails=request.POST.getlist('bankdetails')
        submit_type=request.POST.get('submit_type')
        i=0
        currenttaxids=[]
        while i < len(get_taxids):
            if (get_taxids[i]):
                VendorTaxDetails.objects.filter(id=get_taxids[i]).update(company_id=companyid,vendor=pk,Tax_Name=get_taxname[i],Tax_Number=get_taxnum[i])
                currenttaxids.append(get_taxids[i])
                # print("old",get_taxids[i],get_taxname[i],get_taxnum[i])
            else:
                if (get_taxname[i] != '' or get_taxnum[i] != ''):
                    create_vendortax=VendorTaxDetails.objects.create(company_id=companyid,vendor_id=pk,Tax_Name=get_taxname[i],Tax_Number=get_taxnum[i])
                    currenttaxids.append(create_vendortax.id)
            i+=1
        VendorTaxDetails.objects.filter(company_id=companyid,vendor_id=pk).exclude(id__in=currenttaxids).update(status=0)

        j=0
        currentbankids=[]
        while j < len(bankdetailids):
            if (bankdetailids[j]):
                BankDetails.objects.update_bankdetails(bankdetailids[j],companyid,pk,bankdetails[j],currencybank[j],bankname[j],accno[j])
                currentbankids.append(bankdetailids[j])
            else:
                if (currencybank[j] != '' or bankdetails[j] != '' or bankname[j] != '' or accno[j] != ''):
                    create_vendortax=BankDetails.objects.create_bankdetails(companyid,pk,bankdetails[j],currencybank[j],bankname[j],accno[j])
                    currentbankids.append(create_vendortax.id)
            j+=1
        BankDetails.objects.updateinactive_bankdetails(companyid,pk,currentbankids)
          
        
        if (submit_type == '0'):
            User.objects.filter(id=request.user.id).update(userfirsttimelogin=2)
            ContractMasterVendor.objects.update_approvestatus(pk,1)
            sender = User.objects.getUser(request.user.id)
            recipient = User.objects.filter(company=request.company.id,roles_id=2).first()
            vendorname = ContractMasterVendor.objects.filter(id=pk).values_list('vendor_name',flat=True).first()
            scheme=request.scheme
            gethost=request.get_host()  
            url=f"{scheme}://{gethost}/projects/vendormasterlist"
            notify.send(sender, recipient=recipient,data=url, verb='Vendor Registration Details Updated', description=f'{vendorname} has updated their registration details')
            return redirect('projects:vendordetailpage')
        else: 
            return redirect('projects:editinsidevendorregistration',pk=pk)
    # get full url from request object
    scheme=request.scheme
    data={"vendorform":vendorform,
        "countries":countrylist,
        "maincountry":maincountry,
        "states":states,
        "vendortaxdetails":vendortaxdetails,'taxcount':taxcount,
        'vendorbankdetails':vendorbankdetails,'bankcount':bankcount,
        'taxtypelist':taxtypelist,'getinclusivetax':getinclusivetax,'getexclusivetax':getexclusivetax,'companytaxlist':companytaxlist,'companycurrency':companycurrency}
    return render(request,'editinsidevendorform1.html',data)


def editinsidevendorregistration(request,pk):
    companyid=request.company.id
    print('aa',companyid)
    userid=request.user.id
    titles=['Mr','Mrs','Ms']
    phonecodelist=[]
    phonecode = Basecountries.objects.all().values('iso3','id','phonecode').order_by('phonecode').distinct()
    for codes in phonecode:
        id=codes.get('id')
        num=codes.get('phonecode')
        iso=codes.get('iso3')
        x = re.findall("\+",num)
        if x:
            phonecodelist.append({'id':id,'phonecode':str(num),'iso':iso})
        else:
            phonecodelist.append({'id':str(id),'phonecode':'+'+str(num),'iso':iso})
    newlist=sorted(phonecodelist, key = lambda i:  (i['phonecode']))
    vendor=ContractMasterVendor.objects.get(company_id=companyid,id=pk)
    if (request.POST):
        print(request.POST)
        contact_primary_title=request.POST.get('Tile')
        contact_primary_first_name=request.POST.get('firstname')
        contact_primary_middle_name=request.POST.get('middlename')
        contact_primary_last_name=request.POST.get('lastname')
        
        contact_secondary_title=request.POST.get('Title')
        contact_secondary_first_name=request.POST.get('alternatefirstname')
        contact_secondary_middle_name=request.POST.get('alternatemiddlename')
        contact_secondary_last_name=request.POST.get('alternatelastname')

        secondary_designation=request.POST.get('alternateDesignation')
        primary_designation=request.POST.get('Designation')
        contact_primary_phone_code=request.POST.get('phonecountrycode')
        contact_primary_area_code=request.POST.get('phoneareacode')
        contact_primary_phone=request.POST.get('phone')
        contact_primary_mobile_code=request.POST.get('mobilecoutrycode')
        contact_primary_mobile=request.POST.get('mobile')
        contact_primary_phone_number_extenstion=request.POST.get('phonenumber_extenstion')

        contact_secondary_phone_code=request.POST.get('phonecountryalternatecode')
        contact_secondary_phone=request.POST.get('alternatephone')
        contact_secondary_mobile_code=request.POST.get('alternate_country_code')
        contact_secondary_mobile=request.POST.get('alternatemobile')
        contact_secondary_area_code=request.POST.get('alternatephoneareacode')
        contact_secondary_phone_number_extenstion=request.POST.get('alternatephonenumber_extenstion')


        official_primary_email=request.POST.get('Email')
        official_secondary_email=request.POST.get('alternateEmail')
        submit_type=request.POST.get('submit_type')
        create_registrtion_vendor=ContractMasterVendor.objects.filter(id=pk).update(contact_primary_first_name=contact_primary_first_name,contact_primary_middle_name=contact_primary_middle_name,contact_primary_last_name=contact_primary_last_name,contact_secondary_first_name=contact_secondary_first_name,contact_secondary_middle_name= contact_secondary_middle_name,contact_secondary_last_name=contact_secondary_last_name,secondary_designation=secondary_designation,primary_designation=primary_designation,contact_primary_phone_code=contact_primary_phone_code,contact_primary_phone=contact_primary_phone,contact_primary_mobile_code=contact_primary_mobile_code,contact_primary_mobile=contact_primary_mobile, contact_secondary_phone_code=contact_secondary_phone_code,contact_secondary_phone=contact_secondary_phone,contact_secondary_mobile_code=contact_secondary_mobile_code,contact_secondary_mobile=contact_secondary_mobile,official_primary_email=official_primary_email,official_secondary_email=official_secondary_email,contact_primary_area_code=contact_primary_area_code,contact_secondary_area_code=contact_secondary_area_code,contact_primary_title=contact_primary_title,contact_secondary_title=contact_secondary_title,
        contact_primary_phone_number_extenstion=contact_primary_phone_number_extenstion,
        contact_secondary_phone_number_extenstion=contact_secondary_phone_number_extenstion)
        vendorids=[]
        primarystatus=[]
        secondarystatus=[]
        password = make_password('Hello@123')
        # print('create_registrtion_vendor',create_registrtion_vendor)
        if (official_primary_email != ''):
            if (official_primary_email == request.user.email):
                print('p')
                primary_user=User.objects.filter(company=request.company,email=official_primary_email,cin_number=vendor.vin,status=1).first()
                if (primary_user == None):
                    createuser=User.objects.create(email=official_primary_email,name=contact_primary_first_name,middlename=contact_primary_middle_name,lastname=contact_primary_last_name,phone = contact_primary_phone,mobile=contact_primary_mobile,phone_countrycode=contact_primary_phone_code,mobile_countrycode=contact_primary_mobile_code,areacode=contact_primary_area_code, phone_number_extenstion=contact_primary_phone_number_extenstion,company_id=request.company.id,roles_id=4,Title=contact_primary_title,password=password,designation_role=primary_designation,cin_number=vendor.vin,is_primary=1,status=1)
                    vendorids.append(createuser.id)
                    primarystatus.append(createuser.id)
                else:
                    # here
                    primary_user.is_primary=1
                    primary_user.Title=contact_primary_title
                    primary_user.name=contact_primary_first_name
                    primary_user.middlename=contact_primary_middle_name
                    primary_user.lastname=contact_primary_last_name
                    primary_user.designation_role=primary_designation
                    primary_user.phone=contact_primary_phone
                    primary_user.mobile=contact_primary_mobile
                    primary_user.mobile_countrycode=contact_primary_mobile_code
                    primary_user.phone_countrycode=contact_primary_phone_code
                    primary_user.areacode=contact_primary_area_code
                    primary_user.phone_number_extenstion=contact_primary_phone_number_extenstion
                    primary_user.save()
                    vendorids.append(primary_user.id)
                    primarystatus.append(primary_user.id)
                secondary_user=User.objects.filter(company=request.company,email=official_secondary_email,cin_number=vendor.vin,status=1).first()
                print('a',secondary_user)
                if (secondary_user == None):
                    createuser=User.objects.create(email=official_secondary_email,name=contact_secondary_first_name,middlename=contact_secondary_middle_name,lastname=contact_secondary_last_name,phone_countrycode=contact_secondary_phone_code,phone=contact_secondary_phone,mobile_countrycode=contact_secondary_mobile_code,mobile=contact_secondary_mobile,areacode=contact_secondary_area_code,
                    phone_number_extenstion=contact_secondary_phone_number_extenstion,company_id=request.company.id,roles_id=4,Title=contact_secondary_title,password=password,designation_role=secondary_designation,cin_number=vendor.vin,is_secondary=1,status=1)
                    vendorids.append(createuser.id)
                    secondarystatus.append(createuser.id)
                else:
                    #here
                    secondary_user.is_secondary=1
                    secondary_user.Title=contact_secondary_title
                    secondary_user.name=contact_secondary_first_name
                    secondary_user.middlename=contact_secondary_middle_name
                    secondary_user.lastname=contact_secondary_last_name
                    secondary_user.designation_role=secondary_designation
                    secondary_user.phone_countrycode=contact_secondary_phone_code
                    secondary_user.phone=contact_secondary_phone
                    secondary_user.mobile_countrycode=contact_secondary_mobile_code
                    secondary_user.mobile=contact_secondary_mobile
                    secondary_user.areacode=contact_secondary_area_code
                    secondary_user.phone_number_extenstion=contact_secondary_phone_number_extenstion
                    secondary_user.save()
                    vendorids.append(secondary_user.id)
                    secondarystatus.append(secondary_user.id)
            else:
                print('no')
                primary_user=User.objects.filter(company=request.company,email=official_primary_email,cin_number=vendor.vin,status=1).first()
                if (primary_user == None):
                    createuser=User.objects.create(email=official_primary_email,name=contact_primary_first_name,middlename=contact_primary_middle_name,lastname=contact_primary_last_name,phone = contact_primary_phone,mobile=contact_primary_mobile,phone_countrycode=contact_primary_phone_code,mobile_countrycode=contact_primary_mobile_code,areacode=contact_primary_area_code,phone_number_extenstion=contact_primary_phone_number_extenstion,company_id=request.company.id,roles_id=4,Title=contact_primary_title,password=password,designation_role='Vendor',cin_number=vendor.vin,is_primary=1,status=1)
                    vendorids.append(createuser.id)
                    primarystatus.append(createuser.id)
                else:
                    #here
                    primary_user.is_primary=1
                    primary_user.Title=contact_primary_title
                    primary_user.name=contact_primary_first_name
                    primary_user.middlename=contact_primary_middle_name
                    primary_user.lastname=contact_primary_last_name
                    primary_user.designation_role=primary_designation
                    primary_user.phone=contact_primary_phone
                    primary_user.mobile=contact_primary_mobile
                    primary_user.mobile_countrycode=contact_primary_mobile_code
                    primary_user.phone_countrycode=contact_primary_phone_code
                    primary_user.areacode=contact_primary_area_code
                    primary_user.phone_number_extenstion=contact_primary_phone_number_extenstion
                    primary_user.save()
                    vendorids.append(primary_user.id)
                    primarystatus.append(primary_user.id)
                secondary_user=User.objects.filter(company=request.company,email=official_secondary_email,cin_number=vendor.vin,status=1).first()
                if (secondary_user == None):
                    print('no.s')
                    createuser=User.objects.create(email=official_secondary_email,name=contact_secondary_first_name,middlename=contact_secondary_middle_name,lastname=contact_secondary_last_name,phone_countrycode=contact_secondary_phone_code,phone=contact_secondary_phone,mobile_countrycode=contact_secondary_mobile_code,mobile=contact_secondary_mobile,areacode=contact_secondary_area_code, phone_number_extenstion=contact_secondary_phone_number_extenstion,company_id=request.company.id,roles_id=4,Title=contact_secondary_title,password=password,designation_role='Vendor',cin_number=vendor.vin,is_secondary=1,status=1)
                    vendorids.append(createuser.id)
                    secondarystatus.append(createuser.id)
                else:
                    #here
                    secondary_user.is_secondary=1
                    secondary_user.Title=contact_secondary_title
                    secondary_user.name=contact_secondary_first_name
                    secondary_user.middlename=contact_secondary_middle_name
                    secondary_user.lastname=contact_secondary_last_name
                    secondary_user.designation_role=secondary_designation
                    secondary_user.phone_countrycode=contact_secondary_phone_code
                    secondary_user.phone=contact_secondary_phone
                    secondary_user.mobile_countrycode=contact_secondary_mobile_code
                    secondary_user.mobile=contact_secondary_mobile
                    secondary_user.areacode=contact_secondary_area_code
                    secondary_user.phone_number_extenstion=contact_secondary_phone_number_extenstion
                    secondary_user.save()
                    vendorids.append(secondary_user.id)
                    secondarystatus.append(secondary_user.id)
        elif (official_secondary_email != ''):
            if(official_secondary_email == request.user.email):
                print('s')
                primary_user=User.objects.filter(company=request.company,email=official_primary_email,cin_number=vendor.vin,status=1).first()
                if (primary_user == None):
                    createuser=User.objects.create(email=official_primary_email,name=contact_primary_first_name,middlename=contact_primary_middle_name,lastname=contact_primary_last_name,phone = contact_primary_phone,mobile=contact_primary_mobile,phone_countrycode=contact_primary_phone_code,mobile_countrycode=contact_primary_mobile_code,areacode=contact_primary_area_code,phone_number_extenstion=contact_primary_phone_number_extenstion,company_id=request.company.id,roles_id=4,Title=contact_primary_title,password=password,designation_role='Vendor',cin_number=vendor.vin,is_primary=1,status=1)
                    vendorids.append(createuser.id)
                    primarystatus.append(createuser.id)
                else:
                    #here
                    primary_user.is_primary=1
                    primary_user.Title=contact_primary_title
                    primary_user.name=contact_primary_first_name
                    primary_user.middlename=contact_primary_middle_name
                    primary_user.lastname=contact_primary_last_name
                    primary_user.designation_role=primary_designation
                    primary_user.phone=contact_primary_phone
                    primary_user.mobile=contact_primary_mobile
                    primary_user.mobile_countrycode=contact_primary_mobile_code
                    primary_user.phone_countrycode=contact_primary_phone_code
                    primary_user.areacode=contact_primary_area_code
                    primary_user.phone_number_extenstion=contact_primary_phone_number_extenstion
                    primary_user.save()
                    vendorids.append(primary_user.id)
                    primarystatus.append(primary_user.id)

                secondary_user=User.objects.filter(company=request.company,email=official_secondary_email,cin_number=vendor.vin,status=1).first()
                if (secondary_user == None):
                    createuser=User.objects.create(email=official_secondary_email,name=contact_secondary_first_name,middlename=contact_secondary_middle_name,lastname=contact_secondary_last_name,phone_countrycode=contact_secondary_phone_code,phone=contact_secondary_phone,mobile_countrycode=contact_secondary_mobile_code,mobile=contact_secondary_mobile,areacode=contact_secondary_area_code, phone_number_extenstion=contact_secondary_phone_number_extenstion,company_id=request.company.id,roles_id=4,Title=contact_secondary_title,password=password,designation_role='Vendor',cin_number=vendor.vin,is_secondary=1,status=1)
                    vendorids.append(createuser.id)
                    secondarystatus.append(createuser.id)
                else:
                    # here
                    secondary_user.is_secondary=1
                    secondary_user.Title=contact_secondary_title
                    secondary_user.name=contact_secondary_first_name
                    secondary_user.middlename=contact_secondary_middle_name
                    secondary_user.lastname=contact_secondary_last_name
                    secondary_user.designation_role=secondary_designation
                    secondary_user.phone_countrycode=contact_secondary_phone_code
                    secondary_user.phone=contact_secondary_phone
                    secondary_user.mobile_countrycode=contact_secondary_mobile_code
                    secondary_user.mobile=contact_secondary_mobile
                    secondary_user.areacode=contact_secondary_area_code
                    secondary_user.phone_number_extenstion=contact_secondary_phone_number_extenstion
                    secondary_user.save()
                    vendorids.append(secondary_user.id)
                    secondarystatus.append(secondary_user.id)
            else:

                secondary_user=User.objects.filter(company=request.company,email=official_secondary_email,cin_number=vendor.vin,status=1).first()
                if (secondary_user == None):
                    createuser=User.objects.create(email=official_secondary_email,name=contact_secondary_first_name,middlename=contact_secondary_middle_name,lastname=contact_secondary_last_name,phone_countrycode=contact_secondary_phone_code,phone=contact_secondary_phone,mobile_countrycode=contact_secondary_mobile_code,mobile=contact_secondary_mobile,areacode=contact_secondary_area_code, phone_number_extenstion=contact_secondary_phone_number_extenstion,company_id=request.company.id,roles_id=4,Title=contact_secondary_title,password=password,designation_role='Vendor',cin_number=vendor.vin,is_secondary=1,status=1)
                    vendorids.append(createuser.id)
                    secondarystatus.append(createuser.id)
                else:
                    # here
                    secondary_user.is_secondary=1
                    secondary_user.Title=contact_secondary_title
                    secondary_user.name=contact_secondary_first_name
                    secondary_user.middlename=contact_secondary_middle_name
                    secondary_user.lastname=contact_secondary_last_name
                    secondary_user.designation_role=secondary_designation
                    secondary_user.phone_countrycode=contact_secondary_phone_code
                    secondary_user.phone=contact_secondary_phone
                    secondary_user.mobile_countrycode=contact_secondary_mobile_code
                    secondary_user.mobile=contact_secondary_mobile
                    secondary_user.areacode=contact_secondary_area_code
                    secondary_user.phone_number_extenstion=contact_secondary_phone_number_extenstion
                    secondary_user.save()
                    vendorids.append(secondary_user.id)
                    secondarystatus.append(secondary_user.id)
        vendorids.append(request.user.id)
        User.objects.filter(company=request.company,cin_number=vendor.vin).exclude(id__in=vendorids).update(status=0)
        User.objects.filter(company=request.company,cin_number=vendor.vin).exclude(id__in=primarystatus).update(is_primary=0)
        User.objects.filter(company=request.company,cin_number=vendor.vin).exclude(id__in=secondarystatus).update(is_secondary=0)
        primaryname=contact_primary_first_name+' '+contact_primary_middle_name+' '+contact_primary_last_name
        secondaryname=contact_secondary_first_name+' '+contact_secondary_middle_name+' '+contact_secondary_last_name
        current_url = f'{request.scheme}://{request.get_host()}'
        user_role = request.user.roles_id
        if (official_primary_email != vendor.official_primary_email and official_primary_email != ''):
            vendorsemail(pk, contact_primary_title, primaryname, current_url, request.user.id, companyid, official_primary_email,vendorstatus='Primary')
        if (official_secondary_email != vendor.official_secondary_email and official_secondary_email != ''):
            vendorsemail(pk, contact_secondary_title, secondaryname, current_url, request.user.id, companyid, official_secondary_email, vendorstatus='Secondary')

        # for guidelines 
        # detail_to_pdf(request,companyid)
        if (official_primary_email != '' and official_secondary_email != ''):
            User.objects.filter(id=request.user.id).update(userfirsttimelogin=2)
            ContractMasterVendor.objects.filter(id=pk).update(approver_status=1)

        data={'vendorid':"success"}
        sender = User.objects.get(id=request.user.id)
        recipient = User.objects.filter(company=request.company.id,roles_id=2).first()
        scheme=request.scheme
        gethost=request.get_host()  
        getvendordetails=ContractMasterVendor.objects.filter(vin=request.user.cin_number,company=request.company,status=1).first()
        url=f"{scheme}://{gethost}/projects/vendormasterlist"
        notify.send(sender, recipient=recipient,data=url, verb='Vendor has updated their registration details ', description=f'{getvendordetails.vendor_name} has updated their registration details')
        if (submit_type == '0'):
            User.objects.filter(id=request.user.id).update(userfirsttimelogin=2)
            ContractMasterVendor.objects.filter(id=pk).update(approver_status=1)

            return redirect('projects:vendordetailpage')
        else:
            return redirect('projects:vendordetailpage')

    data={"vendor":vendor,
          "titles":titles,
          "phone":newlist,
          'companyid':companyid}
    return render(request,'editinsidevendorform2.html',data)





# def vendorcheckcontract(request):
#     values=request.session['getcheckedvalue']
#     if (request.POST or request.FILES):
#         for val in values:   
#             reason=request.POST.get('reason'+str(val))
#             cfile=request.FILES.get('cfile'+str(val),'')
#             pfile=request.FILES.get('pfile'+str(val),'')
#             create_checkedvendor=CheckVendorContract.objects.create(refnum=val,reason=reason,c_file=cfile,p_file=pfile,company_id=request.company.id,user_id=request.user.id)
#     data={'checkedvalues':values}
#     return render(request,'vendorcheckcontract.html',data)

def getcontracts_byvendor(request):
    vendor_id=request.POST.get('vendor_id')
    allcontract=getvendorcontracts(vendor_id)
    return JsonResponse({'contracts':allcontract})

def vendorcontracchecklist(request):
    request.session['mainmenu']='vendor'
    request.session['submenu']='checkcontracts'
    querylist=CheckVendorContract.objects.filter(user_id=request.user.id,company=request.company).order_by('-id')
    data={'querylists':querylist,'len_querylist':len(querylist)}
    return render(request,'vendorcontracchecklist.html',data)

def newcontractlist(request):
    request.session['mainmenu']='vendor'
    request.session['submenu']='newcontractslist'
    vendorid=ContractMasterVendor.objects.filter(vin=request.user.cin_number).first()
    listnewcontract=AddNewContracts.objects.filter(user_id=request.user.id,company=request.company).order_by('-id')
    totalcount=listnewcontract.count()
    data={'listnewcontract':listnewcontract,'totalcount':totalcount,'vendorid':vendorid}
    return render(request,'addnewcontractslist.html',data)

def addnewcontract(request):
    utc_timezone = pytz.utc
    utc_time = datetime.now(utc_timezone)
    getvin=request.user.cin_number
    getvendordetails=ContractMasterVendor.objects.getvendor_byvin(getvin,request.company)
    get_original_contract=list(ContractMaster.objects.filter(company_id=request.company.id,contractvendor_id=getvendordetails.id,status=1).values_list('reference_number',flat=True))
    time_stamp = utc_time.strftime('%Y-%m-%d %H:%M:%S')
    if (request.POST or request.FILES):
        try:
            totalrowcount=request.POST.getlist('totalrowcount')
            for count in totalrowcount:
                contracttype=request.POST.get('contracttype'+str(count))
                nameservice=request.POST.get('nameservice'+str(count))
                refnum=request.POST.get('refnum'+str(count))
                amendmentrefnum=request.POST.get('amendmentrefnum'+str(count))
                contaractfile=request.FILES.get('contractfile'+str(count))
                original_filename=contaractfile.name
                if (contracttype == 'Original'):
                    create_newcontrcat=AddNewContracts.objects.create(company=request.company,user_id=request.user.id,contract_type=contracttype,name_service=nameservice,contract_refnum=refnum,c_file=contaractfile,original_file_name=original_filename,vendor_id=getvendordetails.id)
                    AddNewContractsHistory.objects.create(query_id=create_newcontrcat.id,deniedreason="",date=time_stamp,user=request.user,file=contaractfile,original_file_name=original_filename,vendor_id=getvendordetails.id)
                else:

                    create_newcontrcat=AddNewContracts.objects.create(company=request.company,user_id=request.user.id,contract_type=contracttype,name_service=nameservice,contract_refnum=refnum,c_file=contaractfile,original_file_name=original_filename,amendment_refnum=amendmentrefnum,vendor_id=getvendordetails.id)
                    AddNewContractsHistory.objects.create(query_id=create_newcontrcat.id,deniedreason="",date=time_stamp,user=request.user,file=contaractfile,original_file_name=original_filename,vendor_id=getvendordetails.id)
            sender = User.objects.get(id=request.user.id)
            scheme=request.scheme
            gethost=request.get_host()  
            url=f"{scheme}://{gethost}/projects/newcontracthistory/{create_newcontrcat.id}"
            description=f'Missed Contract received from {getvendordetails.vendor_name} '
            getCompanyUserlist(request,sender,url,'Missed Contract Received',description)
            return redirect('projects:newcontractlist')
        except Exception as e:
            return redirect('projects:newcontractlist')
    data={'original_contracts':get_original_contract}
    return render(request,'addnewcontract.html',data)


def vendorquerylist(request):
    request.session['mainmenu'] = 'vendors'
    request.session['submenu']='vendorquery'
    querylist=CheckVendorContract.objects.filter(company=request.company).order_by('-id')
    data={'querylists':querylist,'len_querylist':len(querylist)}
    return render(request,'vendorquerylist.html',data)

def queryhistory(request,pk):
    markas_read_status(request.get_full_path())
    query=CheckVendorContract.objects.get(id=pk)
    print(f'CheckVendorContract {query.refnum}')
    contract_reference_number = None
    print("hsdhgfasds",query.contract_type)
    if (query.contract_type != 'Original'):
        # Check if the contract field is not empty
        contract_reference_number = query.amendment.service.reference_number
        print("reference number",contract_reference_number)
    getvendordetails=ContractMasterVendor.objects.filter(vin=query.user.cin_number,company=request.company,status=1).first()
    querylist=CheckVendorContractHistory.objects.filter(query_id=pk)
    if (query.contract_type == 'Original'):
        get_contract_id = query.contract_id
        contractType = 'original'
    else:
        get_contract_id = query.amendment_id
        contractType = 'amendment'
    print('querylist',query.contract_type)
    utc_timezone = pytz.utc
    utc_time = datetime.now(utc_timezone)
    # print(utc_time.strftime('%Y-%m-%d %H:%M:%S'))
    time_stamp = utc_time.strftime('%Y-%m-%d %H:%M:%S')
    if (request.POST or request.FILES):
        print(request.POST)
        newfilerep=request.POST.get('replace_filename')
        ref_num=request.POST.get('ref_num')
        contractId=request.POST.get('contract_id')
        ammendmentId=request.POST.get('amendment_id')
        contract_type=request.POST.get('contract_types')
        deniedreason=request.POST.get('message')
        file_name=request.FILES.get('file')
        get_query_type=request.POST.get('filetype')
        original_filename = None  # Initialize original_filename variable
        file = None  

        if (file_name != None):
            CheckVendorContract.objects.filter(id=pk).update(c_file=file_name , contractfile=newfilerep,file_name=file_name.name)
            fs = FileSystemStorage()
            file = fs.save(file_name.name, file_name)
            original_filename=file_name.name

        # if (contract_type == 'Original'):
        #         create_checkedvendor=CheckVendorContract.objects.create(refnum=ref_num,reason=deniedreason,c_file=file_name,contract_type=contract_type,company_id=request.company.id,user_id=request.user.id,vendor_id=getvendordetails.id,contract_id=contractId,contractfile_id=newfilerep)
        # else:
        #         create_checkedvendor=CheckVendorContract.objects.create(refnum=ref_num,reason=deniedreason,c_file=file_name,contract_type=contract_type,company_id=request.company.id,user_id=request.user.id,vendor_id=getvendordetails.id,amendment_id=ammendmentId,contractfile_id=newfilerep)
        create_data=CheckVendorContractHistory.objects.create(query_id=pk,deniedreason=deniedreason,user=request.user,file=file,original_file_name=original_filename,date=time_stamp,vendor_id=getvendordetails.id,query_status=int(get_query_type))
        # Send notification to client adminstrator when query received from vendor
        scheme=request.scheme
        gethost=request.get_host()
        if request.user.roles_id == 4:
            sender = User.objects.filter(id=request.user.id).first()
            url=f"{scheme}://{gethost}/projects/queryhistory/{pk}"
            description=f'Reply received from {getvendordetails.vendor_name}'
            create_edit_delete_notfy(request,sender,url,'Reply Received for Query',description)
        else:
            print('To Vendor')
            sender = User.objects.filter(id=request.user.id).first()
            recipients = User.objects.filter(cin_number=getvendordetails.vin,contactpersonstatus=1).first()
            url=f"{scheme}://{gethost}/projects/queryhistory/{pk}"
            notify.send(sender, recipient=recipients,data=url, verb='Reply Received for Query', description=f'Reply received from {sender.name} {sender.lastname if sender.lastname != None else ""}')
        return redirect('projects:queryhistory',pk=pk)
    data={'querylists':querylist,'query':query,'vendor':getvendordetails,'pk':pk,'contractquery':get_contract_id,'contract_type':query.contract_type,'vendor_id':getvendordetails.id,'contractType':contractType, 'contract_reference_number': contract_reference_number}
    print(f'data {data}')
    return render(request,'queryhistory.html',data)

def create_edit_delete_notfy(request,sender,url,verb,description):
    clinet_admin_id=User.objects.filter(company_id=request.company.id).first()
    company_users=User.objects.filter(company_id=request.company.id,roles_id=3,status=1).values_list('id',flat=True)
    check_rights_list=list(UserRights.objects.filter(user_id__in=company_users,module_id=4).filter(Q(create=1) | Q(edit=1) | Q(delete=1)).values_list('user_id',flat=True))
    # | Q(view=1)
    # check_rights_list.append(clinet_admin_id.id)
    for user_id in check_rights_list:
        recipient=User.objects.get(id=user_id) 
        send_url=url
        notify.send(sender, recipient=recipient,data=send_url, verb=verb, description=description)
    return True

def querydenied(request,pk):
    query=CheckVendorContract.objects.filter(id=pk).update(status=0)
    vendor_name = CheckVendorContract.objects.filter(id=pk).values_list('vendor__vendor_name',flat=True).first()
    # Send notification to client adminstrator when query closed.
    sender = User.objects.get(id=request.user.id)
    # recipient = User.objects.filter(company=request.company.id,roles_id=2).first()
    scheme=request.scheme
    gethost=request.get_host()  
    url=f"{scheme}://{gethost}/projects/queryhistory/{pk}"
    description=f'Query closed by {vendor_name}'
    create_edit_delete_notfy(request,sender,url,'Query Closed',description)
    return redirect('projects:vendorcontracchecklist')

        # return redirect('projects:vendorquerylist')

def vendormissedcontractlist(request):
    request.session['mainmenu'] = 'vendors'
    request.session['submenu']='vendormissedcontract'
    listnewcontract=AddNewContracts.objects.filter(company=request.company).order_by('-id')
    totalcount=listnewcontract.count()
    data={'listnewcontract':listnewcontract,'totalcount':totalcount}
    return render(request,'vendormissedcontractlist.html',data)


def newcontracthistory(request,pk):
    markas_read_status(request.get_full_path())
    query=AddNewContracts.objects.get(id=pk)
    getvendordetails=ContractMasterVendor.objects.filter(vin=query.user.cin_number,company=request.company,status=1).first()
    querylist=AddNewContractsHistory.objects.filter(query_id=pk)
    utc_timezone = pytz.utc
    utc_time = datetime.now(utc_timezone)
    time_stamp = utc_time.strftime('%Y-%m-%d %H:%M:%S')
    if (request.POST):
        print(request.POST)
        deniedreason=request.POST.get('message')
        uploaded_file=request.FILES.get('file')
        if (uploaded_file!= None):
            AddNewContracts.objects.filter(id=pk).update(c_file=uploaded_file)
            fs = FileSystemStorage()
            file = fs.save(uploaded_file.name, uploaded_file )
            original_filename = uploaded_file.name
        else:
            file = None
            original_filename = None
        AddNewContractsHistory.objects.create(query_id=pk,deniedreason=deniedreason,user=request.user,file=file,original_file_name=original_filename,  date=time_stamp,vendor_id=getvendordetails.id)
        scheme=request.scheme
        gethost=request.get_host()
        if request.user.roles_id == 4:
            sender = User.objects.filter(id=request.user.id).first()
            url=f"{scheme}://{gethost}/projects/newcontracthistory/{pk}"
            description=f'Reply received from {getvendordetails.vendor_name}'
            create_edit_delete_notfy(request,sender,url,'Reply Received for Missed Contract',description)
        else:
            print('To Vendor')
            sender = User.objects.filter(id=request.user.id).first()
            recipients = User.objects.filter(cin_number=getvendordetails.vin,contactpersonstatus=1).first()
            url=f"{scheme}://{gethost}/projects/newcontracthistory/{pk}"
            notify.send(sender, recipient=recipients,data=url, verb='Reply Received for Missed Contract', description=f'Reply received from {sender.name} {sender.lastname if sender.lastname != None else ""}')
        return redirect('projects:newcontracthistory',pk=pk)
    data={'querylists':querylist,'query':query,'vendor':getvendordetails,'pk':pk}
    return render(request,'newcontracthistory.html',data)

def newcontractdenied(request,pk):
    listnewcontract=AddNewContracts.objects.filter(id=pk).update(status=0)
    # Send notification to client adminstrator when missed contract closed.
    vendor_name = AddNewContracts.objects.filter(id=pk).values_list('vendor__vendor_name',flat=True).first()
    sender = User.objects.get(id=request.user.id)
    # recipient = User.objects.filter(company=request.company.id,roles_id=2).first()
    scheme=request.scheme
    gethost=request.get_host()  
    url=f"{scheme}://{gethost}/projects/newcontracthistory/{pk}"
    # notify.send(sender, recipient=recipient,data=url, verb='Missed Contract Closed', description=f'Missed Contract from {vendor_name} is closed')
    description=f'Missed Contract closed by {vendor_name}'
    create_edit_delete_notfy(request,sender,url,'Missed Contract Closed',description)
    return redirect('projects:newcontractlist')



import base64

def vendormasterpdf(request):
    template = 'vendormasterpdf.html'
    list_vendor_master=ContractMasterVendor.objects.filter(company_id=request.company.id,status=1).order_by('id')
    company= Companies.objects.filter(id=request.company.id).first()
    if company.image:
        imageurl = company.image.url
    else:
        imageurl = ''
    companyName = company
    context = {'vendormaster' :list_vendor_master,'imageurl':imageurl,'company':companyName}
    getcompanyname=request.company.company_name
    filename=getcompanyname+' Vendor master report.pdf'
    response=render_to_pdf_response(request,template,context,filename=filename) 
    response['Content-Disposition'] = 'inline;filename='+filename+''
    return response


def invoiceGuidLinePdf(request,pk):
    companyid=pk
    template = 'invoiceguidelines.html'
    company = Companies.objects.get(id=companyid)
    company_setting=Settings.objects.filter(company=company).first()
    get_invoice_processdate=company_setting.vendor_registartion_time
    company= Companies.objects.filter(id=company.id).first()
    if company.image:
        imageurl=company.image.url
    else:
        imageurl = None
    context = {'company' :company,'processdate':get_invoice_processdate,'imageurl':imageurl}
    return render_to_pdf_response(request,template,context)


def contractmasterreport(request):
    template = 'contractmasterlistpdf.html'
    vendormasterlist=ContractMaster.objects.filter(company_id=request.company.id,status=1).order_by('executed_date')
    company = Companies.objects.filter(id=request.company.id).first()
    if company.image:
        imageurl = company.image.url
    else:
        imageurl = None
    context = {'company' :company,'imageurl':imageurl,'vendormasterlist':vendormasterlist,'value':0}
    getcompanyname=request.company.company_name
    filename=getcompanyname+' Contract master report.pdf'
    response=render_to_pdf_response(request,template,context,filename=filename) 
    response['Content-Disposition'] = 'inline;filename='+filename+''
    return response
# template = 'specificcontractpdf.html'
# contractmasterlist=ContractMaster.objects.filter(contractvendor_id=pk,status=1).order_by('executed_date')
# totalcount=contractmasterlist.count()
# company = Companies.objects.filter(id=request.company.id).first()
# if company.image:
#     imageurl = company.image.url
#     a = company.image.path
#     img = Image.new("RGB", (800, 1280), (255, 255, 255))
#     pngimage = img.save("image.png", "PNG")
#     with open('image.png', "rb") as image_file:
#         encoded_string = base64.b64encode(image_file.read())
#         decodeimage = encoded_string.decode("utf-8")
#         add_base = 'data:image/png;base64,' + decodeimage
#     os.remove('image.png')
# else:
#     imageurl = None
#     add_base = None
# getcompanyname=request.company.company_name
# filename=getcompanyname+' Specific Vendor Contract master report.pdf'
# context = {'company': company, 'imageurl': imageurl, 'vendormasterlist': contractmasterlist, 'countlist': totalcount}
# response=render_to_pdf_response(request,template,context,filename=filename)
# response['Content-Disposition'] = 'inline;filename='+filename+''
# return response
def specificreportgenerate(request):
    allcontractmaster= ContractMaster.objects.filter(company_id=request.company.id,status=1)
    allnameservice= list(allcontractmaster.distinct().order_by('name_service').values_list('name_service',flat=True))
    contractcurrencieslist= list(allcontractmaster.values_list('currency',flat=True).distinct())
    get_currency_object=Basecountries.objects.filter(id__in=contractcurrencieslist)
    allprojects=list(allcontractmaster.values_list('projects',flat=True).distinct())
    projectcreations=Projectcreation.objects.filter(id__in=allprojects)
    data={'allnameservice':allnameservice,'currencies':get_currency_object,'allprojects':projectcreations}
    companys=''
    if (request.GET):
        print(request.GET)
        typeservice=request.GET.get('typeservice','')
        nameservice=request.GET.get('nameservice','')
        project=request.GET.get('projectid','')
        disciplinename=request.GET.get('discipline','')
        if (typeservice != '' and nameservice == '' and project == ''):
            print(1)
            allnameservice= list(allcontractmaster.filter(types_service=typeservice,status=1).distinct().values_list('name_service',flat=True))
            projectlist= list(allcontractmaster.filter(types_service=typeservice,status=1).distinct().values_list('projects',flat=True))
            remove_duplicate=tuple(projectlist)
            convert_list=list(remove_duplicate)
            get_projects=Projectcreation.objects.filter(id__in=convert_list)
            newlist=[]
            for project in get_projects:
                newlist.append({'id':project.id,'projectname':project.projectname.name})
            return JsonResponse({'data':allnameservice,'projects':newlist})
        elif (typeservice != '' and nameservice == '' and project != ''):
            disciplines= list(allcontractmaster.filter(types_service=typeservice,projects_id=project,status=1).distinct().values_list('projectdiscipline',flat=True))
            remove_duplicate=tuple(disciplines)
            convert_list=list(remove_duplicate)
            projectdisciplinelist=ProjectDevelopmentDiscipline.objects.filter(id__in=convert_list)
            disciplinelist=[]
            for discipline in projectdisciplinelist:
                if (discipline.project_discipline == '1'):
                    name='Green Field Development'
                elif (discipline.project_discipline == '2'):
                    name='Brown Field Development'
                else:
                    name='Others'
                fieldname=discipline.development_type.cluster.environment.field.field.field_name.title()
                disciplinelist.append({'id':discipline.id,'projectdiscipline':name+'-'+fieldname})
            # print(disciplinelist)
            print('only service with project')  
            return JsonResponse({'data':disciplinelist})


        
        if (typeservice != '' and nameservice != '' and project == '' and disciplinename == ''):
            print(2,'a')
            projectlist= list(allcontractmaster.filter(types_service=typeservice,name_service=nameservice,status=1).distinct().values_list('projects',flat=True))
            remove_duplicate=tuple(projectlist)
            convert_list=list(remove_duplicate)
            get_projects=Projectcreation.objects.filter(id__in=convert_list)
            newlist=[]
            for project in get_projects:
                newlist.append({'id':project.id,'projectname':project.projectname.name})
            print('2',newlist)
            return JsonResponse({'data':newlist})
        elif (typeservice == '' and nameservice != '' and project == '' and disciplinename == ''):
            print(3,'a')
            projectlist= list(allcontractmaster.filter(name_service=nameservice,status=1).distinct().values_list('projects',flat=True))
            remove_duplicate=tuple(projectlist)
            convert_list=list(remove_duplicate)
            get_projects=Projectcreation.objects.filter(id__in=convert_list)
            newlist=[]
            for project in get_projects:
                newlist.append({'id':project.id,'projectname':project.projectname.name})
            print('2 elif',newlist)
            return JsonResponse({'data':newlist})
        
        if (typeservice != '' and nameservice != '' and project != '' and disciplinename == ''):
            print(5)
            print(project)
            disciplines= list(allcontractmaster.filter(types_service=typeservice,name_service=nameservice,projects_id=project,status=1).distinct().values_list('projectdiscipline',flat=True))
            remove_duplicate=tuple(disciplines)
            convert_list=list(remove_duplicate)
            projectdisciplinelist=ProjectDevelopmentDiscipline.objects.filter(id__in=convert_list)
            disciplinelist=[]
            for discipline in projectdisciplinelist:
                if (discipline.project_discipline == '1'):
                    name='Green Field Development'
                elif (discipline.project_discipline == '2'):
                    name='Brown Field Development'
                else:
                    name='Others'
                fieldname=discipline.development_type.cluster.environment.field.field.field_name.title()
                disciplinelist.append({'id':discipline.id,'projectdiscipline':name+'-'+fieldname})
            # print(disciplinelist)
            # print('a')
            return JsonResponse({'data':disciplinelist})    
        elif (typeservice != '' and nameservice == '' and project != '' and disciplinename == ''):
            # print('project',project)
            disciplines= list(allcontractmaster.filter(types_service=typeservice,projects_id=project,status=1).distinct().values_list('projectdiscipline',flat=True))
            remove_duplicate=tuple(disciplines)
            convert_list=list(remove_duplicate)
            projectdisciplinelist=ProjectDevelopmentDiscipline.objects.filter(id__in=convert_list)
            disciplinelist=[]
            for discipline in projectdisciplinelist:
                if (discipline.project_discipline == '1'):
                    name='Green Field Development'
                elif (discipline.project_discipline == '2'):
                    name='Brown Field Development'
                else:
                    name='Others'
                fieldname=discipline.development_type.cluster.environment.field.field.field_name.title()
                disciplinelist.append({'id':discipline.id,'projectdiscipline':name+'-'+fieldname})
            # print('b')
            return JsonResponse({'data':disciplinelist,'s':'q'})    


        if (typeservice != ''  and nameservice != '' and project != '' and disciplinename != ''):
            # print(5)
            # print('dsffs')
            project_discipline_type= list(allcontractmaster.filter(types_service=typeservice,name_service=nameservice,projectdiscipline=disciplinename,status=1).distinct().values_list('projectdisciplinetype',flat=True))
            remove_duplicate=tuple(project_discipline_type)
            convert_list=list(remove_duplicate)
            disciplinetype=ProjectDevelopmentSubType.objects.filter(id__in=convert_list)
            disciplinetypelist=[]
            for disciplinetypes in disciplinetype:
                disciplinetypelist.append({'id':disciplinetypes.id,'disciplinetype':disciplinetypes.discipline_subtype.discipline_subtype})
            print('3 if',disciplinetypelist)
            return JsonResponse({'data':disciplinetypelist})

    elif (request.POST):
        print(request.POST)
        postcontracts= ContractMaster.objects.filter(company_id=request.company.id,status=1)
        services=request.POST.getlist('type_services')
        name_contract=request.POST.getlist('name_contract')
        project=request.POST.getlist('project')
        discipline_list=request.POST.getlist('discipline')
        disciplinesubtype=request.POST.getlist('disciplinesubtype','')
        currency=request.POST.getlist('currency')
        
        if (services[0] != ''):
            postcontracts=postcontracts.filter(types_service__in=services)
        # print('1',postcontracts)
        if (name_contract[0] != ''):
            postcontracts=postcontracts.filter(name_service__in=name_contract)
        # print('2',postcontracts)
        if (project[0] != ''):
            postcontracts=postcontracts.filter(projects__in=project)
        # print('2 project',postcontracts)    
        if (len(disciplinesubtype) > 0):
            if (discipline_list[0] != ''):
                postcontracts=postcontracts.filter(discipline__in=discipline_list)
            # print('3',postcontracts)
        if (len(disciplinesubtype) > 0):
            if (disciplinesubtype[0] != ''):
                postcontracts=postcontracts.filter(disciplinetype__in=disciplinesubtype)
            # print('4',postcontracts)
        if (currency[0] != ''):
            postcontracts=postcontracts.filter(currency__in=currency)
        # print('5',postcontracts)
        if (len(postcontracts) > 0):
            template = 'specificcontractpdf.html'
            companys= Companies.objects.filter(id=request.company.id).first()
            if companys.image:
                imageurl=companys.image.url
            else:
                imageurl=None
            getcompanyname=request.company.company_name
            filename=getcompanyname+' Specific Contract master report.pdf'
            context = {'company' :companys,'imageurl':imageurl,'vendormasterlist':postcontracts,'countlist':len(postcontracts)}
            response=render_to_pdf_response(request,template,context,filename=filename) 
            response['Content-Disposition'] = 'inline;filename='+filename+''
            return response
        else:
            messages.add_message(request, messages.ERROR, 'No Results Found')
            return redirect('projects:specificreportgenerate')
    return render(request,'specificreportgenerate.html',data) 



def generatespecificvendor(request,pk):
    template = 'specificcontractpdf.html'
    contractmasterlist=ContractMaster.objects.filter(contractvendor_id=pk,status=1).order_by('executed_date')
    totalcount=contractmasterlist.count()
    company = Companies.objects.filter(id=request.company.id).first()
    if company.image:
        imageurl = company.image.url
    else:
        imageurl = None
    getcompanyname=request.company.company_name
    filename=getcompanyname+' Specific Vendor Contract master report.pdf'
    context = {'company': company, 'imageurl': imageurl, 'vendormasterlist': contractmasterlist, 'countlist': totalcount}
    response=render_to_pdf_response(request,template,context,filename=filename)
    response['Content-Disposition'] = 'inline;filename='+filename+''
    return response

def taxlist(request):
    if (request.POST):
        context = {}
        print('post method----------')
        request.session['mainmenu'] = 'masterprojects'
        request.session['submenu'] = 'tax_list'
        alltaxes=CompanyTaxDetails.objects.filter(company=request.company,status=1).order_by('-id')
        data={}
        if (request.user.roles_id == 3):
            userrights=UserRights.objects.get(user_id=request.user.id,module_id=5)
            context['rights']=userrights
        search_query = request.POST.get('q',False)
        page = request.POST.get('page', 1)
        pageper_data = request.POST.get('pageperdata',10)
        print('pageper_data---',pageper_data)
        if search_query:
            tax_list = CompanyTaxDetails.objects.filter(company=request.user.company,status=True,Tax_Name__icontains=search_query).order_by('-id')
           
        else:
            tax_list = CompanyTaxDetails.objects.filter(company=request.user.company,status=True).order_by('-id')
        context['query'] = search_query
        paginator = Paginator(tax_list, pageper_data)
        context['alltaxes'] = paginator.page(page)
        context['pageper_data'] = pageper_data
        context['scheme']=request.scheme
        context['gethost']=request.get_host()
        # render to template string
       
        html = render_to_string('search_tax.html',context,request)
        print(context)
        return JsonResponse({'status':True,'html':html})
        
       
    else:
        print('Get ')
        request.session['mainmenu'] = 'masterprojects'
        request.session['submenu'] = 'tax_list'
        alltaxes=CompanyTaxDetails.objects.filter(company=request.company,status=1).order_by('-id')
        data={}
        if (request.user.roles_id == 3):
            userrights=UserRights.objects.get(user_id=request.user.id,module_id=5)
            data['rights']=userrights
        
        page = request.GET.get('page', 1)
        pageper_data = request.POST.get('pageperdata',10)
        paginator = Paginator(alltaxes, pageper_data)
        data['alltaxes'] = paginator.page(page)
        data['pageper_data'] = pageper_data
        data['scheme']=request.scheme
        data['gethost']=request.get_host()
        
        return render(request,'taxlist.html',data)

def createtax(request):
    if (request.POST):
        postvalues=request.POST
        taxname=postvalues.getlist('taxname')
        taxtype=postvalues.get('taxtype')

        for tax in taxname:
            if (CompanyTaxDetails.objects.filter(company=request.company,Tax_Name__iexact=tax,status=1).exists()):
                # messages.error(request,'Tax name already exists')
                pass
                # return redirect('projects:createtax')
            else:
                createtax=CompanyTaxDetails.objects.create(Tax_Name=tax,company=request.company,Tax_Type=taxtype)
            usercreate = request.user.roles_id
            create_user_log(request,createtax.Tax_Name,'Tax Master','Create','Tax Master Created',usercreate)    
        return redirect('projects:taxlist')
    return render(request,'createtax.html')


def checktaxname(request):
    taxname=request.GET.get('taxname',None)
    taxId = request.GET.get('pk',None)
    if (taxname != None):
        if (CompanyTaxDetails.objects.filter(company=request.company,Tax_Name__iexact=taxname,status=1).exclude(id=taxId).exists()):
            data={'data':'exists'}
        else:
            data={'data':'new'}
    else:
        data={'data':'removecls'}
    return JsonResponse(data)

def viewcompanytax(request,pk):
    getdata=CompanyTaxDetails.objects.get(id=pk,status=1)
    data={'getdata':getdata}
    return render(request,'viewcompanytax.html',data)

def editcompanytax(request,pk):
    getdata=CompanyTaxDetails.objects.get(id=pk,status=1)
    if (request.POST):
        CompanyTaxDetails.objects.filter(id=pk).update(Tax_Name=request.POST.get('taxname'))
        return redirect('projects:taxlist')
    data={'getdata':getdata,'pk':pk}
    return render(request,'editcompanytax.html',data)

def deletecompanytax(request,pk):
    CompanyTaxDetails.objects.filter(id=pk).update(status=0)
    return redirect('projects:taxlist')



def vendor_active_status(request):
    vendor_id=request.GET.get('vendor_id')
    currentactive_status=ContractMasterVendor.objects.get(id=vendor_id)
    if (currentactive_status.active_status == 1):
        ContractMasterVendor.objects.filter(id=vendor_id).update(active_status=2)
    else:
        ContractMasterVendor.objects.filter(id=vendor_id).update(active_status=1)
    data={'data':'success'}
    return JsonResponse(data)

def getoriginalcontract(request):
    contractid=request.GET.get('contractid')
    contractmaster=ContractMaster.objects.get(id=contractid)
    companyid=contractmaster.company.id
    vendorcompanytaxdetails=VendorCompanyTaxDetails.objects.filter(company_id=companyid,vendor_id=contractmaster.contractvendor.id,contract_id=contractmaster.id,status=1)#add contractid
    if (vendorcompanytaxdetails.count() > 0):
        getinclusivetax=vendorcompanytaxdetails.filter(Tax_Type="Inclusive")
        list_inclusive_tax=[]
        for inclusivetax in getinclusivetax:
            if (inclusivetax.tax_id == None):
                tax_id=None
                tax_name=None
            else:
                tax_id=inclusivetax.tax.id
                tax_name=inclusivetax.tax.Tax_Name
            list_inclusive_tax.append({'id':inclusivetax.id,'tax_type':inclusivetax.Tax_Type,'tax_id':tax_id,'tax_name':tax_name,'vendor_id':inclusivetax.vendor.id,'contract_id':inclusivetax.contract.id})
        getexclusivetax=vendorcompanytaxdetails.filter(Tax_Type="Exclusive")
        list_exclusive_tax=[]
        for exclusivetax in getexclusivetax:
            if (exclusivetax.tax_id == None):
                tax_id=None
                tax_name=None
            else:
                tax_id=exclusivetax.tax.id
                tax_name=exclusivetax.tax.Tax_Name
            list_exclusive_tax.append({'id':exclusivetax.id,'tax_type':exclusivetax.Tax_Type,'tax_id':tax_id,'tax_name':tax_name,'vendor_id':exclusivetax.vendor.id,'contract_id':exclusivetax.contract.id})
        vendorpaymentterms=VendorPaymentTerms.objects.filter(company_id=companyid,vendor_id=contractmaster.contractvendor.id,contract_id=contractid,status=1).values()
        invoicesplitlist=VendorInvoiceSplit.objects.filter(contract_id=contractid,status=1)
        print("invoicesplitlist",invoicesplitlist)
        if (invoicesplitlist.count() > 1):
            invoices=invoicesplitlist.values()
        else:
            invoices=invoicesplitlist.values()
        data={'getinclusivetax':list_inclusive_tax,'getexclusivetax':list_exclusive_tax,'invoice_count':invoicesplitlist.count(),'invoices':list(invoices),'vendorpaymentterms':list(vendorpaymentterms),'vendorpaymenttermscount':vendorpaymentterms.count()}
    else:
        data={'status':'no_data'}
    return JsonResponse(data)

def getalltaxes(request):
    tax_id=request.GET.get('tax_id')
    vendor_percentage=VendorCompanyTaxPercentage.objects.filter(vendortax_id=tax_id,status=1).values()
    data={'data':list(vendor_percentage)}
    return JsonResponse(data)


def checkcontractreferencenum(request):
    refnum=request.GET.get('refnum')
    if ContractMaster.objects.filter(company=request.company,reference_number__exact=refnum,status=1).exists() or Amendment.objects.filter(company=request.company,amendment_reference_number__exact=refnum,status=1).exists():
        data={'data':'exists'}
        print(data)
    else:
        data={'data':'success'}
    return JsonResponse(data)


def vendorcontractcheck(request):
    vendor_id=request.GET.get('vendor_id')
    service_type=request.GET.get('type')
    get_contracts_id=list(ContractMaster.objects.filter(contractvendor_id=vendor_id,types_service=service_type,status=1).values_list('id',flat=True))
    get_amendment_ids=list(Amendment.objects.filter(service_id__in=get_contracts_id,status=1).values_list('id',flat=True))
    get_contract_invoice=Invoice.objects.filter(contractid__in=get_contracts_id,contracttype='original',status=1).count()
    get_amendment_invoice=Invoice.objects.filter(contractid__in=get_amendment_ids,contracttype__startswith='a',status=1).count()
    if (get_contract_invoice > 0):
        data={'data':'exsists'}
    elif (get_amendment_invoice > 0):
        data={'data':'exsists'}
    else:
        data={'data':'no_data'}
    return JsonResponse(data)
def checkcontractinvoice(request):
    contract_type=request.GET.get('contract_type')
    contract_id=request.GET.get('contract_id')
    get_invoice=Invoice.objects.filter(contractid=contract_id,contracttype=contract_type,status=1).count()
    if (contract_type == 'original'):
        get_amendment_ids=list(Amendment.objects.filter(service_id=contract_id,status=1).values_list('id',flat=True))
        # print("get_amendment_ids",get_amendment_ids)
        get_amendment_invoice=Invoice.objects.filter(contractid__in=get_amendment_ids,contracttype__startswith='a',status=1).count()
        if (get_invoice > 0):
            data={'data':'exsists'}
        elif (get_amendment_invoice > 0):
            data={'data':'exsists','a':get_amendment_invoice}
        else:
            data={'data':'no_data','a':get_amendment_invoice}
    else:
        if (get_invoice > 0):
            data={'data':'exsists'}
        else:
            data={'data':'no_data'}
    return JsonResponse(data)
    
# #editmasterblock from ajax call when edit master block
# def checkeditmasterblock(request):
#     # get_company_id=request.GET.get("company_id")
#     get_block_id=request.GET.get("block_id")
#     projects=Projectcreation.objects.filter(company_id=request.company.id,status=0).values_list('id',flat=True)
#     print('projects',projects)
#     blocks_id=ProjectBlock.objects.filter(project_id__in=projects,block_id=get_block_id,status=1)
#     print('blocks_id',blocks_id)
#     if (blocks_id.count() > 0):
#         data={'data':'exists'}
#     else:
#         data={'data':'new'}
#     return JsonResponse(data)


#  # editmasterfield from Aajax call when Edit master field
# def checkeditmasterfield(request):
#     # get_company_id=request.GET.get("company_id")
#     get_field_id=request.GET.get("field_id")
#     projects=Projectcreation.objects.filter(company_id=request.company.id,status=0).values_list('id',flat=True)
#     print('projects',projects)
#     field_id=ProjectField.objects.filter(project_id__in=projects,field_id=get_field_id,status=1).count()
#     print('field_id',field_id)
#     if (field_id > 0):
#         data={'data':'exists'}
#     else:
#         data={'data':'new'}
#     return JsonResponse(data)

#  # editmasterfieldenvironment from Aajax call when Edit master field environment
# def checkeditmasterfieldenv(request):
#     get_field_environment_id=request.GET.get('field_environment_id')
#     projects=Projectcreation.objects.filter(company_id=request.company.id,status=0).values_list('id',flat=True)
#     print('projects',projects)
#     field_environment_id=ProjectEnvironment.objects.filter(project_id__in=projects,field_environment_id=get_field_environment_id,status=1).count()
#     print('field_environment_id',field_environment_id)
#     if (field_environment_id > 0):
#         data={'data':'exists'}
#     else:
#         data={'data':'new'}
#     return JsonResponse(data)

#  # editmastercluster from Aajax call when Edit mastercluster
# def checkeditmastercluster(request):
#     get_clustersubname_id=request.GET.get('clustersubname_id')
#     projects=Projectcreation.objects.filter(company_id=request.company.id,status=0).values_list('id',flat=True)
#     clustersubname_id=ProjectCluster.objects.filter(project_id__in=projects,clustersubname_id=get_clustersubname_id,status=1).count()
#     if (clustersubname_id > 0):
#         data={'data':'exists'}
#     else:
#         data={'data':'new'}
#     return JsonResponse(data)


#  # editmasterwell from Aajax call when Edit masterwell
# def checkeditmasterwell(request):
#     get_wellname_id=request.GET.get('wellname_id')
#     projects=Projectcreation.objects.filter(company_id=request.company.id,status=0).values_list('id',flat=True)
#     wellname_id=ProjectWellName.objects.filter(project_id__in=projects,wellname_id=get_wellname_id,status=1).count()
#     if (wellname_id > 0):
#         data={'data':'exists'}
#     else:
#         data={'data':'new'}
#     return JsonResponse(data)

def filechange(request):
    user_type= request.GET.get('user_type')
    button_type= request.GET.get('button_type')
    file = request.GET.get('file')
    print(f'file {file}')
    vendor_id = request.GET.get('vendor_id')
    contract_type = request.GET.get('contract_type')
    contractfile_id = request.GET.get('contractfile_id')
    contractType = request.GET.get('contractType')
    utc_timezone = pytz.utc
    utc_time = datetime.now(utc_timezone)
    time_stamp = utc_time.strftime('%Y-%m-%d %H:%M:%S')
    if (user_type == 'vendor'):
        CheckVendorContractHistory.objects.filter(id=file).update(query_status=int(button_type))
    else:
        get_type = request.GET.get('type')
        pk = request.GET.get('pk')
        sender=User.objects.get(id=request.user.id)
        if (get_type == 'file'):
            filename=CheckVendorContractHistory.objects.filter(id=file).values_list('file',flat=True)
            VendorContractPriceTable.objects.filter(id=contractfile_id).update(file_name = filename,original_file_name=filename[0])
            #     # contractmaster=ContractMaster.objects.filter(id=contract_id).update(upload_contract = filename)
            # else:
            #     VendorContractPriceTable.objects.filter(id=contractfile_id).update(file_name = filename,original_file_name=filename[0])
            #     # Amendment.objects.filter(id=contract_id).update(amendment_upload_contract = filename)
            CheckVendorContractHistory.objects.create(query_id=pk,deniedreason='Replacement file accepted',user=request.user,date=time_stamp,vendor_id=vendor_id,query_status=4)
            description =f"Reply received from {sender.name} {sender.lastname if sender.lastname != None else ''} - Replacement file accepted"
            messages.success(request, 'Contract File replaced successfully, please update Contract and Price table details') 
        else:
            CheckVendorContractHistory.objects.create(query_id=pk,deniedreason='Replacement file not accepted',user=request.user,date=time_stamp,vendor_id=vendor_id,query_status=4)
            description =f"Reply received from {sender.name} {sender.lastname if sender.lastname != None else ''} - Replacement file not accepted"

        CheckVendorContractHistory.objects.filter(id=file).update(query_status=2)
        vinnum = ContractMasterVendor.objects.filter(id=int(vendor_id)).values_list('vin',flat=True).first()
        recipient=User.objects.filter(cin_number=vinnum,contactpersonstatus=1).first()
        scheme=request.scheme
        gethost=request.get_host()  
        url=f"{scheme}://{gethost}/projects/queryhistory/{pk}"
        notify.send(sender,recipient=recipient,data=url ,verb='Reply Received for Query', description=description)
    data={'status':'success'}
    return JsonResponse(data)

class CheckAuthorizedUser(View):
    def post(self,request):
        user_id = request.user.id
        if UserRights.objects.filter(module_id=5,user_id=user_id,create=1).exists():
            return JsonResponse({'status':True})
        else:
            return JsonResponse({'status':False})

class editProjectUser(View): 
    def get(self,request,pk):
        selected_data=Projectcreation.objects.get(pk=pk)
        selected_data=selected_data.signatory_type
        datas = [{'id':'1','settings_type':'Default Signatories'},{'id':'2','settings_type':'Project Signatories'}]
        signatory=SignatoriesSettings.objects.filter(project=pk,status=True)
        default_users=SignatoriesSettings.objects.filter(status=True,company=request.company,signatory_type='1')
        project_user = ProjectUser.objects.filter(company=request.company,project_id=pk)
        all_users = User.objects.filter(company=request.company,status=1,roles_id=3).exclude(id=request.user.id,roles_id=2)
        user = User.objects.get_by_company(request.company,[2,3],1)
        get_settings = Settings.objects.get_company(request.company).values_list('currency',flat=True).first()
        currency = Basecountries.objects.get_by_id(literal_eval(get_settings))
        data={}
        if (request.user.roles_id == 3):
            userrights=UserRights.objects.get(user_id=request.user.id,module_id=16)
            data['rights']=userrights
        data.update({'users':project_user.filter(status=True).values_list('user_id',flat=True),'pk':pk,'all_users':all_users,'signatory':signatory,'signatory_count':signatory.count(),'currency':currency,'user':user,'datatypes':datas,'default_users':default_users,'selected_data':selected_data})
        return render(request,'editprojectuser.html',data)

    def post(self,request,pk):
        print('usersadata',request.POST)
        project_users = request.POST.getlist('project_user',False)
        company_client_admin=User.objects.filter(company=request.company,status=1,roles_id=2).first()
        print('project_users----',project_users)
        if project_users:
            print('project_users22----',project_users)
            project_users.append(str(company_client_admin.id))
            print('project_users33----',project_users)
            for i in project_users:
                print('i------',i)
                project_user, created = ProjectUser.objects.update_or_create(
                        project_id=pk,
                        user_id=i,
                        defaults={'company': request.company,'status':True},
                        )
                UserRights.objects.filter(user_id=i,module__module_name='Projects').update(view=1)
            revmove_status = ProjectUser.objects.filter(project_id=pk).exclude(user_id__in=project_users)
            if revmove_status.exists():
                for i in revmove_status:
                    if ProjectUser.objects.filter(user_id=i.user_id,status=True).count() == 1:
                        exists=UserRights.objects.filter(Q(create__isnull=True) |Q(edit__isnull=True),user_id=i.user_id,module__module_name='Projects')
                        if exists.exists():
                            exists.update(view=None)
                revmove_status.update(status=False)
        else:
            remove_all = ProjectUser.objects.filter(project_id=pk).exclude(user_id=request.user.id)
            if remove_all.exists():
                for i in remove_all:
                    if ProjectUser.objects.filter(user_id=i.user_id,status=True).count() == 1:
                            exists=UserRights.objects.filter(Q(create__isnull=True) |Q(edit__isnull=True),user_id=i.user_id,module__module_name='Projects')
                            if exists.exists():
                                exists.update(view=None)
                remove_all.update(status=False)
        return redirect('projects:projectlist')

class CheckUserhasProject(View):
    def post(self,request):
        user_id = request.POST['user_id']
        check_user_has_project = True if ProjectUser.objects.filter(company=request.company,user_id=user_id,status=True).exists() else False
        return JsonResponse({'status':check_user_has_project})

class MigrateUser(View):
    def get(self,request,pk):
        user_projects = ProjectUser.objects.filter(company=request.company,user_id=pk,status=True)
        all_users = User.objects.filter(company=request.company,status=1,roles_id__in=[2,3]).exclude(id=pk)
        migrate_user = User.objects.get(id=pk)
        default_users=SignatoriesSettings.objects.filter(status=True,company=request.company,signatory_type='1')
        data = {'project_user':user_projects,'all_users':all_users,'migrate_user':migrate_user,'default_users':default_users}
        return render(request,'migrateuser.html',data)

    def post(self,request,pk):
        get_rights=UserRights.objects.filter(user_id=pk,status=1)
        # check_with_user=list(ProjectUser.objects.filter(company=request.company,user_id=pk).values_list('project_id',flat=True))
        user_projects = ProjectUser.objects.filter(company=request.company,user_id=pk)
        for i in user_projects:
            new_user = request.POST.get(f'user_id{i.id}',False)
            if new_user:
                get_user=User.objects.filter(id=new_user).first()
                check_user= ProjectUser.objects.filter(project_id=i.project_id,status=True,user_id__in=list(User.objects.filter(id=new_user).values_list('id',flat=True))).exists()
                if not check_user:                
                    if get_user.roles_id != 2:
                        for right in get_rights:
                            get_rights_new=UserRights.objects.filter(user_id=new_user,status=1,module_id=right.module_id).first()
                            if get_rights_new !=None:
                                create='1' if(right.create == '1' or get_rights_new.create =='1')  else None
                                view='1' if(right.view == '1' or get_rights_new.view =='1')  else None
                                edit='1' if(right.edit == '1' or get_rights_new.edit =='1')  else None
                                delete='1' if(right.delete == '1' or get_rights_new.delete =='1')  else None
                                lock='1' if(right.lock == '1' or get_rights_new.lock =='1')  else None
                                unlock='1' if(right.unlock == '1' or get_rights_new.unlock =='1')  else None
                                UserRights.objects.filter(user_id=new_user,status=1,module_id=right.module_id).update(create=create,view=view,edit=edit,delete=delete,lock=lock,unlock=unlock)
                            else:
                                UserRights.objects.create(user_id=new_user,status=1,module_id=right.module_id,create=right.create,view=right.view,edit=right.edit,delete=right.delete,lock=right.lock,unlock=right.unlock)
                    ProjectUser.objects.filter(id=i.id).update(user_id=str(request.POST.get(f'user_id{i.id}',False)).replace(',',''))
                else:
                    ProjectUser.objects.filter(id=i.id).update(status=False)
        return redirect('projects:userlist')



class validatesignatory(View):
    def post(self,request,pk):
        pk=request.POST.get('pk')
        print(pk,'pk.pk')
        type_val=request.POST.get('changed_val')
        print(type_val,'type_val')
        Projectcreation.objects.filter(id=pk).update(signatory_type=type_val)
        if type_val=='1' or type_val=='':
            SignatoriesSettings.objects.filter(project_id=pk).update(status=False)
        return JsonResponse({'status':True})


class editprojectsignatory(View):
    def get(self,request,pk):
        key_val = [{'id':'1','settings_type':'Default Signatories'},{'id':'2','settings_type':'Project Signatories'}]
        datas_type=Projectcreation.objects.get(id=pk)
        datas=datas_type.signatory_type
        signatory_with_invoice=SignatoriesSettings.objects.filter(project_id=datas_type.projectname_id,status=True,invoice_type='1')
        signatory_without_invoice=SignatoriesSettings.objects.filter(project_id=datas_type.projectname_id,status=True,invoice_type='2')
        company_signatory = SignatoriesSettings.objects.get_by_company(request.company,True,'1').count()
        with_invoice = SignatoriesSettings.objects.get_by_company_type(request.company,True,'1','1')
        without_invoice = SignatoriesSettings.objects.get_by_company_type(request.company,True,'2','1')
        all_users = User.objects.filter(company=request.company,status=1,roles_id__in=[2,3]).exclude(id=request.user.id)
        get_settings = Settings.objects.get_company(request.company).values_list('currency',flat=True).first()
        currency = Basecountries.objects.get_by_id(literal_eval(get_settings))
        project_user = ProjectUser.objects.filter(company=request.company,project_id=pk)
        return render(request,'editprojectsignatory.html',{'datas':datas,'users':project_user.filter(status=True).values_list('user_id',flat=True),'pk':pk,'all_users':all_users,'multiuser':project_user.filter(status=True),'key_val':key_val,'with_invoice':with_invoice,'without_invoice':without_invoice,'signatory_count2':company_signatory,'signatory_with_invoice':signatory_with_invoice,'currency':currency,'datas_type':datas_type,'signatory_without_invoice':signatory_without_invoice})
    
    def post(self,request,pk):
        print('request',request.POST)
        # datas=Projectcreation.objects.get(id=pk)
        # datas=datas.signatory_type
        type_check=request.POST.get('sign_settings')
        print('type_check',type_check)
        update_value=Projectcreation.objects.get(id=pk)
        update_value.signatory_type=type_check
        update_value.save()
        project_users = request.POST.getlist('project_user',False)
        if project_users:
            for i in project_users:
                project_user, created = ProjectUser.objects.update_or_create(
                        project_id=pk,
                        user_id=i,
                        defaults={'company': request.company,'status':True},
                        )
                UserRights.objects.filter(user_id=i,module__module_name='Projects').update(view=1)
            revmove_status = ProjectUser.objects.filter(project_id=pk).exclude(user_id__in=project_users).exclude(user_id=request.user.id)#
            if revmove_status.exists():
                for i in revmove_status:
                    if ProjectUser.objects.filter(user_id=i.user_id,status=True).count() == 1:
                        exists=UserRights.objects.filter(Q(create__isnull=True) |Q(edit__isnull=True),user_id=i.user_id,module__module_name='Projects')
                        if exists.exists():
                            exists.update(view=None)
                revmove_status.update(status=False)
            else:
                remove_all = ProjectUser.objects.filter(project_id=pk).exclude(user_id=request.user.id)
                if remove_all.exists():
                    for i in remove_all:
                        if ProjectUser.objects.filter(user_id=i.user_id,status=True).count() == 1:
                                exists=UserRights.objects.filter(Q(create__isnull=True) |Q(edit__isnull=True),user_id=i.user_id,module__module_name='Projects')
                                if exists.exists():
                                    exists.update(view=None)
                    remove_all.update(status=False)

        print("request.POST.getlist('signatory_id')",request.POST.getlist('signatory_id'))
        # if datas=='2':
        #     print(1)
        #     Projectcreation.objects.filter(id=pk).update(signatory_type=2)
            # for i in request.POST.getlist('signatory_id'):
            #     signatory, created = SignatoriesSettings.objects.update_or_create(
            #     id=i, project_id=pk,
            #     defaults={
            #         'min_amount': None if request.POST.get(f'min_amount{i}') == '' else request.POST.get(f'min_amount{i}').replace(',',''),
            #         'max_amount': None if request.POST.get(f'max_amount{i}') == '' else request.POST.get(f'max_amount{i}').replace(',',''),
            #         'currency': None if request.POST.get(f'currency{i}') == '' else Basecountries.objects.get(id=request.POST.get(f'currency{i}')),
            #         'status':True,
            #         }
            #     )
            #     if created:
            #         signatory.company = request.company
            #         signatory.save()
            #     signatory_user_exists = SignatoriesUsers.objects.get_by_signatory(signatory)
            #     if signatory_user_exists.exists():
            #         for k in signatory_user_exists:
            #             update_signatory_user = SignatoriesUsers.objects.get_by_id(k.id)
            #             update_signatory_user.user = User.objects.get_by_id(request.POST.get(f'user{k.id}').replace(',',''))
            #             update_signatory_user.save()
            #     new_user = request.POST.getlist(f'newuser{i}',False)
            #     print('new_user',new_user)
            #     if new_user:
            #         print(new_user,'new_user')
            #         for new in new_user:
            #             print('SignatoriesSettings.objects.get_by_id(i)',SignatoriesSettings.objects.get_by_id(i))
            #             SignatoriesUsers.objects.create(signatory=SignatoriesSettings.objects.get_by_id(i),user=User.objects.get_by_id(new.replace(',',''))) if new!='' else False
            # new_tr = request.POST.getlist('newtr',False)
            # print('new_tr',new_tr)
            # if new_tr:
            #     for i in range(len(new_tr)):
            #         print('i',len(new_tr))
            #         currency_val=request.POST.get(f'new_currency{i+1}')
            #         print(currency_val,'req')
            #         signatory_create = SignatoriesSettings.objects.create(
            #             currency=Basecountries.objects.get(id=int(currency_val)),
            #             min_amount=None if request.POST.get(f'new_min_amount{i+1}') == '' else request.POST.get(f'new_min_amount{i+1}').replace(',',''),
            #             max_amount=None if request.POST.get(f'new_max_amount{i+1}') == '' else request.POST.get(f'new_max_amount{i+1}').replace(',',''),
            #             company = request.company,
            #             invoice_type = None if request.POST.get(f'new_invoice_type{i+1}') == '' else request.POST.get(f'new_invoice_type{i+1}'),
            #             signatory_type='2',
            #             project_id=pk,
            #         )
            #         new_user = request.POST.getlist(f'new_newuser{i+1}',False)
            #         print('new_user',new_user)
            #         if new_user:
            #             for new in new_user:
            #                 SignatoriesUsers.objects.create(signatory=signatory_create,user=User.objects.get_by_id(new.replace(',',''))) if new!='' else False
        # if datas == '1':
        #     print(2)
        #     Projectcreation.objects.filter(id=pk).update(signatory_type=1)
        return redirect('projects:projectlist')

class usercheck(View):
    def post(self,request,pk):
        users_in=request.POST.getlist('users_in[]')
        empty_li=[]
        for user in users_in:
            empty_li.append(int(user))
        company_users=User.objects.filter(id__in=empty_li,company_id=request.company,roles_id=3,status=1)
        if company_users.count() != 0:
            user_data=list(User.objects.filter(id__in=empty_li,company_id=request.company,roles_id=3,status=1).values_list('id',flat=True))
            pro_user=ProjectUser.objects.filter(project_id=pk).exclude(user_id__in=user_data).update(status=False)
            for i in company_users:
                if ProjectUser.objects.filter(project_id=pk,user_id=i.id).exists():
                    ProjectUser.objects.filter(project_id=pk,user_id=i.id).update(status=True)
                else:
                    ProjectUser.objects.create(project_id=pk,user_id=i.id,status=True,company=request.company)
            return JsonResponse({'status':True})    
        else:
            return JsonResponse({'status':False})

@login_required(login_url='/')
def ProjectWcc(request):
    markas_read_status(request.get_full_path())
    data={}
    request.session['mainmenu'] = 'main_project'
    request.session['submenu'] = 'wcc_flow'
    if (request.user.roles_id == 3):
        userrights=UserRights.objects.get(user_id=request.user.id,module_id=13)
        data['rights']=userrights
        assigned_projects = ProjectUser.objects.filter(user=request.user,status=True).values_list('project_id',flat=True)
        # below comments  is said ,by client maheswari for reason wcc project flow should be shown to all
        
        # get_countries=Projectcreation.objects.filter(company=request.company,id__in=assigned_projects).values('country').annotate(dcount=Count('country'))
        # projects=Projectcreation.objects.filter(company=request.company,id__in=assigned_projects).order_by('-id')
        get_countries=Projectcreation.objects.filter(company=request.company).values('country').annotate(dcount=Count('country'))
        projects=Projectcreation.objects.filter(company=request.company).order_by('-id')
    else:
        get_countries=Projectcreation.objects.filter(company=request.company).values('country').annotate(dcount=Count('country'))
        projects=Projectcreation.objects.filter(company=request.company).order_by('-id')
    print(f'projects {projects}')
    # contract_master extract project with wcc is 1 
    contract_master = list(ContractMaster.objects.filter(company=request.company,status=1,wcc=1).values_list('projects_id',flat=True).distinct())
    projects_count=projects.count()
    countries=[]
    country=''
    project_name=''
    project_creation=Projectcreation.objects.filter(company=request.company)
    projects_count=project_creation.count()
    project_country=[]
    project_id=''
    if (len(get_countries) > 0):
        for country in get_countries:
                country_id=country['country']
                # print("dsf",country)
                print(f'country {country_id}')
                try:
                    countr=Countries.objects.get(id=country_id)
                    countries.append({"country_id":countr.id,"country_name":countr.nicename})
                except Countries.DoesNotExist:
                    print()
    # print(countries)
    if request.method=="POST":
        print(request.POST)
        countval=request.POST['country_name']
        if (countval == ' '):
            country=''
        else:
            country=int(countval)
        project_name=request.POST['project_name']
        projects=Projectcreation.objects.filter(company=request.company)
        print("s")

        if(country):
            project_country=Projectcreation.objects.filter(country_id=country,company=request.company)
            projects=projects.filter(country_id=country)
        else:
            projects=projects

        if(project_name!=' '):
            project_id=int(project_name)
            projects=projects.filter(id=project_name)
        else:
            project_id=project_name
    # projects = projects.filter(id__in=contract_master)
    data.update({
    'projects':projects,
    'projects_count':projects_count,
    'get_countries':countries,
    'post_country':country,
    'post_project':project_id,
    'project_country':project_country,
    # 'projectcreation_status':projectcreation_status,
    })
    # print(data)
    return render(request,"projectwcc.html",data)


@login_required(login_url='/')
def ProjectInvoiceFlow(request):
    markas_read_status(request.get_full_path())
    data={}
    request.session['mainmenu'] = 'main_project'
    request.session['submenu'] = 'invoice_flow'
    # if (request.user.roles_id == 3):
    #     userrights=UserRights.objects.get(user_id=request.user.id,module_id=1)
    #     # print("wew",userrights)
    #     data['rights']=userrights
    #     assigned_projects = ProjectUser.objects.filter(user=request.user,status=True).values_list('project_id',flat=True)
    #     get_countries=Projectcreation.objects.filter(company=request.company,id__in=assigned_projects).values('country').annotate(dcount=Count('country'))
    #     projects=Projectcreation.objects.filter(company=request.company,id__in=assigned_projects).order_by('-id')
    # else:
    get_countries=Projectcreation.objects.filter(company=request.company).values('country').annotate(dcount=Count('country'))
    projects=Projectcreation.objects.filter(company=request.company).order_by('-id')
      # contract_master extract project with wcc is 0
    contract_master = list(ContractMaster.objects.filter(company=request.company,status=1).filter(Q(wcc__isnull=True) | Q(wcc=0)).values_list('projects_id',flat=True).distinct())
    projects_count=projects.count()
    # print("aaa",len(get_countries))
    countries=[]
    country=''
    project_name=''
    project_creation=Projectcreation.objects.filter(company=request.company)
    projects_count=project_creation.count()
    project_country=[]
    project_id=''
    if (len(get_countries) > 0):
        for country in get_countries:
                country_id=country['country']
                # print("dsf",country)
                print(f'country {country_id}')
                try:
                    countr=Countries.objects.get(id=country_id)
                    countries.append({"country_id":countr.id,"country_name":countr.nicename})
                except Countries.DoesNotExist:
                    print()
    # print(countries)
    if request.method=="POST":
        print(request.POST)
        countval=request.POST['country_name']
        if (countval == ' '):
            country=''
        else:
            country=int(countval)
        project_name=request.POST['project_name']
        projects=Projectcreation.objects.filter(company=request.company)
        print("s")

        if(country):
            project_country=Projectcreation.objects.filter(country_id=country,company=request.company)
            projects=projects.filter(country_id=country)
        else:
            projects=projects

        if(project_name!=' '):
            project_id=int(project_name)
            projects=projects.filter(id=project_name)
        else:
            project_id=project_name
    # projects = projects.filter(id__in=contract_master)
    data.update({
    'projects':projects,
    'projects_count':projects_count,
    'get_countries':countries,
    'post_country':country,
    'post_project':project_id,
    'project_country':project_country,
    # 'projectcreation_status':projectcreation_status,
    })
    # print(data)
    return render(request,"projectinvoiceflow.html",data)

class getRights(View):
    def post(self,request):
        pk = request.POST.get('role_id')
        # getrightsid=RoleRight.objects.filter(role_id=pk,status=1)
        right_ids = Right.objects.filter(id__in=RoleRight.objects.filter(role_id=pk,status=1).values_list('right', flat=True))
        print('pk',pk)
        print(list(right_ids),'right_ids')
        rights_data = serialize('json', right_ids) 
        return JsonResponse({'datas': rights_data}, safe=False)
        
class checkProjectStaus(View):
    def get(self, request):
        project_id=request.GET.get('project_id')
        get_project=Projectcreation.objects.get(id=project_id)
        if(get_project.active_status == True):
            Projectcreation.objects.filter(id=project_id).update(active_status = False)
        else:
            Projectcreation.objects.filter(id=project_id).update(active_status = True)
        return JsonResponse({'status':True})

class projectnameduplications(View):
    def get(self, request):
        projectname=(request.GET.get('projectname')).replace(" ", "")
        if (Projects.objects.filter(company=request.company,status=False,name=projectname)).exists():
            return JsonResponse({'status':True})
        else:
            return JsonResponse({'status':False})

class EditContract(View):
    def get(self,request,contracttype,pk):
        company=Settings.objects.filter(company_id=request.company.id).first()
        getcurrencylist=ast.literal_eval(company.currency)
        companycurrency=Basecountries.objects.filter(id__in=getcurrencylist)
        discipline=['green_field_development','brown_field_development','others']
        disciplinetype=['subsurface_and_reservoir_engineering','drilling_and_completions','facilities_and_projects','production_and_operations','human_resources_and_administration','legal','ICT','finance','QHSE_&_community_development','business_development_&_sales_and_marketing','insurance','miscellaneous']
        types=['Amendment','Addendum']
        data={'company':company,
            'companycurrency':companycurrency,
            'projectdiscipline':discipline,
            'disciplinetype':disciplinetype,
            'types':types,
            'contracttype':contracttype,
        }
    #     company=models.ForeignKey("custom_auth.Companies",on_delete=models.CASCADE,blank=True, null=True)
    # vendor=models.ForeignKey("ContractMasterVendor",on_delete=models.CASCADE,blank=True, null=True)
    # contract=models.ForeignKey("ContractMaster",on_delete=models.CASCADE,blank=True, null=True)
    # amendment_addendum=models.ForeignKey("Amendment",on_delete=models.CASCADE,blank=True, null=True)
    # file_name=models.FileField(blank=True,null=True)
    # #file type=1 if contract file or file type=2 if price table
    # file_type=models.IntegerField(blank=True, null=True,default=0)
    # original_file_name=models.CharField(max_length=255,blank=True,null=True)
        if (contracttype == "original"):
            contractmaster=ContractMaster.objects.get(id=pk)
            projectdevelopment=ProjectDevelopmentDiscipline.objects.filter(project_id=contractmaster.projects_id,status=1)
            developmentlist=[]
            for development in projectdevelopment:
                convert_title=development.development_type.cluster.environment.field.field.field_name.title()
                convert_cluster_title=development.development_type.cluster.clustersubname.cluster_subname.title()
                # print('q',convert_cluster_title)
                if (development.project_discipline == '1'):
                    developmentlist.append({'id':development.id,'fieldname':'Green Field Development-'+convert_title+'-'+convert_cluster_title})
                elif (development.project_discipline == '2'):
                    developmentlist.append({'id':development.id,'fieldname':'Brown Field Development-'+convert_title+'-'+convert_cluster_title})
                elif (development.project_discipline == '3'):
                    developmentlist.append({'id':development.id,'fieldname':'Others-'+convert_title+'-'+convert_cluster_title})
            projectdisciplinesubtype=ProjectDevelopmentSubType.objects.filter(project_discipline_id=contractmaster.projectdiscipline_id,status=1)
            projects=Projectcreation.objects.filter(company=request.company,status=0)
            contract_files=VendorContractPriceTable.objects.filter(contract_id=pk,file_type=1,status=1)
            price_files=VendorContractPriceTable.objects.filter(contract_id=pk,file_type=2,status=1)
            formatted_date = contractmaster.executed_date.strftime('%Y-%m-%d')
            data['projects']=projects
            data['contractmaster']=contractmaster
            data['contract_files']=contract_files
            data['developmentlist']=developmentlist
            data['projectdisciplinesubtype']=projectdisciplinesubtype
            print(f'price_files {price_files} , contract_files {contract_files}')
            data['price_files']=price_files
            data['contract_date']=formatted_date
        else:
            ammendment=Amendment.objects.filter(id=pk).first()
            contract_files=VendorContractPriceTable.objects.filter(vendor_id=ammendment.contractvendor_id,amendment_addendum_id=pk,file_type=1,status=1)
            price_files=VendorContractPriceTable.objects.filter(vendor_id=ammendment.contractvendor_id,amendment_addendum_id=pk,file_type=2,status=1)
            previous_pricefiles=list(VendorContractPriceTable.objects.filter(amendment_addendum_id=pk,status=1,file_type=2).values_list('id',flat=True))
            contract_ids=ContractMaster.objects.filter(id=ammendment.service.id,status=1).first()
            formatted_date = contract_ids.executed_date.strftime('%Y-%m-%d')
            data['ammendment']=ammendment
            data['contract_files']=contract_files
            data['price_files']=price_files
            data['contract_date']=formatted_date
            
            print(f'ammendment {contract_files}, price_files {price_files}')
        return render(request,'editcontract.html',data)
    
    def post(self,request,contracttype,pk):
        company=Settings.objects.filter(company_id=request.company.id).first()
        url = reverse('projects:contract_detailed_view', kwargs={'contract_id': pk})
        contract_value_id=''
        if (contracttype == "original"):
            print("yes came")
            types_service=request.POST.get('types_service') or None
            nameservice=request.POST.get('name_service') or None
            projects=request.POST.get('project') or None
            discipline=request.POST.get('project_discipline') or None
            disciplinetype1=request.POST.get('disciplinetype') or None
            referencenumber=request.POST.get('refnumber') or None
            execute_date=request.POST.get('execute_date') or None
            currency=request.POST.get('currency') or None
            amount=request.POST.get('amount') or None
            wcc=request.POST.get('wcc') or None
            checkcontractfile=request.FILES.getlist('con_files') or []
            price_table=request.FILES.getlist('price_files') or []
            submit_type=request.POST.get('submit_type') or None
            save_type=1
           # if disciplinetype1:
            #     try:
            #         disciplinetype = ProjectDevelopmentDiscipline.objects.get(project_discipline=disciplinetype1)
            #     except ProjectDevelopmentDiscipline.DoesNotExist:
            #         disciplinetype = None
            # else:
            #     disciplinetype = None
            disciplinetype = disciplinetype1
            if submit_type == '1' or submit_type == 1:
                save_type=2
            if execute_date:
                convert_date= getinvoiceDate(company.dateformat,execute_date)
            else:
                convert_date=None
            contract_master=ContractMaster.objects.filter(id=pk).first()
            contract_value_id=ContractMaster.objects.filter(id=pk).first()
            contract_file_ids=request.POST.getlist('contract_file_ids') or []
            price_file_ids=request.POST.getlist('price_file_ids') or []
            VendorContractPriceTable.objects.filter(contract_id=pk,file_type=1,status=1).exclude(id__in=contract_file_ids).update(status=0)
            VendorContractPriceTable.objects.filter(contract_id=pk,file_type=2,status=1).exclude(id__in=price_file_ids).update(status=0)
            for i in contract_file_ids:
                specific_file=request.FILES.get('confiles'+str(i),None)
                if specific_file:
                    VendorContractPriceTable.objects.filter(id=i).update(file_name=specific_file,original_file_name=specific_file.name)
            for i in price_file_ids:
                specific_file=request.FILES.get('pricefiles'+str(i),None)
                if specific_file:
                    VendorContractPriceTable.objects.filter(id=i).update(file_name=specific_file,original_file_name=specific_file.name)
            for contrct_file in checkcontractfile:
                VendorContractPriceTable.objects.create(company_id=request.company.id,vendor_id=contract_master.contractvendor_id,contract_id=pk,file_name=contrct_file,file_type=1,original_file_name=contrct_file.name)
            for price_file in price_table:
                VendorContractPriceTable.objects.create(company_id=request.company.id,vendor_id=contract_master.contractvendor_id,contract_id=pk,file_name=price_file,file_type=2,original_file_name=price_file.name)
            ContractMaster.objects.filter(id=pk).update(types_service=types_service,name_service=nameservice,projects_id=projects,projectdiscipline_id=discipline,projectdisciplinetype_id=disciplinetype,reference_number=referencenumber,executed_date=convert_date,currency_id=currency,amount=amount,wcc=wcc,save_type=save_type)
            contract_ids=request.POST.getlist('contract_id')
            ContractMaster.objects.filter(id__in=contract_ids).update(list_items=False,status=1)
            new_con_notfi=ContractMaster.objects.filter(id__in=contract_ids, save_type=2 ,status=1)
            for new_con_not in new_con_notfi:
                sender=request.user
                vinnum = ContractMasterVendor.objects.filter(id=new_con_not.contractvendor.id).values_list('vin',flat=True).first()
                recipient=User.objects.filter(cin_number=vinnum)
                print("recipient",recipient)
                data='projects/contractlist'
                verb='New Contract Added'
                description=f'New Contract  {new_con_not.reference_number} has been added by {sender.name} {sender.lastname}'
                notify_invoice_flow(request,recipient,data,verb,description)
               
        else:
            ammendment_type=request.POST.get('ammendment_type') or None
            ref_num=request.POST.get('ref_num') or None
            execute_date=request.POST.get('dateformat') or None
            currency=request.POST.get('currency') or None
            amount=request.POST.get('amount') or None
            wcc=request.POST.get('wcc') or None
      
            checkcontractfile=request.FILES.getlist('con_files') or []
            price_table=request.FILES.getlist('price_files') or []
            submit_type=request.POST.get('submit_type') or None
            save_type=1
            if submit_type == '1' or submit_type == 1:
                print("wwww",submit_type)
                amendment_id=Amendment.objects.filter(id=pk).first()
                # amendment_data=Amendment.objects.create(amendment_type=ammendment_type,amendment_reference_number=ref_num,amendment_executed_date=convert_date,amendment_amount=amount,amendment_currency_id=currency,save_type=save_type,wcc=wcc,service_id=pk,company_id=request.company.id,contractvendor_id=contractmaster.contractvendor.id,updated_by_id=request.user.id,reference_type=master_name,reference_id=master_id)
                # Get vendor ID from contract master
                vendor_id = amendment_id.contractvendor_id
                contract_value_id=ContractMaster.objects.filter(id=amendment_id.service.id,status=1).first()
                # Get VIN number associated with the vendor
                vinnum = ContractMasterVendor.objects.filter(id=vendor_id).values_list('vin', flat=True).first()
                print("VIN Number:", vinnum)  # Debug print statement

                # Get recipients based on the VIN number
                recipients = User.objects.filter(cin_number=vinnum)
                print("Recipients:", recipients)  # Debug print statement  
                # verb = 'New Amendment Added'
                verb = f'New {ammendment_type} Added'

                try:
                    scheme=request.scheme
                    gethost=request.get_host()
                    url1=f"{scheme}://{gethost}/projects/contractlist"
                    sender = request.user
                    # Ensure recipients are found before sending notifications
                    if recipients.exists():
                        for recipient in recipients:
                            # description = f'New Amendment {amendment_id.amendment_reference_number} has been added by {sender.name} {sender.lastname}.'
                            description = f'New {ammendment_type} {amendment_id.amendment_reference_number} has been added by {sender.name} {sender.lastname}.'
                            notify.send(sender, recipient=recipient, data=url1, verb=verb, description=description)
                    else:
                        print("No recipients found for VIN:", vinnum)
                except Exception as e:
                    # Handle the exception gracefully
                    print(f"An error occurred while sending notifications: {e}")
                save_type=2
            try:
                convert_date= getinvoiceDate(company.dateformat,execute_date)
            except:
                convert_date=None
            amendment_id=Amendment.objects.filter(id=pk).first()
            Amendment.objects.filter(id=pk).update(amendment_type=ammendment_type,amendment_reference_number=ref_num,amendment_executed_date=convert_date,amendment_amount=amount,amendment_currency_id=currency,save_type=save_type,wcc=wcc)
            contract_file_ids=request.POST.getlist('contract_file_ids') or []
            price_file_ids=request.POST.getlist('price_file_ids') or []
            VendorContractPriceTable.objects.filter(amendment_addendum_id=pk,file_type=1,status=1).exclude(id__in=contract_file_ids).update(status=0)
            VendorContractPriceTable.objects.filter(amendment_addendum_id=pk,file_type=2,status=1).exclude(id__in=price_file_ids).update(status=0)
            for i in contract_file_ids:
                specific_file=request.FILES.get('confiles'+str(i),None)
                if specific_file:
                    VendorContractPriceTable.objects.filter(id=i).update(file_name=specific_file,original_file_name=specific_file.name)
            for i in price_file_ids:
                specific_file=request.FILES.get('pricefiles'+str(i),None)
                if specific_file:
                    VendorContractPriceTable.objects.filter(id=i).update(file_name=specific_file,original_file_name=specific_file.name)
            for contrct_file in checkcontractfile:
                VendorContractPriceTable.objects.create(company_id=request.company.id,vendor_id=amendment_id.contractvendor_id,amendment_addendum_id=pk,file_name=contrct_file,file_type=1,original_file_name=contrct_file.name)
            for price_file in price_table:
                VendorContractPriceTable.objects.create(company_id=request.company.id,vendor_id=amendment_id.contractvendor_id,amendment_addendum_id=pk,file_name=price_file,file_type=2,original_file_name=price_file.name)
        submit_type=request.POST.get('submit_type') or None
        contractmaster = ContractMaster.objects.filter(id=pk).first()
        if submit_type == '1' or submit_type == 1:
            return redirect(reverse('projects:contract_detailed_view', kwargs={'contract_id': contract_value_id.id}))
        return redirect(reverse('projects:contractmasterlist'))


class AddAmmendment(View):
    def get(self,request,pk):
        company=Settings.objects.filter(company_id=request.company.id).first()
        contractmaster=ContractMaster.objects.filter(id=pk).first()
        getcurrencylist=ast.literal_eval(company.currency)
        companycurrency=Basecountries.objects.filter(id__in=getcurrencylist)
        contract=ContractMaster.objects.filter(id=pk,status=1).first()
        formatted_date = contract.executed_date.strftime('%Y-%m-%d')
        types=['Amendment','Addendum']
        data={'company':company,
            'companycurrency':companycurrency,
            'types':types,
            'contractmaster':contractmaster,'pk':pk,'contract_date':formatted_date
        }
        return render(request,'createcontract.html',data)
    
    def post(self,request,pk):
        print(f'request.POST {request.POST}')
        contract_id=pk
        contractmaster=ContractMaster.objects.filter(id=pk).first()
        company=Settings.objects.filter(company_id=request.company.id).first()
        ammendment_type=request.POST.get('ammendment_type') or None
        ref_num=request.POST.get('ref_num') or None
        execute_date=request.POST.get('dateformat') or None
        currency=request.POST.get('currency') or None
        amount=request.POST.get('amount') or None
        wcc=request.POST.get('wcc') or None
        checkcontractfile=request.FILES.getlist('con_file') or []
        price_table=request.FILES.getlist('price_file') or []
        submit_type=request.POST.get('submit_type') or None
        save_type=1
        if submit_type == '1':
            save_type=2
        try:
            convert_date= getinvoiceDate(company.dateformat,execute_date)
        except:
            convert_date=None
        master_id=pk
        master_name='original'
        addendum_wcc=contractmaster.wcc
        check_amendment=Amendment.objects.filter(service_id=pk,amendment_type='Amendment',status=1).last()
        if check_amendment:
            master_id=check_amendment.id
            master_name='amendment'
            addendum_wcc=check_amendment.wcc
        # contract_file=checkcontractfile
        # if checkcontractfile:
        #     contract_file=checkcontractfile.name
        # price_file=price_table
        # if price_table:
        #     price_file=price_table.name
        amendment_data=Amendment.objects.create(amendment_type=ammendment_type,amendment_reference_number=ref_num,amendment_executed_date=convert_date,amendment_amount=amount,amendment_currency_id=currency,save_type=save_type,wcc=wcc,service_id=pk,company_id=request.company.id,contractvendor_id=contractmaster.contractvendor.id,updated_by_id=request.user.id,reference_type=master_name,reference_id=master_id)
        # if amendment_data.amendment_type == 'Addendum':
        #     amendment_data.wcc= addendum_wcc
        #     amendment_data.save()
        if(len(checkcontractfile) > 0):
            for contract_file in checkcontractfile:
                if contract_file:
                    VendorContractPriceTable.objects.create(company_id=request.company.id,vendor_id=contractmaster.contractvendor.id,amendment_addendum_id=amendment_data.id,file_name=contract_file,original_file_name=contract_file.name,file_type=1)
        if(len(price_table) > 0):
            for price_file in price_table:
                if price_file:
                    VendorContractPriceTable.objects.create(company_id=request.company.id,vendor_id=contractmaster.contractvendor.id,amendment_addendum_id=amendment_data.id,file_name=price_file,original_file_name=price_file.name,file_type=2)
        url = reverse('projects:contract_detailed_view', kwargs={'contract_id': pk})
        if submit_type == '1':
            vendor_id=contractmaster.contractvendor.id
            # url = reverse('projects:contract_detailed_view', kwargs={'contract_id': pk})
            # Get VIN number associated with the vendor
            vinnum = ContractMasterVendor.objects.filter(id=vendor_id).values_list('vin', flat=True).first()
            print("VIN Number:", vinnum)  # Debug print statement

            # Get recipients based on the VIN number
            recipients = User.objects.filter(cin_number=vinnum)
            print("Recipients:", recipients)  # Debug print statement

            # verb = 'New Amendment Added'
            verb = f'New {ammendment_type} Added'
            

            try:
                sender = request.user
                scheme=request.scheme
                gethost=request.get_host()
                #url1=f"{scheme}://{gethost}/dashboard/dashboard"
                url1=f"{scheme}://{gethost}/projects/contractlist"

                # url1="dashboard:dashboard"
                
                # Ensure recipients are found before sending notifications
                if recipients.exists():

                    for recipient in recipients:
                        # description = f'New Amendment {amendment_data.amendment_reference_number} has been added by {sender.name} {sender.lastname}.'
                        description = f'New {ammendment_type} {amendment_data.amendment_reference_number} has been added by {sender.name} {sender.lastname}.'
                        notify.send(sender, recipient=recipient, data=url1, verb=verb, description=description)
                else:
                    print("No recipients found for VIN:", vinnum)
            except Exception as e:
                # Handle the exception gracefully
                print(f"An error occurred while sending notifications: {e}")

        return redirect(url)
class CreateContractMaster(View):
    def get(self,request):
        company=Settings.objects.filter(company_id=request.company.id).first()
        if (company != None):
            getcurrencylist=ast.literal_eval(company.currency)
            companycurrency=Basecountries.objects.filter(id__in=getcurrencylist)
        else:
            companycurrency=None
        vendormasterlist=ContractMasterVendor.objects.filter(company_id=request.company.id,status=1)
        request.session['submenu']='contract_projects'
        reference_number = []
        projects=Projectcreation.objects.filter(company=request.company,status=0)
        contracts=ContractMaster.objects.filter(company_id=request.company.id,created_by_id=request.user.id,status=0,save_type__in=[1,2],list_items=True)
        vendor_id=0
        if contracts.count() > 0:
            try:
                vendor_id=contracts.last().contractvendor_id
            except:
                vendor_id=0
        data={'company':company,'companycurrency':companycurrency,'vendormasterlist':vendormasterlist,'projects':projects,'contracts':contracts,'contracts_count':contracts.count(),'vendor_id':vendor_id,'create':True}

        return render(request,'createnewcontractmaster.html',data)
    def post(self,request):
        contract_ids=request.POST.getlist('contract_ids')
        print(f'contract master {request.POST} contract_ids  {contract_ids}')
        ContractMaster.objects.filter(id__in=contract_ids).update(list_items=False,status=1)
        return redirect('projects:contractmasterlist')

class assignContractMasterToVendor(View):
    def post(self, request, *args, **kwargs):
        vendor_id = request.POST.get('vendorid')
        contract_id = request.POST.get('contract_id')
        company=Settings.objects.filter(company_id=request.company.id).first()
        getcurrencylist=ast.literal_eval(company.currency)
        companycurrency=Basecountries.objects.filter(id__in=getcurrencylist)
        projects=Projectcreation.objects.filter(company=request.company,status=0)
        print(f"vendor_id   {vendor_id}")
        
        if vendor_id:
            contractmasterview =  render_to_string('components/contractmasterview.html',{'vendor_data':vendor_id,'projects':projects,'companycurrency':companycurrency,'company':company},request)
            return JsonResponse({'contractmasterview':contractmasterview})
        elif contract_id:
            print(f"contract_id   {contract_id}")
            contract_master=ContractMaster.objects.filter(id=contract_id).first()
            projectdevelopment=ProjectDevelopmentDiscipline.objects.filter(project_id=contract_master.projects_id,status=1)
            developmentlist=[]
            for development in projectdevelopment:
                convert_title=development.development_type.cluster.environment.field.field.field_name.title()
                convert_cluster_title=development.development_type.cluster.clustersubname.cluster_subname.title()
                # print('q',convert_cluster_title)
                if (development.project_discipline == '1'):
                    developmentlist.append({'id':development.id,'fieldname':'Green Field Development-'+convert_title+'-'+convert_cluster_title})
                elif (development.project_discipline == '2'):
                    developmentlist.append({'id':development.id,'fieldname':'Brown Field Development-'+convert_title+'-'+convert_cluster_title})
                elif (development.project_discipline == '3'):
                    developmentlist.append({'id':development.id,'fieldname':'Others-'+convert_title+'-'+convert_cluster_title})
            projectdisciplinesubtype=ProjectDevelopmentSubType.objects.filter(project_discipline_id=contract_master.projectdiscipline_id,status=1)
            contract_files=VendorContractPriceTable.objects.filter(company_id=request.company.id,contract_id=contract_id,file_type=1,status=1)
            print(f'contract_files {contract_files.count()}')
            price_files=VendorContractPriceTable.objects.filter(company_id=request.company.id,contract_id=contract_id,file_type=2,status=1)
            # disciplinetypelist=[]
            # for disciplinetype in projectdisciplinesubtype:
            #     # disctype=disciplinetype.discipline_subtype.discipline_subtype
            #     # replace_word=disctype.replace(' ','_')
            #     disciplinetypelist.append({'value':disciplinetype.id,'text':disciplinetype.discipline_subtype.discipline_subtype})
            contractmasterview =  render_to_string('components/contractmasterview.html',{'projects':projects,'companycurrency':companycurrency,'edit':True,'contract_master':contract_master,'developmentlist':developmentlist,'projectdisciplinesubtype':projectdisciplinesubtype,'contract_files':contract_files,'price_files':price_files,'company':company},request)
            return JsonResponse({'contractmasterview':contractmasterview})

class ContractMasterCreation(View):
    def post(self, request,*args, **kwargs):
        contract_id=request.POST.get('contract_id') or None
        company=Settings.objects.filter(company_id=request.company.id).first()
        types_service=request.POST.get('types_service') or None
        nameservice=request.POST.get('name_service') or None
        projects=request.POST.get('project') or None
        discipline=request.POST.get('project_discipline') or None
        disciplinetype=request.POST.get('disciplinetype') or None
        referencenumber=request.POST.get('ref_number') or None
        execute_date=request.POST.get('execute_date') or None
        vendor_id=request.POST.get('vendor_id') or None
        convert_date=None
        if execute_date:
            convert_date= getinvoiceDate(company.dateformat,execute_date)
        currency=request.POST.get('currency') or None
        amount=request.POST.get('amount') or None
        wcc=request.POST.get('wcc') or None
        wcc = int(wcc) if wcc is not None and wcc != 'none' else None
        submit_type=request.POST.get('save_type') or None
        checkcontractfile=request.FILES.getlist('con_files') or []
        price_table=request.FILES.getlist('price_files') or []
        print(f'request.POST {request.POST}')
        save_type=1
        if submit_type == '2' or submit_type == 2:
            save_type=2
        if contract_id:
            contract_file_ids=request.POST.getlist('contract_file_ids') or []
            price_file_ids=request.POST.getlist('price_file_ids') or []
            VendorContractPriceTable.objects.filter(contract_id=contract_id,file_type=1,status=1).exclude(id__in=contract_file_ids).update(status=0)
            VendorContractPriceTable.objects.filter(contract_id=contract_id,file_type=2,status=1).exclude(id__in=price_file_ids).update(status=0)
            ContractMaster.objects.filter(id=contract_id).update(types_service=types_service,name_service=nameservice,projects_id=projects,projectdiscipline_id=discipline,projectdisciplinetype_id=disciplinetype,reference_number=referencenumber,executed_date=convert_date,currency_id=currency,amount=amount,wcc=wcc,save_type=save_type,status=0,list_items=True,updated_by_id=request.user.id)
            for i in contract_file_ids:
                specific_file=request.FILES.get('confiles'+str(i),None)
                if specific_file:
                    VendorContractPriceTable.objects.filter(id=i).update(file_name=specific_file,original_file_name=specific_file.name)
            for i in price_file_ids:
                specific_file=request.FILES.get('pricefiles'+str(i),None)
                if specific_file:
                    VendorContractPriceTable.objects.filter(id=i).update(file_name=specific_file,original_file_name=specific_file.name)
            for contrct_file in checkcontractfile:
                VendorContractPriceTable.objects.create(company_id=request.company.id,vendor_id=vendor_id,contract_id=contract_id,file_name=contrct_file,file_type=1,original_file_name=contrct_file.name)
            for price_file in price_table:
                VendorContractPriceTable.objects.create(company_id=request.company.id,vendor_id=vendor_id,contract_id=contract_id,file_name=price_file,file_type=2,original_file_name=price_file.name)

        else:
            print(f'vbnm {request.POST}, request.FILES {request.FILES} vendor_id') 
            
            print(f'price_table {price_table}, checkcontractfile {checkcontractfile}')
            if vendor_id:
                contract_id=ContractMaster.objects.create(types_service=types_service,name_service=nameservice,projects_id=projects,projectdiscipline_id=discipline,projectdisciplinetype_id=disciplinetype,reference_number=referencenumber,executed_date=convert_date,currency_id=currency,amount=amount,wcc=wcc,save_type=save_type,status=0,list_items=True,contractvendor_id=vendor_id,company_id=request.company.id,created_by_id=request.user.id,updated_by_id=request.user.id)
                print(f'contract_id {contract_id}, id {contract_id.id}')
                for contrct_file in checkcontractfile:
                    VendorContractPriceTable.objects.create(company_id=request.company.id,vendor_id=vendor_id,contract_id=contract_id.id,file_name=contrct_file,file_type=1,original_file_name=contrct_file.name)
                for price_file in price_table:
                    VendorContractPriceTable.objects.create(company_id=request.company.id,vendor_id=vendor_id,contract_id=contract_id.id,file_name=price_file,file_type=2,original_file_name=price_file.name)
        return JsonResponse({'status':True})


        

class DelContractMaster(View):
    def post(self, request, *args, **kwargs):
        print(f'request.POST {request.POST}, request {request}')
        contract_id = request.POST.get('contract_id')
        print(f'contract_id {contract_id}')
        ContractMaster.objects.filter(id=contract_id).update(status=0,list_items=False)
        return JsonResponse({'status':True})

class CreateVendorMasterContract(View):
    def get(self,request,pk):
        company=Settings.objects.filter(company_id=request.company.id).first()
        if (company != None):
            getcurrencylist=ast.literal_eval(company.currency)
            companycurrency=Basecountries.objects.filter(id__in=getcurrencylist)
        else:
            companycurrency=None
        vendormasterlist=ContractMasterVendor.objects.filter(company_id=request.company.id,status=1,id=pk).first()
        request.session['submenu']='contract_projects'
        reference_number = []
        projects=Projectcreation.objects.filter(company=request.company,status=0)
        contracts=ContractMaster.objects.filter(company_id=request.company.id,created_by_id=request.user.id,status=0,save_type__in=[1,2],list_items=True,contractvendor_id=pk)
        vendor_id=0
        edit=True

        if contracts.count() > 0:
            try:
                vendor_id=contracts.last().contractvendor_id
            except:
                vendor_id=0
        data={'company':company,'companycurrency':companycurrency,'vendormasterlist':vendormasterlist,'projects':projects,'contracts':contracts,'contracts_count':contracts.count(),'vendor_id':vendor_id,'edit':edit,'pk':pk}

        return render(request,'createnewcontractmaster.html',data)
    def post(self,request,pk):
        contract_ids=request.POST.getlist('contract_ids')
        contract_action=request.POST.get('contract_action')
        print(f'contract master {request.POST} contract_ids  {contract_ids}')
        ContractMaster.objects.filter(id__in=contract_ids).update(list_items=False,status=1)
        new_con_notfi=ContractMaster.objects.filter(id__in=contract_ids, save_type=2 ,status=1)
        for new_con_not in new_con_notfi:
            sender=request.user
            vinnum = ContractMasterVendor.objects.filter(id=new_con_not.contractvendor.id).values_list('vin',flat=True).first()
            recipient=User.objects.filter(cin_number=vinnum)
            data='projects/contractlist'
            verb='New Contract Added'
            description=f'New Contract  {new_con_not.reference_number} has been added by {sender.name} {sender.lastname}'
            notify_invoice_flow(request,recipient,data,verb,description)
            print("done submit")
        
        return redirect('projects:contractmasterlist')

from .templatetags.custom_tags import convert_projectdiscipline,checktime
def getContracterDataAjax(request):
    print(request)
    vendor_id=request.GET.get('selected_vendor')
    contracts=ContractMaster.objects.filter(company_id=request.company.id,status=0,save_type__in=[1,2],list_items=True,contractvendor_id=vendor_id)
    contract_data=[]
    for contract in contracts:
        contract_items={}
        contract_items['id']=contract.id
        contract_items['name_service']=contract.name_service if contract.name_service else ''
        try:
            contract_items['project_name']=contract.projects.projectname if contract.projects.projectname else ''
        except:
            contract_items['project_name']=''
        contract_items['projectdiscipline']=convert_projectdiscipline(contract.projectdiscipline)
        try:
            contract_items['discipline_subtype']=contract.projectdisciplinetype.discipline_subtype
        except:
            contract_items['discipline_subtype']=''
        contract_items['reference_number']=contract.reference_number if contract.reference_number else ''
        contract_items['executed_date']=checktime(contract.executed_date,request.company.id)
        try:
            contract_items['currency']=contract.currency.currency_symbol
        except:
            contract_items['currency']=''
        contract_data.append(contract_items)
        print(f'contract_data {contract_data}')
    return JsonResponse({'status':True,'contract_data':contract_data,'length':contracts.count()})


def get_relation_file(request):
    contract_type=request.GET.get('contract_type')
    id=request.GET.get('id')
    file_type=request.GET.get('file_type')
    file=get_contract_file(id,contract_type,file_type)
    files=file[0]
    data_files = {}
    count=0
    for data  in files:
        total_files = []
        file_types = {}
        file_types['id'] = data.id
        file_types['file_name'] = data.file_name.url
        file_types['original_file_name'] = data.original_file_name
        total_files.append(file_types)
        data_files[count] = total_files
        count+=1

    print(f'data_files: {data_files}')

    return JsonResponse({'status': True, 'data': data_files})



def get_contractrelation_file(request):
    contract_type=request.GET.get('contract_type')
    id=request.GET.get('id')
    file_type=request.GET.get('file_type')
    file=get_contract_file(id,contract_type,file_type)
    files=file[0]
    
    total_files = []
    for data  in files:
        file_types = {'id':data.id ,'file_name': data.file_name.url,'original_file_name': data.original_file_name}
        total_files.append(file_types)
    
    print(f'total_files: {total_files}')

    return JsonResponse({'status': True, 'data': total_files})
        
        
def checkreferencenumber(request):
    print(f'request.GET {request.GET}contract_id, ammendment_id')
    ref_num=request.GET.get('ref_num')
    contract_id=request.GET.get('contract_id') or None
    ammendment_id=request.GET.get('ammendment_id') or None
    ref_num_check=0
    if ContractMaster.objects.filter(company_id=request.company.id,reference_number__iexact=ref_num , status=1).exclude(id=contract_id).exists():
        ref_num_check=1
    elif Amendment.objects.filter(company_id=request.company.id,amendment_reference_number__iexact=ref_num, status=1).exclude(id=ammendment_id).exists():
        ref_num_check=1
    else:
        ref_num_check=0
    print(f' ref_num {ref_num}')
    return JsonResponse({'status': True, 'data': ref_num_check})
def delete_amendment(request):
    if request.method == 'POST':
        amendment_id = request.POST.get('amendment_id')
        try:
            amendment = Amendment.objects.get(pk=amendment_id)
            amendment.delete()
            return JsonResponse({'success': True})
        except Amendment.DoesNotExist:
            return JsonResponse({'success': False})
    else:
        return JsonResponse({'success': False})


def check_account_number(request) :
    try:
        bank_status = 0
        accountnumber=request.GET.get('account_no')
        bank_details=BankDetails.objects.filter(company_id=request.company.id ,accountnumber=accountnumber , status=1 ).count()
        
        if bank_details == 0:
            bank_acc_details=UserBankAccountno.objects.filter(userbank__company_id=request.company.id  ,accountno=accountnumber ,  userbank__status=1 ).count()
            if bank_acc_details == 0:
                bank_status = 0
            else :
                bank_status = 1
        else :
            bank_status=1

        return JsonResponse({'account_num':bank_status})

    except Exception as e:
        # Log the error (optional)
        print(f"Error checking account number: {e}")
        # You can set an error status or return a specific response
        bank_status = 0
        
        return JsonResponse({'account_num':bank_status})



    
   
        
    